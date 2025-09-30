# -*- mode: python ; coding: utf-8 -*-
from flask import Flask, request, jsonify, send_from_directory, Response
from flask_cors import CORS
from pymongo import MongoClient
from pymongo.errors import ConnectionFailure, ServerSelectionTimeoutError
from bson.objectid import ObjectId
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
import os
import sys
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from werkzeug.utils import secure_filename
import json
import secrets
import hashlib
import bcrypt
import socket
import uuid
import time
import threading
import traceback
from functools import wraps
import jwt
import requests
from io import BytesIO
# Optional dependencies
try:
    import openpyxl
except ImportError:
    openpyxl = None
    logging.warning("openpyxl library not found. Excel export/backup features will be disabled.")
try:
    import schedule
except ImportError:
    schedule = None
    logging.warning("schedule library not found. Automatic tasks will be disabled.")
# --- Configuration Management ---
def get_base_dir():
    """Determine the base directory, handling both development and frozen executable cases."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))
CONFIG_DIR = os.getenv('CONFIG_DIR', get_base_dir())
os.makedirs(CONFIG_DIR, exist_ok=True)
CONFIG_FILE_PATH = os.path.join(CONFIG_DIR, 'config.json')
BASE_DIR = get_base_dir()
STATIC_FOLDER_PATH = os.path.join(BASE_DIR, 'dist')
app = Flask(__name__, static_folder=STATIC_FOLDER_PATH, static_url_path='/')
CORS(app, resources={r"/api/*": {"origins": "*"}})
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(CONFIG_DIR, 'pos_app.log'))
    ]
)
logger = logging.getLogger(__name__)
JWT_SECRET = os.getenv('JWT_SECRET', secrets.token_hex(32))
JWT_ALGORITHM = 'HS256'
JWT_EXP_DELTA_SECONDS = 3600
def load_config():
    """Loads the configuration from config.json, creating a default if it doesn't exist."""
    try:
        if not os.path.exists(CONFIG_FILE_PATH):
            logger.warning(f"config.json not found at {CONFIG_FILE_PATH}. Creating default configuration.")
            default_config = {
                "mode": "server",
                "mongo_uri": "mongodb://localhost:27017/",
                "server_ip": "127.0.0.1"
            }
            save_config(default_config)
            return default_config
        with open(CONFIG_FILE_PATH, 'r') as f:
            logger.info(f"Loading configuration from {CONFIG_FILE_PATH}")
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError, PermissionError) as e:
        logger.error(f"Could not load or parse config.json: {e}. Using default server config.")
        return {"mode": "server", "mongo_uri": "mongodb://localhost:27017/", "server_ip": "127.0.0.1"}
def save_config(config_data):
    """Saves the configuration to config.json."""
    try:
        with open(CONFIG_FILE_PATH, 'w') as f:
            json.dump(config_data, f, indent=4)
        logger.info(f"Configuration saved to {CONFIG_FILE_PATH}")
    except PermissionError as e:
        logger.error(f"Permission denied when saving config.json to {CONFIG_FILE_PATH}: {e}")
        raise
    except Exception as e:
        logger.error(f"Error saving config.json: {e}")
        raise
config = load_config()
# Updated to lowercase 'uploads' for consistency
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', os.path.join(BASE_DIR, 'static', 'uploads'))
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
ALLOWED_JSON_EXTENSIONS = {'json'}
MAX_BACKUPS = 10
def create_directory(directory):
    """Ensure directory exists."""
    try:
        os.makedirs(directory, exist_ok=True)
        logger.info(f"Ensured directory exists: {directory}")
    except Exception as e:
        logger.error(f"Error creating directory {directory}: {e}")
create_directory(app.config['UPLOAD_FOLDER'])
client = None
db = None
items_collection = None
customers_collection = None
sales_collection = None
tables_collection = None
users_collection = None
settings_collection = None
email_tokens_collection = None
opening_collection = None
pos_closing_collection = None
kitchens_collection = None
item_groups_collection = None
kitchen_saved_collection = None
picked_up_collection = None
variants_collection = None
employees_collection = None
activeorders_collection = None
order_counters_collection = None
tripreports_collection = None
email_settings_collection = None
purchase_items_collection = None
suppliers_collection = None
purchase_orders_collection = None
purchase_receipts_collection = None
purchase_invoices_collection = None
def connect_to_mongodb():
    """Connect to MongoDB and initialize all collections."""
    global client, db, items_collection, customers_collection, sales_collection, tables_collection, users_collection, settings_collection, email_tokens_collection, opening_collection, pos_closing_collection, kitchens_collection, item_groups_collection, kitchen_saved_collection, picked_up_collection, variants_collection, employees_collection, activeorders_collection, order_counters_collection, tripreports_collection, email_settings_collection, purchase_items_collection, suppliers_collection, purchase_orders_collection, purchase_receipts_collection, purchase_invoices_collection
   
    mongo_uri = config.get("mongo_uri", "mongodb://localhost:27017/")
    mode = config.get("mode", "server")
   
    if mode == 'client':
        server_ip = config.get("server_ip")
        if not server_ip:
            logger.error("Server IP not provided in client mode")
            return False
        mongo_uri = f"mongodb://{server_ip}:27017/"
        logger.info(f"Client mode: Attempting to connect to MongoDB at {mongo_uri}")
    else:
        logger.info(f"Server mode: Attempting to connect to MongoDB at {mongo_uri}")
    try:
        client = MongoClient(mongo_uri, serverSelectionTimeoutMS=5000)
        client.admin.command('ping')
        logger.info(f"Successfully connected to MongoDB at {mongo_uri}")
        db = client['restaurant']
       
        # Initialize collections
        items_collection = db['items']
        customers_collection = db['customers']
        sales_collection = db['sales']
        tables_collection = db['tables']
        users_collection = db['users']
        settings_collection = db['system_settings']
        email_tokens_collection = db['email_tokens']
        opening_collection = db['pos_opening_entries']
        pos_closing_collection = db['pos_closing_entries']
        kitchens_collection = db['kitchens']
        item_groups_collection = db['item_groups']
        kitchen_saved_collection = db['kitchen_saved_orders']
        picked_up_collection = db['picked_up_items']
        variants_collection = db['variants']
        employees_collection = db['employees']
        activeorders_collection = db['active_orders']
        order_counters_collection = db['order_counters']
        tripreports_collection = db['trip_reports']
        email_settings_collection = db['email_settings']
        purchase_items_collection = db['purchase_items']
        suppliers_collection = db['suppliers']
        purchase_orders_collection = db['purchase_orders']
        purchase_receipts_collection = db['purchase_receipts']
        purchase_invoices_collection = db['purchase_invoices']
        ensure_test_users()
        return True
    except (ConnectionFailure, ServerSelectionTimeoutError) as e:
        logger.error(f"Failed to connect to MongoDB at {mongo_uri}: {e}")
        client = None
        db = None
        return False
    except Exception as e:
        logger.error(f"Unexpected error connecting to MongoDB: {e}")
        return False
def ensure_test_users():
    for test_user_template in TEST_USERS:
        email = test_user_template['email']
        user = users_collection.find_one({"email": email})
        new_hash = bcrypt.hashpw("123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        test_user = {**test_user_template, "password": new_hash}
        if user:
            # Fix missing 'firstName' if absent
            if 'firstName' not in user:
                users_collection.update_one({"email": email}, {"$set": {"firstName": test_user['firstName']}})
                logger.warning(f"Added missing 'firstName' for user {email}")
            try:
                # Test if valid hash
                bcrypt.checkpw(b"123", user['password'].encode('utf-8'))
            except ValueError as ve:
                # Invalid hash, rehash
                logger.warning(f"Invalid hash for user {email}, rehashing: {ve}")
                users_collection.update_one({"email": email}, {"$set": {"password": new_hash}})
            except Exception as e:
                logger.error(f"Error checking hash for {email}: {e}")
        else:
            users_collection.insert_one(test_user)
            logger.info(f"Inserted test user {email}")
def db_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if db is None:
            server_ip = config.get("server_ip", "unknown")
            mode = config.get("mode", "server")
            error_msg = (
                f"Cannot connect to MongoDB at {config.get('mongo_uri', 'unknown')}. "
                f"Please check the {'server IP' if mode == 'client' else 'MongoDB configuration'}."
            )
            logger.error(error_msg)
            return jsonify({"error": error_msg, "message": "Database not connected."}), 503
        return f(*args, **kwargs)
    return decorated_function
def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions
def handle_image_upload(file):
    if file and allowed_file(file.filename, ALLOWED_IMAGE_EXTENSIONS):
        ext = file.filename.rsplit('.', 1)[1].lower()
        unique_filename = f"{uuid.uuid4()}.{ext}"
        filename = secure_filename(unique_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        logger.info(f"Image saved: {filename}")
        return filename
    logger.warning(f"Invalid file for upload: {file.filename}")
    return None
def sanitize_image_fields(data):
    fields_to_sanitize = ['image', 'addon_image', 'combo_image', 'variant_image']
    for field in fields_to_sanitize:
        if field in data and isinstance(data[field], str):
            data[field] = os.path.basename(data[field])
    for addon in data.get("addons", []):
        if 'addon_image' in addon and isinstance(addon['addon_image'], str):
            addon['addon_image'] = os.path.basename(addon['addon_image'])
    for combo in data.get("combos", []):
        if 'combo_image' in combo and isinstance(combo['combo_image'], str):
            combo['combo_image'] = os.path.basename(combo['combo_image'])
    return data
def convert_objectid_to_str(item):
    if isinstance(item, list):
        return [convert_objectid_to_str(i) for i in item]
    if isinstance(item, dict):
        return {key: convert_objectid_to_str(value) for key, value in item.items()}
    if isinstance(item, ObjectId):
        return str(item)
    return item
def get_system_settings():
    if settings_collection is None:
        logger.warning("Settings collection not available, returning default settings")
        return {
            "_id": "system_settings",
            "disableUserPassLogin": False,
            "allowLoginUsingMobileNumber": True,
            "allowLoginUsingUsername": True,
            "loginWithEmailLink": False,
            "sessionExpiry": "06:00"
        }
    settings = settings_collection.find_one({"_id": "system_settings"})
    if not settings:
        default_settings = {
            "_id": "system_settings",
            "disableUserPassLogin": False,
            "allowLoginUsingMobileNumber": True,
            "allowLoginUsingUsername": True,
            "loginWithEmailLink": False,
            "sessionExpiry": "06:00"
        }
        settings_collection.insert_one(default_settings)
        logger.info("Inserted default system settings")
        return default_settings
    return settings
def save_system_settings(settings):
    settings["_id"] = "system_settings"
    settings_collection.replace_one({"_id": "system_settings"}, settings, upsert=True)
    logger.info("System settings saved")
TEST_USERS = [
    {
        "email": "admin@gmail.com",
        "password": bcrypt.hashpw("123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8'),
        "phone_number": "1234567890",
        "username": "admin",
        "role": "admin",
        "firstName": "Admin",
        "company": "POS 8",
        "pos_profile": "POS-001",
        "status": "Active",
        "created_at": datetime.now(ZoneInfo("UTC")).isoformat(),
        "is_test": True
    },
    {
        "email": "bearer@gmail.com",
        "password": bcrypt.hashpw("123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8'),
        "phone_number": "0987654321",
        "username": "bearer",
        "role": "bearer",
        "firstName": "Bearer",
        "company": "POS 8",
        "pos_profile": "POS-001",
        "status": "Active",
        "created_at": datetime.now(ZoneInfo("UTC")).isoformat(),
        "is_test": True
    }
]
# --- API Endpoints ---
@app.route('/api/test', methods=['GET'])
def test_endpoint():
    return jsonify({"status": "success", "message": "Server is running"}), 200
@app.route('/api/network_info', methods=['GET'])
def get_network_info():
    current_config = load_config()
    db_status = "Connected" if db is not None else "Disconnected"
    local_ip = "127.0.0.1"
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(0.1)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except Exception:
        try:
            hostname = socket.gethostname()
            local_ip = socket.gethostbyname(hostname)
        except Exception:
            logger.warning("Could not determine local IP")
    return jsonify({
        "local_ip": local_ip,
        "config": current_config,
        "database_status": db_status
    }), 200
@app.route('/api/configure', methods=['POST'])
def configure_app():
    global config
    try:
        data = request.get_json()
        mode = data.get('mode')
        server_ip = data.get('server_ip')
        mongo_uri = data.get('mongo_uri')
        if mode not in ['server', 'client']:
            logger.error(f"Invalid mode specified: {mode}")
            return jsonify({"error": "Invalid mode specified. Must be 'server' or 'client'"}), 400
        if mode == 'client' and not server_ip:
            logger.error("Server IP not provided in client mode")
            return jsonify({"error": "Server IP is required for client mode"}), 400
        new_config = {"mode": mode}
        if mode == 'client':
            new_config['server_ip'] = server_ip
            new_config['mongo_uri'] = f"mongodb://{server_ip}:27017/"
        else:
            new_config['mongo_uri'] = mongo_uri or "mongodb://localhost:27017/"
            new_config['server_ip'] = "127.0.0.1"
        save_config(new_config)
        config = new_config
        success = connect_to_mongodb()
        if not success:
            logger.error("Failed to connect to MongoDB after configuration change")
            return jsonify({"error": "Failed to connect to MongoDB with new configuration"}), 503
        logger.info(f"Application configured as {mode} with MongoDB URI: {new_config['mongo_uri']}")
        return jsonify({"message": "Configuration saved successfully. The application will now use the new settings."}), 200
    except Exception as e:
        logger.error(f"Configuration error: {e}\n{traceback.format_exc()}")
        return jsonify({"error": f"An internal server error occurred during configuration: {str(e)}"}), 500
@app.route('/api/login', methods=['POST'])
@db_required
def login():
    try:
        data = request.get_json()
        settings = get_system_settings()
        identifier = data.get('identifier')
        password = data.get('password')
        login_type = data.get('type', 'mobile_or_username')
        if not identifier or not password:
            logger.error("Identifier or password missing in login request")
            return jsonify({"error": "Identifier and password are required"}), 400
        user = None
        if login_type == 'mobile_or_username':
            query = {"$or": [
                {"phone_number": identifier},
                {"username": identifier},
                {"email": identifier}
            ]}
            user = users_collection.find_one(query)
        else:
            logger.error(f"Invalid login type: {login_type}")
            return jsonify({"error": "Invalid login type"}), 400
        if not user:
            for test_user in TEST_USERS:
                if test_user['email'] == identifier or test_user['phone_number'] == identifier or test_user['username'] == identifier:
                    if bcrypt.checkpw(password.encode('utf-8'), test_user['password'].encode('utf-8')):
                        existing_user = users_collection.find_one({"email": test_user['email']})
                        if not existing_user:
                            users_collection.insert_one(test_user)
                            existing_user = users_collection.find_one({"email": test_user['email']})
                        user = existing_user
                        break
        if not user:
            logger.warning(f"Invalid login attempt: {identifier}")
            return jsonify({"error": "Invalid credentials"}), 401
        try:
            if not bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):
                logger.warning(f"Invalid password for: {identifier}")
                return jsonify({"error": "Invalid credentials"}), 401
        except ValueError as ve:
            if "Invalid salt" in str(ve):
                logger.error(f"Invalid salt for user {user.get('email')}, rehashing")
                new_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                users_collection.update_one({"_id": user['_id']}, {"$set": {"password": new_hash}})
                user['password'] = new_hash
                if not bcrypt.checkpw(password.encode('utf-8'), new_hash.encode('utf-8')):
                    return jsonify({"error": "Invalid credentials"}), 401
            else:
                raise
        token_payload = {
            'user_id': str(user['_id']),
            'exp': datetime.now(timezone.utc) + timedelta(seconds=JWT_EXP_DELTA_SECONDS)
        }
        token = jwt.encode(token_payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
        user = convert_objectid_to_str(user)
        requires_opening_entry = opening_collection.find_one({
            "user_id": user['_id'],
            "date": {"$gte": datetime.now(ZoneInfo("UTC")).replace(hour=0, minute=0, second=0, microsecond=0)}
        }) is None
        response = {
            "message": "Login successful",
            "token": token,
            "user": {
                "id": user['_id'],
                "username": user.get('username', user.get('firstName', user.get('email', 'Unknown'))),
                "role": user.get('role', 'bearer'),
                "email": user.get('email', ''),
                "phone_number": user.get('phone_number', ''),
                "pos_profile": user.get('pos_profile', 'POS-001'),
                "company": user.get('company', 'POS 8'),
                "is_test": user.get('is_test', False)
            },
            "requires_opening_entry": requires_opening_entry
        }
        logger.info(f"User logged in: {identifier}")
        return jsonify(response), 200
    except KeyError as ke:
        logger.error(f"KeyError during login response building: {str(ke)}")
        return jsonify({"error": f"Internal error: Missing user field {str(ke)}"}), 500
    except Exception as e:
        logger.error(f"Login error: {e}\n{traceback.format_exc()}")
        return jsonify({"error": f"An internal server error occurred during login: {str(e)}"}), 500
@app.route('/api/register', methods=['POST'])
@db_required
def register():
    try:
        data = request.get_json()
        required_fields = ['email', 'password', 'firstName', 'phone_number']
        if not all(field in data for field in required_fields):
            return jsonify({"message": "Missing required fields"}), 400
        if users_collection.find_one({"email": data['email']}):
            return jsonify({"message": "Email already registered"}), 400
        hashed_password = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        new_user = {
            "email": data['email'],
            "password": hashed_password,
            "role": data.get('role', 'bearer'),
            "username": data.get('username', data['firstName']),
            "firstName": data['firstName'],
            "phone_number": data['phone_number'],
            "company": data.get('company', 'POS 8'),
            "pos_profile": data.get('pos_profile', 'POS-001'),
            "status": "Active",
            "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
        }
        result = users_collection.insert_one(new_user)
        logger.info(f"User registered: {data['email']}")
        return jsonify({"message": "Registration successful", "userId": str(result.inserted_id)}), 201
    except Exception as e:
        logger.error(f"Registration failed: {str(e)}")
        return jsonify({"message": f"Registration failed: {str(e)}"}), 500
@app.route('/api/users', methods=['GET'])
@db_required
def get_users():
    try:
        users = list(users_collection.find({}, {"password": 0}))
        return jsonify(convert_objectid_to_str(users)), 200
    except Exception as e:
        logger.error(f"Error fetching users: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/users/<email>', methods=['DELETE'])
@db_required
def delete_user(email):
    try:
        if email in [u['email'] for u in TEST_USERS]:
            return jsonify({"message": "Cannot delete test users"}), 400
        result = users_collection.delete_one({"email": email})
        if result.deleted_count == 0:
            return jsonify({"message": "User not found"}), 404
        logger.info(f"User deleted: {email}")
        return jsonify({"message": "User deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting user {email}: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/settings', methods=['GET'])
@db_required
def get_settings_route():
    try:
        settings = get_system_settings()
        return jsonify(convert_objectid_to_str(settings)), 200
    except Exception as e:
        logger.error(f"Error fetching settings: {e}")
        return jsonify({"error": "An internal server error occurred."}), 500
@app.route('/api/settings', methods=['POST'])
@db_required
def update_settings():
    try:
        data = request.get_json()
        save_system_settings(data)
        logger.info("System settings updated")
        return jsonify({"message": "Settings updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating settings: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/items', methods=['GET'])
@db_required
def get_items():
    try:
        items = list(items_collection.find())
        items_list = []
        current_time = datetime.now(ZoneInfo("UTC"))
        placeholder_url = 'https://placehold.co/100x100/EFEFEF/AAAAAA?text=No+Image'
       
        for item in items:
            item = convert_objectid_to_str(item)
            is_offer_active = False
            if 'offer_start_time' in item and item['offer_start_time'] and 'offer_end_time' in item and item['offer_end_time']:
                try:
                    offer_start_time = datetime.fromisoformat(str(item['offer_start_time']).replace('Z', '+00:00'))
                    offer_end_time = datetime.fromisoformat(str(item['offer_end_time']).replace('Z', '+00:00'))
                    if offer_start_time <= current_time <= offer_end_time:
                        is_offer_active = True
                    else:
                        items_collection.update_one(
                            {'_id': ObjectId(item['_id'])},
                            {'$unset': {'offer_price': "", 'offer_start_time': "", 'offer_end_time': ""}}
                        )
                        item.pop('offer_price', None)
                        item.pop('offer_start_time', None)
                        item.pop('offer_end_time', None)
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid offer time format for item {item['_id']}: {str(e)}")
                    items_collection.update_one(
                        {'_id': ObjectId(item['_id'])},
                        {'$unset': {'offer_price': "", 'offer_start_time': "", 'offer_end_time': ""}}
                    )
                    item.pop('offer_price', None)
                    item.pop('offer_start_time', None)
                    item.pop('offer_end_time', None)
            if not is_offer_active:
                item.pop('offer_price', None)
                item.pop('offer_start_time', None)
                item.pop('offer_end_time', None)
            # Updated to use /api/images/
            if item.get('image'):
                item['image'] = f"/api/images/{os.path.basename(item['image'])}"
            else:
                item['image'] = placeholder_url
           
            for addon in item.get("addons", []):
                if addon.get('addon_image'):
                    addon['addon_image'] = f"/api/images/{os.path.basename(addon['addon_image'])}"
                else:
                    addon['addon_image'] = placeholder_url
           
            for combo in item.get("combos", []):
                if combo.get('combo_image'):
                    combo['combo_image'] = f"/api/images/{os.path.basename(combo['combo_image'])}"
                else:
                    combo['combo_image'] = placeholder_url
           
            items_list.append(item)
       
        logger.info(f"Fetched {len(items_list)} items")
        return jsonify(items_list), 200
    except Exception as e:
        logger.error(f"Error fetching items: {str(e)}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/items/<identifier>', methods=['GET'])
@db_required
def get_item(identifier):
    try:
        item = None
        try:
            object_id = ObjectId(identifier)
            item = items_collection.find_one({'_id': object_id})
        except Exception:
            item = items_collection.find_one({'item_name': identifier})
        if not item:
            logger.warning(f"Item not found: {identifier}")
            return jsonify({"error": "Item not found"}), 404
        item = convert_objectid_to_str(item)
        current_time = datetime.now(ZoneInfo("UTC"))
        placeholder_url = 'https://placehold.co/100x100/EFEFEF/AAAAAA?text=No+Image'
       
        is_offer_active = False
        if 'offer_start_time' in item and item['offer_start_time'] and 'offer_end_time' in item and item['offer_end_time']:
            try:
                offer_start_time = datetime.fromisoformat(str(item['offer_start_time']).replace('Z', '+00:00'))
                offer_end_time = datetime.fromisoformat(str(item['offer_end_time']).replace('Z', '+00:00'))
                if offer_start_time <= current_time <= offer_end_time:
                    is_offer_active = True
            except (ValueError, TypeError):
                items_collection.update_one(
                    {'_id': ObjectId(item['_id'])},
                    {'$unset': {'offer_price': "", 'offer_start_time': "", 'offer_end_time': ""}}
                )
                item.pop('offer_price', None)
                item.pop('offer_start_time', None)
                item.pop('offer_end_time', None)
       
        if not is_offer_active:
            item.pop('offer_price', None)
            item.pop('offer_start_time', None)
            item.pop('offer_end_time', None)
        # Updated to use /api/images/
        if item.get('image'):
            item['image'] = f"/api/images/{os.path.basename(item['image'])}"
        else:
            item['image'] = placeholder_url
           
        for addon in item.get('addons', []):
            if addon.get('addon_image'):
                addon['addon_image'] = f"/api/images/{os.path.basename(addon['addon_image'])}"
            else:
                addon['addon_image'] = placeholder_url
               
        for combo in item.get('combos', []):
            if combo.get('combo_image'):
                combo['combo_image'] = f"/api/images/{os.path.basename(combo['combo_image'])}"
            else:
                combo['combo_image'] = placeholder_url
       
        logger.info(f"Fetched item: {identifier}")
        return jsonify(item), 200
    except Exception as e:
        logger.error(f"Error fetching item {identifier}: {str(e)}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/items', methods=['POST'])
@db_required
def create_item():
    try:
        data = request.json
        if not data:
            logger.error("No data provided for item creation")
            return jsonify({"error": "No data provided"}), 400
        required_fields = ['item_name', 'item_code', 'item_group', 'price_list_rate']
        for field in required_fields:
            if field not in data or not data[field]:
                logger.error(f"Missing or empty required field: {field}")
                return jsonify({"error": f"Missing or empty required field: {field}"}), 400
        if 'offer_start_time' in data and data['offer_start_time'] and 'offer_end_time' in data and data['offer_end_time']:
            try:
                offer_start_time = datetime.fromisoformat(str(data['offer_start_time']).replace('Z', '+00:00'))
                offer_end_time = datetime.fromisoformat(str(data['offer_end_time']).replace('Z', '+00:00'))
                if offer_start_time >= offer_end_time:
                    logger.error("offer_start_time must be before offer_end_time")
                    return jsonify({"error": "Offer start time must be before offer end time"}), 400
            except (ValueError, TypeError) as e:
                logger.error(f"Invalid offer time format: {str(e)}")
                return jsonify({"error": f"Invalid offer time format: {str(e)}"}), 400
        data = sanitize_image_fields(data)
        data.setdefault('custom_addon_applicable', False)
        data.setdefault('custom_combo_applicable', False)
        data.setdefault('custom_total_calories', 0)
        data.setdefault('custom_total_protein', 0)
        data.setdefault('kitchen', "")
        data.setdefault('has_variant_pricing', False)
        data.setdefault('variant_prices', {"small_price": 0, "medium_price": 0, "large_price": 0})
        data.setdefault('variant_quantities', {"small_quantity": 0, "medium_quantity": 0, "large_quantity": 0})
        data.setdefault('sold_quantities', {"small_sold": 0, "medium_sold": 0, "large_sold": 0})
        data.setdefault('ice_preference', "without_ice")
        data.setdefault('ice_price', 0)
        data.setdefault('addons', [])
        data.setdefault('combos', [])
        data.setdefault('ingredients', [])
        data.setdefault('variants', [])
        data['created_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
        item_id = items_collection.insert_one(data).inserted_id
        logger.info(f"Item created with ID: {item_id}")
        return jsonify({'message': 'Item created successfully!', 'id': str(item_id)}), 201
    except Exception as e:
        logger.error(f"Error creating item: {str(e)}")
        return jsonify({'error': str(e)}), 500
@app.route('/api/items/<item_id>', methods=['PUT'])
@db_required
def update_item(item_id):
    try:
        data = request.json
        if not data:
            logger.error("No data provided for item update")
            return jsonify({"error": "No data provided"}), 400
        try:
            object_id = ObjectId(item_id)
        except Exception:
            logger.error(f"Invalid item ID: {item_id}")
            return jsonify({"error": "Invalid item ID"}), 400
        if '_id' in data:
            del data['_id']
        if 'offer_start_time' in data and data['offer_start_time'] and 'offer_end_time' in data and data['offer_end_time']:
            try:
                offer_start_time = datetime.fromisoformat(str(data['offer_start_time']).replace('Z', '+00:00'))
                offer_end_time = datetime.fromisoformat(str(data['offer_end_time']).replace('Z', '+00:00'))
                if offer_start_time >= offer_end_time:
                    logger.error("offer_start_time must be before offer_end_time")
                    return jsonify({"error": "Offer start time must be before offer end time"}), 400
            except (ValueError, TypeError) as e:
                logger.error(f"Invalid offer time format: {str(e)}")
                return jsonify({"error": f"Invalid offer time format: {str(e)}"}), 400
        data = sanitize_image_fields(data)
        data['modified_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
        result = items_collection.update_one({'_id': object_id}, {'$set': data})
        if result.matched_count == 0:
            logger.warning(f"Item not found for update: {item_id}")
            return jsonify({"error": "Item not found"}), 404
        logger.info(f"Item updated: {item_id}")
        return jsonify({"message": "Item updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating item {item_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/items/<item_id>', methods=['PATCH'])
@db_required
def patch_item(item_id):
    try:
        data = request.json
        if not data:
            logger.error("No data provided for item patch")
            return jsonify({"error": "No data provided"}), 400
        try:
            object_id = ObjectId(item_id)
        except Exception:
            logger.error(f"Invalid item ID: {item_id}")
            return jsonify({"error": "Invalid item ID"}), 400
        if '_id' in data:
            del data['_id']
       
        data = sanitize_image_fields(data)
        data['modified_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
        result = items_collection.update_one({'_id': object_id}, {'$set': data})
        if result.matched_count == 0:
            logger.warning(f"Item not found for patch: {item_id}")
            return jsonify({"error": "Item not found"}), 404
        logger.info(f"Item patched: {item_id}")
        return jsonify({"message": "Item updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error patching item {item_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/items/<item_id>', methods=['DELETE'])
@db_required
def delete_item(item_id):
    try:
        try:
            object_id = ObjectId(item_id)
        except Exception:
            logger.error(f"Invalid item ID: {item_id}")
            return jsonify({"error": "Invalid item ID"}), 400
        result = items_collection.delete_one({'_id': object_id})
        if result.deleted_count == 0:
            logger.warning(f"Item not found for deletion: {item_id}")
            return jsonify({"error": "Item not found"}), 404
        logger.info(f"Item deleted: {item_id}")
        return jsonify({"message": "Item deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting item {item_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/items/<item_id>/offer', methods=['PUT'])
@db_required
def update_item_offer(item_id):
    try:
        offer_data = request.json
        try:
            object_id = ObjectId(item_id)
        except Exception:
            logger.error(f"Invalid item ID: {item_id}")
            return jsonify({"error": "Invalid item ID"}), 400
        if 'offer_price' not in offer_data or 'offer_start_time' not in offer_data or 'offer_end_time' not in offer_data:
            return jsonify({"error": "Offer price, start time, and end time are required"}), 400
        try:
            offer_start_time = datetime.fromisoformat(str(offer_data['offer_start_time']).replace('Z', '+00:00'))
            offer_end_time = datetime.fromisoformat(str(offer_data['offer_end_time']).replace('Z', '+00:00'))
            if offer_start_time >= offer_end_time:
                logger.error("offer_start_time must be before offer_end_time")
                return jsonify({"error": "Offer start time must be before offer end time"}), 400
        except (ValueError, TypeError) as e:
            logger.error(f"Invalid offer time format: {str(e)}")
            return jsonify({"error": f"Invalid offer time format: {str(e)}"}), 400
        offer_data['modified_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
        result = items_collection.update_one({'_id': object_id}, {'$set': offer_data})
        if result.matched_count == 0:
            logger.warning(f"Item not found for offer update: {item_id}")
            return jsonify({"error": "Item not found"}), 404
        logger.info(f"Offer updated for item: {item_id}")
        return jsonify({"message": "Offer updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating offer for item {item_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/customers', methods=['GET'])
@db_required
def get_all_customers():
    try:
        customers = list(customers_collection.find())
        customers = [convert_objectid_to_str(customer) for customer in customers]
        logger.info(f"Fetched {len(customers)} customers")
        return jsonify(customers), 200
    except Exception as e:
        logger.error(f"Error fetching customers: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/customers/<customer_id>', methods=['GET', 'PUT', 'DELETE'])
@db_required
def customer_operations(customer_id):
    try:
        if not customer_id or customer_id == "undefined":
            logger.error("Invalid customer_id: 'undefined' or empty")
            return jsonify({"error": "Invalid customer ID"}), 400
        try:
            object_id = ObjectId(customer_id)
        except Exception:
            logger.error(f"Invalid ObjectId: {customer_id}")
            return jsonify({"error": "Invalid customer ID format"}), 400
        if request.method == 'GET':
            customer = customers_collection.find_one({'_id': object_id})
            if not customer:
                logger.warning(f"Customer not found: {customer_id}")
                return jsonify({"error": "Customer not found"}), 404
            customer = convert_objectid_to_str(customer)
            logger.info(f"Fetched customer: {customer_id}")
            return jsonify(customer), 200
        elif request.method == 'PUT':
            customer_data = request.get_json()
            if not customer_data:
                logger.error("No data provided for customer update")
                return jsonify({"error": "No data provided"}), 400
            result = customers_collection.update_one(
                {'_id': object_id},
                {'$set': {
                    'customer_name': customer_data.get('customer_name', ''),
                    'phone_number': customer_data.get('phone_number', ''),
                    'whatsapp_number': customer_data.get('whatsapp_number', ''),
                    'email': customer_data.get('email', ''),
                    'building_name': customer_data.get('building_name', ''),
                    'flat_villa_no': customer_data.get('flat_villa_no', ''),
                    'location': customer_data.get('location', ''),
                    'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()
                }}
            )
            if result.matched_count == 0:
                logger.warning(f"Customer not found for update: {customer_id}")
                return jsonify({"error": "Customer not found"}), 404
            logger.info(f"Customer updated: {customer_id}")
            return jsonify({"message": "Customer updated successfully"}), 200
        elif request.method == 'DELETE':
            result = customers_collection.delete_one({'_id': object_id})
            if result.deleted_count == 0:
                logger.warning(f"Customer not found for deletion: {customer_id}")
                return jsonify({"error": "Customer not found"}), 404
            logger.info(f"Customer deleted: {customer_id}")
            return jsonify({"message": "Customer deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error in customer operations for {customer_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/customers', methods=['POST'])
@db_required
def create_customer():
    try:
        customer_data = request.get_json()
        if not customer_data or 'customer_name' not in customer_data or 'phone_number' not in customer_data:
            logger.error("Invalid customer data provided")
            return jsonify({"error": "Customer name and phone number are required"}), 400
       
        existing_customer = customers_collection.find_one({'phone_number': customer_data['phone_number']})
        if existing_customer:
            logger.warning(f"Duplicate phone number: {customer_data['phone_number']}")
            return jsonify({"error": "Phone number already exists", "customer_name": existing_customer['customer_name']}), 409
        customer_data['created_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
        customer_data['modified_at'] = customer_data['created_at']
        result = customers_collection.insert_one(customer_data)
        new_customer_id = str(result.inserted_id)
        logger.info(f"Customer created: {new_customer_id}")
        return jsonify({"id": new_customer_id, "message": "Customer created successfully"}), 201
    except Exception as e:
        logger.error(f"Error creating customer: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/sales', methods=['POST'])
@db_required
def create_sales_invoice():
    try:
        sales_data = request.json
        required_fields = ['customer', 'items', 'total', 'userId']
        missing_fields = [field for field in required_fields if field not in sales_data or sales_data[field] is None]
        if missing_fields:
            error_msg = f"Missing required fields: {', '.join(missing_fields)}"
            logger.error(error_msg)
            return jsonify({"error": error_msg}), 400
       
        user = users_collection.find_one({"email": sales_data['userId']})
        if not user:
            logger.error(f"Invalid userId: {sales_data['userId']}")
            return jsonify({"error": "Invalid userId"}), 400
       
        sales_data['date'] = sales_data.get('date', datetime.now().strftime("%Y-%m-%d"))
        sales_data['time'] = sales_data.get('time', datetime.now().strftime("%H:%M:%S"))
       
        net_total = float(sales_data['total'])
        vat_amount = float(sales_data.get('vat_amount', net_total * 0.10))
        grand_total = float(sales_data.get('grand_total', net_total + vat_amount))
        sales_data['vat_amount'] = round(vat_amount, 2)
        sales_data['grand_total'] = round(grand_total, 2)
       
        sales_data['invoice_no'] = sales_data.get('invoice_no', f"INV-{int(datetime.now().timestamp())}")
        sales_data['status'] = sales_data.get('status', 'Draft')
       
        processed_items = []
        for item in sales_data.get('items', []):
            if not all(key in item for key in ['item_name', 'basePrice', 'quantity']):
                logger.error("Invalid item structure in sales invoice")
                return jsonify({"error": "Each item must include item_name, basePrice, and quantity"}), 400
            processed_addons = []
            for addon in item.get('addons', []):
                if not all(key in addon for key in ['name1', 'addon_price', 'addon_quantity']):
                    logger.error("Invalid addon structure in sales invoice")
                    return jsonify({"error": "Each addon must include name1, addon_price, and addon_quantity"}), 400
                processed_addons.append({
                    "addon_name": addon['name1'],
                    "addon_price": float(addon['addon_price']),
                    "addon_quantity": int(addon['addon_quantity']),
                    "addon_image": addon.get('addon_image', ''),
                    "size": addon.get('size', 'M'),
                    "kitchen": addon.get('kitchen', 'Main Kitchen'),
                })
            processed_combos = []
            for combo in item.get('selectedCombos', []):
                if not all(key in combo for key in ['name1', 'combo_price']):
                    logger.error("Invalid combo structure in sales invoice")
                    return jsonify({"error": "Each combo must include name1 and combo_price"}), 400
                processed_combos.append({
                    "name1": combo['name1'],
                    "combo_price": float(combo['combo_price']),
                    "combo_quantity": int(combo.get('combo_quantity', 1)),
                    "combo_image": combo.get('combo_image', ''),
                    "size": combo.get('size', 'M'),
                    "spicy": combo.get('spicy', False),
                    "kitchen": combo.get('kitchen', 'Main Kitchen'),
                })
            processed_items.append({
                "item_name": item['item_name'],
                "basePrice": float(item['basePrice']),
                "quantity": int(item['quantity']),
                "amount": float(item.get('amount', item['basePrice'])),
                "icePreference": item.get('icePreference', 'without_ice'),
                "isSpicy": item.get('isSpicy', False),
                "kitchen": item.get('kitchen', 'Main Kitchen'),
                "selectedSize": item.get('selectedSize', 'M'),
                "ingredients": item.get('ingredients', []),
                "addons": processed_addons,
                "selectedCombos": processed_combos,
            })
        sales_data['items'] = processed_items
        sales_data['created_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
       
        sales_id = sales_collection.insert_one(sales_data).inserted_id
        logger.info(f"Sale saved successfully: {sales_data['invoice_no']}")
        return jsonify({
            "id": str(sales_id),
            "invoice_no": sales_data['invoice_no'],
            "net_total": sales_data['total'],
            "vat_amount": sales_data['vat_amount'],
            "grand_total": sales_data['grand_total'],
            "userId": sales_data['userId']
        }), 201
    except Exception as e:
        logger.error(f"Error creating sales invoice: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/sales', methods=['GET'])
@db_required
def get_all_sales():
    try:
        sales = list(sales_collection.find())
        sales = [convert_objectid_to_str(sale) for sale in sales]
        logger.info(f"Fetched {len(sales)} sales invoices")
        return jsonify(sales), 200
    except Exception as e:
        logger.error(f"Error fetching sales: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/sales/<invoice_no>', methods=['GET'])
@db_required
def get_sale_by_invoice_no(invoice_no):
    try:
        sale = sales_collection.find_one({'invoice_no': invoice_no.strip()})
        if not sale:
            logger.warning(f"Sale not found: {invoice_no}")
            return jsonify({"error": "Invoice not found"}), 404
        sale = convert_objectid_to_str(sale)
        logger.info(f"Fetched sale: {invoice_no}")
        return jsonify(sale), 200
    except Exception as e:
        logger.error(f"Error fetching sale {invoice_no}: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/sales/<invoice_no>/status', methods=['PUT'])
@db_required
def update_sale_status(invoice_no):
    try:
        data = request.get_json()
        status = data.get('status')
        if not status:
            return jsonify({"error": "Status is required"}), 400
        result = sales_collection.update_one(
            {'invoice_no': invoice_no.strip()},
            {'$set': {'status': status, 'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
        )
        if result.matched_count == 0:
            logger.warning(f"Sale not found for status update: {invoice_no}")
            return jsonify({"error": "Invoice not found"}), 404
        logger.info(f"Sale status updated: {invoice_no} to {status}")
        return jsonify({"message": "Sale status updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating sale status {invoice_no}: {str(e)}")
        return jsonify({"error": str(e)}), 500
## Tables
@app.route('/api/tables', methods=['GET'])
@db_required
def get_tables():
    try:
        tables = list(tables_collection.find())
        tables = [convert_objectid_to_str(table) for table in tables]
        logger.info(f"Fetched {len(tables)} tables")
        return jsonify({"message": tables}), 200
    except Exception as e:
        logger.error(f"Error fetching tables: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/tables', methods=['POST'])
@db_required
def add_table():
    try:
        data = request.get_json()
        table_number = data.get("table_number")
        number_of_chairs = data.get("number_of_chairs")
        if not table_number or not number_of_chairs:
            logger.error("Missing table_number or number_of_chairs")
            return jsonify({"error": "Table number and number of chairs are required"}), 400
        if tables_collection.find_one({"table_number": table_number}):
            logger.warning(f"Table number already exists: {table_number}")
            return jsonify({"error": "Table number already exists"}), 400
        new_table = {
            "table_number": table_number,
            "number_of_chairs": int(number_of_chairs),
            "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
        }
        tables_collection.insert_one(new_table)
        logger.info(f"Table added: {table_number}")
        return jsonify({"message": "Table added successfully"}), 201
    except Exception as e:
        logger.error(f"Error adding table: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/tables/<table_number>', methods=['PUT'])
@db_required
def update_table(table_number):
    try:
        data = request.get_json()
        number_of_chairs = data.get("number_of_chairs")
        if not number_of_chairs:
            return jsonify({"error": "Number of chairs is required"}), 400
        result = tables_collection.update_one(
            {"table_number": table_number},
            {"$set": {"number_of_chairs": int(number_of_chairs), "modified_at": datetime.now(ZoneInfo("UTC")).isoformat()}}
        )
        if result.matched_count == 0:
            logger.warning(f"Table not found for update: {table_number}")
            return jsonify({"error": "Table not found"}), 404
        logger.info(f"Table updated: {table_number}")
        return jsonify({"message": "Table updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating table {table_number}: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/tables/<table_number>', methods=['DELETE'])
@db_required
def delete_table(table_number):
    try:
        result = tables_collection.delete_one({"table_number": table_number})
        if result.deleted_count == 0:
            logger.warning(f"Table not found: {table_number}")
            return jsonify({"error": "Table not found"}), 404
        logger.info(f"Table deleted: {table_number}")
        return jsonify({"message": "Table deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting table {table_number}: {str(e)}")
        return jsonify({"error": str(e)}), 500
## POS Entries
@app.route('/api/create_opening_entry', methods=['POST'])
@db_required
def create_opening_entry():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"message": "No data provided", "status": "error"}), 400
        required_fields = ['period_start_date', 'posting_date', 'company', 'user', 'balance_details']
        missing_fields = [field for field in required_fields if field not in data or not data[field]]
        if missing_fields:
            return jsonify({"message": f"Missing fields: {', '.join(missing_fields)}", "status": "error"}), 400
        balance_details = data['balance_details']
        if not isinstance(balance_details, list):
            return jsonify({"message": "balance_details must be a list", "status": "error"}), 400
        for detail in balance_details:
            if not all(key in detail for key in ['mode_of_payment', 'opening_amount']):
                return jsonify({"message": "Each balance detail must have mode_of_payment and opening_amount", "status": "error"}), 400
            detail['opening_amount'] = float(detail['opening_amount'])
        data['creation'] = datetime.now(ZoneInfo("UTC")).isoformat()
        data['modified'] = data['creation']
        data['name'] = f"OPEN-{int(datetime.now().timestamp())}"
        data['status'] = data.get('status', 'Draft')
        data['docstatus'] = data.get('docstatus', 0)
        data['pos_profile'] = data.get('pos_profile', 'POS-001')
        result = opening_collection.insert_one(data)
        username = data['user']
        user = users_collection.find_one({"firstName": username})
        if user:
            users_collection.update_one(
                {"_id": user['_id']},
                {"$set": {"last_opening_entry_time": datetime.now(ZoneInfo("UTC")).isoformat()}}
            )
            logger.info(f"Updated last_opening_entry_time for user: {username}")
        logger.info(f"POS opening entry created: {data['name']}")
        return jsonify({"message": {"name": data['name'], "status": "success"}}), 201
    except Exception as e:
        logger.error(f"Error in create_opening_entry: {str(e)}")
        return jsonify({"message": f"Server error: {str(e)}", "status": "error"}), 500
@app.route('/api/get_pos_opening_entries', methods=['POST'])
@db_required
def get_pos_opening_entries():
    try:
        data = request.get_json()
        if not data or 'pos_profile' not in data:
            return jsonify({"message": "POS profile is required", "status": "error"}), 400
        pos_profile = data['pos_profile']
        entries = list(opening_collection.find({"pos_profile": pos_profile}))
        entries = [convert_objectid_to_str(entry) for entry in entries]
        logger.info(f"Fetched {len(entries)} POS opening entries for profile: {pos_profile}")
        return jsonify({"message": entries, "status": "success"}), 200
    except Exception as e:
        logger.error(f"Error in get_pos_opening_entries: {str(e)}")
        return jsonify({"message": f"Server error: {str(e)}", "status": "error"}), 500
@app.route('/api/get_pos_invoices', methods=['POST'])
@db_required
def get_pos_invoices():
    try:
        data = request.get_json()
        pos_opening_entry = data.get('pos_opening_entry')
        if not pos_opening_entry:
            return jsonify({"message": "POS opening entry is required", "status": "error"}), 400
        opening_entry = opening_collection.find_one({"name": pos_opening_entry})
        if not opening_entry:
            return jsonify({"message": "Opening entry not found", "status": "error"}), 404
        period_start = opening_entry['period_start_date']
        invoices = list(sales_collection.find({"date": {"$gte": period_start}}))
        invoices = [convert_objectid_to_str(inv) for inv in invoices]
        total = sum(float(inv['grand_total']) for inv in invoices)
        net_total = sum(float(inv['total']) for inv in invoices)
        total_qty = sum(sum(item['quantity'] for item in inv['items']) for inv in invoices)
        taxes = [{"account_head": "VAT", "rate": 10, "amount": total - net_total}]
        response = {
            "invoices": [{"pos_invoice": inv['invoice_no'], "grand_total": inv['grand_total'], "posting_date": inv['date'], "customer": inv['customer']} for inv in invoices],
            "taxes": taxes,
            "grand_total": total,
            "net_total": net_total,
            "total_quantity": total_qty,
            "status": "success"
        }
        logger.info(f"Fetched POS invoices for opening entry: {pos_opening_entry}")
        return jsonify({"message": response}), 200
    except Exception as e:
        logger.error(f"Error in get_pos_invoices: {str(e)}")
        return jsonify({"message": f"Error: {str(e)}", "status": "error"}), 500
@app.route('/api/create_closing_entry', methods=['POST'])
@db_required
def create_closing_entry():
    try:
        data = request.get_json()
        required_fields = ['pos_opening_entry', 'posting_date', 'period_end_date', 'pos_transactions', 'payment_reconciliation', 'taxes', 'grand_total', 'net_total', 'total_quantity']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"message": f"Missing fields: {', '.join([f for f in required_fields if f not in data])}", "status": "error"}), 400
        opening_entry = opening_collection.find_one({"name": data['pos_opening_entry']})
        if not opening_entry:
            return jsonify({"message": "Opening entry not found", "status": "error"}), 404
        data['creation'] = datetime.now(ZoneInfo("UTC")).isoformat()
        data['modified'] = data['creation']
        data['name'] = f"CLOSE-{int(datetime.now().timestamp())}"
        data['status'] = 'Draft'
        data['docstatus'] = 0
        result = pos_closing_collection.insert_one(data)
        logger.info(f"POS closing entry created: {data['name']}")
        return jsonify({"message": {"name": data['name'], "status": "success", "message": "Closing Entry created"}}), 201
    except Exception as e:
        logger.error(f"Error in create_closing_entry: {str(e)}")
        return jsonify({"message": f"Error: {str(e)}", "status": "error"}), 500
## Image Handling
@app.route('/api/upload-image', methods=['POST', 'OPTIONS'])
@db_required
def upload_image():
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    try:
        if 'files' not in request.files:
            logger.error("No files part in request")
            return jsonify({"error": "No files provided"}), 400
        files = request.files.getlist('files')
        if not files or all(file.filename == '' for file in files):
            logger.error("No valid files selected")
            return jsonify({"error": "No valid files selected"}), 400
        urls = []
        for file in files:
            if file and allowed_file(file.filename, ALLOWED_IMAGE_EXTENSIONS):
                filename = handle_image_upload(file)
                if filename:
                    # Updated to return /api/images/ URLs
                    urls.append(f"/api/images/{filename}")
                    logger.info(f"Uploaded image: {filename}")
                else:
                    logger.warning(f"Failed to upload image: {file.filename}")
            else:
                logger.warning(f"Invalid file type: {file.filename}")
        if not urls:
            return jsonify({"error": "No valid images uploaded"}), 400
        return jsonify({"urls": urls}), 200
    except Exception as e:
        logger.error(f"Error uploading images: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500
@app.route('/api/delete-image/<filename>', methods=['DELETE', 'OPTIONS'])
@db_required
def delete_image(filename):
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    try:
        item_id = request.args.get('item_id')
        field = request.args.get('field', 'image')
        valid_fields = {'image', 'images', 'addon_image', 'combo_image', 'variant_image'}
        if field not in valid_fields:
            logger.error(f"Invalid field specified: {field}")
            return jsonify({"error": f"Invalid field: {field}. Must be one of {valid_fields}"}), 400
        if not item_id:
            logger.error("Item ID is required for image deletion")
            return jsonify({"error": "Item ID is required"}), 400
        try:
            object_id = ObjectId(item_id)
        except Exception:
            logger.error(f"Invalid item ID: {item_id}")
            return jsonify({"error": "Invalid item ID"}), 400
        filename = secure_filename(filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if field == 'images':
            result = items_collection.update_one(
                {"_id": object_id, "images": filename},
                {"$pull": {"images": filename}}
            )
            if result.modified_count == 0:
                logger.warning(f"Image {filename} not found in images array for item {item_id}")
                return jsonify({"error": "Image not found in item"}), 404
        elif field == 'image':
            result = items_collection.update_one(
                {"_id": object_id, "image": filename},
                {"$set": {"image": None}}
            )
            if result.modified_count == 0:
                logger.warning(f"Image {filename} not found in image field for item {item_id}")
                return jsonify({"error": "Image not found in item"}), 404
        elif field == 'addon_image':
            result = items_collection.update_one(
                {"_id": object_id, "addons.addon_image": filename},
                {"$set": {"addons.$[elem].addon_image": None}},
                array_filters=[{"elem.addon_image": filename}]
            )
            if result.modified_count == 0:
                logger.warning(f"Image {filename} not found in addons for item {item_id}")
                return jsonify({"error": "Image not found in addons"}), 404
        elif field == 'combo_image':
            result = items_collection.update_one(
                {"_id": object_id, "combos.combo_image": filename},
                {"$set": {"combos.$[elem].combo_image": None}},
                array_filters=[{"elem.combo_image": filename}]
            )
            if result.modified_count == 0:
                logger.warning(f"Image {filename} not found in combos for item {item_id}")
                return jsonify({"error": "Image not found in combos"}), 404
        elif field == 'variant_image':
            result = items_collection.update_one(
                {"_id": object_id, "variants.variant_image": filename},
                {"$set": {"variants.$[elem].variant_image": None}},
                array_filters=[{"elem.variant_image": filename}]
            )
            if result.modified_count == 0:
                logger.warning(f"Image {filename} not found in variants for item {item_id}")
                return jsonify({"error": "Image not found in variants"}), 404
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
                logger.info(f"Image deleted from filesystem: {filename}")
            except PermissionError as e:
                logger.error(f"Permission denied deleting image {filename}: {str(e)}")
                return jsonify({"error": "Permission denied deleting image"}), 403
            except Exception as e:
                logger.error(f"Error deleting image {filename}: {str(e)}")
                return jsonify({"error": "Error deleting image"}), 500
        else:
            logger.warning(f"Image file not found on filesystem: {filename}")
        logger.info(f"Image {filename} deleted from {field} for item {item_id}")
        return jsonify({"message": "Image deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting image {filename} for item {item_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500
# Updated to use /api/images/ route
@app.route('/api/images/<path:filename>', methods=['GET'])
def serve_uploaded_image(filename):
    logger.debug(f"Serving image: {filename}")
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except Exception as e:
        logger.error(f"Error serving image {filename}: {str(e)}")
        return jsonify({"error": "Image not found"}), 404
## Import MongoDB Data
@app.route('/api/import-mongodb', methods=['POST', 'OPTIONS'])
@db_required
def import_mongodb():
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Accept'
        return response, 200
    try:
        if 'file' not in request.files:
            logger.error("No file part in request")
            return jsonify({"error": "No file uploaded"}), 400
        file = request.files['file']
        if file.filename == '':
            logger.error("No selected file")
            return jsonify({"error": "No selected file"}), 400
        if not allowed_file(file.filename, ALLOWED_JSON_EXTENSIONS):
            logger.error(f"Invalid file type: {file.filename}")
            return jsonify({"error": "Only JSON files are allowed"}), 400
        filename = secure_filename(file.filename)
        collection_name = filename.rsplit('.', 1)[0].split('.')[-1]
        valid_collections = [
            'users', 'tables', 'items', 'customers', 'sales',
            'picked_up_items', 'pos_opening_entries', 'pos_closing_entries',
            'kitchens', 'item_groups'
        ]
        if collection_name not in valid_collections:
            logger.error(f"Invalid collection name: {collection_name}")
            return jsonify({"error": f"Unsupported collection name: {collection_name}"}), 400
        target_collection = db[collection_name]
        data = json.loads(file.read().decode('utf-8'))
        if not isinstance(data, list):
            logger.error("JSON data must be an array")
            return jsonify({"error": "JSON data must be an array"}), 400
        inserted_count = 0
        for record in data:
            if '_id' in record and isinstance(record['_id'], dict) and '$oid' in record['_id']:
                record['_id'] = ObjectId(record['_id']['$oid'])
            record['imported_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
            unique_key = (
                {'_id': record.get('_id')} if '_id' in record else
                {'email': record.get('email')} if collection_name == 'users' else
                {'table_number': record.get('table_number')} if collection_name == 'tables' else
                {'item_name': record.get('item_name')} if collection_name == 'items' else
                {'phone_number': record.get('phone_number')} if collection_name == 'customers' else
                {'invoice_no': record.get('invoice_no')} if collection_name == 'sales' else
                {'customerName': record.get('customerName')} if collection_name == 'picked_up_items' else
                {'name': record.get('name')} if collection_name in ['pos_opening_entries', 'pos_closing_entries'] else
                {'kitchen_name': record.get('kitchen_name')} if collection_name == 'kitchens' else
                {'group_name': record.get('group_name')} if collection_name == 'item_groups' else
                {}
            )
            if not unique_key:
                logger.error(f"No unique key defined for record in collection {collection_name}")
                return jsonify({"error": f"No unique key defined for record in collection {collection_name}"}), 400
            target_collection.replace_one(unique_key, record, upsert=True)
            inserted_count += 1
        logger.info(f"Imported {inserted_count} records into {collection_name}")
        return jsonify({"message": f"Successfully imported {inserted_count} records into {collection_name}"}), 200
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON format in file {filename}: {str(e)}")
        return jsonify({"error": f"Invalid JSON format: {str(e)}"}), 400
    except Exception as e:
        logger.error(f"Error importing data: {str(e)}")
        return jsonify({"error": str(e)}), 500
## Kitchens
@app.route('/api/kitchens', methods=['GET'])
@db_required
def get_kitchens():
    try:
        kitchens = list(kitchens_collection.find())
        kitchens = [convert_objectid_to_str(kitchen) for kitchen in kitchens]
        logger.info(f"Fetched {len(kitchens)} kitchens")
        return jsonify(kitchens), 200
    except Exception as e:
        logger.error(f"Error fetching kitchens: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/kitchens', methods=['POST'])
@db_required
def create_kitchen():
    try:
        data = request.get_json()
        if not data or 'kitchen_name' not in data:
            logger.error("Missing kitchen_name in request")
            return jsonify({"error": "Kitchen name is required"}), 400
        kitchen_name = data['kitchen_name']
        if kitchens_collection.find_one({"kitchen_name": kitchen_name}):
            logger.warning(f"Kitchen already exists: {kitchen_name}")
            return jsonify({"error": "Kitchen name already exists"}), 400
        new_kitchen = {
            "kitchen_name": kitchen_name,
            "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
        }
        result = kitchens_collection.insert_one(new_kitchen)
        logger.info(f"Kitchen created: {kitchen_name}")
        return jsonify({"message": "Kitchen created successfully", "id": str(result.inserted_id)}), 201
    except Exception as e:
        logger.error(f"Error creating kitchen: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/kitchens/<kitchen_id>', methods=['PUT'])
@db_required
def update_kitchen(kitchen_id):
    try:
        data = request.get_json()
        if not data or 'kitchen_name' not in data:
            logger.error("Missing kitchen_name in request")
            return jsonify({"error": "Kitchen name is required"}), 400
        try:
            object_id = ObjectId(kitchen_id)
        except Exception:
            logger.error(f"Invalid kitchen ID: {kitchen_id}")
            return jsonify({"error": "Invalid kitchen ID"}), 400
        result = kitchens_collection.update_one(
            {'_id': object_id},
            {'$set': {'kitchen_name': data['kitchen_name'], 'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
        )
        if result.matched_count == 0:
            logger.warning(f"Kitchen not found for update: {kitchen_id}")
            return jsonify({"error": "Kitchen not found"}), 404
        logger.info(f"Kitchen updated: {kitchen_id}")
        return jsonify({"message": "Kitchen updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating kitchen {kitchen_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/kitchens/<kitchen_id>', methods=['DELETE'])
@db_required
def delete_kitchen(kitchen_id):
    try:
        try:
            object_id = ObjectId(kitchen_id)
        except Exception:
            logger.error(f"Invalid kitchen ID: {kitchen_id}")
            return jsonify({"error": "Invalid kitchen ID"}), 400
        result = kitchens_collection.delete_one({'_id': object_id})
        if result.deleted_count == 0:
            logger.warning(f"Kitchen not found for deletion: {kitchen_id}")
            return jsonify({"error": "Kitchen not found"}), 404
        logger.info(f"Kitchen deleted: {kitchen_id}")
        return jsonify({"message": "Kitchen deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting kitchen {kitchen_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500
## Item Groups
@app.route('/api/item-groups', methods=['GET'])
@db_required
def get_item_groups():
    try:
        item_groups = list(item_groups_collection.find())
        item_groups = [convert_objectid_to_str(group) for group in item_groups]
        logger.info(f"Fetched {len(item_groups)} item groups")
        return jsonify(item_groups), 200
    except Exception as e:
        logger.error(f"Error fetching item groups: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/item-groups', methods=['POST'])
@db_required
def create_item_group():
    try:
        data = request.get_json()
        if not data or 'group_name' not in data:
            logger.error("Missing group_name in request")
            return jsonify({"error": "Group name is required"}), 400
        group_name = data['group_name']
        if item_groups_collection.find_one({"group_name": group_name}):
            logger.warning(f"Item group already exists: {group_name}")
            return jsonify({"error": "Item group name already exists"}), 400
        new_group = {
            "group_name": group_name,
            "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
        }
        result = item_groups_collection.insert_one(new_group)
        logger.info(f"Item group created: {group_name}")
        return jsonify({"message": "Item group created successfully", "id": str(result.inserted_id)}), 201
    except Exception as e:
        logger.error(f"Error creating item group: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/item-groups/<group_id>', methods=['PUT'])
@db_required
def update_item_group(group_id):
    try:
        data = request.get_json()
        if not data or 'group_name' not in data:
            logger.error("Missing group_name in request")
            return jsonify({"error": "Group name is required"}), 400
        try:
            object_id = ObjectId(group_id)
        except Exception:
            logger.error(f"Invalid group ID: {group_id}")
            return jsonify({"error": "Invalid group ID"}), 400
        result = item_groups_collection.update_one(
            {'_id': object_id},
            {'$set': {'group_name': data['group_name'], 'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
        )
        if result.matched_count == 0:
            logger.warning(f"Item group not found for update: {group_id}")
            return jsonify({"error": "Item group not found"}), 404
        logger.info(f"Item group updated: {group_id}")
        return jsonify({"message": "Item group updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error updating item group {group_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/item-groups/<group_id>', methods=['DELETE'])
@db_required
def delete_item_group(group_id):
    try:
        try:
            object_id = ObjectId(group_id)
        except Exception:
            logger.error(f"Invalid group ID: {group_id}")
            return jsonify({"error": "Invalid group ID"}), 400
        result = item_groups_collection.delete_one({'_id': object_id})
        if result.deleted_count == 0:
            logger.warning(f"Item group not found for deletion: {group_id}")
            return jsonify({"error": "Item group not found"}), 404
        logger.info(f"Item group deleted: {group_id}")
        return jsonify({"message": "Item group deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting item group {group_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500
## Item Nutrition
@app.route('/api/items/nutrition', methods=['POST', 'OPTIONS'])
@db_required
def save_item_nutrition():
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    try:
        data = request.get_json()
        if not data:
            logger.error("No data provided for saving ingredients")
            return jsonify({"error": "No data provided"}), 400
        required_fields = ['item_name', 'type', 'instances', 'ingredients']
        for field in required_fields:
            if field not in data or data[field] is None:
                logger.error(f"Missing or null required field: {field}")
                return jsonify({"error": f"Missing or null required field: {field}"}), 400
        item_name = data['item_name']
        item_type = data['type']
        instances = data['instances']
        ingredients = data['ingredients']
        if item_type not in ['item', 'addon', 'combo']:
            logger.error(f"Invalid type: {item_type}")
            return jsonify({"error": "Invalid type, must be 'item', 'addon', or 'combo'"}), 400
        if not isinstance(ingredients, list):
            logger.error("Ingredients must be a list of objects")
            return jsonify({"error": "Ingredients must be a list of objects"}), 400
        for ingredient in ingredients:
            if not isinstance(ingredient, dict) or not all(key in ingredient for key in ['name', 'small', 'medium', 'large', 'weight', 'nutrition']):
                logger.error("Each ingredient must be an object with required fields")
                return jsonify({"error": "Each ingredient must be an object with required fields"}), 400
        filtered_ingredients = [ing for ing in ingredients if ing['name'].strip()]
        updated_count = 0
        for instance in instances:
            item_id = instance['item_id']
            index = instance.get('index')
            item = items_collection.find_one({'_id': ObjectId(item_id)})
            if not item:
                logger.warning(f"Item not found: {item_id}")
                continue
            update_query = {'_id': ObjectId(item_id)}
            update_data = {'$set': {'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
            if item_type == 'item':
                update_data['$set']['ingredients'] = filtered_ingredients
            elif item_type == 'addon':
                if index is None or not isinstance(index, int):
                    logger.error(f"Index is required for addons in item_id: {item_id}")
                    continue
                update_data['$set'][f'addons.{index}.ingredients'] = filtered_ingredients
            elif item_type == 'combo':
                if index is None or not isinstance(index, int):
                    logger.error(f"Index is required for combos in item_id: {item_id}")
                    continue
                update_data['$set'][f'combos.{index}.ingredients'] = filtered_ingredients
            result = items_collection.update_one(update_query, update_data)
            if result.matched_count > 0:
                updated_count += 1
        if updated_count == 0:
            logger.error(f"No items updated for {item_type}: {item_name}")
            return jsonify({"error": "No items updated, please check instance data"}), 400
        logger.info(f"Ingredients updated for {item_type}: {item_name} across {updated_count} instances")
        return jsonify({"message": "Ingredients saved successfully"}), 200
    except Exception as e:
        logger.error(f"Error saving ingredients for {item_name}: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500
@app.route('/api/items/nutrition/<item_name>', methods=['GET', 'DELETE', 'OPTIONS'])
@db_required
def handle_item_nutrition(item_name):
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, X-Instances'
        return response, 200
    try:
        item_type = request.args.get('type')
        item_id = request.args.get('item_id')
        index = request.args.get('index')
        if request.method == 'GET':
            if not item_type or not item_id:
                logger.error("Type and item_id are required")
                return jsonify({"error": "Type and item_id are required"}), 400
            if item_type not in ['item', 'addon', 'combo']:
                logger.error(f"Invalid type: {item_type}")
                return jsonify({"error": "Invalid type, must be 'item', 'addon', or 'combo'"}), 400
            item = items_collection.find_one({'_id': ObjectId(item_id)})
            if not item:
                logger.warning(f"Item not found: {item_id}")
                return jsonify({"error": "Item not found"}), 404
            ingredients = []
            if item_type == 'item':
                ingredients = item.get('ingredients', [])
            elif item_type == 'addon':
                if index is None or not index.isdigit():
                    logger.error("Index is required for addons")
                    return jsonify({"error": "Index is required for addons"}), 400
                index = int(index)
                if index < len(item.get('addons', [])):
                    ingredients = item['addons'][index].get('ingredients', [])
            elif item_type == 'combo':
                if index is None or not index.isdigit():
                    logger.error("Index is required for combos")
                    return jsonify({"error": "Index is required for combos"}), 400
                index = int(index)
                if index < len(item.get('combos', [])):
                    ingredients = item['combos'][index].get('ingredients', [])
            response_data = {'ingredients': ingredients, 'nutrition': {}}
            logger.info(f"Fetched nutrition data for {item_type}: {item_name}")
            return jsonify(response_data), 200
        if request.method == 'DELETE':
            instances = request.headers.get('X-Instances')
            if not instances:
                logger.error("Instances header is required for DELETE")
                return jsonify({"error": "Instances header is required"}), 400
            try:
                instances = json.loads(instances)
            except json.JSONDecodeError:
                logger.error("Invalid instances header format")
                return jsonify({"error": "Invalid instances header format"}), 400
            if item_type not in ['item', 'addon', 'combo']:
                logger.error(f"Invalid type: {item_type}")
                return jsonify({"error": "Invalid type, must be 'item', 'addon', or 'combo'"}), 400
            deleted_count = 0
            for instance in instances:
                item_id = instance['item_id']
                index = instance.get('index')
                item = items_collection.find_one({'_id': ObjectId(item_id)})
                if not item:
                    logger.warning(f"Item not found: {item_id}")
                    continue
                update_query = {'_id': ObjectId(item_id)}
                update_data = {'$unset': {}, '$set': {'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
                if item_type == 'item':
                    update_data['$unset']['ingredients'] = ''
                elif item_type == 'addon':
                    if index is None or not isinstance(index, int):
                        logger.error(f"Index is required for addons in item_id: {item_id}")
                        continue
                    update_data['$unset'][f'addons.{index}.ingredients'] = ''
                elif item_type == 'combo':
                    if index is None or not isinstance(index, int):
                        logger.error(f"Index is required for combos in item_id: {item_id}")
                        continue
                    update_data['$unset'][f'combos.{index}.ingredients'] = ''
                result = items_collection.update_one(update_query, update_data)
                if result.matched_count > 0:
                    deleted_count += 1
            if deleted_count == 0:
                logger.error(f"No items updated for deletion of {item_type}: {item_name}")
                return jsonify({"error": "No items updated for deletion, please check instance data"}), 400
            logger.info(f"Nutrition cleared for {item_type}: {item_name} across {deleted_count} instances")
            return jsonify({"message": "Nutrition and ingredients cleared successfully"}), 200
    except Exception as e:
        logger.error(f"Error handling nutrition for {item_name}: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500
@app.route('/api/kitchen-saved', methods=['POST'])
@db_required
def save_kitchen_order():
    try:
        data = request.get_json()
        if not data or 'orderId' not in data:
            return jsonify({'success': False, 'error': 'No data or orderId provided'}), 400
        order_id = data['orderId']
        existing_order = kitchen_saved_collection.find_one({'orderId': order_id})
       
        cart_items = data.get('cartItems', [])
        for item in cart_items:
            required_kitchens = set()
            if 'kitchen' in item:
                required_kitchens.add(item['kitchen'])
            for addon_name, qty in item.get('addonQuantities', {}).items():
                if qty > 0 and 'addonVariants' in item and addon_name in item['addonVariants']:
                    addon = item['addonVariants'][addon_name]
                    if 'kitchen' in addon:
                        required_kitchens.add(addon['kitchen'])
            for combo_name, qty in item.get('comboQuantities', {}).items():
                if qty > 0 and 'comboVariants' in item and combo_name in item['comboVariants']:
                    combo = item['comboVariants'][combo_name]
                    if 'kitchen' in combo:
                        required_kitchens.add(combo['kitchen'])
            item['requiredKitchens'] = list(required_kitchens)
            item['kitchenStatuses'] = item.get('kitchenStatuses', {kitchen: 'Pending' for kitchen in required_kitchens})
        order = {
            'orderId': order_id,
            'customerName': data.get('customerName', 'N/A'),
            'tableNumber': data.get('tableNumber', 'N/A'),
            'chairsBooked': data.get('chairsBooked', []),
            'phoneNumber': data.get('phoneNumber', ''),
            'deliveryAddress': data.get('deliveryAddress', {}),
            'whatsappNumber': data.get('whatsappNumber', ''),
            'email': data.get('email', ''),
            'cartItems': cart_items,
            'timestamp': data.get('timestamp', datetime.utcnow().isoformat()),
            'orderType': data.get('orderType', 'Dine In'),
            'createdAt': datetime.utcnow(),
            'status': data.get('status', 'Pending'),
            'pickedUpTime': data.get('pickedUpTime', None)
        }
        if existing_order:
            kitchen_saved_collection.update_one(
                {'orderId': order_id},
                {'$set': order}
            )
            logger.info(f"Updated kitchen order: {order_id}")
        else:
            kitchen_saved_collection.insert_one(order)
            logger.info(f"Created kitchen order: {order_id}")
        return jsonify({'success': True, 'order_id': order_id}), 201
    except Exception as e:
        logger.error(f"Error in /api/kitchen-saved POST: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/api/kitchen-saved', methods=['GET'])
@db_required
def get_kitchen_orders():
    try:
        orders = list(kitchen_saved_collection.find({}, {'_id': 0}))
        return jsonify({'success': True, 'orders': orders}), 200
    except Exception as e:
        logger.error(f"Error in /api/kitchen-saved GET: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/api/kitchen-saved/<order_id>', methods=['DELETE'])
@db_required
def delete_kitchen_order(order_id):
    try:
        result = kitchen_saved_collection.delete_one({'orderId': order_id})
        if result.deleted_count == 0:
            logger.warning(f"Order not found: {order_id}")
            return jsonify({'success': False, 'error': 'Order not found'}), 404
        logger.info(f"Order deleted: {order_id}")
        return jsonify({'success': True, 'message': 'Order deleted successfully'}), 200
    except Exception as e:
        logger.error(f"Error deleting order {order_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/api/kitchen-saved/<order_id>/items/<item_id>/mark-prepared', methods=['POST'])
@db_required
def mark_item_prepared(order_id, item_id):
    try:
        data = request.get_json()
        kitchen = data.get('kitchen')
        if not kitchen:
            return jsonify({'success': False, 'error': 'Kitchen not provided'}), 400
        order = kitchen_saved_collection.find_one({'orderId': order_id})
        if not order:
            return jsonify({'success': False, 'error': 'Order not found'}), 404
        item = next((item for item in order['cartItems'] if item['id'] == item_id), None)
        if not item:
            return jsonify({'success': False, 'error': 'Item not found'}), 404
        if not item.get('requiredKitchens') or kitchen not in item['requiredKitchens']:
            return jsonify({'success': False, 'error': 'Kitchen not required for this item'}), 400
        if not item.get('kitchenStatuses'):
            item['kitchenStatuses'] = {k: 'Pending' for k in item['requiredKitchens']}
        if item['kitchenStatuses'][kitchen] in ['Prepared', 'PickedUp']:
            return jsonify({'success': False, 'error': 'Kitchen already marked as prepared or picked up'}), 400
        item['kitchenStatuses'][kitchen] = 'Prepared'
        kitchen_saved_collection.update_one(
            {'orderId': order_id, 'cartItems.id': item_id},
            {'$set': {'cartItems.$.kitchenStatuses': item['kitchenStatuses']}}
        )
        activeorders_collection.update_one(
            {'orderId': order_id, 'cartItems.id': item_id},
            {'$set': {'cartItems.$.kitchenStatuses': item['kitchenStatuses']}}
        )
        return jsonify({'success': True, 'status': 'Prepared'}), 200
    except Exception as e:
        logger.error(f"Error in /api/kitchen-saved/{order_id}/items/{item_id}/mark-prepared: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
# Save picked-up item (POST /api/picked-up-items)
@app.route('/api/picked-up-items', methods=['POST'])
@db_required
def save_picked_up_item():
    try:
        item_data = request.get_json()
        if not item_data:
            logger.error("No data provided for picked-up items")
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        customer_name = item_data.get('customerName', 'Unknown')
        table_number = item_data.get('tableNumber', 'N/A')
        pickup_time = datetime.utcnow().isoformat()
        new_item = {
            'itemName': item_data.get('itemName', 'Unknown'),
            'quantity': item_data.get('quantity', 0),
            'category': item_data.get('category', 'N/A'),
            'kitchen': item_data.get('kitchen', 'Unknown'),
            'addonCounts': item_data.get('addonCounts', []),
            'selectedCombos': item_data.get('selectedCombos', [])
        }
        existing_entry = picked_up_collection.find_one({
            'customerName': customer_name,
            'tableNumber': table_number
        })
        if existing_entry:
            updated_items = existing_entry.get('items', [])
            updated_items.append(new_item)
            result = picked_up_collection.update_one(
                {'_id': existing_entry['_id']},
                {
                    '$set': {
                        'items': updated_items,
                        'pickupTime': pickup_time,
                        'modified_at': datetime.utcnow().isoformat()
                    }
                }
            )
            logger.info(f"Picked-up items updated for customer: {customer_name}, table: {table_number}")
            return jsonify({
                'success': True,
                'message': 'Picked-up items updated successfully',
                'id': str(existing_entry['_id'])
            }), 200
        else:
            picked_up_data = {
                'customerName': customer_name,
                'tableNumber': table_number,
                'items': [new_item],
                'pickupTime': pickup_time,
                'created_at': datetime.utcnow().isoformat(),
                'orderType': item_data.get('orderType', 'N/A')
            }
            result = picked_up_collection.insert_one(picked_up_data)
            logger.info(f"Picked-up items saved with ID: {result.inserted_id}")
            return jsonify({
                'success': True,
                'message': 'Picked-up items saved successfully',
                'id': str(result.inserted_id)
            }), 201
    except Exception as e:
        logger.error(f"Error saving picked-up items: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500
# Fetch all picked-up items (GET /api/picked-up-items)
@app.route('/api/picked-up-items', methods=['GET'])
@db_required
def get_picked_up_items():
    try:
        picked_up_items = list(picked_up_collection.find({}))
        # Convert ObjectId to string for JSON serialization
        picked_up_items = convert_objectid_to_str(picked_up_items)
        logger.info(f"Fetched {len(picked_up_items)} picked-up item entries")
        return jsonify({'success': True, 'pickedUpItems': picked_up_items}), 200
    except Exception as e:
        logger.error(f"Error fetching picked-up items: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500
# Delete a picked-up item entry (DELETE /api/picked-up-items/<entry_id>)
@app.route('/api/picked-up-items/<entry_id>', methods=['DELETE'])
@db_required
def delete_picked_up_item(entry_id):
    try:
        result = picked_up_collection.delete_one({'_id': ObjectId(entry_id)})
        if result.deleted_count == 0:
            logger.warning(f"Picked-up entry not found: {entry_id}")
            return jsonify({"error": "Picked-up entry not found"}), 404
        logger.info(f"Picked-up entry deleted: {entry_id}")
        return jsonify({"message": "Picked-up entry deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting picked-up entry {entry_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500
@app.route('/api/variants', methods=['POST'])
@db_required
def create_variants():
    try:
        # Get the variant data from the request
        data = request.get_json()
        if not data or not data.get('heading') or not isinstance(data.get('subheadings'), list):
            return jsonify({'error': 'Variant must have a heading and a list of subheadings'}), 400
        # Validate data
        for subheading in data['subheadings']:
            if not subheading.get('name'):
                return jsonify({'error': 'Each subheading must have a name'}), 400
            # Validate price if provided
            if 'price' in subheading and subheading['price'] is not None:
                try:
                    subheading['price'] = float(subheading['price'])
                except (ValueError, TypeError):
                    return jsonify({'error': f"Invalid price for subheading {subheading['name']}"}), 400
            # Validate image if provided
            if 'image' in subheading and subheading['image'] is not None:
                if not isinstance(subheading['image'], str):
                    return jsonify({'error': f"Image for subheading {subheading['name']} must be a string"}), 400
            # Validate dropdown if provided
            if 'dropdown' in subheading and not isinstance(subheading['dropdown'], bool):
                return jsonify({'error': f"Dropdown for subheading {subheading['name']} must be a boolean"}), 400
        # Insert variant into MongoDB
        result = variants_collection.insert_one(data)
        return jsonify({
            'message': 'Variant created successfully',
            'inserted_id': str(result.inserted_id)
        }), 201
    except Exception as e:
        return jsonify({'error': f"Server error: {str(e)}"}), 500
# Route to get all variants
@app.route('/api/variants', methods=['GET'])
@db_required
def get_variants():
    try:
        placeholder_url = 'https://placehold.co/100x100/EFEFEF/AAAAAA?text=No+Image'
        variants = list(variants_collection.find({}, {'_id': 1, 'heading': 1, 'subheadings': 1, 'activeSection': 1}))
       
        for variant in variants:
            variant['_id'] = str(variant['_id'])
            for subheading in variant.get('subheadings', []):
                # Updated to use /api/images/
                if subheading.get('image'):
                    subheading['image'] = f"/api/images/{os.path.basename(subheading['image'])}"
                else:
                    subheading['image'] = placeholder_url
                   
        return jsonify(variants), 200
    except Exception as e:
        return jsonify({'error': f"Server error: {str(e)}"}), 500
# Route to get a specific variant by ID
@app.route('/api/variants/<id>', methods=['GET'])
@db_required
def get_variant(id):
    try:
        # Validate ObjectId
        if not ObjectId.is_valid(id):
            return jsonify({'error': 'Invalid variant ID'}), 400
        variant = variants_collection.find_one({'_id': ObjectId(id)}, {'_id': 0, 'heading': 1, 'subheadings': 1, 'activeSection': 1})
        if not variant:
            return jsonify({'error': 'Variant not found'}), 404
        return jsonify(variant), 200
    except Exception as e:
        return jsonify({'error': f"Server error: {str(e)}"}), 500
# Route to update a variant
@app.route('/api/variants/<id>', methods=['PUT'])
@db_required
def update_variant(id):
    try:
        # Validate ObjectId
        if not ObjectId.is_valid(id):
            return jsonify({'error': 'Invalid variant ID'}), 400
        data = request.get_json()
        if not data or not data.get('heading') or not isinstance(data.get('subheadings'), list):
            return jsonify({'error': 'Variant must have a heading and a list of subheadings'}), 400
        # Validate data
        for subheading in data['subheadings']:
            if not subheading.get('name'):
                return jsonify({'error': 'Each subheading must have a name'}), 400
            if 'price' in subheading and subheading['price'] is not None:
                try:
                    subheading['price'] = float(subheading['price'])
                except (ValueError, TypeError):
                    return jsonify({'error': f"Invalid price for subheading {subheading['name']}"}), 400
            if 'image' in subheading and subheading['image'] is not None:
                if not isinstance(subheading['image'], str):
                    return jsonify({'error': f"Image for subheading {subheading['name']} must be a string"}), 400
            if 'dropdown' in subheading and not isinstance(subheading['dropdown'], bool):
                return jsonify({'error': f"Dropdown for subheading {subheading['name']} must be a boolean"}), 400
        # Update variant in MongoDB
        result = variants_collection.update_one(
            {'_id': ObjectId(id)},
            {'$set': data}
        )
        if result.matched_count == 0:
            return jsonify({'error': 'Variant not found'}), 404
        return jsonify({'message': 'Variant updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Server error: {str(e)}"}), 500
# Route to delete a variant by ID
@app.route('/api/variants/<id>', methods=['DELETE'])
@db_required
def delete_variant(id):
    try:
        # Validate ObjectId
        if not ObjectId.is_valid(id):
            return jsonify({'error': 'Invalid variant ID'}), 400
        result = variants_collection.delete_one({'_id': ObjectId(id)})
        if result.deleted_count == 0:
            return jsonify({'error': 'Variant not found'}), 404
        return jsonify({'message': 'Variant deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Server error: {str(e)}"}), 500
# Route to delete a variant by heading
@app.route('/api/variants/heading/<heading>', methods=['DELETE'])
@db_required
def delete_variant_by_heading(heading):
    try:
        result = variants_collection.delete_one({'heading': heading})
        return jsonify({'message': 'Variant deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Server error: {str(e)}"}), 500
VALID_COUNTRY_CODES = [
    '+91', # India
    '+1', # USA
    '+971', # UAE (Dubai)
    '+44', # UK
    '+61', # Australia
    # Add more country codes as needed
]
def generate_employee_id():
    """Generate a unique employee ID."""
    return str(uuid.uuid4())[:8] # Simple 8-character UUID for employee ID
def validate_email(email):
    """Validate email format."""
    import re
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(pattern, email) is not None
@app.route('/api/employees', methods=['GET'])
@db_required
def get_employees():
    try:
        employees = list(employees_collection.find({}, {'_id': 0}))
        logger.info(f"Fetched {len(employees)} employees")
        return jsonify(employees), 200
    except Exception as e:
        logger.error(f"Error fetching employees: {str(e)}")
        return jsonify({'error': str(e)}), 500
@app.route('/api/employees', methods=['POST'])
@db_required
def create_employee():
    try:
        data = request.get_json()
        required_fields = ['name', 'phoneNumber', 'vehicleNumber', 'role', 'email']
        if not data or not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400
        phone_number = data['phoneNumber']
        if not any(phone_number.startswith(code) for code in VALID_COUNTRY_CODES):
            return jsonify({'error': 'Phone number must include a valid country code (e.g., +91, +1, +971)'}), 400
        code_length = len(next(code for code in VALID_COUNTRY_CODES if phone_number.startswith(code)))
        if len(phone_number) < code_length + 7:
            return jsonify({'error': 'Phone number is too short'}), 400
        email = data['email']
        if not validate_email(email):
            return jsonify({'error': 'Invalid email format'}), 400
        # Check if email is unique
        if employees_collection.find_one({'email': email}):
            return jsonify({'error': 'Email already exists'}), 400
        employee_id = generate_employee_id()
        employee = {
            'employeeId': employee_id,
            'name': data['name'],
            'phoneNumber': phone_number,
            'vehicleNumber': data['vehicleNumber'],
            'role': data['role'],
            'email': email
        }
        # Insert employee into employees_collection
        employees_collection.insert_one(employee)
        # Ensure email exists in users_collection for /api/sales compatibility
        if not users_collection.find_one({'email': email}):
            users_collection.insert_one({
                'email': email,
                'name': data['name'],
                'role': data['role'],
                'created_at': datetime.now(ZoneInfo("UTC")).isoformat()
            })
        employee.pop('_id', None)
        logger.info(f"Created employee: {employee_id} with email: {email}")
        return jsonify({'message': 'Employee created successfully', 'employee': employee}), 201
    except Exception as e:
        logger.error(f"Error creating employee: {str(e)}")
        return jsonify({'error': str(e)}), 500
@app.route('/api/employees/<employee_id>', methods=['PUT'])
@db_required
def update_employee(employee_id):
    try:
        data = request.get_json()
        required_fields = ['name', 'phoneNumber', 'vehicleNumber', 'role', 'email']
        if not data or not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400
        phone_number = data['phoneNumber']
        if not any(phone_number.startswith(code) for code in VALID_COUNTRY_CODES):
            return jsonify({'error': 'Phone number must include a valid country code (e.g., +91, +1, +971)'}), 400
        code_length = len(next(code for code in VALID_COUNTRY_CODES if phone_number.startswith(code)))
        if len(phone_number) < code_length + 7:
            return jsonify({'error': 'Phone number is too short'}), 400
        email = data['email']
        if not validate_email(email):
            return jsonify({'error': 'Invalid email format'}), 400
        # Check if email is unique (excluding current employee)
        existing_employee = employees_collection.find_one({'email': email, 'employeeId': {'$ne': employee_id}})
        if existing_employee:
            return jsonify({'error': 'Email already exists'}), 400
        updated_employee = {
            'name': data['name'],
            'phoneNumber': phone_number,
            'vehicleNumber': data['vehicleNumber'],
            'role': data['role'],
            'email': email
        }
        result = employees_collection.update_one(
            {'employeeId': employee_id},
            {'$set': updated_employee}
        )
        if result.matched_count == 0:
            return jsonify({'error': 'Employee not found'}), 404
        # Update or insert user in users_collection
        users_collection.update_one(
            {'email': email},
            {'$set': {
                'email': email,
                'name': data['name'],
                'role': data['role'],
                'updated_at': datetime.now(ZoneInfo("UTC")).isoformat()
            }},
            upsert=True
        )
        logger.info(f"Updated employee: {employee_id} with email: {email}")
        return jsonify({'message': 'Employee updated successfully'}), 200
    except Exception as e:
        logger.error(f"Error updating employee: {str(e)}")
        return jsonify({'error': str(e)}), 500
@app.route('/api/employees/<employee_id>', methods=['DELETE'])
@db_required
def delete_employee(employee_id):
    try:
        employee = employees_collection.find_one({'employeeId': employee_id}, {'_id': 0})
        if not employee:
            return jsonify({'error': 'Employee not found'}), 404
        result = employees_collection.delete_one({'employeeId': employee_id})
        if result.deleted_count == 0:
            return jsonify({'error': 'Employee not found'}), 404
        # Optionally, remove from users_collection if no other employees use this email
        if not employees_collection.find_one({'email': employee['email']}):
            users_collection.delete_one({'email': employee['email']})
        logger.info(f"Deleted employee: {employee_id}")
        return jsonify({'message': 'Employee deleted successfully'}), 200
    except Exception as e:
        logger.error(f"Error deleting employee: {str(e)}")
        return jsonify({'error': str(e)}), 500
def generate_unique_id():
    return str(uuid.uuid4())
def generate_order_number(order_type):
    """Generate order number based on order type (e.g., T0001, D0001, ON001)."""
    prefix = {'Dine In': 'D', 'Take Away': 'T', 'Online Delivery': 'ON'}.get(order_type, 'D')
    counter = order_counters_collection.find_one_and_update(
        {'order_type': order_type},
        {'$inc': {'counter': 1}},
        upsert=True,
        return_document=True
    )
    number = counter['counter']
    if prefix == 'ON':
        return f'{prefix}{number:03d}' # e.g., ON001
    return f'{prefix}{number:04d}' # e.g., D0001, T0001
# Route to fetch a specific active order by orderId
@app.route('/api/activeorders/<order_id>', methods=['GET'])
@db_required
def get_active_order(order_id):
    try:
        order = activeorders_collection.find_one({'orderId': order_id}, {'_id': 0})
        if order:
            logger.info(f"Fetched order: {order_id}")
            return jsonify(order), 200
        else:
            logger.warning(f"Order not found: {order_id}")
            return jsonify({'error': 'Order not found'}), 404
    except Exception as e:
        logger.error(f"Error fetching order: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
# Route to create a new active order
@app.route('/api/activeorders', methods=['POST'])
@db_required
def save_active_order():
    try:
        data = request.get_json()
        order_id = generate_unique_id()
        order_type = data.get('orderType', 'Dine In')
        order_no = generate_order_number(order_type)
        cart_items = data.get('cartItems', [])
        for item in cart_items:
            required_kitchens = set()
            if item.get('kitchen'):
                required_kitchens.add(item['kitchen'])
           
            for addon_name, qty in item.get('addonQuantities', {}).items():
                if qty > 0 and 'addonVariants' in item and addon_name in item['addonVariants']:
                    if item['addonVariants'][addon_name].get('kitchen'):
                        required_kitchens.add(item['addonVariants'][addon_name]['kitchen'])
           
            for combo_name, qty in item.get('comboQuantities', {}).items():
                if qty > 0 and 'comboVariants' in item and combo_name in item['comboVariants']:
                    if item['comboVariants'][combo_name].get('kitchen'):
                        required_kitchens.add(item['comboVariants'][combo_name]['kitchen'])
            item['requiredKitchens'] = list(required_kitchens)
            item['kitchenStatuses'] = {kitchen: 'Pending' for kitchen in required_kitchens}
        active_order = {
            'orderId': order_id,
            'orderNo': order_no,
            'customerName': data.get('customerName', 'N/A'),
            'tableNumber': data.get('tableNumber', 'N/A'),
            'chairsBooked': data.get('chairsBooked', []),
            'phoneNumber': data.get('phoneNumber', ''),
            'deliveryAddress': data.get('deliveryAddress', {}),
            'whatsappNumber': data.get('whatsappNumber', ''),
            'email': data.get('email', ''),
            'cartItems': cart_items,
            'timestamp': data.get('timestamp', datetime.utcnow().isoformat()),
            'orderType': order_type,
            'status': data.get('status', 'Pending'),
            'created_at': datetime.utcnow(),
            'deliveryPersonId': data.get('deliveryPersonId', ''),
            'deliveryPersonName': data.get('deliveryPersonName', ''),
            'pickedUpTime': None
        }
        activeorders_collection.insert_one(active_order)
        kitchen_saved_collection.insert_one(active_order.copy())
        logger.info(f"Created order: {order_id} with order number: {order_no}")
        return jsonify({'success': True, 'orderId': order_id, 'orderNo': order_no}), 201
    except Exception as e:
        logger.error(f"Error saving active order: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
# Route to fetch all active orders
@app.route('/api/activeorders', methods=['GET'])
@db_required
def get_active_orders():
    try:
        orders = list(activeorders_collection.find({}, {'_id': 0}))
        logger.info(f"Fetched {len(orders)} active orders")
        return jsonify(orders), 200
    except Exception as e:
        logger.error(f"Error fetching active orders: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
# Route to mark an item as prepared
@app.route('/api/activeorders/<order_id>/items/<item_id>/mark-prepared', methods=['POST'])
@db_required
def mark_item_prepared_active(order_id, item_id):
    try:
        data = request.get_json()
        kitchen = data.get('kitchen')
        if not kitchen:
            return jsonify({'success': False, 'error': 'Kitchen not provided'}), 400
        for collection in [activeorders_collection, kitchen_saved_collection]:
            collection.update_one(
                {'orderId': order_id, 'cartItems.id': item_id},
                {'$set': {f'cartItems.$.kitchenStatuses.{kitchen}': 'Prepared'}}
            )
       
        logger.info(f"Marked item {item_id} in order {order_id} as Prepared for kitchen {kitchen}")
        return jsonify({'success': True, 'status': 'Prepared'}), 200
    except Exception as e:
        logger.error(f"Error in mark-prepared: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
# Route to mark an item as picked up
@app.route('/api/activeorders/<order_id>/items/<item_id>/mark-pickedup', methods=['POST'])
@db_required
def mark_item_pickedup_active(order_id, item_id):
    try:
        data = request.get_json()
        kitchen = data.get('kitchen')
        if not kitchen:
            return jsonify({'success': False, 'error': 'Kitchen not provided'}), 400
        order = activeorders_collection.find_one({'orderId': order_id}, {'_id': 0})
        if not order:
            return jsonify({'success': False, 'error': 'Order not found'}), 404
        item = next((item for item in order['cartItems'] if item['id'] == item_id), None)
        if not item:
            return jsonify({'success': False, 'error': 'Item not found'}), 404
       
        if item.get('kitchenStatuses', {}).get(kitchen) != 'Prepared':
            return jsonify({'success': False, 'error': 'Item must be prepared before picking up'}), 400
        for collection in [activeorders_collection, kitchen_saved_collection]:
            collection.update_one(
                {'orderId': order_id, 'cartItems.id': item_id},
                {'$set': {f'cartItems.$.kitchenStatuses.{kitchen}': 'PickedUp'}}
            )
        picked_up_data = {
            'customerName': order.get('customerName', 'Unknown'),
            'tableNumber': order.get('tableNumber', 'N/A'),
            'orderType': order.get('orderType', 'Dine In'),
            'pickupTime': datetime.utcnow().isoformat(),
            'items': [{
                'itemName': item.get('name', 'Unknown'),
                'quantity': item.get('quantity', 0),
                'category': item.get('category', 'N/A'),
                'kitchen': kitchen,
                'addonCounts': [
                    {'name': name, 'quantity': qty}
                    for name, qty in item.get('addonQuantities', {}).items()
                    if qty > 0 and item.get('addonVariants', {}).get(name, {}).get('kitchen') == kitchen
                ],
                'selectedCombos': [
                    {'name': name, 'size': item['comboVariants'][name]['size'], 'quantity': qty}
                    for name, qty in item.get('comboQuantities', {}).items()
                    if qty > 0 and item.get('comboVariants', {}).get(name, {}).get('kitchen') == kitchen
                ]
            }]
        }
        picked_up_collection.insert_one(picked_up_data)
       
        logger.info(f"Marked item {item_id} in order {order_id} as PickedUp for kitchen {kitchen}")
        return jsonify({'success': True, 'status': 'PickedUp'}), 200
    except Exception as e:
        logger.error(f"Error in mark-pickedup: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
# Route to update an active order
@app.route('/api/activeorders/<order_id>', methods=['PUT'])
@db_required
def update_active_order(order_id):
    try:
        data = request.get_json()
        if '_id' in data:
            del data['_id']
        order_in_db = activeorders_collection.find_one({'orderId': order_id}, {'_id': 0})
        if not order_in_db:
            logger.warning(f"Order not found for update: {order_id}")
            return jsonify({'error': 'Order not found'}), 404
        old_statuses_map = {
            item['id']: item.get('kitchenStatuses', {})
            for item in order_in_db.get('cartItems', []) if 'id' in item
        }
        if 'cartItems' in data:
            for item in data['cartItems']:
                required_kitchens = set()
                if item.get('kitchen'):
                    required_kitchens.add(item['kitchen'])
                for addon_name, qty in item.get('addonQuantities', {}).items():
                    if qty > 0 and item.get('addonVariants', {}).get(addon_name, {}).get('kitchen'):
                        required_kitchens.add(item['addonVariants'][addon_name]['kitchen'])
                for combo_name, qty in item.get('comboQuantities', {}).items():
                    if qty > 0 and item.get('comboVariants', {}).get(combo_name, {}).get('kitchen'):
                        required_kitchens.add(item['comboVariants'][combo_name]['kitchen'])
               
                item['requiredKitchens'] = list(required_kitchens)
                item_id = item.get('id')
                old_item_statuses = old_statuses_map.get(item_id, {})
               
                new_kitchen_statuses = {}
                for kitchen in required_kitchens:
                    if kitchen in old_item_statuses:
                        new_kitchen_statuses[kitchen] = old_item_statuses[kitchen]
                    else:
                        new_kitchen_statuses[kitchen] = 'Pending'
               
                item['kitchenStatuses'] = new_kitchen_statuses
        if 'deliveryPersonId' in data and data['deliveryPersonId']:
            employee = employees_collection.find_one({'employeeId': data['deliveryPersonId']}, {'_id': 0})
            if not employee:
                logger.warning(f"Delivery person not found: {data['deliveryPersonId']}")
                return jsonify({'error': 'Delivery person not found'}), 404
            # Save to trip reports before deleting
            trip_report = {
                'tripId': generate_unique_id(),
                'orderId': order_in_db['orderId'],
                'orderNo': order_in_db['orderNo'],
                'customerName': order_in_db.get('customerName', 'N/A'),
                'tableNumber': order_in_db.get('tableNumber', 'N/A'),
                'chairsBooked': order_in_db.get('chairsBooked', []),
                'phoneNumber': order_in_db.get('phoneNumber', ''),
                'deliveryAddress': order_in_db.get('deliveryAddress', {}),
                'whatsappNumber': order_in_db.get('whatsappNumber', ''),
                'email': order_in_db.get('email', ''),
                'cartItems': order_in_db.get('cartItems', []),
                'timestamp': order_in_db.get('timestamp', datetime.utcnow().isoformat()),
                'orderType': order_in_db.get('orderType', 'Dine In'),
                'status': order_in_db.get('status', 'Pending'),
                'deliveryPersonId': data['deliveryPersonId'],
                'deliveryPersonName': data.get('deliveryPersonName', employee.get('name', 'N/A')),
                'pickedUpTime': order_in_db.get('pickedUpTime', None),
                'paymentMethods': order_in_db.get('paymentMethods', []),
                'cardDetails': order_in_db.get('cardDetails', ''),
                'upiDetails': order_in_db.get('upiDetails', ''),
                'created_at': datetime.utcnow()
            }
            tripreports_collection.insert_one(trip_report)
            logger.info(f"Saved trip report for order {order_id} with delivery person {data['deliveryPersonId']}")
            activeorders_collection.delete_one({'orderId': order_id})
            kitchen_saved_collection.delete_one({'orderId': order_id})
            logger.info(f"Deleted order {order_id} from active orders after delivery person assignment")
            return jsonify({'success': True, 'message': 'Delivery person assigned and order moved to trip reports', 'order': order_in_db}), 200
        result = activeorders_collection.update_one({'orderId': order_id}, {'$set': data})
        kitchen_result = kitchen_saved_collection.update_one({'orderId': order_id}, {'$set': data})
        updated_order = activeorders_collection.find_one({'orderId': order_id}, {'_id': 0})
        if result.modified_count > 0 or kitchen_result.modified_count > 0:
            logger.info(f"Updated order: {order_id}")
            return jsonify({'success': True, 'message': 'Order updated', 'order': updated_order}), 200
       
        logger.info(f"No changes made to order: {order_id}")
        return jsonify({'success': True, 'message': 'No changes made', 'order': updated_order}), 200
    except Exception as e:
        logger.error(f"Error updating active order: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
# Route to delete an active order
@app.route('/api/activeorders/<order_id>', methods=['DELETE'])
@db_required
def delete_order(order_id):
    try:
        result = activeorders_collection.delete_one({'orderId': order_id})
        kitchen_result = kitchen_saved_collection.delete_one({'orderId': order_id})
        if result.deleted_count > 0 or kitchen_result.deleted_count > 0:
            logger.info(f"Deleted order: {order_id}")
            return jsonify({'success': True}), 200
        logger.warning(f"Order not found for deletion: {order_id}")
        return jsonify({'error': 'Order not found'}), 404
    except Exception as e:
        logger.error(f"Error deleting order: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
# Route to fetch trip reports for an employee
@app.route('/api/tripreports/<employee_id>', methods=['GET'])
@db_required
def get_trip_reports(employee_id):
    try:
        trip_reports = list(tripreports_collection.find({'deliveryPersonId': employee_id}, {'_id': 0}))
        logger.info(f"Fetched {len(trip_reports)} trip reports for employee: {employee_id}")
        return jsonify(trip_reports), 200
    except Exception as e:
        logger.error(f"Error fetching trip reports: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
shutdown_flag = False
# Shutdown endpoint
@app.route('/api/shutdown', methods=['POST'])
@db_required
def shutdown():
    global shutdown_flag
    logger.info("Shutdown requested")
   
    try:
        # For development (Werkzeug server)
        func = request.environ.get('werkzeug.server.shutdown')
        if func:
            func()
            logger.info("Werkzeug server shutdown initiated")
            return jsonify({"message": "Server shutting down"}), 200
       
        # For production (Waitress server)
        shutdown_flag = True
        logger.info("Setting shutdown flag for Waitress")
       
        # Start a thread to exit the process after a short delay
        def exit_process():
            time.sleep(1) # Reduced delay to 1 second for faster shutdown
            logger.info("Exiting Python process")
            os._exit(0) # Forcefully exit the process
       
        threading.Thread(target=exit_process, daemon=True).start()
        return jsonify({"message": "Server shutting down"}), 200
   
    except Exception as e:
        logger.error(f"Shutdown error: {str(e)}")
        return jsonify({"message": "Error during shutdown", "error": str(e)}), 500
@app.route('/api/save-email-settings', methods=['POST'])
@db_required
def save_email_settings():
    """Save email settings to MongoDB."""
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        from_email = data.get('from_email')
       
        if not all([email, password, from_email]):
            return jsonify({"success": False, "error": "Missing required fields: email, password, from_email"}), 400
        # Update or insert email settings
        email_settings_collection.update_one(
            {},
            {
                '$set': {
                    'email': email,
                    'password': password,
                    'from_email': from_email,
                    'updated_at': datetime.now(ZoneInfo("UTC"))
                }
            },
            upsert=True
        )
        logger.info(f"Email settings saved for {email}")
        return jsonify({"success": True, "message": "Email settings saved successfully"}), 200
    except Exception as e:
        logger.error(f"Error saving email settings: {str(e)}")
        return jsonify({"success": False, "error": f"Failed to save email settings: {str(e)}"}), 500
@app.route('/api/get-email-settings', methods=['GET'])
@db_required
def get_email_settings():
    """Retrieve email settings from MongoDB."""
    try:
        settings = email_settings_collection.find_one({}, {'_id': 0, 'password': 0})
        if not settings:
            return jsonify({"success": False, "error": "No email settings found"}), 404
        return jsonify({"success": True, "email": settings.get('email'), "from_email": settings.get('from_email')}), 200
    except Exception as e:
        logger.error(f"Error retrieving email settings: {str(e)}")
        return jsonify({"success": False, "error": f"Failed to retrieve email settings: {str(e)}"}), 500
@app.route('/api/test-email-settings', methods=['POST'])
@db_required
def test_email_settings():
    """Test email settings by attempting to authenticate with SMTP."""
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        if not all([email, password]):
            return jsonify({"success": False, "error": "Missing required fields: email, password"}), 400
        # Attempt SMTP login
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(email, password)
        logger.info(f"SMTP authentication successful for {email}")
        return jsonify({"success": True, "message": "Email settings are valid"}), 200
    except smtplib.SMTPAuthenticationError as e:
        logger.error(f"SMTP Authentication Error during test: {str(e)}")
        return jsonify({"success": False, "error": "Invalid email or app password. Please check your credentials and ensure an App Password is used for Gmail."}), 401
    except smtplib.SMTPException as e:
        logger.error(f"SMTP Error during test: {str(e)}")
        return jsonify({"success": False, "error": f"SMTP error: {str(e)}"}), 500
    except Exception as e:
        logger.error(f"Unexpected error testing email settings: {str(e)}")
        return jsonify({"success": False, "error": f"Failed to test email settings: {str(e)}"}), 500
@app.route('/api/send-email', methods=['POST', 'OPTIONS'])
@db_required
def send_email():
    """Send an email with HTML content."""
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    try:
        data = request.get_json()
        logger.info(f"Received email request: {data}")
        if not data:
            logger.error("No data received in send-email request")
            return jsonify({"success": False, "message": "No data provided"}), 400
        to_email = data.get('to')
        subject = data.get('subject')
        html_content = data.get('html')
        if not all([to_email, subject, html_content]):
            logger.error("Missing required email fields")
            return jsonify({"success": False, "message": "Missing required fields: to, subject, html"}), 400
       
        # Fetch email settings
        settings = email_settings_collection.find_one()
        if not settings:
            logger.error("No email settings configured")
            return jsonify({"success": False, "message": "Email settings not configured. Please configure in Email Settings."}), 500
        email_user = settings.get('email')
        email_pass = settings.get('password')
        from_email = settings.get('from_email')
        msg = MIMEMultipart('alternative')
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(html_content, 'html'))
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(email_user, email_pass)
            server.send_message(msg)
            logger.info(f"Email sent successfully to {to_email}")
        return jsonify({"success": True, "message": "Email sent successfully"}), 200
    except smtplib.SMTPAuthenticationError as e:
        logger.error(f"SMTP Authentication Error: {str(e)}")
        return jsonify({"success": False, "message": "Invalid email or app password. Please check your Email Settings and ensure an App Password is used for Gmail."}), 401
    except smtplib.SMTPException as e:
        logger.error(f"SMTP Error: {str(e)}")
        return jsonify({"success": False, "message": f"SMTP error: {str(e)}"}), 500
    except Exception as e:
        logger.error(f"Unexpected error sending email: {str(e)}")
        return jsonify({"success": False, "message": f"Failed to send email: {str(e)}"}), 500
@app.route('/api/export-all-to-excel', methods=['GET'])
@db_required
def export_all_to_excel():
    """Export all data to an Excel file."""
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        collections = {
            'customers': customers_collection,
            'items': items_collection,
            'sales': sales_collection,
            'tables': tables_collection,
            'users': users_collection,
            'picked_up_items': picked_up_collection,
            'pos_opening_entries': opening_collection,
            'pos_closing_entries': pos_closing_collection,
            'system_settings': settings_collection,
            'kitchens': kitchens_collection,
            'item_groups': item_groups_collection
        }
        for collection_name, collection in collections.items():
            ws = wb.create_sheet(title=collection_name)
            data = list(collection.find())
            if not data:
                ws.append(['No data'])
                continue
            sample_doc = data[0]
            headers = list(sample_doc.keys())
            ws.append(headers)
            for doc in data:
                row = [str(doc.get(header, '')) if isinstance(doc.get(header), (ObjectId, list, dict)) else doc.get(header, '') for header in headers]
                ws.append(row)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f'restaurant_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        logger.info(f"Exported data to Excel: {filename}")
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500
def manage_backup_limit():
    """Manage the number of backup files to keep only the latest MAX_BACKUPS."""
    try:
        backup_files = [f for f in os.listdir(app.config['UPLOAD_FOLDER']) if f.endswith('.xlsx')]
        backup_files = sorted(
            backup_files,
            key=lambda x: os.path.getctime(os.path.join(app.config['UPLOAD_FOLDER'], x)),
            reverse=True
        )
        for old_file in backup_files[MAX_BACKUPS:]:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], old_file))
            logger.info(f"Deleted old old backup: {old_file}")
    except Exception as e:
        logger.error(f"Error managing backup limit: {str(e)}")
def create_backup():
    """Create a backup and send it via email using settings from MongoDB."""
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        collections = {
            'customers': customers_collection,
            'items': items_collection,
            'sales': sales_collection,
            'tables': tables_collection,
            'users': users_collection,
            'picked_up_items': picked_up_collection,
            'pos_opening_entries': opening_collection,
            'pos_closing_entries': pos_closing_collection,
            'system_settings': settings_collection,
            'kitchens': kitchens_collection,
            'item_groups': item_groups_collection
        }
        for collection_name, collection in collections.items():
            ws = wb.create_sheet(title=collection_name)
            data = list(collection.find())
            if not data:
                ws.append(['No data'])
                continue
            sample_doc = data[0]
            headers = list(sample_doc.keys())
            ws.append(headers)
            for doc in data:
                row = [str(doc.get(header, '')) if isinstance(doc.get(header), (ObjectId, list, dict)) else doc.get(header, '') for header in headers]
                ws.append(row)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'backup_restaurant_data_{timestamp}.xlsx'
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        with open(file_path, 'wb') as f:
            f.write(buffer.getvalue())
        manage_backup_limit()
        # Fetch email settings
        settings = email_settings_collection.find_one()
        if not settings:
            logger.error("No email settings configured")
            return False, "Email settings not configured. Please configure in Email Settings."
        email_user = settings.get('email')
        email_pass = settings.get('password')
        from_email = settings.get('from_email')
        # Send email with backup attachment
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = email_user
        msg['Subject'] = f'Restaurant Data Backup - {timestamp}'
        body = f'Backup of restaurant data generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}.'
        msg.attach(MIMEText(body, 'plain'))
        with open(file_path, 'rb') as f:
            attachment = MIMEBase('application', 'octet-stream')
            attachment.set_payload(f.read())
            encoders.encode_base64(attachment)
            attachment.add_header('Content-Disposition', f'attachment; filename={filename}')
            msg.attach(attachment)
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(email_user, email_pass)
            server.send_message(msg)
        logger.info(f"Backup created and emailed: {filename}")
        return True, f"Backup created successfully: {filename}"
    except smtplib.SMTPAuthenticationError as e:
        logger.error(f"SMTP Authentication Error: {str(e)}")
        return False, f"Invalid email or app password. Please check your Email Settings and ensure an App Password is used for Gmail."
    except smtplib.SMTPException as e:
        logger.error(f"SMTP Error: {str(e)}")
        return False, f"SMTP error: {str(e)}"
    except Exception as e:
        logger.error(f"Error in backup: {str(e)}")
        return False, str(e)
@app.route('/api/backup-to-excel', methods=['GET'])
@db_required
def backup_to_excel():
    """Create a backup and serve it as a download."""
    try:
        success, message = create_backup()
        if not success:
            return jsonify({"error": message}), 500
        filename = message.split(': ')[1]
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        with open(file_path, 'rb') as f:
            file_data = f.read()
        return Response(
            file_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        logger.error(f"Error serving backup file: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500
@app.route('/api/backup-info', methods=['GET'])
@db_required
def backup_info():
    """Retrieve information about existing backups."""
    try:
        backup_files = [f for f in os.listdir(app.config['UPLOAD_FOLDER']) if f.endswith('.xlsx')]
        backups = []
        for filename in backup_files:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            stat = os.stat(file_path)
            backups.append({
                'filename': filename,
                'date': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                'size': f"{stat.st_size / 1024:.2f} KB"
            })
        backups.sort(key=lambda x: x['date'], reverse=True)
        return jsonify(backups)
    except Exception as e:
        logger.error(f"Error retrieving backup info: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500
@app.route('/api/download-backup', methods=['POST'])
@db_required
def download_backup():
    """Download a specific backup file."""
    try:
        data = request.get_json()
        filename = data.get('filename')
        if not filename:
            return jsonify({"error": "Filename not provided"}), 400
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if not os.path.exists(file_path):
            return jsonify({"error": "Backup file not found"}), 404
        with open(file_path, 'rb') as f:
            file_data = f.read()
        return Response(
            file_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        logger.error(f"Error downloading backup {filename}: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500
def manage_offers():
    """Check all items and update offer status based on current time."""
    try:
        current_time = datetime.now(ZoneInfo("UTC"))
        items = items_collection.find({
            '$or': [
                {'offer_start_time': {'$exists': True}},
                {'offer_end_time': {'$exists': True}}
            ]
        })
        for item in items:
            item_id = item['_id']
            offer_start_time = item.get('offer_start_time')
            offer_end_time = item.get('offer_end_time')
            should_unset = False
            if offer_start_time and offer_end_time:
                try:
                    start_time = datetime.fromisoformat(str(offer_start_time).replace('Z', '+00:00'))
                    end_time = datetime.fromisoformat(str(offer_end_time).replace('Z', '+00:00'))
                    if current_time > end_time:
                        should_unset = True
                        logger.info(f"Offer expired for item {item.get('item_name')} (ID: {item_id})")
                    elif start_time > end_time:
                        should_unset = True
                        logger.warning(f"Invalid offer times for item {item_id}: start_time after end_time")
                    else:
                        logger.debug(f"Offer for item {item.get('item_name')} (ID: {item_id}) is active or pending")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid offer time format for item {item_id}: {str(e)}")
                    should_unset = True
            elif offer_end_time:
                try:
                    end_time = datetime.fromisoformat(str(offer_end_time).replace('Z', '+00:00'))
                    if current_time > end_time:
                        should_unset = True
                        logger.info(f"Offer expired for item {item.get('item_name')} (ID: {item_id})")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid offer_end_time for item {item_id}: {str(e)}")
                    should_unset = True
            if should_unset:
                items_collection.update_one(
                    {'_id': item_id},
                    {'$unset': {'offer_price': "", 'offer_start_time': "", 'offer_end_time': ""}}
                )
                logger.info(f"Unset offer fields for item {item.get('item_name')} (ID: {item_id})")
    except Exception as e:
        logger.error(f"Error in manage_offers: {str(e)}")
def schedule_tasks():
    """Schedule backups and offer management."""
    schedule.every(6).hours.do(create_backup)
    schedule.every(1).minutes.do(manage_offers)
    while True:
        schedule.run_pending()
        time.sleep(60)
def start_scheduler():
    scheduler_thread = threading.Thread(target=schedule_tasks, daemon=True)
    scheduler_thread.start()
    logger.info("Automatic backup and offer scheduler started")
def document_to_dict(doc):
    doc['_id'] = str(doc['_id'])
    if 'created_at' in doc and isinstance(doc['created_at'], datetime):
        doc['created_at'] = doc['created_at'].isoformat()
    if 'date' in doc and isinstance(doc['date'], datetime):
        doc['date'] = doc['date'].isoformat()
    return doc
# --- Purchase Items Routes ---
@app.route('/api/purchase_items', methods=['GET'])
@db_required
def get_purchase_items():
    try:
        items = list(purchase_items_collection.find())
        return jsonify([document_to_dict(item) for item in items]), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch items: {str(e)}"}), 500

@app.route('/api/purchase_items', methods=['POST'])
@db_required
def add_purchase_item():
    try:
        data = request.json
        required_fields = ['name', 'boxToMaster', 'masterUnit', 'outerUnit', 'nosUnit', 'masterToOuter', 'outerToNos']
        if not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400
        if not data['name'] or not data['masterUnit'] or not data['outerUnit'] or not data['nosUnit'] or float(data['boxToMaster']) <= 0 or float(data['masterToOuter']) <= 0 or float(data['outerToNos']) <= 0:
            return jsonify({'error': 'Invalid input data'}), 400
        
        conversion_factor = float(data['masterToOuter']) * float(data['outerToNos'])
        total_stock = float(data['boxToMaster']) * conversion_factor
        item = {
            'name': data['name'],
            'boxToMaster': float(data['boxToMaster']),
            'masterUnit': data['masterUnit'],
            'outerUnit': data['outerUnit'],
            'nosUnit': data['nosUnit'],
            'masterToOuter': float(data['masterToOuter']),
            'outerToNos': float(data['outerToNos']),
            'conversionFactor': conversion_factor,
            'stockMaster': float(data['boxToMaster']),
            'stockOuter': 0,
            'stockNos': 0,
            'soldNos': 0,
            'totalStock': total_stock,
            'created_at': datetime.utcnow()
        }
        result = purchase_items_collection.insert_one(item)
        item['_id'] = str(result.inserted_id)
        return jsonify({'message': 'Item added successfully', 'item': document_to_dict(item)}), 201
    except Exception as e:
        return jsonify({'error': f"Failed to add item: {str(e)}"}), 500

@app.route('/api/purchase_items/<id>', methods=['PUT'])
@db_required
def update_purchase_item(id):
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No input data provided'}), 400
        old_item = purchase_items_collection.find_one({'_id': ObjectId(id)})
        if not old_item:
            return jsonify({'error': 'Item not found'}), 404
        update_fields = {}
        if 'name' in data:
            update_fields['name'] = data['name']
        if 'boxToMaster' in data:
            update_fields['boxToMaster'] = float(data['boxToMaster'])
        if 'masterUnit' in data:
            update_fields['masterUnit'] = data['masterUnit']
        if 'outerUnit' in data:
            update_fields['outerUnit'] = data['outerUnit']
        if 'nosUnit' in data:
            update_fields['nosUnit'] = data['nosUnit']
        if 'masterToOuter' in data:
            update_fields['masterToOuter'] = float(data['masterToOuter'])
        if 'outerToNos' in data:
            update_fields['outerToNos'] = float(data['outerToNos'])
        change_conversion = 'masterToOuter' in data or 'outerToNos' in data
        change_box = 'boxToMaster' in data
        if change_conversion or change_box:
            mto = update_fields.get('masterToOuter', old_item['masterToOuter'])
            otn = update_fields.get('outerToNos', old_item['outerToNos'])
            conv = mto * otn
            update_fields['conversionFactor'] = conv
            stock_master = old_item['stockMaster']
            if change_box:
                old_box = old_item['boxToMaster']
                new_box = update_fields['boxToMaster']
                diff = new_box - old_box
                stock_master += diff
                update_fields['stockMaster'] = stock_master
            total_stock = stock_master * conv + old_item['stockOuter'] * otn + old_item['stockNos']
            update_fields['totalStock'] = total_stock
        if not update_fields:
            return jsonify({'error': 'No fields to update'}), 400
        result = purchase_items_collection.update_one({'_id': ObjectId(id)}, {'$set': update_fields})
        if result.modified_count == 0:
            return jsonify({'error': 'Item not found or no changes made'}), 404
        return jsonify({'message': 'Item updated successfully'}), 200
    except ValueError:
        return jsonify({'error': 'Invalid input data'}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to update item: {str(e)}"}), 500

@app.route('/api/purchase_items/<id>', methods=['DELETE'])
@db_required
def delete_purchase_item(id):
    try:
        result = purchase_items_collection.delete_one({'_id': ObjectId(id)})
        if result.deleted_count == 0:
            return jsonify({'error': 'Item not found'}), 404
        return jsonify({'message': 'Item deleted successfully'}), 200
    except ValueError:
        return jsonify({'error': 'Invalid item ID'}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to delete item: {str(e)}"}), 500

# --- Suppliers Routes ---
@app.route('/api/suppliers', methods=['GET'])
@db_required
def get_suppliers():
    try:
        suppliers = list(suppliers_collection.find())
        return jsonify([document_to_dict(supplier) for supplier in suppliers]), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch suppliers: {str(e)}"}), 500

@app.route('/api/suppliers', methods=['POST'])
@db_required
def add_supplier():
    try:
        data = request.json
        required_fields = ['name', 'shopName', 'address', 'phone', 'email']
        if not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400
        if not all(data[key] for key in required_fields):
            return jsonify({'error': 'Invalid input data'}), 400
        
        supplier = {
            'name': data['name'],
            'shopName': data['shopName'],
            'address': data['address'],
            'phone': data['phone'],
            'email': data['email'],
            'created_at': datetime.utcnow()
        }
        result = suppliers_collection.insert_one(supplier)
        supplier['_id'] = str(result.inserted_id)
        return jsonify({'message': 'Supplier added successfully', 'supplier': document_to_dict(supplier)}), 201
    except Exception as e:
        return jsonify({'error': f"Failed to add supplier: {str(e)}"}), 500

@app.route('/api/suppliers/<id>', methods=['PUT'])
@db_required
def update_supplier(id):
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No input data provided'}), 400
        update_fields = {}
        if 'name' in data:
            update_fields['name'] = data['name']
        if 'shopName' in data:
            update_fields['shopName'] = data['shopName']
        if 'address' in data:
            update_fields['address'] = data['address']
        if 'phone' in data:
            update_fields['phone'] = data['phone']
        if 'email' in data:
            update_fields['email'] = data['email']
        if not update_fields:
            return jsonify({'error': 'No fields to update'}), 400
        result = suppliers_collection.update_one({'_id': ObjectId(id)}, {'$set': update_fields})
        if result.modified_count == 0:
            return jsonify({'error': 'Supplier not found or no changes made'}), 404
        return jsonify({'message': 'Supplier updated successfully'}), 200
    except ValueError:
        return jsonify({'error': 'Invalid input data'}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to update supplier: {str(e)}"}), 500

@app.route('/api/suppliers/<id>', methods=['DELETE'])
@db_required
def delete_supplier(id):
    try:
        result = suppliers_collection.delete_one({'_id': ObjectId(id)})
        if result.deleted_count == 0:
            return jsonify({'error': 'Supplier not found'}), 404
        return jsonify({'message': 'Supplier deleted successfully'}), 200
    except ValueError:
        return jsonify({'error': 'Invalid supplier ID'}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to delete supplier: {str(e)}"}), 500

# --- Purchase Orders Routes ---
@app.route('/api/purchase_orders', methods=['GET'])
@db_required
def get_purchase_orders():
    try:
        orders = list(purchase_orders_collection.find())
        return jsonify([document_to_dict(order) for order in orders]), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch purchase orders: {str(e)}"}), 500

@app.route('/api/purchase_orders', methods=['POST'])
@db_required
def add_purchase_order():
    try:
        data = request.json
        required_fields = ['supplierId', 'name', 'shopName', 'address', 'phone', 'email', 'date', 'items', 'status', 'total']
        if not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400
        if not data['supplierId'] or not data['date'] or not isinstance(data['items'], list) or not data['items'] or data['status'] not in ['Draft', 'Submitted'] or float(data['total']) < 0:
            return jsonify({'error': 'Invalid input data'}), 400
        
        supplier = suppliers_collection.find_one({'_id': ObjectId(data['supplierId'])})
        if not supplier:
            return jsonify({'error': 'Supplier not found'}), 404
        
        for idx, item in enumerate(data['items']):
            item_required = ['itemId', 'quantity', 'unit', 'tax']
            if not all(key in item for key in item_required):
                return jsonify({'error': f"Invalid item data in row {idx+1}: missing fields"}), 400
            if not item['itemId'] or float(item['quantity']) <= 0 or item['unit'] not in ['master', 'outer', 'nos'] or float(item['tax']) < 0:
                return jsonify({'error': f"Invalid item data in row {idx+1}: invalid quantity, ID, unit, or tax"}), 400
            if 'rate' in item and item['rate'] and float(item['rate']) < 0:
                return jsonify({'error': f"Invalid item data in row {idx+1}: rate cannot be negative"}), 400
            if not purchase_items_collection.find_one({'_id': ObjectId(item['itemId'])}):
                return jsonify({'error': f"Item {item['itemId']} not found in row {idx+1}"}), 404
        
        order_id = f"PO-{str(uuid.uuid4())[:8].upper()}"
        order = {
            'id': order_id,
            'supplierId': data['supplierId'],
            'name': data['name'],
            'shopName': data['shopName'],
            'address': data['address'],
            'phone': data['phone'],
            'email': data['email'],
            'date': datetime.strptime(data['date'], '%Y-%m-%d'),
            'items': data['items'],
            'status': data['status'],
            'total': float(data['total']),
            'created_at': datetime.utcnow()
        }
        result = purchase_orders_collection.insert_one(order)
        order['_id'] = str(result.inserted_id)
        return jsonify({'message': 'Purchase Order created successfully', 'order': document_to_dict(order)}), 201
    except ValueError as e:
        return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to create purchase order: {str(e)}"}), 500

@app.route('/api/purchase_orders/<id>', methods=['PUT'])
@db_required
def update_purchase_order(id):
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No input data provided'}), 400
        old_order = purchase_orders_collection.find_one({'id': id})
        if not old_order:
            return jsonify({'error': 'Purchase Order not found'}), 404
        update_fields = {}
        if 'name' in data:
            update_fields['name'] = data['name']
        if 'shopName' in data:
            update_fields['shopName'] = data['shopName']
        if 'address' in data:
            update_fields['address'] = data['address']
        if 'phone' in data:
            update_fields['phone'] = data['phone']
        if 'email' in data:
            update_fields['email'] = data['email']
        if 'date' in data:
            update_fields['date'] = datetime.strptime(data['date'], '%Y-%m-%d')
        if 'items' in data:
            for idx, item in enumerate(data['items']):
                item_required = ['itemId', 'quantity', 'unit', 'tax']
                if not all(key in item for key in item_required):
                    return jsonify({'error': f"Invalid item data in row {idx+1}: missing fields"}), 400
                if not item['itemId'] or float(item['quantity']) <= 0 or item['unit'] not in ['master', 'outer', 'nos'] or float(item['tax']) < 0:
                    return jsonify({'error': f"Invalid item data in row {idx+1}: invalid quantity, ID, unit, or tax"}), 400
                if 'rate' in item and item['rate'] and float(item['rate']) < 0:
                    return jsonify({'error': f"Invalid item data in row {idx+1}: rate cannot be negative"}), 400
                if not purchase_items_collection.find_one({'_id': ObjectId(item['itemId'])}):
                    return jsonify({'error': f"Item {item['itemId']} not found in row {idx+1}"}), 404
            update_fields['items'] = data['items']
        if 'status' in data:
            if data['status'] not in ['Draft', 'Submitted']:
                return jsonify({'error': 'Invalid status'}), 400
            update_fields['status'] = data['status']
        if 'total' in data:
            update_fields['total'] = float(data['total'])
        if not update_fields:
            return jsonify({'error': 'No fields to update'}), 400
        result = purchase_orders_collection.update_one({'id': id}, {'$set': update_fields})
        if result.modified_count == 0:
            return jsonify({'error': 'Purchase Order not found or no changes made'}), 404
        return jsonify({'message': 'Purchase Order updated successfully'}), 200
    except ValueError as e:
        return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to update purchase order: {str(e)}"}), 500

@app.route('/api/purchase_orders/<id>', methods=['DELETE'])
@db_required
def delete_purchase_order(id):
    try:
        result = purchase_orders_collection.delete_one({'id': id})
        if result.deleted_count == 0:
            return jsonify({'error': 'Purchase Order not found'}), 404
        return jsonify({'message': 'Purchase Order deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to delete purchase order: {str(e)}"}), 500

# --- Purchase Receipts Routes ---
@app.route('/api/purchase_receipts', methods=['GET'])
@db_required
def get_purchase_receipts():
    try:
        receipts = list(purchase_receipts_collection.find())
        return jsonify([document_to_dict(receipt) for receipt in receipts]), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch purchase receipts: {str(e)}"}), 500

@app.route('/api/purchase_receipts', methods=['POST'])
@db_required
def add_purchase_receipt():
    try:
        data = request.json
        required_fields = ['poId', 'name', 'shopName', 'address', 'phone', 'email', 'date', 'items', 'status', 'total']
        if not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400
        if not data['poId'] or not data['date'] or not isinstance(data['items'], list) or not data['items'] or data['status'] not in ['Draft', 'Submitted'] or float(data['total']) < 0:
            return jsonify({'error': 'Invalid input data'}), 400
        
        po = purchase_orders_collection.find_one({'id': data['poId']})
        if not po:
            return jsonify({'error': 'Purchase Order not found'}), 404
        
        for idx, item in enumerate(data['items']):
            item_required = ['itemId', 'quantity', 'unit', 'status', 'tax']
            if not all(key in item for key in item_required):
                return jsonify({'error': f"Invalid item data in row {idx+1}: missing fields"}), 400
            if not item['itemId'] or float(item['quantity']) <= 0 or item['status'] not in ['Accepted', 'Rejected'] or item['unit'] not in ['master', 'outer', 'nos'] or float(item['tax']) < 0:
                return jsonify({'error': f"Invalid item data in row {idx+1}: invalid quantity, ID, unit, status, or tax"}), 400
            if 'rate' in item and item['rate'] and float(item['rate']) < 0:
                return jsonify({'error': f"Invalid item data in row {idx+1}: rate cannot be negative"}), 400
            if not purchase_items_collection.find_one({'_id': ObjectId(item['itemId'])}):
                return jsonify({'error': f"Item {item['itemId']} not found in row {idx+1}"}), 404
        
        receipt_id = f"PR-{str(uuid.uuid4())[:8].upper()}"
        receipt = {
            'id': receipt_id,
            'poId': data['poId'],
            'name': data['name'],
            'shopName': data['shopName'],
            'address': data['address'],
            'phone': data['phone'],
            'email': data['email'],
            'date': datetime.strptime(data['date'], '%Y-%m-%d'),
            'items': data['items'],
            'status': data['status'],
            'total': float(data['total']),
            'created_at': datetime.utcnow()
        }
        result = purchase_receipts_collection.insert_one(receipt)
        receipt['_id'] = str(result.inserted_id)
        
        if data['status'] == 'Submitted':
            for item in data['items']:
                if item['status'] == 'Accepted':
                    item_obj = purchase_items_collection.find_one({'_id': ObjectId(item['itemId'])})
                    add_master = 0
                    add_outer = 0
                    add_nos = 0
                    if item['unit'] == 'master':
                        add_master = float(item['quantity'])
                    elif item['unit'] == 'outer':
                        add_outer = float(item['quantity'])
                    elif item['unit'] == 'nos':
                        add_nos = float(item['quantity'])
                    add_total = add_master * item_obj['conversionFactor'] + add_outer * item_obj['outerToNos'] + add_nos
                    purchase_items_collection.update_one(
                        {'_id': ObjectId(item['itemId'])},
                        {'$inc': {
                            'stockMaster': add_master,
                            'stockOuter': add_outer,
                            'stockNos': add_nos,
                            'totalStock': add_total
                        }}
                    )
        
        return jsonify({'message': 'Purchase Receipt created successfully', 'receipt': document_to_dict(receipt)}), 201
    except ValueError as e:
        return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to create purchase receipt: {str(e)}"}), 500

@app.route('/api/purchase_receipts/<id>', methods=['PUT'])
@db_required
def update_purchase_receipt(id):
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No input data provided'}), 400
        old_receipt = purchase_receipts_collection.find_one({'id': id})
        if not old_receipt:
            return jsonify({'error': 'Purchase Receipt not found'}), 404
        update_fields = {}
        if 'name' in data:
            update_fields['name'] = data['name']
        if 'shopName' in data:
            update_fields['shopName'] = data['shopName']
        if 'address' in data:
            update_fields['address'] = data['address']
        if 'phone' in data:
            update_fields['phone'] = data['phone']
        if 'email' in data:
            update_fields['email'] = data['email']
        if 'date' in data:
            update_fields['date'] = datetime.strptime(data['date'], '%Y-%m-%d')
        if 'items' in data:
            for idx, item in enumerate(data['items']):
                item_required = ['itemId', 'quantity', 'unit', 'status', 'tax']
                if not all(key in item for key in item_required):
                    return jsonify({'error': f"Invalid item data in row {idx+1}: missing fields"}), 400
                if not item['itemId'] or float(item['quantity']) <= 0 or item['status'] not in ['Accepted', 'Rejected'] or item['unit'] not in ['master', 'outer', 'nos'] or float(item['tax']) < 0:
                    return jsonify({'error': f"Invalid item data in row {idx+1}: invalid quantity, ID, unit, status, or tax"}), 400
                if 'rate' in item and item['rate'] and float(item['rate']) < 0:
                    return jsonify({'error': f"Invalid item data in row {idx+1}: rate cannot be negative"}), 400
                if not purchase_items_collection.find_one({'_id': ObjectId(item['itemId'])}):
                    return jsonify({'error': f"Item {item['itemId']} not found in row {idx+1}"}), 404
                update_fields['items'] = data['items']
        if 'status' in data:
            if data['status'] not in ['Draft', 'Submitted']:
                return jsonify({'error': 'Invalid status'}), 400
            update_fields['status'] = data['status']
        if 'total' in data:
            update_fields['total'] = float(data['total'])
        if not update_fields:
            return jsonify({'error': 'No fields to update'}), 400
        result = purchase_receipts_collection.update_one({'id': id}, {'$set': update_fields})
        if result.modified_count == 0:
            return jsonify({'error': 'Purchase Receipt not found or no changes made'}), 404
        
        if 'status' in update_fields and update_fields['status'] == 'Submitted' and old_receipt['status'] != 'Submitted':
            items = data.get('items', old_receipt['items'])
            for item in items:
                if item['status'] == 'Accepted':
                    item_obj = purchase_items_collection.find_one({'_id': ObjectId(item['itemId'])})
                    add_master = 0
                    add_outer = 0
                    add_nos = 0
                    if item['unit'] == 'master':
                        add_master = float(item['quantity'])
                    elif item['unit'] == 'outer':
                        add_outer = float(item['quantity'])
                    elif item['unit'] == 'nos':
                        add_nos = float(item['quantity'])
                    add_total = add_master * item_obj['conversionFactor'] + add_outer * item_obj['outerToNos'] + add_nos
                    purchase_items_collection.update_one(
                        {'_id': ObjectId(item['itemId'])},
                        {'$inc': {
                            'stockMaster': add_master,
                            'stockOuter': add_outer,
                            'stockNos': add_nos,
                            'totalStock': add_total
                        }}
                    )
        
        return jsonify({'message': 'Purchase Receipt updated successfully'}), 200
    except ValueError as e:
        return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to update purchase receipt: {str(e)}"}), 500

@app.route('/api/purchase_receipts/<id>', methods=['DELETE'])
@db_required
def delete_purchase_receipt(id):
    try:
        result = purchase_receipts_collection.delete_one({'id': id})
        if result.deleted_count == 0:
            return jsonify({'error': 'Purchase Receipt not found'}), 404
        return jsonify({'message': 'Purchase Receipt deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to delete purchase receipt: {str(e)}"}), 500

# --- Purchase Invoices Routes ---
@app.route('/api/purchase_invoices', methods=['GET'])
@db_required
def get_purchase_invoices():
    try:
        invoices = list(purchase_invoices_collection.find())
        return jsonify([document_to_dict(invoice) for invoice in invoices]), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch purchase invoices: {str(e)}"}), 500

@app.route('/api/purchase_invoices', methods=['POST'])
@db_required
def add_purchase_invoice():
    try:
        data = request.json
        required_fields = ['poId', 'name', 'shopName', 'address', 'phone', 'email', 'date', 'items', 'status', 'total']
        if not all(key in data for key in required_fields):
            return jsonify({'error': 'Missing required fields'}), 400
        if not data['poId'] or not data['date'] or not isinstance(data['items'], list) or not data['items'] or data['status'] not in ['Draft', 'Submitted'] or float(data['total']) < 0:
            return jsonify({'error': 'Invalid input data'}), 400
        
        po = purchase_orders_collection.find_one({'id': data['poId']})
        if not po:
            return jsonify({'error': 'Purchase Order not found'}), 404
        
        pr = None
        if data.get('prId'):
            pr = purchase_receipts_collection.find_one({'id': data['prId']})
            if not pr:
                return jsonify({'error': 'Purchase Receipt not found'}), 404
        
        for idx, item in enumerate(data['items']):
            item_required = ['itemId', 'quantity', 'unit', 'tax']
            if not all(key in item for key in item_required):
                return jsonify({'error': f"Invalid item data in row {idx+1}: missing fields"}), 400
            if not item['itemId'] or float(item['quantity']) <= 0 or float(item['tax']) < 0 or item['unit'] not in ['master', 'outer', 'nos']:
                return jsonify({'error': f"Invalid item data in row {idx+1}: invalid quantity, tax, or unit"}), 400
            if 'rate' in item and item['rate'] and float(item['rate']) < 0:
                return jsonify({'error': f"Invalid item data in row {idx+1}: rate cannot be negative"}), 400
            if not purchase_items_collection.find_one({'_id': ObjectId(item['itemId'])}):
                return jsonify({'error': f"Item {item['itemId']} not found in row {idx+1}"}), 404
        
        invoice_id = f"PI-{str(uuid.uuid4())[:8].upper()}"
        invoice = {
            'id': invoice_id,
            'poId': data['poId'],
            'prId': data.get('prId', ''),
            'name': data['name'],
            'shopName': data['shopName'],
            'address': data['address'],
            'phone': data['phone'],
            'email': data['email'],
            'date': datetime.strptime(data['date'], '%Y-%m-%d'),
            'items': data['items'],
            'status': data['status'],
            'total': float(data['total']),
            'created_at': datetime.utcnow()
        }
        result = purchase_invoices_collection.insert_one(invoice)
        invoice['_id'] = str(result.inserted_id)
        
        if data['status'] == 'Submitted':
            for item in data['items']:
                item_obj = purchase_items_collection.find_one({'_id': ObjectId(item['itemId'])})
                quantity = float(item['quantity'])
                sold_nos = 0
                dec_master = 0
                dec_outer = 0
                dec_nos = 0
                if item['unit'] == 'master':
                    sold_nos = quantity * item_obj['conversionFactor']
                    dec_master = quantity
                elif item['unit'] == 'outer':
                    sold_nos = quantity * item_obj['outerToNos']
                    dec_outer = quantity
                elif item['unit'] == 'nos':
                    sold_nos = quantity
                    dec_nos = quantity
                if sold_nos > item_obj['totalStock']:
                    return jsonify({'error': f"Insufficient stock for {item_obj['name']}. Available: {item_obj['totalStock']} {item_obj['nosUnit']}, Requested: {sold_nos} {item_obj['nosUnit']}"}), 400
                purchase_items_collection.update_one(
                    {'_id': ObjectId(item['itemId'])},
                    {'$inc': {
                        'stockMaster': -dec_master,
                        'stockOuter': -dec_outer,
                        'stockNos': -dec_nos,
                        'totalStock': -sold_nos,
                        'soldNos': sold_nos
                    }}
                )
        
        return jsonify({'message': 'Purchase Invoice created successfully', 'invoice': document_to_dict(invoice)}), 201
    except ValueError as e:
        return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to create purchase invoice: {str(e)}"}), 500

@app.route('/api/purchase_invoices/<id>', methods=['PUT'])
@db_required
def update_purchase_invoice(id):
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No input data provided'}), 400
        old_invoice = purchase_invoices_collection.find_one({'id': id})
        if not old_invoice:
            return jsonify({'error': 'Purchase Invoice not found'}), 404
        update_fields = {}
        if 'name' in data:
            update_fields['name'] = data['name']
        if 'shopName' in data:
            update_fields['shopName'] = data['shopName']
        if 'address' in data:
            update_fields['address'] = data['address']
        if 'phone' in data:
            update_fields['phone'] = data['phone']
        if 'email' in data:
            update_fields['email'] = data['email']
        if 'date' in data:
            update_fields['date'] = datetime.strptime(data['date'], '%Y-%m-%d')
        if 'items' in data:
            for idx, item in enumerate(data['items']):
                item_required = ['itemId', 'quantity', 'unit', 'tax']
                if not all(key in item for key in item_required):
                    return jsonify({'error': f"Invalid item data in row {idx+1}: missing fields"}), 400
                if not item['itemId'] or float(item['quantity']) <= 0 or float(item['tax']) < 0 or item['unit'] not in ['master', 'outer', 'nos']:
                    return jsonify({'error': f"Invalid item data in row {idx+1}: invalid quantity, tax, or unit"}), 400
                if 'rate' in item and item['rate'] and float(item['rate']) < 0:
                    return jsonify({'error': f"Invalid item data in row {idx+1}: rate cannot be negative"}), 400
                if not purchase_items_collection.find_one({'_id': ObjectId(item['itemId'])}):
                    return jsonify({'error': f"Item {item['itemId']} not found in row {idx+1}"}), 404
            update_fields['items'] = data['items']
        if 'status' in data:
            if data['status'] not in ['Draft', 'Submitted']:
                return jsonify({'error': 'Invalid status'}), 400
            update_fields['status'] = data['status']
        if 'total' in data:
            update_fields['total'] = float(data['total'])
        if 'prId' in data:
            update_fields['prId'] = data['prId']
        if not update_fields:
            return jsonify({'error': 'No fields to update'}), 400
        result = purchase_invoices_collection.update_one({'id': id}, {'$set': update_fields})
        if result.modified_count == 0:
            return jsonify({'error': 'Purchase Invoice not found or no changes made'}), 404
        
        if 'status' in update_fields and update_fields['status'] == 'Submitted' and old_invoice['status'] != 'Submitted':
            items = data.get('items', old_invoice['items'])
            for item in items:
                item_obj = purchase_items_collection.find_one({'_id': ObjectId(item['itemId'])})
                quantity = float(item['quantity'])
                sold_nos = 0
                dec_master = 0
                dec_outer = 0
                dec_nos = 0
                if item['unit'] == 'master':
                    sold_nos = quantity * item_obj['conversionFactor']
                    dec_master = quantity
                elif item['unit'] == 'outer':
                    sold_nos = quantity * item_obj['outerToNos']
                    dec_outer = quantity
                elif item['unit'] == 'nos':
                    sold_nos = quantity
                    dec_nos = quantity
                if sold_nos > item_obj['totalStock']:
                    return jsonify({'error': f"Insufficient stock for {item_obj['name']}. Available: {item_obj['totalStock']} {item_obj['nosUnit']}, Requested: {sold_nos} {item_obj['nosUnit']}"}), 400
                purchase_items_collection.update_one(
                    {'_id': ObjectId(item['itemId'])},
                    {'$inc': {
                        'stockMaster': -dec_master,
                        'stockOuter': -dec_outer,
                        'stockNos': -dec_nos,
                        'totalStock': -sold_nos,
                        'soldNos': sold_nos
                    }}
                )
        
        return jsonify({'message': 'Purchase Invoice updated successfully'}), 200
    except ValueError as e:
        return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
    except Exception as e:
        return jsonify({'error': f"Failed to update purchase invoice: {str(e)}"}), 500

@app.route('/api/purchase_invoices/<id>', methods=['DELETE'])
@db_required
def delete_purchase_invoice(id):
    try:
        result = purchase_invoices_collection.delete_one({'id': id})
        if result.deleted_count == 0:
            return jsonify({'error': 'Purchase Invoice not found'}), 404
        return jsonify({'message': 'Purchase Invoice deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to delete purchase invoice: {str(e)}"}), 500
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_react_app(path):
    if path.startswith('api/'):
        return jsonify({"error": "API route not found"}), 404
   
    if path != "" and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    else:
        if os.path.exists(os.path.join(app.static_folder, 'index.html')):
            return send_from_directory(app.static_folder, 'index.html')
        else:
            return "<h1>Backend is running</h1><p>Frontend not found. Ensure the 'dist' folder contains the built React app.</p>", 404
if __name__ == '__main__':
    connect_to_mongodb()
    if db and schedule:
        start_scheduler()
   
    logger.info(f"Serving static files from: {app.static_folder}")
    if getattr(sys, 'frozen', False):
        logger.info("Running as frozen executable, using Waitress")
        import waitress
        waitress.serve(app, host='0.0.0.0', port=8000, threads=8)
    else:
        logger.info("Running in development mode, using Flask")
        app.run(host='0.0.0.0', port=8000, debug=True)