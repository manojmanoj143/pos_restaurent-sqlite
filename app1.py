# -*- mode: python ; coding: utf-8 -*-
from flask import Flask, request, jsonify, send_from_directory, Response
from flask_cors import CORS
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
import os
import sys
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from werkzeug.utils import secure_filename
import json
import secrets
import hashlib
import bcrypt
import socket
import uuid
import time
import threading
import traceback
from functools import wraps
import jwt
import requests
from io import BytesIO
# Optional dependencies
try:
    import openpyxl
except ImportError:
    openpyxl = None
    logging.warning("openpyxl library not found. Excel export/backup features will be disabled.")
try:
    import schedule
except ImportError:
    schedule = None
    logging.warning("schedule library not found. Automatic tasks will be disabled.")
# SQLite import
import sqlite3

COMBO_IMAGES_DIR = os.path.join(
    os.environ.get('UPLOAD_FOLDER', 'static'),  # Fallback to 'static' in dev
    'combo_images'
)


def is_valid_secret_key(secret_key):
    """Validate secret key is exactly 6 digits."""
    return secret_key.isdigit() and len(secret_key) == 6
def convert_to_24h(time_str):
    if not time_str:
        return ""
    try:
        # Check if already HH:mm
        datetime.strptime(time_str, '%H:%M')
        return time_str
    except ValueError:
        pass
    
    try:
        # Try 12-hour format with AM/PM
        t = datetime.strptime(time_str, '%I:%M %p')
        return t.strftime('%H:%M')
    except ValueError:
        pass
        
    return time_str


# --- Configuration Management ---
def get_base_dir():
    """Determine the base directory, handling both development and frozen executable cases."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))
CONFIG_DIR = os.getenv('CONFIG_DIR', get_base_dir())
os.makedirs(CONFIG_DIR, exist_ok=True)
CONFIG_FILE_PATH = os.path.join(CONFIG_DIR, 'config.json')
BASE_DIR = get_base_dir()
STATIC_FOLDER_PATH = os.path.join(BASE_DIR, 'dist')
app = Flask(__name__, static_folder=STATIC_FOLDER_PATH, static_url_path='/')
CORS(app, resources={r"/api/*": {"origins": "*"}})
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(CONFIG_DIR, 'pos_app.log'))
    ]
)
logger = logging.getLogger(__name__)
JWT_SECRET = os.getenv('JWT_SECRET', secrets.token_hex(32))
JWT_ALGORITHM = 'HS256'
JWT_EXP_DELTA_SECONDS = 3600
def load_config():
    """Loads the configuration from config.json, creating a default if it doesn't exist."""
    try:
        if not os.path.exists(CONFIG_FILE_PATH):
            logger.warning(f"config.json not found at {CONFIG_FILE_PATH}. Creating default configuration.")
            default_config = {
                "mode": "server",
                "server_ip": "127.0.0.1"
            }
            save_config(default_config)
            return default_config
        with open(CONFIG_FILE_PATH, 'r') as f:
            logger.info(f"Loading configuration from {CONFIG_FILE_PATH}")
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError, PermissionError) as e:
        logger.error(f"Could not load or parse config.json: {e}. Using default server config.")
        return {"mode": "server", "server_ip": "127.0.0.1"}
def save_config(config_data):
    """Saves the configuration to config.json."""
    try:
        with open(CONFIG_FILE_PATH, 'w') as f:
            json.dump(config_data, f, indent=4)
        logger.info(f"Configuration saved to {CONFIG_FILE_PATH}")
    except PermissionError as e:
        logger.error(f"Permission denied when saving config.json to {CONFIG_FILE_PATH}: {e}")
        raise
    except Exception as e:
        logger.error(f"Error saving config.json: {e}")
        raise
config = load_config()
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', os.path.join(BASE_DIR, 'static', 'uploads'))
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp','jfif','ico'}
ALLOWED_JSON_EXTENSIONS = {'json'}
MAX_BACKUPS = 10
def create_directory(directory):
    """Ensure directory exists."""
    try:
        os.makedirs(directory, exist_ok=True)
        logger.info(f"Ensured directory exists: {directory}")
    except Exception as e:
        logger.error(f"Error creating directory {directory}: {e}")
create_directory(app.config['UPLOAD_FOLDER'])
conn = None
class SQLiteCollection:
    def __init__(self, conn, name):
        self.conn = conn
        self.name = name

    def matches_filter(self, d, filter_):
        if not filter_:
            return True
        for key, value in filter_.items():
            if key == '$or':
                if not any(self.matches_filter(d, subfilter) for subfilter in value):
                    return False
            else:
                if d.get(key) != value:
                    return False
        return True

    def insert_one(self, doc):
        if '_id' not in doc:
            doc['_id'] = str(uuid.uuid4())
        json_doc = json.dumps(document_to_dict(doc))
        cur = self.conn.cursor()
        cur.execute(f"INSERT INTO {self.name} (id, data) VALUES (?, ?)", (doc['_id'], json_doc))
        self.conn.commit()
        return type('InsertResult', (), {'inserted_id': doc['_id']})()

    def find(self, filter=None):
        cur = self.conn.cursor()
        rows = cur.execute(f"SELECT data FROM {self.name}").fetchall()
        docs = [json.loads(row[0]) for row in rows]
        if filter:
            docs = [d for d in docs if self.matches_filter(d, filter)]
        return docs

    def find_one(self, filter):
        rows = self.conn.execute(f"SELECT data FROM {self.name}").fetchall()
        for row in rows:
            d = json.loads(row[0])
            if self.matches_filter(d, filter):
                return d
        return None

    def update_one(self, filter, update, array_filters=None, upsert=False):  # FIXED: Added upsert=False param
        cur = self.conn.cursor()
        rows = cur.execute(f"SELECT id, data FROM {self.name}").fetchall()
        matched = False
        for row_id, row_data in rows:
            d = json.loads(row_data)
            if self.matches_filter(d, filter):
                matched = True
                if '$set' in update:
                    for k, v in update['$set'].items():
                        if '.' in k:
                            parts = k.split('.')
                            current = d
                            for part in parts[:-1]:
                                if part.isdigit():
                                    part = int(part)
                                if part not in current:
                                    current[part] = {} if not part.isdigit() else []
                                current = current[part]
                            current[parts[-1]] = v
                        else:
                            d[k] = v
                if '$unset' in update:
                    for k in update['$unset']:
                        if '.' in k:
                            parts = k.split('.')
                            current = d
                            for part in parts[:-1]:
                                if part.isdigit():
                                    part = int(part)
                                current = current[part]
                            current.pop(parts[-1], None)
                        else:
                            d.pop(k, None)
                if '$inc' in update:
                    for k, v in update['$inc'].items():
                        if '.' in k:
                            parts = k.split('.')
                            current = d
                            for part in parts[:-1]:
                                if part.isdigit():
                                    part = int(part)
                                if part not in current:
                                    current[part] = {} if not part.isdigit() else []
                                current = current[part]
                            current[parts[-1]] = current.get(parts[-1], 0) + v
                        else:
                            d[k] = d.get(k, 0) + v
                if '$pull' in update:
                    for k, v in update['$pull'].items():
                        if isinstance(d.get(k), list):
                            d[k] = [i for i in d[k] if i != v]
                if array_filters:
                    for uk, uv in update['$set'].items():
                        if '$[elem]' in uk:
                            array_name, rest = uk.split('.$[elem].', 1)
                            array = d.get(array_name, [])
                            af = array_filters[0]
                            af_key = list(af.keys())[0].split('.')[-1]
                            af_value = af[list(af.keys())[0]]
                            for elem in array:
                                if elem.get(af_key) == af_value:
                                    elem[rest] = uv
                json_doc = json.dumps(document_to_dict(d))
                cur.execute(f"UPDATE {self.name} SET data = ? WHERE id = ?", (json_doc, row_id))
                self.conn.commit()
                return type('UpdateResult', (), {'matched_count': 1, 'modified_count': 1})()
        if not matched and upsert:
            # FIXED: Implement upsert logic
            new_doc = dict(filter)  # Start with filter as base
            if '$set' in update:
                for k, v in update['$set'].items():
                    if '.' in k:
                        parts = k.split('.')
                        current = new_doc
                        for part in parts[:-1]:
                            if part.isdigit():
                                part = int(part)
                            if part not in current:
                                current[part] = {} if not str(part).isdigit() else []
                            current = current[part]
                        current[parts[-1]] = v
                    else:
                        new_doc[k] = v
            if '$inc' in update:
                for k, v in update['$inc'].items():
                    if '.' in k:
                        parts = k.split('.')
                        current = new_doc
                        for part in parts[:-1]:
                            if part.isdigit():
                                part = int(part)
                            if part not in current:
                                current[part] = {} if not str(part).isdigit() else []
                            current = current[part]
                        current[parts[-1]] = current.get(parts[-1], 0) + v
                    else:
                        new_doc[k] = new_doc.get(k, 0) + v
            # Insert the new doc (add _id if missing)
            self.insert_one(new_doc)
            return type('UpdateResult', (), {'matched_count': 0, 'modified_count': 1})()
        return type('UpdateResult', (), {'matched_count': 0, 'modified_count': 0})()
    def update_many(self, filter, update):
        cur = self.conn.cursor()
        rows = cur.execute(f"SELECT id, data FROM {self.name}").fetchall()
        modified_count = 0
        for row_id, row_data in rows:
            d = json.loads(row_data)
            if self.matches_filter(d, filter):
                if '$set' in update:
                    for k, v in update['$set'].items():
                        if '.' in k:
                            parts = k.split('.')
                            current = d
                            for part in parts[:-1]:
                                if part.isdigit():
                                    part = int(part)
                                if part not in current:
                                    current[part] = {} if not part.isdigit() else []
                                current = current[part]
                            current[parts[-1]] = v
                        else:
                            d[k] = v
                json_doc = json.dumps(document_to_dict(d))
                cur.execute(f"UPDATE {self.name} SET data = ? WHERE id = ?", (json_doc, row_id))
                modified_count += 1
        self.conn.commit()
        return type('UpdateResult', (), {'modified_count': modified_count})()

    def delete_one(self, filter):
        cur = self.conn.cursor()
        rows = cur.execute(f"SELECT id, data FROM {self.name}").fetchall()
        for row_id, row_data in rows:
            d = json.loads(row_data)
            if self.matches_filter(d, filter):
                cur.execute(f"DELETE FROM {self.name} WHERE id = ?", (row_id,))
                self.conn.commit()
                return type('DeleteResult', (), {'deleted_count': 1})()
        return type('DeleteResult', (), {'deleted_count': 0})()

    def replace_one(self, filter, replacement, upsert=False):
        cur = self.conn.cursor()
        rows = cur.execute(f"SELECT id, data FROM {self.name}").fetchall()
        matched = False
        for row_id, row_data in rows:
            d = json.loads(row_data)
            if self.matches_filter(d, filter):
                matched = True
                json_doc = json.dumps(document_to_dict(replacement))
                cur.execute(f"UPDATE {self.name} SET data = ? WHERE id = ?", (json_doc, row_id))
                self.conn.commit()
                return type('UpdateResult', (), {'matched_count': 1, 'modified_count': 1})()
        if upsert:
            self.insert_one(replacement)
            return type('UpdateResult', (), {'matched_count': 0, 'modified_count': 1})()
        return type('UpdateResult', (), {'matched_count': 0, 'modified_count': 0})()

    def find_one_and_update(self, filter, update, upsert=False, return_document=True):
        cur = self.conn.cursor()
        rows = cur.execute(f"SELECT id, data FROM {self.name}").fetchall()
        matched = False
        for row_id, row_data in rows:
            d = json.loads(row_data)
            if self.matches_filter(d, filter):
                matched = True
                old_d = d.copy()
                if '$inc' in update:
                    for k, v in update['$inc'].items():
                        d[k] = d.get(k, 0) + v
                json_doc = json.dumps(document_to_dict(d))
                cur.execute(f"UPDATE {self.name} SET data = ? WHERE id = ?", (json_doc, row_id))
                self.conn.commit()
                if return_document:
                    return d
                else:
                    return old_d
        if upsert:
            doc = filter.copy()
            if '$inc' in update:
                for k, v in update['$inc'].items():
                    doc[k] = doc.get(k, 0) + v
            self.insert_one(doc)
            return doc
        return None
def connect_to_sqlite():
    global conn, items_collection, customers_collection, sales_collection, tables_collection, users_collection, settings_collection, email_tokens_collection, opening_collection, pos_closing_collection, kitchens_collection, item_groups_collection, kitchen_saved_collection, picked_up_collection, variants_collection, employees_collection, activeorders_collection, order_counters_collection, tripreports_collection, email_settings_collection, purchase_items_collection, suppliers_collection, purchase_orders_collection, purchase_receipts_collection, purchase_invoices_collection, uoms_collection, purchase_sales_collection, print_settings_collection, combo_offers_collection, vat_collection, customer_groups_collection, company_details_collection, logo_details_collection, supplier_group_collection,address_structures_collection,worker_collection, employee_designations_collection,employee_type_collection,working_days_collection,attendance_collection,brands_collection,shift_master_collection, schedule_master_collection, employee_schedule_assign_collection
    mode = config.get("mode", "server")
    if mode == 'server':
        db_path = os.path.join(CONFIG_DIR, 'restaurant.db')
        conn = sqlite3.connect(db_path, check_same_thread=False)
        cur = conn.cursor()
        tables = [
            'active_orders', 'combo_offers', 'customers', 'email_settings', 'email_tokens', 'employees', 'item_groups', 'items', 'kitchen_saved_orders', 'kitchens',
            'order_counters', 'picked_up_items', 'pos_closing_entries', 'pos_opening_entries', 'print_settings', 'purchase_invoices', 'purchase_items', 'purchase_orders',
            'purchase_receipts', 'purchase_sales', 'sales', 'suppliers', 'system_settings', 'tables', 'trip_reports', 'uoms', 'users', 'variants', 'vat', 'customer_groups', 'company_details', 'logo_details', 'supplier_groups','address_structures','new_employee','employee_designations','employee_types','working_days','attendance','brands',
            'shift_master', 'schedule_master', 'employee_schedule_assign'
        ]
        
        for table in tables:
            cur.execute(f"CREATE TABLE IF NOT EXISTS {table} (id TEXT PRIMARY KEY, data TEXT)")
        conn.commit()
        logger.info(f"Successfully connected to SQLite at {db_path}")
        items_collection = SQLiteCollection(conn, 'items')
        customers_collection = SQLiteCollection(conn, 'customers')
        sales_collection = SQLiteCollection(conn, 'sales')
        tables_collection = SQLiteCollection(conn, 'tables')
        users_collection = SQLiteCollection(conn, 'users')
        settings_collection = SQLiteCollection(conn, 'system_settings')
        email_tokens_collection = SQLiteCollection(conn, 'email_tokens')
        opening_collection = SQLiteCollection(conn, 'pos_opening_entries')
        pos_closing_collection = SQLiteCollection(conn, 'pos_closing_entries')
        kitchens_collection = SQLiteCollection(conn, 'kitchens')
        item_groups_collection = SQLiteCollection(conn, 'item_groups')
        kitchen_saved_collection = SQLiteCollection(conn, 'kitchen_saved_orders')
        picked_up_collection = SQLiteCollection(conn, 'picked_up_items')
        variants_collection = SQLiteCollection(conn, 'variants')
        employees_collection = SQLiteCollection(conn, 'employees')
        activeorders_collection = SQLiteCollection(conn, 'active_orders')
        order_counters_collection = SQLiteCollection(conn, 'order_counters')
        tripreports_collection = SQLiteCollection(conn, 'trip_reports')
        email_settings_collection = SQLiteCollection(conn, 'email_settings')
        purchase_items_collection = SQLiteCollection(conn, 'purchase_items')
        suppliers_collection = SQLiteCollection(conn, 'suppliers')
        purchase_orders_collection = SQLiteCollection(conn, 'purchase_orders')
        purchase_receipts_collection = SQLiteCollection(conn, 'purchase_receipts')
        purchase_invoices_collection = SQLiteCollection(conn, 'purchase_invoices')
        uoms_collection = SQLiteCollection(conn, 'uoms')
        purchase_sales_collection = SQLiteCollection(conn, 'purchase_sales')
        print_settings_collection = SQLiteCollection(conn, 'print_settings')
        combo_offers_collection = SQLiteCollection(conn, 'combo_offers')
        vat_collection = SQLiteCollection(conn, 'vat')
        customer_groups_collection = SQLiteCollection(conn, 'customer_groups')
        company_details_collection = SQLiteCollection(conn, 'company_details')
        logo_details_collection = SQLiteCollection(conn, 'logo_details')
        supplier_group_collection = SQLiteCollection(conn, 'supplier_groups')
        address_structures_collection = SQLiteCollection(conn, 'address_structures')
        worker_collection = SQLiteCollection(conn, 'new_employee')
        employee_designations_collection = SQLiteCollection(conn, 'employee_designations')
        employee_type_collection = SQLiteCollection(conn, 'employee_types')
        working_days_collection = SQLiteCollection(conn, 'working_days')
        attendance_collection = SQLiteCollection(conn, 'attendance')
        brands_collection = SQLiteCollection(conn, 'brands')  # NEW: Brands collection
        shift_master_collection = SQLiteCollection(conn, 'shift_master')
        schedule_master_collection = SQLiteCollection(conn, 'schedule_master')
        employee_schedule_assign_collection = SQLiteCollection(conn, 'employee_schedule_assign')
        ensure_test_users()
        return True
    else:
        logger.info(f"Client mode: No local database connection")
        conn = None
        return True # Allow client mode without local DB
def ensure_test_users():
    for test_user_template in TEST_USERS:
        email = test_user_template['email']
        user = users_collection.find_one({"email": email})
        new_hash = bcrypt.hashpw("123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        test_user = {**test_user_template, "password": new_hash}
        if user:
            if 'firstName' not in user:
                users_collection.update_one({"email": email}, {"$set": {"firstName": test_user['firstName']}})
                logger.warning(f"Added missing 'firstName' for user {email}")
            try:
                bcrypt.checkpw(b"123", user['password'].encode('utf-8'))
            except ValueError as ve:
                logger.warning(f"Invalid hash for user {email}, rehashing: {ve}")
                users_collection.update_one({"email": email}, {"$set": {"password": new_hash}})
            except Exception as e:
                logger.error(f"Error checking hash for {email}: {e}")
        else:
            users_collection.insert_one(test_user)
            logger.info(f"Inserted test user {email}")
def db_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        mode = config.get("mode", "server")
        if mode == 'client':
            return f(*args, **kwargs)
        if conn is None:
            error_msg = "Cannot connect to database. Please check the database configuration."
            logger.error(error_msg)
            return jsonify({"error": error_msg, "message": "Database not connected."}), 503
        return f(*args, **kwargs)
    return decorated_function
# Proxy for client mode - Add this to proxy all /api/* except local ones
if config.get('mode') == 'client':
    @app.route('/api/<path:path>', methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'])
    def proxy_api(path):
        if path in ['configure', 'network_info']: # Handle local routes without proxy
            return "Local route in client mode", 200
        server_url = f"http://{config['server_ip']}:8000/api/{path}"
        if request.method == 'OPTIONS':
            response = jsonify({"success": True})
            response.headers.add('Access-Control-Allow-Origin', '*')
            response.headers.add('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
            response.headers.add('Access-Control-Allow-Headers', 'Content-Type, Authorization')
            return response
        try:
            resp = requests.request(
                method=request.method,
                url=server_url,
                headers={k: v for k, v in request.headers if k != 'Host'},
                data=request.get_data(),
                cookies=request.cookies,
                allow_redirects=False,
                stream=True
            )
            excluded_headers = ['content-encoding', 'content-length', 'transfer-encoding', 'connection']
            headers = [(name, value) for (name, value) in resp.raw.headers.items() if name.lower() not in excluded_headers]
            response = Response(resp.content, resp.status_code, headers)
            return response
        except Exception as e:
            logger.error(f"Proxy error: {str(e)}")
            return jsonify({"error": str(e)}), 503 # Change to 503 for service unavailable
def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions
def handle_image_upload(file):
    if file and allowed_file(file.filename, ALLOWED_IMAGE_EXTENSIONS):
        ext = file.filename.rsplit('.', 1)[1].lower()
        unique_filename = f"{uuid.uuid4()}.{ext}"
        filename = secure_filename(unique_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        logger.info(f"Image saved: {filename}")
        return filename
    logger.warning(f"Invalid file for upload: {file.filename}")
    return None
def sanitize_image_fields(data):
    fields_to_sanitize = ['image', 'addon_image', 'combo_image', 'variant_image']
    for field in fields_to_sanitize:
        if field in data and isinstance(data[field], str):
            data[field] = os.path.basename(data[field])
    for addon in data.get("addons", []):
        if 'addon_image' in addon and isinstance(addon['addon_image'], str):
            addon['addon_image'] = os.path.basename(addon['addon_image'])
    for combo in data.get("combos", []):
        if 'combo_image' in combo and isinstance(combo['combo_image'], str):
            combo['combo_image'] = os.path.basename(combo['combo_image'])
    return data
def convert_objectid_to_str(item):
    if isinstance(item, list):
        return [convert_objectid_to_str(i) for i in item]
    if isinstance(item, dict):
        return {key: convert_objectid_to_str(value) for key, value in item.items()}
    if isinstance(item, datetime):
        return item.isoformat()
    return item
def get_system_settings():
    if settings_collection is None:
        logger.warning("Settings collection not available, returning default settings")
        return {
            "_id": "system_settings",
            "disableUserPassLogin": False,
            "allowLoginUsingMobileNumber": True,
            "allowLoginUsingUsername": True,
            "loginWithEmailLink": False,
            "sessionExpiry": "06:00",
            "backup_interval_hours": 6,
            "country": 'United Arab Emirates',
            "language": 'English',
            "timeZone": 'Asia/Dubai',
            "currency": '',
            "dateFormat": 'dd-mm-yyyy',
            "timeFormat": 'HH:mm:ss',
            "numberFormat": '#,##,###.##',
            "useNumberFormatFromCurrency": False,
            "firstDayOfWeek": 'Monday',
            "floatPrecision": 3,
            "currencyPrecision": '',
            "roundingMethod": '',
        }
    settings = settings_collection.find_one({"_id": "system_settings"})
    if not settings:
        default_settings = {
            "_id": "system_settings",
            "disableUserPassLogin": False,
            "allowLoginUsingMobileNumber": True,
            "allowLoginUsingUsername": True,
            "loginWithEmailLink": False,
            "sessionExpiry": "06:00",
            "backup_interval_hours": 6,
            "country": 'United Arab Emirates',
            "language": 'English',
            "timeZone": 'Asia/Dubai',
            "currency": '',
            "dateFormat": 'dd-mm-yyyy',
            "timeFormat": 'HH:mm:ss',
            "numberFormat": '#,##,###.##',
            "useNumberFormatFromCurrency": False,
            "firstDayOfWeek": 'Monday',
            "floatPrecision": 3,
            "currencyPrecision": '',
            "roundingMethod": '',
        }
        settings_collection.insert_one(default_settings)
        logger.info("Inserted default system settings")
        return default_settings
    return settings
def save_system_settings(settings):
    settings["_id"] = "system_settings"
    settings_collection.replace_one({"_id": "system_settings"}, settings, upsert=True)
    logger.info("System settings saved")
TEST_USERS = [
    {
        "email": "admin@gmail.com",
        "password": bcrypt.hashpw("123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8'),
        "phone_number": "1234567890",
        "username": "admin",
        "role": "admin",
        "firstName": "Admin",
        "company": "POS 8",
        "pos_profile": "POS-001",
        "status": "Active",
        "created_at": datetime.now(ZoneInfo("UTC")).isoformat(),
        "is_test": True
    },
    {
        "email": "bearer@gmail.com",
        "password": bcrypt.hashpw("123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8'),
        "phone_number": "0987654321",
        "username": "bearer",
        "role": "bearer",
        "firstName": "Bearer",
        "company": "POS 8",
        "pos_profile": "POS-001",
        "status": "Active",
        "created_at": datetime.now(ZoneInfo("UTC")).isoformat(),
        "is_test": True
    }
]
def to_json_safe(data):
    """Recursively converts documents to JSON-serializable format."""
    if isinstance(data, list):
        return [to_json_safe(item) for item in data]
    if isinstance(data, dict):
        return {key: to_json_safe(value) for key, value in data.items()}
    if isinstance(data, datetime):
        return data.isoformat()
    return data
def document_to_dict(doc):
    if isinstance(doc, list):
        return [document_to_dict(v) for v in doc]
    if isinstance(doc, dict):
        return {key: document_to_dict(value) for key, value in doc.items()}
    if isinstance(doc, datetime):
        return doc.isoformat()
    return doc
# NEW: Function to clean expired combo offers
# NEW: Function to clean expired combo offers
def clean_expired_combo_offers():
    """Delete expired combo offers from the database."""
    try:
        current_time = datetime.now(timezone.utc)
        offers = combo_offers_collection.find()
        deleted_count = 0
        for offer in offers:
            if 'offer_end_time' in offer and offer['offer_end_time']:
                try:
                    end_time_str = str(offer['offer_end_time'])
                    if end_time_str.endswith('Z'):
                        end_time_str = end_time_str.replace('Z', '+00:00')
                    end_time = datetime.fromisoformat(end_time_str)
                    if end_time.tzinfo is None:
                        end_time = end_time.replace(tzinfo=timezone.utc)
                    if current_time > end_time:
                        combo_offers_collection.delete_one({'_id': offer['_id']})
                        logger.info(f"Automatically deleted expired combo offer: {offer['_id']} (ended {end_time})")
                        deleted_count += 1
                except (ValueError, TypeError) as e:
                    logger.error(f"Invalid offer_end_time for combo offer {offer['_id']}: {str(e)}")
            else:
                # If no end time, keep it (always active)
                continue
        logger.info(f"Cleanup job completed: Deleted {deleted_count} expired combo offers")
        return True
    except Exception as e:
        logger.error(f"Error in clean_expired_combo_offers: {str(e)}")
        return False
def has_associated_sales(item_name):
    """Check if any sale contains this item_name in its items."""
    sales = sales_collection.find()
    for sale in sales:
        items = sale.get('items', [])
        if any(sale_item.get('item_name') == item_name for sale_item in items):
            return True
    return False
# NEW: Function to get sales for a specific item
def get_sales_for_item(item_name):
    """Get all sales containing the specified item_name."""
    sales = sales_collection.find()
    relevant_sales = []
    for sale in sales:
        items = sale.get('items', [])
        if any(sale_item.get('item_name') == item_name for sale_item in items):
            relevant_sales.append(sale)
    return relevant_sales
# --- Local Routes (available in both modes) ---
@app.route('/api/test', methods=['GET'])
def test_endpoint():
    return jsonify({"status": "success", "message": "Server is running"}), 200
@app.route('/api/network_info', methods=['GET'])
def get_network_info():
    current_config = load_config()
    mode = current_config.get("mode", "server")
    server_ip = current_config.get("server_ip", "127.0.0.1")
    db_status = "Disconnected"
    if mode == 'server':
        db_status = "Connected" if conn is not None else "Disconnected"
    else: # client mode
        try:
            r = requests.get(f"http://{server_ip}:8000/api/test", timeout=1)
            db_status = "Connected" if r.status_code == 200 else "Disconnected"
        except:
            db_status = "Disconnected"
    local_ip = "127.0.0.1"
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(0.1)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except Exception:
        try:
            hostname = socket.gethostname()
            local_ip = socket.gethostbyname(hostname)
        except Exception:
            logger.warning("Could not determine local IP")
    return jsonify({
        "local_ip": local_ip,
        "config": current_config,
        "database_status": db_status
    }), 200
@app.route('/api/configure', methods=['POST'])
def configure_app():
    global config
    try:
        data = request.get_json()
        mode = data.get('mode')
        server_ip = data.get('server_ip')
        if mode not in ['server', 'client']:
            logger.error(f"Invalid mode specified: {mode}")
            return jsonify({"error": "Invalid mode specified. Must be 'server' or 'client'"}), 400
        if mode == 'client' and not server_ip:
            logger.error("Server IP not provided in client mode")
            return jsonify({"error": "Server IP is required for client mode"}), 400
        old_mode = config.get("mode", "server")
        new_config = {"mode": mode}
        if mode == 'client':
            new_config['server_ip'] = server_ip
        else:
            new_config['server_ip'] = "127.0.0.1"
        save_config(new_config)
        config = new_config
        success = connect_to_sqlite()
        mode_changed = old_mode != mode
        if mode == 'server' and not success:
            logger.error("Failed to connect to SQLite after configuration change")
            return jsonify({"error": "Failed to connect to SQLite with new configuration"}), 503
        logger.info(f"Application configured as {mode}")
        return jsonify({"message": "Configuration saved successfully. The application will now use the new settings.", "mode_changed": mode_changed}), 200
    except Exception as e:
        logger.error(f"Configuration error: {e}\n{traceback.format_exc()}")
        return jsonify({"error": f"An internal server error occurred during configuration: {str(e)}"}), 500
# NEW: Logo upload route - Updated to use logo_details_collection
@app.route('/api/upload-logo', methods=['POST', 'OPTIONS'])
@db_required
def upload_logo():
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    try:
        if 'logo' not in request.files:
            return jsonify({"error": "No logo file provided"}), 400
        file = request.files['logo']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        if not allowed_file(file.filename, ALLOWED_IMAGE_EXTENSIONS):
            return jsonify({"error": "Invalid file type. Only images allowed"}), 400
        filename = handle_image_upload(file)
        if not filename:
            return jsonify({"error": "Failed to upload logo"}), 500
        # Delete old logo if exists
        old_logo_data = logo_details_collection.find_one({"_id": "logo_settings"})
        old_logo = old_logo_data.get('logo') if old_logo_data else None
        if old_logo:
            old_path = os.path.join(app.config['UPLOAD_FOLDER'], old_logo)
            if os.path.exists(old_path):
                os.remove(old_path)
                logger.info(f"Deleted old logo: {old_logo}")
        # Update logo_details
        logo_settings = {"_id": "logo_settings", "logo": filename, "uploaded_at": datetime.now(ZoneInfo("UTC")).isoformat()}
        logo_details_collection.replace_one({"_id": "logo_settings"}, logo_settings, upsert=True)
        logger.info(f"Logo uploaded and saved: {filename}")
        return jsonify({"message": "Logo uploaded successfully", "logo": f"/api/images/{filename}"}), 200
    except Exception as e:
        logger.error(f"Error uploading logo: {str(e)}")
        return jsonify({"error": str(e)}), 500
# NEW: Get logo route - Updated to use logo_details_collection
@app.route('/api/logo', methods=['GET'])
@db_required
def get_logo():
    try:
        logo_data = logo_details_collection.find_one({"_id": "logo_settings"})
        logo_filename = logo_data.get('logo') if logo_data else None
        if not logo_filename:
            return jsonify({"logo": None}), 200
        return jsonify({"logo": f"/api/images/{logo_filename}"}), 200
    except Exception as e:
        logger.error(f"Error fetching logo: {str(e)}")
        return jsonify({"error": str(e)}), 500
# NEW: Delete logo route - Updated to use logo_details_collection
@app.route('/api/delete-logo', methods=['DELETE', 'OPTIONS'])
@db_required
def delete_logo():
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    try:
        logo_data = logo_details_collection.find_one({"_id": "logo_settings"})
        logo_filename = logo_data.get('logo') if logo_data else None
        if logo_filename:
            logo_path = os.path.join(app.config['UPLOAD_FOLDER'], logo_filename)
            if os.path.exists(logo_path):
                os.remove(logo_path)
                logger.info(f"Deleted logo file: {logo_filename}")
            logo_details_collection.replace_one({"_id": "logo_settings"}, {"_id": "logo_settings", "logo": None}, upsert=True)
            logger.info("Logo deleted successfully")
            return jsonify({"message": "Logo deleted successfully"}), 200
        return jsonify({"message": "No logo to delete"}), 200
    except Exception as e:
        logger.error(f"Error deleting logo: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/upload-image', methods=['POST', 'OPTIONS'])
def upload_image():
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    try:
        mode = config.get("mode", "server")
        if mode == 'client':
            files = {}
            for file in request.files.getlist('files'):
                files[file.filename] = (file.filename, file.stream.read(), file.content_type)
            server_url = f"http://{config['server_ip']}:8000/api/upload-image"
            response = requests.post(server_url, files=files)
            if response.status_code == 200:
                data = response.json()
                data['urls'] = [f"http://{config['server_ip']}:8000{url}" for url in data['urls']]
                return jsonify(data), 200
            else:
                logger.error(f"Proxy upload failed: {response.text}")
                return jsonify({"error": "Proxy upload failed"}), response.status_code
        else:
            if 'files' not in request.files:
                logger.error("No files part in request")
                return jsonify({"error": "No files provided"}), 400
            files = request.files.getlist('files')
            if not files or all(file.filename == '' for file in files):
                logger.error("No valid files selected")
                return jsonify({"error": "No valid files selected"}), 400
            urls = []
            for file in files:
                if file and allowed_file(file.filename, ALLOWED_IMAGE_EXTENSIONS):
                    filename = handle_image_upload(file)
                    if filename:
                        urls.append(f"/api/images/{filename}")
                        logger.info(f"Uploaded image: {filename}")
                    else:
                        logger.warning(f"Failed to upload image: {file.filename}")
                else:
                    logger.warning(f"Invalid file type: {file.filename}")
            if not urls:
                return jsonify({"error": "No valid images uploaded"}), 400
            return jsonify({"urls": urls}), 200
    except Exception as e:
        logger.error(f"Error uploading images: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500
@app.route('/api/delete-image/<filename>', methods=['DELETE', 'OPTIONS'])
def delete_image(filename):
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    try:
        mode = config.get("mode", "server")
        if mode == 'client':
            server_url = f"http://{config['server_ip']}:8000/api/delete-image/{filename}"
            response = requests.delete(server_url, params=request.args)
            if response.status_code == 200:
                return jsonify(response.json()), 200
            else:
                logger.error(f"Proxy delete failed: {response.text}")
                return jsonify({"error": "Proxy delete failed"}), response.status_code
        else:
            item_id = request.args.get('item_id')
            field = request.args.get('field', 'image')
            valid_fields = {'image', 'images', 'addon_image', 'combo_image', 'variant_image'}
            if field not in valid_fields:
                logger.error(f"Invalid field specified: {field}")
                return jsonify({"error": f"Invalid field: {field}. Must be one of {valid_fields}"}), 400
            if not item_id:
                logger.error("Item ID is required for image deletion")
                return jsonify({"error": "Item ID is required"}), 400
            filename = secure_filename(filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            if field == 'images':
                result = items_collection.update_one(
                    {"_id": item_id, "images": filename},
                    {"$pull": {"images": filename}}
                )
                if result.modified_count == 0:
                    logger.warning(f"Image {filename} not found in images array for item {item_id}")
                    return jsonify({"error": "Image not found in item"}), 404
            elif field == 'image':
                result = items_collection.update_one(
                    {"_id": item_id, "image": filename},
                    {"$set": {"image": None}}
                )
                if result.modified_count == 0:
                    logger.warning(f"Image {filename} not found in image field for item {item_id}")
                    return jsonify({"error": "Image not found in item"}), 404
            elif field == 'addon_image':
                result = items_collection.update_one(
                    {"_id": item_id, "addons.addon_image": filename},
                    {"$set": {"addons.$[elem].addon_image": None}},
                    array_filters=[{"elem.addon_image": filename}]
                )
                if result.modified_count == 0:
                    logger.warning(f"Image {filename} not found in addons for item {item_id}")
                    return jsonify({"error": "Image not found in addons"}), 404
            elif field == 'combo_image':
                result = items_collection.update_one(
                    {"_id": item_id, "combos.combo_image": filename},
                    {"$set": {"combos.$[elem].combo_image": None}},
                    array_filters=[{"elem.combo_image": filename}]
                )
                if result.modified_count == 0:
                    logger.warning(f"Image {filename} not found in combos for item {item_id}")
                    return jsonify({"error": "Image not found in combos"}), 404
            elif field == 'variant_image':
                result = items_collection.update_one(
                    {"_id": item_id, "variants.variant_image": filename},
                    {"$set": {"variants.$[elem].variant_image": None}},
                    array_filters=[{"elem.variant_image": filename}]
                )
                if result.modified_count == 0:
                    logger.warning(f"Image {filename} not found in variants for item {item_id}")
                    return jsonify({"error": "Image not found in variants"}), 404
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    logger.info(f"Image deleted from filesystem: {filename}")
                except PermissionError as e:
                    logger.error(f"Permission denied deleting image {filename}: {str(e)}")
                    return jsonify({"error": "Permission denied deleting image"}), 403
                except Exception as e:
                    logger.error(f"Error deleting image {filename}: {str(e)}")
                    return jsonify({"error": "Error deleting image"}), 500
            else:
                logger.warning(f"Image file not found on filesystem: {filename}")
            logger.info(f"Image {filename} deleted from {field} for item {item_id}")
            return jsonify({"message": "Image deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting image {filename} for item {item_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/images/<path:filename>', methods=['GET'])
def serve_uploaded_image(filename):
    logger.debug(f"Serving image: {filename}")
    try:
        mode = config.get("mode", "server")
        if mode == 'client':
            server_url = f"http://{config['server_ip']}:8000/api/images/{filename}"
            response = requests.get(server_url, stream=True)
            if response.status_code == 200:
                headers = {k: v for k, v in response.headers.items() if k.lower() in ('content-type', 'content-length')}
                return Response(response.iter_content(chunk_size=8192), headers=headers, status=200)
            else:
                return jsonify({"error": "Image not found"}), 404
        else:
            return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except Exception as e:
        logger.error(f"Error serving image {filename}: {str(e)}")
        return jsonify({"error": "Image not found"}), 404
@app.route('/api/import-mongodb', methods=['POST', 'OPTIONS'])
@db_required
def import_mongodb():
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Accept'
        return response, 200
    try:
        if 'file' not in request.files:
            logger.error("No file part in request")
            return jsonify({"error": "No file uploaded"}), 400
        file = request.files['file']
        if file.filename == '':
            logger.error("No selected file")
            return jsonify({"error": "No selected file"}), 400
        if not allowed_file(file.filename, ALLOWED_JSON_EXTENSIONS):
            logger.error(f"Invalid file type: {file.filename}")
            return jsonify({"error": "Only JSON files are allowed"}), 400
        filename = secure_filename(file.filename)
        collection_name = filename.rsplit('.', 1)[0].split('.')[-1]
        valid_collections = [
            'active_orders', 'combo_offers', 'customers', 'email_settings', 'email_tokens', 'employees', 'item_groups', 'items', 'kitchen_saved_orders', 'kitchens',
            'order_counters', 'picked_up_items', 'pos_closing_entries', 'pos_opening_entries', 'print_settings', 'purchase_invoices', 'purchase_items', 'purchase_orders',
            'purchase_receipts', 'purchase_sales', 'sales', 'suppliers', 'system_settings', 'tables', 'trip_reports', 'uoms', 'users', 'variants', 'vat', 'customer_groups', 'company_details', 'logo_details', 'supplier_groups', 'new_employee','employee_designations','employee_types','working_days'
        ]
        if collection_name not in valid_collections:
            logger.error(f"Invalid collection name: {collection_name}")
            return jsonify({"error": f"Unsupported collection name: {collection_name}"}), 400
        target_collection = SQLiteCollection(conn, collection_name)
        data = json.loads(file.read().decode('utf-8'))
        if not isinstance(data, list):
            logger.error("JSON data must be an array")
            return jsonify({"error": "JSON data must be an array"}), 400
        inserted_count = 0
        for record in data:
            if '_id' in record:
                record['_id'] = str(record['_id'])
            record['imported_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
            unique_key = (
                {'_id': record.get('_id')} if '_id' in record else
                {'email': record.get('email')} if collection_name == 'users' else
                {'table_number': record.get('table_number')} if collection_name == 'tables' else
                {'item_name': record.get('item_name')} if collection_name == 'items' else
                {'phone_number': record.get('phone_number')} if collection_name == 'customers' else
                {'invoice_no': record.get('invoice_no')} if collection_name == 'sales' else
                {'customerName': record.get('customerName')} if collection_name == 'picked_up_items' else
                {'name': record.get('name')} if collection_name in ['pos_opening_entries', 'pos_closing_entries'] else
                {'kitchen_name': record.get('kitchen_name')} if collection_name == 'kitchens' else
                {'group_name': record.get('group_name')} if collection_name == 'item_groups' else
                {'group_name': record.get('group_name')} if collection_name == 'customer_groups' else
                {'group_name': record.get('group_name')} if collection_name == 'supplier_groups' else
                {'order_id': record.get('order_id')} if collection_name == 'active_orders' else
                {'combo_name': record.get('combo_name')} if collection_name == 'combo_offers' else
                {'email': record.get('email')} if collection_name == 'email_settings' else
                {'token': record.get('token')} if collection_name == 'email_tokens' else
                {'employee_id': record.get('employee_id')} if collection_name == 'employees' else
                {'kitchen_order_id': record.get('kitchen_order_id')} if collection_name == 'kitchen_saved_orders' else
                {'counter_id': record.get('counter_id')} if collection_name == 'order_counters' else
                {'print_setting_id': record.get('print_setting_id')} if collection_name == 'print_settings' else
                {'invoice_no': record.get('invoice_no')} if collection_name == 'purchase_invoices' else
                {'item_name': record.get('item_name')} if collection_name == 'purchase_items' else
                {'order_no': record.get('order_no')} if collection_name == 'purchase_orders' else
                {'receipt_no': record.get('receipt_no')} if collection_name == 'purchase_receipts' else
                {'sale_id': record.get('sale_id')} if collection_name == 'purchase_sales' else
                {'supplier_name': record.get('supplier_name')} if collection_name == 'suppliers' else
                {'report_id': record.get('report_id')} if collection_name == 'trip_reports' else
                {'uom_name': record.get('uom_name')} if collection_name == 'uoms' else
                {'variant_name': record.get('variant_name')} if collection_name == 'variants' else
                {'vat_rate': record.get('vat_rate')} if collection_name == 'vat' else
                {'company_name': record.get('company_name')} if collection_name == 'company_details' else
                {'logo': record.get('logo')} if collection_name == 'logo_details' else
                {'name': record.get('name')} if collection_name == 'new_employee' else
                {'name': record.get('name')} if collection_name == 'employee_designations' else
                {'name': record.get('name')} if collection_name == 'employee_types' else
                {'year': record.get('year'), 'month': record.get('month')} if collection_name == 'working_days' else
                {}
            )
            if not unique_key:
                logger.error(f"No unique key defined for record in collection {collection_name}")
                return jsonify({"error": f"No unique key defined for record in collection {collection_name}"}), 400
            target_collection.replace_one(unique_key, record, upsert=True)
            inserted_count += 1
        logger.info(f"Imported {inserted_count} records into {collection_name}")
        return jsonify({"message": f"Successfully imported {inserted_count} records into {collection_name}"}), 200
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON format in file {filename}: {str(e)}")
        return jsonify({"error": f"Invalid JSON format: {str(e)}"}), 400
    except Exception as e:
        logger.error(f"Error importing data: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/save-email-settings', methods=['POST'])
@db_required
def save_email_settings():
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        from_email = data.get('from_email')
        if not all([email, password, from_email]):
            return jsonify({"error": "Missing required fields: email, password, from_email"}), 400
        settings = {
            '_id': 'email_settings',
            'email': email,
            'password': password,
            'from_email': from_email,
            'updated_at': datetime.now(ZoneInfo("UTC")).isoformat()
        }
        email_settings_collection.replace_one({'_id': 'email_settings'}, settings, upsert=True)
        logger.info(f"Email settings saved for {email}")
        return jsonify({"success": True, "message": "Email settings saved successfully"}), 200
    except Exception as e:
        logger.error(f"Error saving email settings: {str(e)}")
        return jsonify({"error": f"Failed to save email settings: {str(e)}"}), 500
@app.route('/api/get-email-settings', methods=['GET'])
@db_required
def get_email_settings():
    try:
        settings = email_settings_collection.find_one({'_id': 'email_settings'})
        if not settings:
            return jsonify({"success": False, "error": "No email settings found"}), 404
        return jsonify({"success": True, "email": settings.get('email'), "from_email": settings.get('from_email')}), 200
    except Exception as e:
        logger.error(f"Error retrieving email settings: {str(e)}")
        return jsonify({"success": False, "error": f"Failed to retrieve email settings: {str(e)}"}), 500
@app.route('/api/test-email-settings', methods=['POST'])
@db_required
def test_email_settings():
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        if not all([email, password]):
            return jsonify({"success": False, "error": "Missing required fields: email, password"}), 400
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(email, password)
        logger.info(f"SMTP authentication successful for {email}")
        return jsonify({"success": True, "message": "Email settings are valid"}), 200
    except smtplib.SMTPAuthenticationError as e:
        logger.error(f"SMTP Authentication Error during test: {str(e)}")
        return jsonify({"success": False, "error": "Invalid email or app password. Please check your credentials and ensure an App Password is used for Gmail."}), 401
    except smtplib.SMTPException as e:
        logger.error(f"SMTP Error during test: {str(e)}")
        return jsonify({"success": False, "error": f"SMTP error: {str(e)}"}), 500
    except Exception as e:
        logger.error(f"Unexpected error testing email settings: {str(e)}")
        return jsonify({"success": False, "error": f"Failed to test email settings: {str(e)}"}), 500
@app.route('/api/send-email', methods=['POST', 'OPTIONS'])
@db_required
def send_email():
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    try:
        data = request.get_json()
        logger.info(f"Received email request: {data}")
        if not data:
            logger.error("No data received in send-email request")
            return jsonify({"success": False, "message": "No data provided"}), 400
        to_email = data.get('to')
        subject = data.get('subject')
        html_content = data.get('html')
        if not all([to_email, subject, html_content]):
            logger.error("Missing required email fields")
            return jsonify({"success": False, "message": "Missing required fields: to, subject, html"}), 400
        settings = email_settings_collection.find_one({'_id': 'email_settings'})
        if not settings:
            logger.error("No email settings configured")
            return jsonify({"success": False, "message": "Email settings not configured. Please configure in Email Settings."}), 500
        email_user = settings.get('email')
        email_pass = settings.get('password')
        from_email = settings.get('from_email')
        msg = MIMEMultipart('alternative')
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(html_content, 'html'))
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(email_user, email_pass)
            server.send_message(msg)
            logger.info(f"Email sent successfully to {to_email}")
        return jsonify({"success": True, "message": "Email sent successfully"}), 200
    except smtplib.SMTPAuthenticationError as e:
        logger.error(f"SMTP Authentication Error: {str(e)}")
        return jsonify({"success": False, "message": "Invalid email or app password. Please check your Email Settings and ensure an App Password is used for Gmail."}), 401
    except smtplib.SMTPException as e:
        logger.error(f"SMTP Error: {str(e)}")
        return jsonify({"success": False, "message": f"SMTP error: {str(e)}"}), 500
    except Exception as e:
        logger.error(f"Unexpected error sending email: {str(e)}")
        return jsonify({"success": False, "message": f"Failed to send email: {str(e)}"}), 500
@app.route('/api/export-all-to-excel', methods=['GET'])
@db_required
def export_all_to_excel():
    """Export all data to an Excel file."""
    try:
        if openpyxl is None:
            logger.error("openpyxl not installed")
            return jsonify({"error": "Excel export not available. Please install openpyxl library."}), 500
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        collections = {
            'active_orders': activeorders_collection,
            'combo_offers': combo_offers_collection,
            'customers': customers_collection,
            'email_settings': email_settings_collection,
            'email_tokens': email_tokens_collection,
            'employees': employees_collection,
            'item_groups': item_groups_collection,
            'items': items_collection,
            'kitchen_saved_orders': kitchen_saved_collection,
            'kitchens': kitchens_collection,
            'order_counters': order_counters_collection,
            'picked_up_items': picked_up_collection,
            'pos_closing_entries': pos_closing_collection,
            'pos_opening_entries': opening_collection,
            'print_settings': print_settings_collection,
            'purchase_invoices': purchase_invoices_collection,
            'purchase_items': purchase_items_collection,
            'purchase_orders': purchase_orders_collection,
            'purchase_receipts': purchase_receipts_collection,
            'purchase_sales': purchase_sales_collection,
            'sales': sales_collection,
            'suppliers': suppliers_collection,
            'system_settings': settings_collection,
            'tables': tables_collection,
            'trip_reports': tripreports_collection,
            'uoms': uoms_collection,
            'users': users_collection,
            'variants': variants_collection,
            'vat': vat_collection,
            'customer_groups': customer_groups_collection,
            'company_details': company_details_collection,
            'logo_details': logo_details_collection,
            'supplier_groups': supplier_group_collection,
            'new_employee': worker_collection,
            'employee_designations': employee_designations_collection,
            'employee_types': employee_type_collection,
            'working_days': working_days_collection
        }
        for collection_name, collection in collections.items():
            ws = wb.create_sheet(title=collection_name)
            data = collection.find()
            if not data:
                ws.append(['No data'])
                continue
            sample_doc = data[0]
            headers = list(sample_doc.keys())
            ws.append(headers)
            for doc in data:
                row = [str(doc.get(header, '')) if isinstance(doc.get(header), (list, dict)) else doc.get(header, '') for header in headers]
                ws.append(row)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f'restaurant_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        logger.info(f"Exported data to Excel: {filename}")
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500
def manage_backup_limit():
    try:
        backup_files = [f for f in os.listdir(app.config['UPLOAD_FOLDER']) if f.endswith('.xlsx')]
        backup_files = sorted(
            backup_files,
            key=lambda x: os.path.getctime(os.path.join(app.config['UPLOAD_FOLDER'], x)),
            reverse=True
        )
        for old_file in backup_files[MAX_BACKUPS:]:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], old_file))
            logger.info(f"Deleted old backup: {old_file}")
    except Exception as e:
        logger.error(f"Error managing backup limit: {str(e)}")
def create_backup():
    try:
        if openpyxl is None:
            return False, "Excel library not available. Please install openpyxl."
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        collections = {
            'active_orders': activeorders_collection,
            'combo_offers': combo_offers_collection,
            'customers': customers_collection,
            'email_settings': email_settings_collection,
            'email_tokens': email_tokens_collection,
            'employees': employees_collection,
            'item_groups': item_groups_collection,
            'items': items_collection,
            'kitchen_saved_orders': kitchen_saved_collection,
            'kitchens': kitchens_collection,
            'order_counters': order_counters_collection,
            'picked_up_items': picked_up_collection,
            'pos_closing_entries': pos_closing_collection,
            'pos_opening_entries': opening_collection,
            'print_settings': print_settings_collection,
            'purchase_invoices': purchase_invoices_collection,
            'purchase_items': purchase_items_collection,
            'purchase_orders': purchase_orders_collection,
            'purchase_receipts': purchase_receipts_collection,
            'purchase_sales': purchase_sales_collection,
            'sales': sales_collection,
            'suppliers': suppliers_collection,
            'system_settings': settings_collection,
            'tables': tables_collection,
            'trip_reports': tripreports_collection,
            'uoms': uoms_collection,
            'users': users_collection,
            'variants': variants_collection,
            'vat': vat_collection,
            'new_employee': worker_collection,
            'employee_designations': employee_designations_collection,
            'employee_types': employee_type_collection,
            'customer_groups': customer_groups_collection,
            'company_details': company_details_collection,
            'logo_details': logo_details_collection,
            'supplier_groups': supplier_group_collection,
            'working_days': working_days_collection
        }
        for collection_name, collection in collections.items():
            ws = wb.create_sheet(title=collection_name)
            data = collection.find()
            if not data:
                ws.append(['No data'])
                continue
            sample_doc = data[0]
            headers = list(sample_doc.keys())
            ws.append(headers)
            for doc in data:
                row = [str(doc.get(header, '')) if isinstance(doc.get(header), (list, dict)) else doc.get(header, '') for header in headers]
                ws.append(row)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'backup_restaurant_data_{timestamp}.xlsx'
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        with open(file_path, 'wb') as f:
            f.write(buffer.getvalue())
        manage_backup_limit()
        settings = email_settings_collection.find_one({'_id': 'email_settings'})
        if not settings:
            logger.error("No email settings configured")
            return False, "Email settings not configured. Please configure in Email Settings."
        email_user = settings.get('email')
        email_pass = settings.get('password')
        from_email = settings.get('from_email')
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = email_user
        msg['Subject'] = f'Restaurant Data Backup - {timestamp}'
        body = f'Backup of restaurant data generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}.'
        msg.attach(MIMEText(body, 'plain'))
        with open(file_path, 'rb') as f:
            attachment = MIMEBase('application', 'octet-stream')
            attachment.set_payload(f.read())
            encoders.encode_base64(attachment)
            attachment.add_header('Content-Disposition', f'attachment; filename={filename}')
            msg.attach(attachment)
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(email_user, email_pass)
            server.send_message(msg)
        logger.info(f"Backup created and emailed: {filename}")
        return True, f"Backup created successfully: {filename}"
    except smtplib.SMTPAuthenticationError as e:
        logger.error(f"SMTP Authentication Error: {str(e)}")
        return False, "Invalid email or app password. Please check your Email Settings and ensure an App Password is used for Gmail."
    except smtplib.SMTPException as e:
        logger.error(f"SMTP Error: {str(e)}")
        return False, f"SMTP error: {str(e)}"
    except Exception as e:
        logger.error(f"Error in backup: {str(e)}")
        return False, str(e)
@app.route('/api/backup-to-excel', methods=['GET'])
@db_required
def backup_to_excel():
    """Create a backup and serve it as a download."""
    try:
        success, message = create_backup()
        if not success:
            return jsonify({"error": message}), 500
        filename = message.split(': ')[1]
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        with open(file_path, 'rb') as f:
            file_data = f.read()
        return Response(
            file_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        logger.error(f"Error serving backup file: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500
@app.route('/api/backup-info', methods=['GET'])
@db_required
def backup_info():
    """Retrieve information about existing backups."""
    try:
        backup_files = [f for f in os.listdir(app.config['UPLOAD_FOLDER']) if f.endswith('.xlsx')]
        backups = []
        for filename in backup_files:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            stat = os.stat(file_path)
            backups.append({
                'filename': filename,
                'date': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                'size': f"{stat.st_size / 1024:.2f} KB"
            })
        backups.sort(key=lambda x: x['date'], reverse=True)
        return jsonify(backups)
    except Exception as e:
        logger.error(f"Error retrieving backup info: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500
@app.route('/api/download-backup', methods=['POST'])
@db_required
def download_backup():
    """Download a specific backup file."""
    try:
        data = request.get_json()
        filename = data.get('filename')
        if not filename:
            return jsonify({"error": "Filename not provided"}), 400
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if not os.path.exists(file_path):
            return jsonify({"error": "Backup file not found"}), 404
        with open(file_path, 'rb') as f:
            file_data = f.read()
        return Response(
            file_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        logger.error(f"Error downloading backup {filename}: {str(e)}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500
@app.route('/api/get-backup-interval', methods=['GET'])
@db_required
def get_backup_interval():
    try:
        settings = get_system_settings()
        interval = settings.get('backup_interval_hours', 6)
        return jsonify({"interval": interval}), 200
    except Exception as e:
        logger.error(f"Error fetching backup interval: {str(e)}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/set-backup-interval', methods=['POST'])
@db_required
def set_backup_interval():
    try:
        data = request.get_json()
        interval = data.get('interval')
        if interval is None or not isinstance(interval, int) or interval <= 0:
            return jsonify({"error": "Invalid interval. Must be a positive integer."}), 400
        settings = get_system_settings()
        settings['backup_interval_hours'] = interval
        save_system_settings(settings)
        # Update scheduler
        if schedule:
            schedule.clear('backup')
            schedule.every(interval).hours.do(create_backup).tag('backup')
            logger.info(f"Backup interval updated to every {interval} hours")
        return jsonify({"message": "Backup interval updated successfully"}), 200
    except Exception as e:
        logger.error(f"Error setting backup interval: {str(e)}")
        return jsonify({"error": str(e)}), 500
def start_scheduler():
    if schedule:
        settings = get_system_settings()
        interval = settings.get('backup_interval_hours', 6)
        schedule.every(interval).hours.do(create_backup).tag('backup')
        threading.Thread(target=run_scheduler, daemon=True).start()
        logger.info(f"Scheduler started with backup every {interval} hours")
def run_scheduler():
    while True:
        schedule.run_pending()
        time.sleep(60)
@app.route('/api/shutdown', methods=['POST'])
@db_required
def shutdown():
    global shutdown_flag
    logger.info("Shutdown requested")
    try:
        func = request.environ.get('werkzeug.server.shutdown')
        if func:
            func()
            logger.info("Werkzeug server shutdown initiated")
            return jsonify({"message": "Server shutting down"}), 200
        shutdown_flag = True
        logger.info("Setting shutdown flag for Waitress")
        def exit_process():
            time.sleep(1)
            logger.info("Exiting Python process")
            os._exit(0)
        threading.Thread(target=exit_process, daemon=True).start()
        return jsonify({"message": "Server shutting down"}), 200
    except Exception as e:
        logger.error(f"Shutdown error: {str(e)}")
        return jsonify({"message": "Error during shutdown", "error": str(e)}), 500
# NEW: Endpoint to get hidden items with sales details
@app.route('/api/hidden-items', methods=['GET'])
@db_required
def get_hidden_items():
    try:
        hidden_items = items_collection.find({'is_hidden': True})
        enhanced_items = []
        for item in hidden_items:
            item_name = item.get('item_name')
            sales = get_sales_for_item(item_name)
            enhanced_item = convert_objectid_to_str(item)
            if sales:
                enhanced_item['sales'] = [convert_objectid_to_str(sale) for sale in sales]
                total_records = len(sales)
                total_qty_sold = sum(sum(item_obj.get('quantity', 0) for item_obj in sale.get('items', []) if item_obj.get('item_name') == item_name) for sale in sales)
                subtotal = sum(float(sale.get('total', 0)) for sale in sales)
                vat_total = sum(float(sale.get('vat_amount', 0)) for sale in sales)
                grand_total = sum(float(sale.get('grand_total', 0)) for sale in sales)
                currency = sales[0].get('invoice_currency', 'AED')
            else:
                enhanced_item['sales'] = []
                total_records = 0
                total_qty_sold = 0
                subtotal = 0
                vat_total = 0
                grand_total = 0
                currency = 'AED'
            enhanced_item['summary'] = {
                'total_records': total_records,
                'total_qty_sold': total_qty_sold,
                'currency': currency,
                'subtotal': round(subtotal, 2),
                'vat': round(vat_total, 2),
                'grand_total': round(grand_total, 2)
            }
            enhanced_items.append(enhanced_item)
        return jsonify(enhanced_items), 200
    except Exception as e:
        logger.error(f"Error fetching hidden items: {str(e)}")
        return jsonify({"error": str(e)}), 500
# NEW: Endpoint to get sales for an item
@app.route('/api/items/<item_id>/sales', methods=['GET'])
@db_required
def get_item_sales(item_id):
    try:
        item = items_collection.find_one({'_id': item_id})
        if not item:
            return jsonify({"error": "Item not found"}), 404
        item_name = item.get('item_name')
        sales = get_sales_for_item(item_name)
        sales = [convert_objectid_to_str(sale) for sale in sales]
        return jsonify(sales), 200
    except Exception as e:
        logger.error(f"Error fetching sales for item {item_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500
# NEW: Endpoint to delete a sale by invoice_no
@app.route('/api/sales/<invoice_no>', methods=['DELETE'])
@db_required
def delete_sale(invoice_no):
    try:
        result = sales_collection.delete_one({'invoice_no': invoice_no.strip()})
        if result.deleted_count == 0:
            return jsonify({"error": "Sale not found"}), 404
        logger.info(f"Sale deleted: {invoice_no}")
        return jsonify({"message": "Sale deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting sale {invoice_no}: {str(e)}")
        return jsonify({"error": str(e)}), 500
# NEW: Force delete item endpoint (deletes regardless of sales)
@app.route('/api/items/<item_id>/force-delete', methods=['DELETE'])
@db_required
def force_delete_item(item_id):
    try:
        item = items_collection.find_one({'_id': item_id})
        if not item:
            return jsonify({"error": "Item not found"}), 404
        result = items_collection.delete_one({'_id': item_id})
        if result.deleted_count == 0:
            return jsonify({"error": "Item not found"}), 404
        logger.info(f"Item force deleted: {item_id} ({item.get('item_name', 'Unknown')})")
        return jsonify({"message": "Item deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error force deleting item {item_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500
# --- Data Routes (only in server mode) ---
if config.get('mode') == 'server':
    @app.route('/api/login', methods=['POST'])
    @db_required
    def login():
        try:
            data = request.get_json()
            settings = get_system_settings()
            identifier = data.get('identifier')
            password = data.get('password')
            login_type = data.get('type', 'mobile_or_username')
            if not identifier or not password:
                logger.error("Identifier or password missing in login request")
                return jsonify({"error": "Identifier and password are required"}), 400
            user = None
            if login_type == 'mobile_or_username':
                query = {"$or": [
                    {"phone_number": identifier},
                    {"username": identifier},
                    {"email": identifier}
                ]}
                user = users_collection.find_one(query)
            else:
                logger.error(f"Invalid login type: {login_type}")
                return jsonify({"error": "Invalid login type"}), 400
            if not user:
                logger.warning(f"Invalid login attempt: {identifier}")
                return jsonify({"error": "Invalid credentials"}), 401
            try:
                if not bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):
                    logger.warning(f"Invalid password for: {identifier}")
                    return jsonify({"error": "Invalid credentials"}), 401
            except ValueError as ve:
                if "Invalid salt" in str(ve):
                    logger.error(f"Invalid salt for user {user.get('email')}, rehashing")
                    new_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                    users_collection.update_one({"_id": user['_id']}, {"$set": {"password": new_hash}})
                    user['password'] = new_hash
                    if not bcrypt.checkpw(password.encode('utf-8'), new_hash.encode('utf-8')):
                        return jsonify({"error": "Invalid credentials"}), 401
                else:
                    raise
            # FIXED: Compute consistent user_identifier to match frontend logic - Use email as key for consistency
            user_identifier = user.get('email', '').split('@')[0] if '@' in user.get('email', '') else user.get('email', '')
            logger.info(f"Computed user_identifier for query: '{user_identifier}' from user doc: email='{user.get('email')}', username='{user.get('username')}'")
            token_payload = {
                'user_id': user['_id'],
                'exp': datetime.now(timezone.utc) + timedelta(seconds=JWT_EXP_DELTA_SECONDS)
            }
            token = jwt.encode(token_payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
            user_converted = convert_objectid_to_str(user)
            # Enhanced logic for requires_opening_entry and requires_closing_entry using consistent user_identifier
            today_start = datetime.now(ZoneInfo("UTC")).replace(hour=0, minute=0, second=0, microsecond=0)
            today_date_str = today_start.strftime("%Y-%m-%d")
            logger.info(f"Today's date for query: {today_date_str}")
            # FIXED: Changed to exact match for posting_date to avoid $gte string comparison issues
            # UPDATED: Filter only open shifts (status="Open") to require new opening after closing
            opening_today = opening_collection.find_one({
                "user": user_identifier,
                "posting_date": today_date_str,
                "status": "Open"  # Only consider open shifts; closed ones are ignored
            })
            logger.info(f"Found opening_today: {opening_today is not None}, details: {opening_today.get('name', 'None') if opening_today else 'None'} (Query: user='{user_identifier}', posting_date='{today_date_str}', status='Open')")
            requires_opening_entry = opening_today is None
            requires_closing_entry = False
            pos_opening_entry = None
            if opening_today:
                # Check if closing exists for this opening
                closing_for_opening = pos_closing_collection.find_one({
                    "pos_opening_entry": opening_today['name']
                })
                logger.info(f"Found closing for opening {opening_today['name']}: {closing_for_opening is not None}")
                if closing_for_opening is None:
                    # FIXED: Only require closing if there are sales associated with this opening
                    # FIXED: Replace count_documents (MongoDB method) with len(list(find(...))) for SQLiteCollection compatibility
                    sales = list(sales_collection.find({
                        "pos_opening_entry": opening_today['name']
                    }))
                    sales_count = len(sales)
                    logger.info(f"Sales count for opening {opening_today['name']}: {sales_count}")
                    requires_closing_entry = sales_count > 0
                    if requires_closing_entry:
                        pos_opening_entry = opening_today['name'] # Provide the opening name for closing page
                        logger.info(f"Setting requires_closing_entry=True, pos_opening_entry={pos_opening_entry}")
            response = {
                "message": "Login successful",
                "token": token,
                "user": {
                    "id": user_converted['_id'],
                    "username": user_identifier,
                    "role": user.get('role', 'bearer'),
                    "email": user.get('email', ''),
                    "phone_number": user.get('phone_number', ''),
                    "pos_profile": user.get('pos_profile', 'POS-001'),
                    "company": user.get('company', 'POS 8'),
                    "is_test": user.get('is_test', False)
                },
                "requires_opening_entry": requires_opening_entry,
                "requires_closing_entry": requires_closing_entry,
                "pos_opening_entry": pos_opening_entry # NEW: Pass opening name if closing required
            }
            logger.info(f"Login response flags - requires_opening: {requires_opening_entry}, requires_closing: {requires_closing_entry}, pos_opening_entry: {pos_opening_entry}")
            logger.info(f"User logged in: {identifier}")
            return jsonify(response), 200
        except KeyError as ke:
            logger.error(f"KeyError during login response building: {str(ke)}")
            return jsonify({"error": f"Internal error: Missing user field {str(ke)}"}), 500
        except Exception as e:
            logger.error(f"Login error: {e}\n{traceback.format_exc()}")
            return jsonify({"error": f"An internal server error occurred during login: {str(e)}"}), 500
    @app.route('/api/register', methods=['POST'])
    @db_required
    def register():
        try:
            data = request.get_json()
            required_fields = ['email', 'password', 'firstName', 'phone_number']
            if not all(field in data for field in required_fields):
                return jsonify({"message": "Missing required fields"}), 400
            if users_collection.find_one({"email": data['email']}):
                return jsonify({"message": "Email already registered"}), 400
            hashed_password = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            new_user = {
                "email": data['email'],
                "password": hashed_password,
                "role": data.get('role', 'bearer'),
                "username": data.get('username', data['firstName']),
                "firstName": data['firstName'],
                "phone_number": data['phone_number'],
                "company": data.get('company', 'POS 8'),
                "pos_profile": data.get('pos_profile', 'POS-001'),
                "status": "Active",
                "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }
            result = users_collection.insert_one(new_user)
            logger.info(f"User registered: {data['email']}")
            return jsonify({"message": "Registration successful", "userId": new_user['_id']}), 201
        except Exception as e:
            logger.error(f"Registration failed: {str(e)}")
            return jsonify({"message": f"Registration failed: {str(e)}"}), 500
    @app.route('/api/users', methods=['GET'])
    @db_required
    def get_users():
        try:
            users = users_collection.find()
            return jsonify(convert_objectid_to_str(users)), 200
        except Exception as e:
            logger.error(f"Error fetching users: {str(e)}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/users/<email>', methods=['DELETE'])
    @db_required
    def delete_user(email):
        try:
            if email in [u['email'] for u in TEST_USERS]:
                return jsonify({"message": "Cannot delete test users"}), 400
            # ADD: Delete from worker_collection if exists (for employees)
            worker = worker_collection.find_one({"email": email})
            if worker:
                worker_collection.delete_one({"email": email})
                logger.info(f"Employee deleted from new_employee: {email}")
            result = users_collection.delete_one({"email": email})
            if result.deleted_count == 0:
                return jsonify({"message": "User not found"}), 404
            logger.info(f"User deleted: {email}")
            return jsonify({"message": "User deleted successfully"}), 200
        except Exception as e:
            logger.error(f"Error deleting user {email}: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/settings', methods=['GET'])
    @db_required
    def get_settings_route():
        try:
            settings = get_system_settings()
            return jsonify(convert_objectid_to_str(settings)), 200
        except Exception as e:
            logger.error(f"Error fetching settings: {e}")
            return jsonify({"error": "An internal server error occurred."}), 500
    @app.route('/api/settings', methods=['POST'])
    @db_required
    def update_settings():
        try:
            data = request.get_json()
            save_system_settings(data)
            logger.info("System settings updated")
            return jsonify({"message": "Settings updated successfully"}), 200
        except Exception as e:
            logger.error(f"Error updating settings: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/items', methods=['GET'])
    @db_required
    def get_items():
        try:
            items = items_collection.find()
            items_list = []
            current_time = datetime.now(ZoneInfo("UTC"))
            placeholder_url = 'https://placehold.co/100x100/EFEFEF/AAAAAA?text=No+Image'
            for item in items:
                if item.get('is_hidden', False):
                    continue # Skip hidden items in main list
                item = convert_objectid_to_str(item)
                is_offer_active = False
                if 'offer_start_time' in item and item['offer_start_time'] and 'offer_end_time' in item and item['offer_end_time']:
                    try:
                        offer_start_time = datetime.fromisoformat(str(item['offer_start_time']).replace('Z', '+00:00'))
                        offer_end_time = datetime.fromisoformat(str(item['offer_end_time']).replace('Z', '+00:00'))
                        if offer_start_time <= current_time <= offer_end_time:
                            is_offer_active = True
                        else:
                            items_collection.update_one(
                                {'_id': item['_id']},
                                {'$unset': {'offer_price': "", 'offer_start_time': "", 'offer_end_time': ""}}
                            )
                            item.pop('offer_price', None)
                            item.pop('offer_start_time', None)
                            item.pop('offer_end_time', None)
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Invalid offer time format for item {item['_id']}: {str(e)}")
                        items_collection.update_one(
                            {'_id': item['_id']},
                            {'$unset': {'offer_price': "", 'offer_start_time': "", 'offer_end_time': ""}}
                        )
                        item.pop('offer_price', None)
                        item.pop('offer_start_time', None)
                        item.pop('offer_end_time', None)
                if not is_offer_active:
                    item.pop('offer_price', None)
                    item.pop('offer_start_time', None)
                    item.pop('offer_end_time', None)
                # Ensure tax_rate is 0 if tax_applicable is False
                if not item.get('tax_applicable', False):
                    item['tax_rate'] = 0
                for addon in item.get('addons', []):
                    if not addon.get('tax_applicable', False):
                        addon['tax_rate'] = 0
                    if addon.get('addon_image'):
                        addon['addon_image'] = f"/api/images/{os.path.basename(addon['addon_image'])}"
                    else:
                        addon['addon_image'] = placeholder_url
                for combo in item.get("combos", []):
                    if not combo.get('tax_applicable', False):
                        combo['tax_rate'] = 0
                    if combo.get('combo_image'):
                        combo['combo_image'] = f"/api/images/{os.path.basename(combo['combo_image'])}"
                    else:
                        combo['combo_image'] = placeholder_url
                if item.get('image'):
                    item['image'] = f"/api/images/{os.path.basename(item['image'])}"
                else:
                    item['image'] = placeholder_url
                items_list.append(item)
                logger.info(f"Fetched {len(items_list)} items")
            return jsonify(items_list), 200
        except Exception as e:
            logger.error(f"Error fetching items: {str(e)}\n{traceback.format_exc()}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/items/<identifier>', methods=['GET'])
    @db_required
    def get_item(identifier):
        try:
            item = items_collection.find_one({'_id': identifier})
            if not item:
                item = items_collection.find_one({'item_name': identifier})
            if not item:
                logger.warning(f"Item not found: {identifier}")
                return jsonify({"error": "Item not found"}), 404
            item = convert_objectid_to_str(item)
            current_time = datetime.now(ZoneInfo("UTC"))
            placeholder_url = 'https://placehold.co/100x100/EFEFEF/AAAAAA?text=No+Image'
            is_offer_active = False
            if 'offer_start_time' in item and item['offer_start_time'] and 'offer_end_time' in item and item['offer_end_time']:
                try:
                    offer_start_time = datetime.fromisoformat(str(item['offer_start_time']).replace('Z', '+00:00'))
                    offer_end_time = datetime.fromisoformat(str(item['offer_end_time']).replace('Z', '+00:00'))
                    if offer_start_time <= current_time <= offer_end_time:
                        is_offer_active = True
                except (ValueError, TypeError):
                    items_collection.update_one(
                        {'_id': item['_id']},
                        {'$unset': {'offer_price': "", 'offer_start_time': "", 'offer_end_time': ""}}
                    )
                    item.pop('offer_price', None)
                    item.pop('offer_start_time', None)
                    item.pop('offer_end_time', None)
            if not is_offer_active:
                item.pop('offer_price', None)
                item.pop('offer_start_time', None)
                item.pop('offer_end_time', None)
            # Ensure tax_rate is 0 if tax_applicable is False
            if not item.get('tax_applicable', False):
                item['tax_rate'] = 0
            for addon in item.get('addons', []):
                if not addon.get('tax_applicable', False):
                    addon['tax_rate'] = 0
                if addon.get('addon_image'):
                    addon['addon_image'] = f"/api/images/{os.path.basename(addon['addon_image'])}"
                else:
                    addon['addon_image'] = placeholder_url
            for combo in item.get("combos", []):
                if not combo.get('tax_applicable', False):
                    combo['tax_rate'] = 0
                if combo.get('combo_image'):
                    combo['combo_image'] = f"/api/images/{os.path.basename(combo['combo_image'])}"
                else:
                    combo['combo_image'] = placeholder_url
            if item.get('image'):
                item['image'] = f"/api/images/{os.path.basename(item['image'])}"
            else:
                item['image'] = placeholder_url
            logger.info(f"Fetched item: {identifier}")
            return jsonify(item), 200
        except Exception as e:
            logger.error(f"Error fetching item {identifier}: {str(e)}\n{traceback.format_exc()}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/items', methods=['POST'])
    @db_required
    def create_item():
        try:
            data = request.json
            if not data:
                logger.error("No data provided for item creation")
                return jsonify({"error": "No data provided"}), 400
            required_fields = ['item_name', 'item_code', 'item_group', 'price_list_rate']
            for field in required_fields:
                if field not in data or not data[field]:
                    logger.error(f"Missing or empty required field: {field}")
                    return jsonify({"error": f"Missing or empty required field: {field}"}), 400
            if 'offer_start_time' in data and data['offer_start_time'] and 'offer_end_time' in data and data['offer_end_time']:
                try:
                    offer_start_time = datetime.fromisoformat(str(data['offer_start_time']).replace('Z', '+00:00'))
                    offer_end_time = datetime.fromisoformat(str(data['offer_end_time']).replace('Z', '+00:00'))
                    if offer_start_time >= offer_end_time:
                        logger.error("offer_start_time must be before offer_end_time")
                        return jsonify({"error": "Offer start time must be before offer end time"}), 400
                except (ValueError, TypeError) as e:
                    logger.error(f"Invalid offer time format: {str(e)}")
                    return jsonify({"error": f"Invalid offer time format: {str(e)}"}), 400
            data = sanitize_image_fields(data)
            # Ensure tax_rate is 0 if tax_applicable is False
            if not data.get('tax_applicable', False):
                data['tax_rate'] = 0
            for addon in data.get('addons', []):
                if not addon.get('tax_applicable', False):
                    addon['tax_rate'] = 0
            for combo in data.get('combos', []):
                if not combo.get('tax_applicable', False):
                    combo['tax_rate'] = 0
            data.setdefault('custom_addon_applicable', False)
            data.setdefault('custom_combo_applicable', False)
            data.setdefault('custom_total_calories', 0)
            data.setdefault('custom_total_protein', 0)
            data.setdefault('kitchen', "")
            data.setdefault('has_variant_pricing', False)
            data.setdefault('variant_prices', {"small_price": 0, "medium_price": 0, "large_price": 0})
            data.setdefault('variant_quantities', {"small_quantity": 0, "medium_quantity": 0, "large_quantity": 0})
            data.setdefault('sold_quantities', {"small_sold": 0, "medium_sold": 0, "large_sold": 0})
            data.setdefault('ice_preference', "without_ice")
            data.setdefault('ice_price', 0)
            data.setdefault('addons', [])
            data.setdefault('combos', [])
            data.setdefault('ingredients', [])
            data.setdefault('variants', [])
            data['created_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
            item_id = items_collection.insert_one(data).inserted_id
            logger.info(f"Item created with ID: {item_id}")
            return jsonify({'message': 'Item created successfully!', 'id': item_id}), 201
        except Exception as e:
            logger.error(f"Error creating item: {str(e)}")
            return jsonify({'error': str(e)}), 500
    @app.route('/api/items/<item_id>', methods=['PUT'])
    @db_required
    def update_item(item_id):
        try:
            data = request.json
            if not data:
                logger.error("No data provided for item update")
                return jsonify({"error": "No data provided"}), 400
            if '_id' in data:
                del data['_id']
            if 'offer_start_time' in data and data['offer_start_time'] and 'offer_end_time' in data and data['offer_end_time']:
                try:
                    offer_start_time = datetime.fromisoformat(str(data['offer_start_time']).replace('Z', '+00:00'))
                    offer_end_time = datetime.fromisoformat(str(data['offer_end_time']).replace('Z', '+00:00'))
                    if offer_start_time >= offer_end_time:
                        logger.error("offer_start_time must be before offer_end_time")
                        return jsonify({"error": "Offer start time must be before offer end time"}), 400
                except (ValueError, TypeError) as e:
                    logger.error(f"Invalid offer time format: {str(e)}")
                    return jsonify({"error": f"Invalid offer time format: {str(e)}"}), 400
            data = sanitize_image_fields(data)
            # Ensure tax_rate is 0 if tax_applicable is False
            if not data.get('tax_applicable', False):
                data['tax_rate'] = 0
            for addon in data.get('addons', []):
                if not addon.get('tax_applicable', False):
                    addon['tax_rate'] = 0
            for combo in data.get('combos', []):
                if not combo.get('tax_applicable', False):
                    combo['tax_rate'] = 0
            data['modified_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
            result = items_collection.update_one({'_id': item_id}, {'$set': data})
            if result.matched_count == 0:
                logger.warning(f"Item not found for update: {item_id}")
                return jsonify({"error": "Item not found"}), 404
            logger.info(f"Item updated: {item_id}")
            return jsonify({"message": "Item updated successfully"}), 200
        except Exception as e:
            logger.error(f"Error updating item {item_id}: {str(e)}")
            return jsonify({"error": str(e)}), 500
    # NEW: Unhide item endpoint
    @app.route('/api/items/<item_id>/unhide', methods=['PATCH'])
    @db_required
    def unhide_item(item_id):
        try:
            result = items_collection.update_one(
                {'_id': item_id},
                {'$set': {'is_hidden': False, 'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
            )
            if result.matched_count == 0:
                return jsonify({"error": "Item not found"}), 404
            logger.info(f"Item unhidden: {item_id}")
            return jsonify({"message": "Item unhidden successfully"}), 200
        except Exception as e:
            logger.error(f"Error unhiding item {item_id}: {str(e)}")
            return jsonify({"error": str(e)}), 500
    # Modified: Delete item - hide if has sales
    @app.route('/api/items/<item_id>', methods=['DELETE'])
    @db_required
    def delete_item(item_id):
        try:
            item = items_collection.find_one({'_id': item_id})
            if not item:
                logger.warning(f"Item not found for deletion: {item_id}")
                return jsonify({"error": "Item not found"}), 404
            item_name = item.get('item_name')
            has_sales = has_associated_sales(item_name)
            if has_sales:
                # Hide instead of delete
                result = items_collection.update_one(
                    {'_id': item_id},
                    {'$set': {'is_hidden': True, 'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
                )
                logger.info(f"Item hidden due to associated sales: {item_id} ({item_name})")
                return jsonify({"message": f"Item hidden because it has associated sales. (Item: {item_name})"}), 200
            else:
                # Delete if no sales
                result = items_collection.delete_one({'_id': item_id})
                if result.deleted_count == 0:
                    logger.warning(f"Item not found for deletion: {item_id}")
                    return jsonify({"error": "Item not found"}), 404
                logger.info(f"Item deleted: {item_id} ({item_name})")
                return jsonify({"message": "Item deleted successfully"}), 200
        except Exception as e:
            logger.error(f"Error handling item {item_id}: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/items/<item_id>/offer', methods=['PUT'])
    @db_required
    def update_item_offer(item_id):
        try:
            offer_data = request.json
            if 'offer_price' not in offer_data or 'offer_start_time' not in offer_data or 'offer_end_time' in offer_data:
                return jsonify({"error": "Offer price, start time, and end time are required"}), 400
            try:
                offer_start_time = datetime.fromisoformat(str(offer_data['offer_start_time']).replace('Z', '+00:00'))
                offer_end_time = datetime.fromisoformat(str(offer_data['offer_end_time']).replace('Z', '+00:00'))
                if offer_start_time >= offer_end_time:
                    logger.error("offer_start_time must be before offer_end_time")
                    return jsonify({"error": "Offer start time must be before offer end time"}), 400
            except (ValueError, TypeError) as e:
                logger.error(f"Invalid offer time format: {str(e)}")
                return jsonify({"error": f"Invalid offer time format: {str(e)}"}), 400
            offer_data['modified_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
            result = items_collection.update_one({'_id': item_id}, {'$set': offer_data})
            if result.matched_count == 0:
                logger.warning(f"Item not found for offer update: {item_id}")
                return jsonify({"error": "Item not found"}), 404
            logger.info(f"Offer updated for item: {item_id}")
            return jsonify({"message": "Offer updated successfully"}), 200
        except Exception as e:
            logger.error(f"Error updating offer for item {item_id}: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/customers', methods=['GET'])
    @db_required
    def get_all_customers():
        try:
            customers = customers_collection.find()
            customers = [convert_objectid_to_str(customer) for customer in customers]
            logger.info(f"Fetched {len(customers)} customers")
            return jsonify(customers), 200
        except Exception as e:
            logger.error(f"Error fetching customers: {str(e)}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/customers/<customer_id>', methods=['GET', 'PUT', 'DELETE'])
    @db_required
    def customer_operations(customer_id):
        try:
            if not customer_id or customer_id == "undefined":
                logger.error("Invalid customer_id: 'undefined' or empty")
                return jsonify({"error": "Invalid customer ID"}), 400
            
            # FIXED: No longer convert to ObjectId; treat customer_id as str (UUID)
            # customer_oid = ObjectId(customer_id)  # REMOVED - causes InvalidId for UUID
            
            if request.method == 'GET':
                customer = customers_collection.find_one({'_id': customer_id})  # Use str customer_id
                if not customer:
                    logger.warning(f"Customer not found: {customer_id}")
                    return jsonify({"error": "Customer not found"}), 404
                customer = convert_objectid_to_str(customer)
                logger.info(f"Fetched customer: {customer_id}")
                return jsonify(customer), 200
            elif request.method == 'PUT':
                customer_data = request.get_json()
                if not customer_data:
                    logger.error("No data provided for customer update")
                    return jsonify({"error": "No data provided"}), 400
                result = customers_collection.update_one(
                    {'_id': customer_id},  # Use str customer_id
                    {'$set': {
                        'customer_name': customer_data.get('customer_name', ''),
                        'phone_number': customer_data.get('phone_number', ''),
                        'whatsapp_number': customer_data.get('whatsapp_number', ''),
                        'email': customer_data.get('email', ''),
                        'building_name': customer_data.get('building_name', ''),
                        'flat_villa_no': customer_data.get('flat_villa_no', ''),
                        'country': customer_data.get('country', ''),
                        'field1': customer_data.get('field1', ''),
                        'field2': customer_data.get('field2', ''),
                        'field3': customer_data.get('field3', ''),
                        'customer_group': customer_data.get('customer_group', ''),
                        'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()
                    }}
                )
                if result.matched_count == 0:
                    logger.warning(f"Customer not found for update: {customer_id}")
                    return jsonify({"error": "Customer not found"}), 404
                # Fetch and return the updated customer
                updated_customer = customers_collection.find_one({'_id': customer_id})
                updated_customer = convert_objectid_to_str(updated_customer)
                logger.info(f"Customer updated: {customer_id}")
                return jsonify(updated_customer), 200
            elif request.method == 'DELETE':
                result = customers_collection.delete_one({'_id': customer_id})  # Use str customer_id
                if result.deleted_count == 0:
                    logger.warning(f"Customer not found for deletion: {customer_id}")
                    return jsonify({"error": "Customer not found"}), 404
                logger.info(f"Customer deleted: {customer_id}")
                return jsonify({"message": "Customer deleted successfully"}), 200
        except Exception as e:
            logger.error(f"Error in customer operations for {customer_id}: {str(e)}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/customers', methods=['POST'])
    @db_required
    def create_customer():
        try:
            customer_data = request.get_json()
            if not customer_data or 'customer_name' not in customer_data or 'phone_number' not in customer_data:
                logger.error("Invalid customer data provided")
                return jsonify({"error": "Customer name and phone number are required"}), 400
            existing_customer = customers_collection.find_one({'phone_number': customer_data['phone_number']})
            if existing_customer:
                logger.warning(f"Duplicate phone number: {customer_data['phone_number']}")
                return jsonify({"error": "Phone number already exists", "customer_name": existing_customer.get('customer_name', 'existing customer')}), 409
            
            # FIXED: Set _id as UUID string for consistency with frontend and other entities
            customer_data['_id'] = str(uuid.uuid4())
            customer_data['created_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
            customer_data['modified_at'] = customer_data['created_at']
            result = customers_collection.insert_one(customer_data)
            new_customer_id = customer_data['_id']  # Use the set UUID
            # Fetch and return the created customer
            new_customer = customers_collection.find_one({'_id': new_customer_id})
            new_customer = convert_objectid_to_str(new_customer)
            logger.info(f"Customer created: {new_customer_id}")
            return jsonify(new_customer), 201
        except Exception as e:
            logger.error(f"Error creating customer: {str(e)}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/customer-groups', methods=['GET'])
    @db_required
    def get_customer_groups():
        try:
            groups = list(customer_groups_collection.find())
            groups = convert_objectid_to_str(groups) # Handles str _id gracefully
            logger.info(f"Fetched {len(groups)} customer groups")
            return jsonify(groups), 200
        except Exception as e:
            logger.error(f"Error fetching customer groups: {str(e)}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/customer-groups', methods=['POST'])
    @db_required
    def create_customer_group():
        try:
            data = request.get_json()
            if not data or 'group_name' not in data:
                logger.error("Missing group_name in request")
                return jsonify({"error": "Group name is required"}), 400
            group_name = data['group_name'].strip()
            if customer_groups_collection.find_one({"group_name": group_name}):
                logger.warning(f"Customer group already exists: {group_name}")
                return jsonify({"error": "Customer group name already exists"}), 400
            new_group_id = str(uuid.uuid4())
            new_group = {
                "_id": new_group_id,
                "group_name": group_name,
                "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }
            result = customer_groups_collection.insert_one(new_group)
            logger.info(f"Customer group created: {group_name} with ID {new_group_id}")
            return jsonify({"message": "Customer group created successfully", "_id": new_group_id}), 201
        except Exception as e:
            logger.error(f"Error creating customer group: {str(e)}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/customer-groups/<group_id>', methods=['PUT'])
    @db_required
    def update_customer_group(group_id):
        try:
            data = request.get_json()
            if not data or 'group_name' not in data:
                logger.error("Missing group_name in request")
                return jsonify({"error": "Group name is required"}), 400
            new_group_name = data['group_name'].strip()
            # Optional: Check if new name already exists (excluding current ID)
            if customer_groups_collection.find_one({"group_name": new_group_name, "_id": {"$ne": group_id}}):
                logger.warning(f"Customer group name already exists: {new_group_name}")
                return jsonify({"error": "Customer group name already exists"}), 400
            result = customer_groups_collection.update_one(
                {'_id': group_id},
                {'$set': {'group_name': new_group_name, 'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
            )
            if result.matched_count == 0:
                logger.warning(f"Customer group not found for update: {group_id}")
                return jsonify({"error": "Customer group not found"}), 404
            logger.info(f"Customer group updated: {group_id} to {new_group_name}")
            return jsonify({"message": "Customer group updated successfully"}), 200
        except Exception as e:
            logger.error(f"Error updating customer group {group_id}: {str(e)}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/customer-groups/<group_id>', methods=['DELETE'])
    @db_required
    def delete_customer_group(group_id):
        try:
            # Optional: Check if group has associated customers before delete (implement if needed)
            # e.g., if customers_collection.count_documents({"customer_group": group_id}) > 0: return error
            result = customer_groups_collection.delete_one({'_id': group_id})
            if result.deleted_count == 0:
                logger.warning(f"Customer group not found for deletion: {group_id}")
                return jsonify({"error": "Customer group not found"}), 404
            logger.info(f"Customer group deleted: {group_id}")
            return jsonify({"message": "Customer group deleted successfully"}), 200
        except Exception as e:
            logger.error(f"Error deleting customer group {group_id}: {str(e)}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/address-structures', methods=['GET'])
    @db_required
    def get_address_structure():
        try:
            doc = address_structures_collection.find_one({'_id': 'global'})
            if doc:
                return jsonify(doc), 200
            else:
                default = {
                    "_id": "global",
                    "structure": {"countries": {}},
                    "linkedValues": {},
                    "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
                }
                address_structures_collection.insert_one(default)
                return jsonify(default), 200
        except Exception as e:
            logger.error(f"Error fetching address structure: {str(e)}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/address-structures', methods=['PUT'])
    @db_required
    def update_address_structure():
        try:
            data = request.get_json()
            if not data:
                return jsonify({"error": "No data provided"}), 400
            update_data = {
                "structure": data.get("structure", {"countries": {}}),
                "linkedValues": data.get("linkedValues", {}),
                "modified_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }
            result = address_structures_collection.replace_one(
                {'_id': 'global'},
                {**update_data, "_id": "global"},
                upsert=True
            )
            return jsonify({"message": "Address structure updated successfully"}), 200
        except Exception as e:
            logger.error(f"Error updating address structure: {str(e)}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/add-address-value', methods=['POST'])
    @db_required
    def add_address_value():
        try:
            data = request.get_json()
            if not data or 'country' not in data or 'field' not in data or 'value' not in data:
                return jsonify({"error": "country, field, and value are required"}), 400
            country = data['country'].strip()
            field = data['field'].strip() # 'field1', 'field2', 'field3'
            value = data['value'].strip()
            parent_value = data.get('parent_value', '').strip()
            if not country or not field or not value:
                return jsonify({"error": "Valid country, field, and value are required"}), 400
            doc = address_structures_collection.find_one({'_id': 'global'})
            if not doc:
                return jsonify({"error": "No address structure found"}), 404
            structure = doc.get('structure', {'countries': {}})
            linked = doc.get('linkedValues', {})
            if country not in structure['countries']:
                return jsonify({"error": f"Country '{country}' structure not defined"}), 400
            country_data = structure['countries'][country]
            if field not in ['field1', 'field2', 'field3'] or not country_data.get(field):
                return jsonify({"error": f"Field '{field}' not defined for country '{country}'"}), 400
            # Add to global values if not exists
            values = country_data[field]['values']
            if value not in values:
                values.append(value)
                # Update structure
                address_structures_collection.update_one(
                    {'_id': 'global'},
                    {'$set': {f'structure.countries.{country}.{field}.values': values}}
                )
                logger.info(f"Added value '{value}' to {field} for {country}")
            # Handle linked values for field2/field3 if parent_value provided
            if field in ['field2', 'field3'] and parent_value:
                if country not in linked:
                    linked[country] = {}
                if parent_value not in linked[country]:
                    linked[country][parent_value] = {'field2': [], 'field3': []}
                links = linked[country][parent_value]
                link_array = links[field]
                if value not in link_array:
                    link_array.append(value)
                    # Update linked
                    address_structures_collection.update_one(
                        {'_id': 'global'},
                        {'$set': {f'linkedValues.{country}.{parent_value}.{field}': link_array}}
                    )
                    logger.info(f"Added linked value '{value}' to {field} for {country}:{parent_value}")
            return jsonify({"message": "Value added successfully"}), 200
        except Exception as e:
            logger.error(f"Error adding address value: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/sales', methods=['POST'])
    @db_required
    def create_sales_invoice():
        try:
            logger.info("Starting sales invoice creation")
            sales_data = request.json
            required_fields = ['customer', 'items', 'total', 'userId']
            missing_fields = [field for field in required_fields if field not in sales_data or sales_data[field] is None]
            if missing_fields:
                error_msg = f"Missing required fields: {', '.join(missing_fields)}"
                logger.error(error_msg)
                return jsonify({"error": error_msg}), 400
            # Log full raw payload for debugging (remove in prod)
            logger.info(f"Raw payload received: {json.dumps(sales_data, default=str)}")
            user = users_collection.find_one({"email": sales_data['userId']})
            if not user:
                logger.error(f"Invalid userId: {sales_data['userId']}")
                return jsonify({"error": "Invalid userId"}), 400
            sales_data['date'] = sales_data.get('date', datetime.now().strftime("%Y-%m-%d"))
            sales_data['time'] = sales_data.get('time', datetime.now().strftime("%H:%M:%S"))
            net_total = float(sales_data['total']) if sales_data['total'] and str(sales_data['total']).strip() != '' else 0.0
            vat_settings = vat_collection.find_one({"_id": "vat_settings"})
            vat_rate = vat_settings.get("vat", 10) / 100 if vat_settings else 0
            vat_amount_raw = sales_data.get('vat_amount', net_total * vat_rate)
            vat_amount = float(vat_amount_raw) if vat_amount_raw and str(vat_amount_raw).strip() != '' else (net_total * vat_rate)
            grand_total_raw = sales_data.get('grand_total', net_total + vat_amount)
            grand_total = float(grand_total_raw) if grand_total_raw and str(grand_total_raw).strip() != '' else (net_total + vat_amount)
            sales_data['vat_amount'] = round(vat_amount, 2)
            sales_data['grand_total'] = round(grand_total, 2)
            sales_data['invoice_no'] = sales_data.get('invoice_no', f"INV-{int(datetime.now().timestamp())}")
            sales_data['status'] = sales_data.get('status', 'Draft')
            # Store current currency and precision
            current_settings = get_system_settings()
            sales_data['invoice_currency'] = current_settings.get('currency', 'INR')
            sales_data['invoice_currency_precision'] = int(current_settings.get('currencyPrecision', 2)) if current_settings.get('currencyPrecision') and str(current_settings.get('currencyPrecision')).strip() != '' else 2
            # UPDATED: For Online Delivery orders, store orderNo and deliveryPersonName if provided - Added explicit logging for debugging
            if sales_data.get('orderType') == 'Online Delivery':
                sales_data['orderNo'] = sales_data.get('orderNo', '') # Store orderNo from frontend payload
                sales_data['deliveryPersonName'] = sales_data.get('deliveryPersonName', '') # Store deliveryPersonName from frontend payload
                logger.info(f"Online Delivery order - Stored orderNo: '{sales_data['orderNo']}', deliveryPersonName: '{sales_data['deliveryPersonName']}' (from payload: '{sales_data.get('deliveryPersonName', 'MISSING')}')")
            else:
                sales_data['orderNo'] = None # Explicitly set to None for non-Online Delivery
                sales_data['deliveryPersonName'] = None # Explicitly set to None for non-Online Delivery
            # NEW: Associate with POS Opening Entry (from localStorage in frontend)
            sales_data['pos_opening_entry'] = sales_data.get('pos_opening_entry', '') # Store the opening entry name (e.g., "OPEN-1234567890")
            logger.info(f"Associated POS Opening Entry: '{sales_data['pos_opening_entry']}' for invoice {sales_data['invoice_no']}")
            processed_items = []
            for item_idx, item in enumerate(sales_data.get('items', [])):
                if not all(key in item for key in ['item_name', 'basePrice', 'quantity']):
                    logger.error(f"Invalid item structure at index {item_idx}: {item}")
                    return jsonify({"error": "Each item must include item_name, basePrice, and quantity"}), 400
                # Safe conversion for item quantity and price
                qty_val = item.get('quantity')
                item['quantity'] = int(qty_val) if qty_val is not None and str(qty_val).strip() != '' and str(qty_val).isdigit() else 1
                price_val = item.get('basePrice')
                item['basePrice'] = float(price_val) if price_val is not None and str(price_val).strip() != '' else 0.0
                logger.info(f"Processed item {item_idx}: {item['item_name']} - Raw Qty: '{qty_val}', Processed Qty: {item['quantity']}, Raw Price: '{price_val}', Processed Price: {item['basePrice']}")
                if item.get('is_combo_offer'):
                    item['offer_description'] = item.get('offer_description', item['item_name'])
                processed_addons = []
                for addon_idx, addon in enumerate(item.get('addons', [])):
                    if not all(key in addon for key in ['name1', 'addon_price', 'addon_quantity']):
                        logger.error(f"Invalid addon structure at item {item_idx}, addon {addon_idx}: {addon}")
                        return jsonify({"error": "Each addon must include name1, addon_price, and addon_quantity"}), 400
                    # FIXED: Safe int() for addon_quantity (handles '', None, non-numeric)
                    addon_qty_raw = addon.get('addon_quantity')
                    addon_qty = int(addon_qty_raw) if addon_qty_raw is not None and str(addon_qty_raw).strip() != '' and str(addon_qty_raw).isdigit() else 1
                    # FIXED: Safe float() for addon_price
                    addon_price_raw = addon.get('addon_price')
                    addon_price = float(addon_price_raw) if addon_price_raw is not None and str(addon_price_raw).strip() != '' else 0.0
                    logger.info(f"Processed addon {addon_idx} for item {item_idx}: {addon['name1']} - Raw Qty: '{addon_qty_raw}', Processed Qty: {addon_qty}, Raw Price: '{addon_price_raw}', Processed Price: {addon_price}")
                    processed_addons.append({
                        "addon_name": addon['name1'],
                        "addon_price": addon_price,
                        "addon_quantity": addon_qty,
                        "addon_image": addon.get('addon_image', ''),
                        "size": addon.get('size', 'M'),
                        "kitchen": addon.get('kitchen', 'Main Kitchen'),
                    })
                processed_combos = []
                for combo_idx, combo in enumerate(item.get('selectedCombos', [])):
                    if not all(key in combo for key in ['name1', 'combo_price']):
                        logger.error(f"Invalid combo structure at item {item_idx}, combo {combo_idx}: {combo}")
                        return jsonify({"error": "Each combo must include name1 and combo_price"}), 400
                    # FIXED: Safe int() for combo_quantity (handles '', None, non-numeric; defaults to 1 even if missing)
                    combo_qty_raw = combo.get('combo_quantity', 1) # Default 1 if missing key
                    combo_qty = int(combo_qty_raw) if combo_qty_raw is not None and str(combo_qty_raw).strip() != '' and str(combo_qty_raw).isdigit() else 1
                    # FIXED: Safe float() for combo_price
                    combo_price_raw = combo.get('combo_price')
                    combo_price = float(combo_price_raw) if combo_price_raw is not None and str(combo_price_raw).strip() != '' else 0.0
                    logger.info(f"Processed combo {combo_idx} for item {item_idx}: {combo['name1']} - Raw Qty: '{combo_qty_raw}', Processed Qty: {combo_qty}, Raw Price: '{combo_price_raw}', Processed Price: {combo_price}")
                    processed_combos.append({
                        "name1": combo['name1'],
                        "combo_price": combo_price,
                        "combo_quantity": combo_qty,
                        "combo_image": combo.get('combo_image', ''),
                        "size": combo.get('size', 'M'),
                        "spicy": combo.get('spicy', False),
                        "kitchen": combo.get('kitchen', 'Main Kitchen'),
                    })
                processed_items.append({
                    "item_name": item['item_name'],
                    "basePrice": item['basePrice'],
                    "quantity": item['quantity'],
                    "amount": float(item.get('amount', item['basePrice'])) if item.get('amount') and str(item.get('amount')).strip() != '' else item['basePrice'],
                    "icePreference": item.get('icePreference', 'without_ice'),
                    "isSpicy": item.get('isSpicy', False),
                    "kitchen": item.get('kitchen', 'Main Kitchen'),
                    "selectedSize": item.get('selectedSize', 'M'),
                    "ingredients": item.get('ingredients', []),
                    "addons": processed_addons,
                    "selectedCombos": processed_combos,
                    "is_combo_offer": item.get('is_combo_offer', False),
                    "offer_description": item.get('offer_description'),
                })
            sales_data['items'] = processed_items
            sales_data['created_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
            sales_id = sales_collection.insert_one(sales_data).inserted_id
            # UPDATED: Additional logging for saved sale details, especially for Online Delivery and POS Opening Entry
            if sales_data.get('orderType') == 'Online Delivery':
                logger.info(f"Sale saved successfully for Online Delivery: {sales_data['invoice_no']} (ID: {sales_id}) - orderNo: '{sales_data['orderNo']}', deliveryPersonName: '{sales_data['deliveryPersonName']}', pos_opening_entry: '{sales_data['pos_opening_entry']}'")
            else:
                logger.info(f"Sale saved successfully: {sales_data['invoice_no']} (ID: {sales_id}) - pos_opening_entry: '{sales_data['pos_opening_entry']}'")
            return jsonify({
                "id": sales_id,
                "invoice_no": sales_data['invoice_no'],
                "net_total": sales_data['total'],
                "vat_amount": sales_data['vat_amount'],
                "grand_total": sales_data['grand_total'],
                "userId": sales_data['userId']
            }), 201
        except ValueError as ve:
            logger.error(f"ValueError in sales invoice (likely qty/price conversion): {str(ve)}\nFull traceback: {traceback.format_exc()}")
            return jsonify({"error": f"Invalid data conversion: {str(ve)}. Check quantities/prices."}), 400
        except Exception as e:
            logger.error(f"Error creating sales invoice: {str(e)}\n{traceback.format_exc()}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/sales', methods=['GET'])
    @db_required
    def get_all_sales():
        try:
            sales = sales_collection.find()
            sales = convert_objectid_to_str(sales)
            sales = [sale for sale in sales if sale.get('status') != 'Cancelled']
            logger.info(f"Fetched {len(sales)} sales invoices")
            return jsonify(sales), 200
        except Exception as e:
            logger.error(f"Error fetching sales: {str(e)}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/sales/<invoice_no>', methods=['GET'])
    @db_required
    def get_sale_by_invoice_no(invoice_no):
        try:
            sale = sales_collection.find_one({"invoice_no": invoice_no.strip()})
            if not sale:
                logger.warning(f"Sale not found: {invoice_no}")
                return jsonify({"error": "Invoice not found"}), 404
            sale = convert_objectid_to_str(sale)
            logger.info(f"Fetched sale: {invoice_no}")
            return jsonify(sale), 200
        except Exception as e:
            logger.error(f"Error fetching sale {invoice_no}: {str(e)}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/sales/<invoice_no>/status', methods=['PUT'])
    @db_required
    def update_sale_status(invoice_no):
        try:
            data = request.get_json()
            status = data.get('status')
            if not status:
                return jsonify({"error": "Status is required"}), 400
            result = sales_collection.update_one(
                {'invoice_no': invoice_no.strip()},
                {'$set': {'status': status, 'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
            )
            if result.matched_count == 0:
                logger.warning(f"Sale not found for status update: {invoice_no}")
                return jsonify({"error": "Invoice not found"}), 404
            logger.info(f"Sale status updated: {invoice_no} to {status}")
            return jsonify({"message": "Sale status updated successfully"}), 200
        except Exception as e:
            logger.error(f"Error updating sale status {invoice_no}: {str(e)}")
            return jsonify({"error": str(e)}), 500

    # NEW: Endpoint to update deliveryPersonName in sales record by orderNo (called from ActiveOrders when marking delivered)
    @app.route('/api/sales/update-delivery', methods=['POST'])
    @db_required
    def update_sale_delivery():
        try:
            data = request.get_json()
            order_no = data.get('orderNo')
            delivery_person_name = data.get('deliveryPersonName')
            if not order_no or not delivery_person_name:
                return jsonify({"error": "orderNo and deliveryPersonName are required"}), 400
            result = sales_collection.update_one(
                {'orderNo': order_no},
                {'$set': {'deliveryPersonName': delivery_person_name, 'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
            )
            if result.matched_count == 0:
                logger.warning(f"No sale found for orderNo: {order_no}")
                return jsonify({"error": "Sale not found"}), 404
            logger.info(f"Updated delivery person for sale with orderNo: {order_no} to {delivery_person_name}")
            return jsonify({"message": "Delivery person updated successfully"}), 200
        except Exception as e:
            logger.error(f"Error updating sale delivery: {str(e)}")
            return jsonify({"error": str(e)}), 500

    # UPDATED: Endpoint to deliver order - Update status and payments by orderNo
    @app.route('/api/sales/deliver-order', methods=['POST'])
    @db_required
    def deliver_sale_order():
        try:
            data = request.get_json()
            order_no = data.get('orderNo')
            status = data.get('status', 'Delivered')
            payments = data.get('payments', [])
            if not order_no:
                return jsonify({"error": "orderNo is required"}), 400
            # Find sale by orderNo
            sale = sales_collection.find_one({'orderNo': order_no})
            if not sale:
                logger.warning(f"No sale found for orderNo: {order_no}")
                return jsonify({"error": "Sale not found"}), 404
            # Update status and payments
            update_data = {
                '$set': {
                    'status': status,
                    'payments': payments,
                    'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()
                }
            }
            result = sales_collection.update_one(
                {'orderNo': order_no},
                update_data
            )
            if result.matched_count == 0:
                return jsonify({"error": "Failed to update order"}), 500
            logger.info(f"Delivered order {order_no} (Invoice: {sale['invoice_no']}) with payments: {payments}")
            return jsonify({"message": "Order delivered successfully", "invoice_no": sale['invoice_no']}), 200
        except Exception as e:
            logger.error(f"Error delivering order: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/tables', methods=['GET'])
    @db_required
    def get_tables():
        try:
            tables = tables_collection.find()
            tables = convert_objectid_to_str(tables)
            logger.info(f"Fetched {len(tables)} tables")
            return jsonify({"message": tables}), 200
        except Exception as e:
            logger.error(f"Error fetching tables: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/tables', methods=['POST'])
    @db_required
    def add_table():
        try:
            data = request.get_json()
            table_number = data.get("table_number")
            floor = data.get("floor")
            if floor:
                floor = floor.strip()
            number_of_chairs = data.get("number_of_chairs")
            type_ = data.get("type", "Round")
            chairs = data.get("chairs", [])
            x = data.get("x", 0)
            y = data.get("y", 0)
            if not table_number or not floor or not number_of_chairs:
                logger.error("Missing table_number, floor, or number_of_chairs")
                return jsonify({"error": "Table number, floor, and number of chairs are required"}), 400
            # Check for duplicate table_number on the same floor only
            if tables_collection.find_one({"table_number": table_number, "floor": floor}):
                logger.warning(f"Table number {table_number} already exists on floor {floor}")
                return jsonify({"error": f"Table number {table_number} already exists on floor {floor}"}), 400
            new_table = {
                "table_number": table_number,
                "floor": floor,
                "number_of_chairs": int(number_of_chairs),
                "type": type_,
                "chairs": chairs,
                "x": x,
                "y": y,
                "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }
            tables_collection.insert_one(new_table)
            logger.info(f"Table added: {table_number} on floor {floor}")
            return jsonify({"message": "Table added successfully"}), 201
        except Exception as e:
            logger.error(f"Error adding table: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/tables/<table_number>', methods=['PUT'])
    @db_required
    def update_table(table_number):
        try:
            data = request.get_json()
            floor = data.get("floor")
            if not floor:
                logger.error("Floor is required for table update")
                return jsonify({"error": "Floor is required"}), 400
            floor = floor.strip()
            update_data = {}
            if "number_of_chairs" in data:
                update_data["number_of_chairs"] = int(data["number_of_chairs"])
            if "type" in data:
                update_data["type"] = data["type"]
            if "chairs" in data:
                update_data["chairs"] = data["chairs"]
            if "x" in data:
                update_data["x"] = data["x"]
            if "y" in data:
                update_data["y"] = data["y"]
            if "floor" in data:
                update_data["floor"] = data["floor"].strip()
            if not update_data:
                return jsonify({"error": "No data provided to update"}), 400
            update_data["modified_at"] = datetime.now(ZoneInfo("UTC")).isoformat()
            # Match by both table_number and floor to ensure floor-specific update
            result = tables_collection.update_one(
                {"table_number": table_number, "floor": floor},
                {"$set": update_data}
            )
            if result.matched_count == 0:
                logger.warning(f"Table not found for update: {table_number} on floor {floor}")
                return jsonify({"error": "Table not found"}), 404
            logger.info(f"Table updated: {table_number} on floor {floor}")
            return jsonify({"message": "Table updated successfully"}), 200
        except Exception as e:
            logger.error(f"Error updating table {table_number}: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/tables/<table_number>', methods=['DELETE'])
    @db_required
    def delete_table(table_number):
        try:
            data = request.get_json() or {}
            floor = data.get("floor")
            if not floor:
                logger.error("Floor is required for table deletion")
                return jsonify({"error": "Floor is required"}), 400
            floor = floor.strip()
            # Delete by both table_number and floor to ensure floor-specific deletion
            result = tables_collection.delete_one({"table_number": table_number, "floor": floor})
            if result.deleted_count == 0:
                logger.warning(f"Table not found: {table_number} on floor {floor}")
                return jsonify({"error": "Table not found"}), 404
            logger.info(f"Table deleted: {table_number} on floor {floor}")
            return jsonify({"message": "Table deleted successfully"}), 200
        except Exception as e:
            logger.error(f"Error deleting table {table_number}: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/create_opening_entry', methods=['POST'])
    @db_required
    def create_opening_entry():
        try:
            data = request.get_json()
            if not data:
                return jsonify({"message": "No data provided", "status": "error"}), 400
            required_fields = ['period_start_date', 'posting_date', 'company', 'user', 'balance_details']
            missing_fields = [field for field in required_fields if field not in data or not data[field]]
            if missing_fields:
                return jsonify({"message": f"Missing fields: {', '.join(missing_fields)}", "status": "error"}), 400
            balance_details = data['balance_details']
            if not isinstance(balance_details, list):
                return jsonify({"message": "balance_details must be a list", "status": "error"}), 400
            for detail in balance_details:
                if not all(key in detail for key in ['mode_of_payment', 'opening_amount']):
                    return jsonify({"message": "Each balance detail must have mode_of_payment and opening_amount", "status": "error"}), 400
                detail['opening_amount'] = float(detail['opening_amount'])
            data['creation'] = datetime.now(ZoneInfo("UTC")).isoformat()
            data['modified'] = data['creation']
            data['name'] = f"OPEN-{int(datetime.now().timestamp())}"
            data['status'] = 'Open'  # UPDATED: Set to 'Open' for POS shifts (instead of 'Draft')
            data['docstatus'] = data.get('docstatus', 0)
            data['pos_profile'] = data.get('pos_profile', 'POS-001')
            result = opening_collection.insert_one(data)
            username = data['user']
            user = users_collection.find_one({"firstName": username})
            if user:
                users_collection.update_one(
                    {"_id": user['_id']},
                    {"$set": {"last_opening_entry_time": datetime.now(ZoneInfo("UTC")).isoformat()}}
                )
                logger.info(f"Updated last_opening_entry_time for user: {username}")
            logger.info(f"POS opening entry created: {data['name']}")
            return jsonify({"message": {"name": data['name'], "status": "success"}}), 201
        except Exception as e:
            logger.error(f"Error in create_opening_entry: {str(e)}")
            return jsonify({"message": f"Server error: {str(e)}", "status": "error"}), 500

    @app.route('/api/get_pos_opening_entries', methods=['POST'])
    @db_required
    def get_pos_opening_entries():
        try:
            data = request.get_json()
            if not data or 'pos_profile' not in data:
                return jsonify({"message": "POS profile is required", "status": "error"}), 400
            pos_profile = data['pos_profile']
            # UPDATED: Filter only open entries (status="Open") for closing dropdown
            entries = opening_collection.find({"pos_profile": pos_profile, "status": "Open"})
            entries = convert_objectid_to_str(entries)
            logger.info(f"Fetched {len(entries)} POS opening entries for profile: {pos_profile}")
            return jsonify({"message": entries, "status": "success"}), 200
        except Exception as e:
            logger.error(f"Error in get_pos_opening_entries: {str(e)}")
            return jsonify({"message": f"Server error: {str(e)}", "status": "error"}), 500

    # FIXED: Updated filter to fetch ALL invoices associated with pos_opening_entry (any status, including Pending/Draft/Delivered, etc. - no $ne "Cancelled" filter to include all as per user request)
    @app.route('/api/get_pos_invoices', methods=['POST'])
    @db_required
    def get_pos_invoices():
        try:
            data = request.get_json()
            pos_opening_entry = data.get('pos_opening_entry')
            if not pos_opening_entry:
                return jsonify({"message": "POS opening entry is required", "status": "error"}), 400
            # FIXED: Filter ONLY by pos_opening_entry (fetch ALL associated invoices, regardless of status - Pending, Draft, Delivered, etc.)
            invoices_cursor = sales_collection.find({
                "pos_opening_entry": pos_opening_entry
            })
            invoices = convert_objectid_to_str(invoices_cursor)
            logger.info(f"DEBUG: Found {len(invoices)} invoices for pos_opening_entry: {pos_opening_entry}. Sample statuses: {[inv.get('status', 'N/A') for inv in invoices[:3]]}") # Debug log
            # Get opening entry for period_start (if needed for other logic, but not for filtering now)
            opening_entry = opening_collection.find_one({"name": pos_opening_entry})
            if not opening_entry:
                return jsonify({"message": "Opening entry not found", "status": "error"}), 404
            # Calculate totals from filtered invoices
            total = sum(float(inv['grand_total']) for inv in invoices)
            net_total = sum(float(inv['total']) for inv in invoices)
            total_qty = sum(sum(item['quantity'] for item in inv['items']) for inv in invoices)
            vat_settings = vat_collection.find_one({"_id": "vat_settings"})
            vat_percentage = vat_settings.get("vat", 10) if vat_settings else 10
            taxes = [{"account_head": "VAT", "rate": vat_percentage, "amount": total - net_total}]
            response = {
                "invoices": [{"pos_invoice": inv['invoice_no'], "grand_total": inv['grand_total'], "posting_date": inv['date'], "customer": inv['customer']} for inv in invoices],
                "taxes": taxes,
                "grand_total": total,
                "net_total": net_total,
                "total_quantity": total_qty,
                "status": "success"
            }
            logger.info(f"Fetched {len(invoices)} POS invoices for opening entry: {pos_opening_entry}")
            return jsonify({"message": response}), 200
        except Exception as e:
            logger.error(f"Error in get_pos_invoices: {str(e)}")
            return jsonify({"message": f"Error: {str(e)}", "status": "error"}), 500

    @app.route('/api/create_closing_entry', methods=['POST'])
    @db_required
    def create_closing_entry():
        try:
            data = request.get_json()
            required_fields = ['pos_opening_entry', 'posting_date', 'period_end_date', 'pos_transactions', 'payment_reconciliation', 'taxes', 'grand_total', 'net_total', 'total_quantity']
            if not data or not all(field in data for field in required_fields):
                return jsonify({"message": f"Missing fields: {', '.join([f for f in required_fields if f not in data])}", "status": "error"}), 400
            opening_entry = opening_collection.find_one({"name": data['pos_opening_entry']})
            if not opening_entry:
                return jsonify({"message": "Opening entry not found", "status": "error"}), 404
            data['creation'] = datetime.now(ZoneInfo("UTC")).isoformat()
            data['modified'] = data['creation']
            data['name'] = f"CLOSE-{int(datetime.now().timestamp())}"
            data['status'] = 'Draft'
            data['docstatus'] = 0
            result = pos_closing_collection.insert_one(data)
            # UPDATED: After creating closing, mark the opening as 'Closed' to require new opening on next login
            opening_collection.update_one(
                {"name": data['pos_opening_entry']},
                {"$set": {"status": "Closed", "modified": datetime.now(ZoneInfo("UTC")).isoformat()}}
            )
            logger.info(f"POS closing entry created: {data['name']} and marked opening {data['pos_opening_entry']} as Closed")
            return jsonify({"message": {"name": data['name'], "status": "success", "message": "Closing Entry created"}}), 201
        except Exception as e:
            logger.error(f"Error in create_closing_entry: {str(e)}")
            return jsonify({"message": f"Error: {str(e)}", "status": "error"}), 500
    @app.route('/api/kitchens', methods=['GET'])
    @db_required
    def get_kitchens():
        try:
            kitchens = kitchens_collection.find()
            kitchens = convert_objectid_to_str(kitchens)
            logger.info(f"Fetched {len(kitchens)} kitchens")
            return jsonify(kitchens), 200
        except Exception as e:
            logger.error(f"Error fetching kitchens: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/kitchens', methods=['POST'])
    @db_required
    def create_kitchen():
        try:
            data = request.get_json()
            if not data or 'kitchen_name' not in data:
                logger.error("Missing kitchen_name in request")
                return jsonify({"error": "Kitchen name is required"}), 400
            kitchen_name = data['kitchen_name']
            if kitchens_collection.find_one({"kitchen_name": kitchen_name}):
                logger.warning(f"Kitchen already exists: {kitchen_name}")
                return jsonify({"error": "Kitchen name already exists"}), 400
            new_kitchen = {
                "kitchen_name": kitchen_name,
                "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }
            result = kitchens_collection.insert_one(new_kitchen)
            logger.info(f"Kitchen created: {kitchen_name}")
            return jsonify({"message": "Kitchen created successfully", "id": result.inserted_id}), 201
        except Exception as e:
            logger.error(f"Error creating kitchen: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/kitchens/<kitchen_id>', methods=['PUT'])
    @db_required
    def update_kitchen(kitchen_id):
        try:
            data = request.get_json()
            if not data or 'kitchen_name' not in data:
                logger.error("Missing kitchen_name in request")
                return jsonify({"error": "Kitchen name is required"}), 400
            result = kitchens_collection.update_one(
                {'_id': kitchen_id},
                {'$set': {'kitchen_name': data['kitchen_name'], 'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
            )
            if result.matched_count == 0:
                logger.warning(f"Kitchen not found for update: {kitchen_id}")
                return jsonify({"error": "Kitchen not found"}), 404
            logger.info(f"Kitchen updated: {kitchen_id}")
            return jsonify({"message": "Kitchen updated successfully"}), 200
        except Exception as e:
            logger.error(f"Error updating kitchen {kitchen_id}: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/kitchens/<kitchen_id>', methods=['DELETE'])
    @db_required
    def delete_kitchen(kitchen_id):
        try:
            result = kitchens_collection.delete_one({'_id': kitchen_id})
            if result.deleted_count == 0:
                logger.warning(f"Kitchen not found for deletion: {kitchen_id}")
                return jsonify({"error": "Kitchen not found"}), 404
            logger.info(f"Kitchen deleted: {kitchen_id}")
            return jsonify({"message": "Kitchen deleted successfully"}), 200
        except Exception as e:
            logger.error(f"Error deleting kitchen {kitchen_id}: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/item-groups', methods=['GET'])
    @db_required
    def get_item_groups():
        try:
            item_groups = item_groups_collection.find()
            item_groups = convert_objectid_to_str(item_groups)
            logger.info(f"Fetched {len(item_groups)} item groups")
            return jsonify(item_groups), 200
        except Exception as e:
            logger.error(f"Error fetching item groups: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/item-groups', methods=['POST'])
    @db_required
    def create_item_group():
        try:
            data = request.get_json()
            if not data or 'group_name' not in data:
                logger.error("Missing group_name in request")
                return jsonify({"error": "Group name is required"}), 400
            group_name = data['group_name']
            if item_groups_collection.find_one({"group_name": group_name}):
                logger.warning(f"Item group already exists: {group_name}")
                return jsonify({"error": "Item group name already exists"}), 400
            new_group = {
                "group_name": group_name,
                "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }
            result = item_groups_collection.insert_one(new_group)
            logger.info(f"Item group created: {group_name}")
            return jsonify({"message": "Item group created successfully", "id": result.inserted_id}), 201
        except Exception as e:
            logger.error(f"Error creating item group: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/item-groups/<group_id>', methods=['PUT'])
    @db_required
    def update_item_group(group_id):
        try:
            data = request.get_json()
            if not data or 'group_name' not in data:
                logger.error("Missing group_name in request")
                return jsonify({"error": "Group name is required"}), 400
            result = item_groups_collection.update_one(
                {'_id': group_id},
                {'$set': {'group_name': data['group_name'], 'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
            )
            if result.matched_count == 0:
                logger.warning(f"Item group not found for update: {group_id}")
                return jsonify({"error": "Item group not found"}), 404
            logger.info(f"Item group updated: {group_id}")
            return jsonify({"message": "Item group updated successfully"}), 200
        except Exception as e:
            logger.error(f"Error updating item group {group_id}: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/item-groups/<group_id>', methods=['DELETE'])
    @db_required
    def delete_item_group(group_id):
        try:
            result = item_groups_collection.delete_one({'_id': group_id})
            if result.deleted_count == 0:
                logger.warning(f"Item group not found for deletion: {group_id}")
                return jsonify({"error": "Item group not found"}), 404
            logger.info(f"Item group deleted: {group_id}")
            return jsonify({"message": "Item group deleted successfully"}), 200
        except Exception as e:
            logger.error(f"Error deleting item group {group_id}: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/items/nutrition', methods=['POST', 'OPTIONS'])
    @db_required
    def save_item_nutrition():
        if request.method == 'OPTIONS':
            response = jsonify({"success": True})
            response.headers['Access-Control-Allow-Origin'] = '*'
            response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
            return response, 200
        try:
            data = request.get_json()
            if not data:
                logger.error("No data provided for saving ingredients")
                return jsonify({"error": "No data provided"}), 400
            required_fields = ['item_name', 'type', 'instances', 'ingredients']
            for field in required_fields:
                if field not in data or data[field] is None:
                    logger.error(f"Missing or null required field: {field}")
                    return jsonify({"error": f"Missing or null required field: {field}"}), 400
            item_name = data['item_name']
            item_type = data['type']
            instances = data['instances']
            ingredients = data['ingredients']
            if item_type not in ['item', 'addon', 'combo']:
                logger.error(f"Invalid type: {item_type}")
                return jsonify({"error": "Invalid type, must be 'item', 'addon', or 'combo'"}), 400
            if not isinstance(ingredients, list):
                logger.error("Ingredients must be a list of objects")
                return jsonify({"error": "Ingredients must be a list of objects"}), 400
            for ingredient in ingredients:
                if not isinstance(ingredient, dict) or not all(key in ingredient for key in ['name', 'small', 'medium', 'large', 'weight', 'nutrition']):
                    logger.error("Each ingredient must be an object with required fields")
                    return jsonify({"error": "Each ingredient must be an object with required fields"}), 400
            filtered_ingredients = [ing for ing in ingredients if ing['name'].strip()]
            updated_count = 0
            for instance in instances:
                item_id = instance['item_id']
                index = instance.get('index')
                item = items_collection.find_one({'_id': item_id})
                if not item:
                    logger.warning(f"Item not found: {item_id}")
                    continue
                update_data = {'$set': {'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
                if item_type == 'item':
                    update_data['$set']['ingredients'] = filtered_ingredients
                elif item_type == 'addon':
                    if index is None or not isinstance(index, int):
                        logger.error(f"Index is required for addons in item_id: {item_id}")
                        continue
                    update_data['$set'][f'addons.{index}.ingredients'] = filtered_ingredients
                elif item_type == 'combo':
                    if index is None or not isinstance(index, int):
                        logger.error(f"Index is required for combos in item_id: {item_id}")
                        continue
                    update_data['$set'][f'combos.{index}.ingredients'] = filtered_ingredients
                result = items_collection.update_one({'_id': item_id}, update_data)
                if result.matched_count > 0:
                    updated_count += 1
            if updated_count == 0:
                logger.error(f"No items updated for {item_type}: {item_name}")
                return jsonify({"error": "No items updated, please check instance data"}), 400
            logger.info(f"Ingredients updated for {item_type}: {item_name} across {updated_count} instances")
            return jsonify({"message": "Ingredients saved successfully"}), 200
        except Exception as e:
            logger.error(f"Error saving ingredients for {item_name}: {str(e)}")
            return jsonify({"error": f"Server error: {str(e)}"}), 500
    @app.route('/api/items/nutrition/<item_name>', methods=['GET', 'DELETE', 'OPTIONS'])
    @db_required
    def handle_item_nutrition(item_name):
        if request.method == 'OPTIONS':
            response = jsonify({"success": True})
            response.headers['Access-Control-Allow-Origin'] = '*'
            response.headers['Access-Control-Allow-Methods'] = 'GET, DELETE, OPTIONS'
            response.headers['Access-Control-Allow-Headers'] = 'Content-Type, X-Instances'
            return response, 200
        try:
            item_type = request.args.get('type')
            item_id = request.args.get('item_id')
            index = request.args.get('index')
            if request.method == 'GET':
                if not item_type or not item_id:
                    logger.error("Type and item_id are required")
                    return jsonify({"error": "Type and item_id are required"}), 400
                if item_type not in ['item', 'addon', 'combo']:
                    logger.error(f"Invalid type: {item_type}")
                    return jsonify({"error": "Invalid type, must be 'item', 'addon', or 'combo'"}), 400
                item = items_collection.find_one({'_id': item_id})
                if not item:
                    logger.warning(f"Item not found: {item_id}")
                    return jsonify({"error": "Item not found"}), 404
                ingredients = []
                if item_type == 'item':
                    ingredients = item.get('ingredients', [])
                elif item_type == 'addon':
                    if index is None or not index.isdigit():
                        logger.error("Index is required for addons")
                        return jsonify({"error": "Index is required for addons"}), 400
                    index = int(index)
                    if index < len(item.get('addons', [])):
                        ingredients = item['addons'][index].get('ingredients', [])
                elif item_type == 'combo':
                    if index is None or not index.isdigit():
                        logger.error("Index is required for combos")
                        return jsonify({"error": "Index is required for combos"}), 400
                    index = int(index)
                    if index < len(item.get('combos', [])):
                        ingredients = item['combos'][index].get('ingredients', [])
                response_data = {'ingredients': ingredients, 'nutrition': {}}
                logger.info(f"Fetched nutrition data for {item_type}: {item_name}")
                return jsonify(response_data), 200
            if request.method == 'DELETE':
                instances = request.headers.get('X-Instances')
                if not instances:
                    logger.error("Instances header is required for DELETE")
                    return jsonify({"error": "Instances header is required"}), 400
                try:
                    instances = json.loads(instances)
                except json.JSONDecodeError:
                    logger.error("Invalid instances header format")
                    return jsonify({"error": "Invalid instances header format"}), 400
                if item_type not in ['item', 'addon', 'combo']:
                    logger.error(f"Invalid type: {item_type}")
                    return jsonify({"error": "Invalid type, must be 'item', 'addon', or 'combo'"}), 400
                deleted_count = 0
                for instance in instances:
                    item_id = instance['item_id']
                    index = instance.get('index')
                    item = items_collection.find_one({'_id': item_id})
                    if not item:
                        logger.warning(f"Item not found: {item_id}")
                        continue
                    update_data = {'$unset': {}, '$set': {'modified_at': datetime.now(ZoneInfo("UTC")).isoformat()}}
                    if item_type == 'item':
                        update_data['$unset']['ingredients'] = ''
                    elif item_type == 'addon':
                        if index is None or not isinstance(index, int):
                            logger.error(f"Index is required for addons in item_id: {item_id}")
                            continue
                        update_data['$unset'][f'addons.{index}.ingredients'] = ''
                    elif item_type == 'combo':
                        if index is None or not isinstance(index, int):
                            logger.error(f"Index is required for combos in item_id: {item_id}")
                            continue
                        update_data['$unset'][f'combos.{index}.ingredients'] = ''
                    result = items_collection.update_one({'_id': item_id}, update_data)
                    if result.matched_count > 0:
                        deleted_count += 1
                if deleted_count == 0:
                    logger.error(f"No items updated for deletion of {item_type}: {item_name}")
                    return jsonify({"error": "No items updated for deletion, please check instance data"}), 400
                logger.info(f"Nutrition cleared for {item_type}: {item_name} across {deleted_count} instances")
                return jsonify({"message": "Nutrition and ingredients cleared successfully"}), 200
        except Exception as e:
            logger.error(f"Error handling nutrition for {item_name}: {str(e)}")
            return jsonify({"error": f"Server error: {str(e)}"}), 500

    @app.route('/api/picked-up-items', methods=['POST'])
    @db_required
    def save_picked_up_item():
        try:
            item_data = request.get_json()
            if not item_data:
                logger.error("No data provided for picked-up items")
                return jsonify({'success': False, 'message': 'No data provided'}), 400
            customer_name = item_data.get('customerName', 'Unknown')
            table_number = item_data.get('tableNumber', 'N/A')
            pickup_time = datetime.now(timezone.utc).isoformat()
            new_item = {
                'itemName': item_data.get('itemName', 'Unknown'),
                'quantity': item_data.get('quantity', 0),
                'category': item_data.get('category', 'N/A'),
                'kitchen': item_data.get('kitchen', 'Unknown'),
                'addonCounts': item_data.get('addonCounts', []),
                'selectedCombos': item_data.get('selectedCombos', [])
            }
            existing_entry = picked_up_collection.find_one({
                'customerName': customer_name,
                'tableNumber': table_number
            })
            if existing_entry:
                updated_items = existing_entry.get('items', [])
                updated_items.append(new_item)
                picked_up_collection.update_one(
                    {'_id': existing_entry['_id']},
                    {
                        '$set': {
                            'items': updated_items,
                            'pickupTime': pickup_time,
                            'modified_at': datetime.now(timezone.utc).isoformat()
                        }
                    }
                )
                logger.info(f"Picked-up items updated for customer: {customer_name}, table: {table_number}")
                return jsonify({
                    'success': True,
                    'message': 'Picked-up items updated successfully',
                    'id': existing_entry['_id']
                }), 200
            else:
                picked_up_data = {
                    'customerName': customer_name,
                    'tableNumber': table_number,
                    'items': [new_item],
                    'pickupTime': pickup_time,
                    'created_at': datetime.now(timezone.utc).isoformat(),
                    'orderType': item_data.get('orderType', 'N/A')
                }
                result = picked_up_collection.insert_one(picked_up_data)
                logger.info(f"Picked-up items saved with ID: {result.inserted_id}")
                return jsonify({
                    'success': True,
                    'message': 'Picked-up items saved successfully',
                    'id': result.inserted_id
                }), 201
        except Exception as e:
            logger.error(f"Error saving picked-up items: {str(e)}")
            logger.error(traceback.format_exc())
            return jsonify({'success': False, 'message': str(e)}), 500
    @app.route('/api/picked-up-items', methods=['GET'])
    @db_required
    def get_picked_up_items():
        try:
            picked_up_items = picked_up_collection.find()
            picked_up_items = convert_objectid_to_str(picked_up_items)
            logger.info(f"Fetched {len(picked_up_items)} picked-up item entries")
            return jsonify({'success': True, 'pickedUpItems': picked_up_items}), 200
        except Exception as e:
            logger.error(f"Error fetching picked-up items: {str(e)}")
            logger.error(traceback.format_exc())
            return jsonify({'success': False, 'message': str(e)}), 500
    @app.route('/api/picked-up-items/<entry_id>', methods=['DELETE'])
    @db_required
    def delete_picked_up_item(entry_id):
        try:
            result = picked_up_collection.delete_one({'_id': entry_id})
            if result.deleted_count == 0:
                logger.warning(f"Picked-up entry not found: {entry_id}")
                return jsonify({"error": "Picked-up entry not found"}), 404
            logger.info(f"Picked-up entry deleted: {entry_id}")
            return jsonify({"message": "Picked-up entry deleted successfully"}), 200
        except Exception as e:
            logger.error(f"Error deleting picked-up entry {entry_id}: {str(e)}")
            logger.error(traceback.format_exc())
            return jsonify({"error": str(e)}), 500
    @app.route('/api/variants', methods=['POST'])
    @db_required
    def create_variants():
        try:
            data = request.get_json()
            if not data or not data.get('heading') or not isinstance(data.get('subheadings'), list):
                return jsonify({'error': 'Variant must have a heading and a list of subheadings'}), 400
            for subheading in data['subheadings']:
                if not subheading.get('name'):
                    return jsonify({'error': 'Each subheading must have a name'}), 400
                if 'price' in subheading and subheading['price'] is not None:
                    try:
                        subheading['price'] = float(subheading['price'])
                    except (ValueError, TypeError):
                        return jsonify({'error': f"Invalid price for subheading {subheading['name']}"}), 400
                if 'image' in subheading and subheading['image'] is not None:
                    if not isinstance(subheading['image'], str):
                        return jsonify({'error': f"Image for subheading {subheading['name']} must be a string"}), 400
                if 'dropdown' in subheading and not isinstance(subheading['dropdown'], bool):
                    return jsonify({'error': f"Dropdown for subheading {subheading['name']} must be a boolean"}), 400
            result = variants_collection.insert_one(data)
            return jsonify({
                'message': 'Variant created successfully',
                'inserted_id': result.inserted_id
            }), 201
        except Exception as e:
            return jsonify({'error': f"Server error: {str(e)}"}), 500
    @app.route('/api/variants', methods=['GET'])
    @db_required
    def get_variants():
        try:
            placeholder_url = 'https://placehold.co/100x100/EFEFEF/AAAAAA?text=No+Image'
            variants = variants_collection.find()
            variants_list = []
            for variant in variants:
                variant = convert_objectid_to_str(variant)
                for subheading in variant.get('subheadings', []):
                    if subheading.get('image'):
                        subheading['image'] = f"/api/images/{os.path.basename(subheading['image'])}"
                    else:
                        subheading['image'] = placeholder_url
                variants_list.append(variant)
            return jsonify(variants_list), 200
        except Exception as e:
            return jsonify({'error': f"Server error: {str(e)}"}), 500
    @app.route('/api/variants/<id>', methods=['GET'])
    @db_required
    def get_variant(id):
        try:
            variant = variants_collection.find_one({'_id': id})
            if not variant:
                return jsonify({'error': 'Variant not found'}), 404
            variant = convert_objectid_to_str(variant)
            placeholder_url = 'https://placehold.co/100x100/EFEFEF/AAAAAA?text=No+Image'
            for subheading in variant.get('subheadings', []):
                if subheading.get('image'):
                    subheading['image'] = f"/api/images/{os.path.basename(subheading['image'])}"
                else:
                    subheading['image'] = placeholder_url
            return jsonify(variant), 200
        except Exception as e:
            return jsonify({'error': f"Server error: {str(e)}"}), 500
    @app.route('/api/variants/<id>', methods=['PUT'])
    @db_required
    def update_variant(id):
        try:
            data = request.get_json()
            if not data or not data.get('heading') or not isinstance(data.get('subheadings'), list):
                return jsonify({'error': 'Variant must have a heading and a list of subheadings'}), 400
            for subheading in data['subheadings']:
                if not subheading.get('name'):
                    return jsonify({'error': 'Each subheading must have a name'}), 400
                if 'price' in subheading and subheading['price'] is not None:
                    try:
                        subheading['price'] = float(subheading['price'])
                    except (ValueError, TypeError):
                        return jsonify({'error': f"Invalid price for subheading {subheading['name']}"}), 400
                if 'image' in subheading and subheading['image'] is not None:
                    if not isinstance(subheading['image'], str):
                        return jsonify({'error': f"Image for subheading {subheading['name']} must be a string"}), 400
                if 'dropdown' in subheading and not isinstance(subheading['dropdown'], bool):
                    return jsonify({'error': f"Dropdown for subheading {subheading['name']} must be a boolean"}), 400
            result = variants_collection.update_one(
                {'_id': id},
                {'$set': data}
            )
            if result.matched_count == 0:
                return jsonify({'error': 'Variant not found'}), 404
            return jsonify({'message': 'Variant updated successfully'}), 200
        except Exception as e:
            return jsonify({'error': f"Server error: {str(e)}"}), 500
    @app.route('/api/variants/<id>', methods=['DELETE'])
    @db_required
    def delete_variant(id):
        try:
            result = variants_collection.delete_one({'_id': id})
            if result.deleted_count == 0:
                return jsonify({'error': 'Variant not found'}), 404
            return jsonify({'message': 'Variant deleted successfully'}), 200
        except Exception as e:
            return jsonify({'error': f"Server error: {str(e)}"}), 500
    @app.route('/api/variants/heading/<heading>', methods=['DELETE'])
    @db_required
    def delete_variant_by_heading(heading):
        try:
            result = variants_collection.delete_one({'heading': heading})
            return jsonify({'message': 'Variant deleted successfully'}), 200
        except Exception as e:
            return jsonify({'error': f"Server error: {str(e)}"}), 500
    VALID_COUNTRY_CODES = [
        '+91', # India
        '+1', # USA
        '+971', # UAE (Dubai)
        '+44', # UK
        '+61', # Australia
    ]

    def generate_employee_id():
        """Generate a unique employee ID."""
        return str(uuid.uuid4())[:8]

    def validate_email(email):
        """Validate email format."""
        import re
        pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
        return re.match(pattern, email) is not None

    VALID_COUNTRY_CODES = ['+91', '+1', '+971', '+44', '+61'] # Add more as needed

    # UPDATED: Endpoint for employees - Manual 6-digit secret key, no auto-generate
    @app.route('/api/employees', methods=['GET'])
    @db_required
    def get_employees():
        try:
            employees = employees_collection.find()
            return jsonify(convert_objectid_to_str(employees)), 200
        except Exception as e:
            logger.error(f"Error fetching employees: {str(e)}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/employees', methods=['POST'])
    @db_required
    def create_employee():
        try:
            data = request.get_json()
            required_fields = ['name', 'phoneNumber', 'vehicleNumber', 'role', 'email', 'secretKey']
            if not data or not all(key in data for key in required_fields):
                return jsonify({'error': 'Missing required fields: name, phoneNumber, vehicleNumber, role, email, secretKey'}), 400
            phone_number = data['phoneNumber']
            if not any(phone_number.startswith(code) for code in VALID_COUNTRY_CODES):
                return jsonify({'error': 'Phone number must include a valid country code (e.g., +91, +1, +971)'}), 400
            code_length = len(next(code for code in VALID_COUNTRY_CODES if phone_number.startswith(code)))
            if len(phone_number) < code_length + 7:
                return jsonify({'error': 'Phone number is too short'}), 400
            email = data['email']
            if not validate_email(email):
                return jsonify({'error': 'Invalid email format'}), 400
            if employees_collection.find_one({'email': email}):
                return jsonify({'error': 'Email already exists'}), 400
            secret_key = data['secretKey']
            if not is_valid_secret_key(secret_key):  # FIXED: Global helper now available
                return jsonify({'error': 'Secret key must be exactly 6 digits'}), 400
            if employees_collection.find_one({'secretKey': secret_key}):
                return jsonify({'error': 'Secret key already exists. Please choose a unique 6-digit key'}), 400
            employee_id = generate_employee_id()
            employee = {
                'employeeId': employee_id,
                'name': data['name'],
                'phoneNumber': phone_number,
                'vehicleNumber': data['vehicleNumber'],
                'role': data['role'],
                'email': email,
                'secretKey': secret_key, # Manual secret key
                'created_at': datetime.now(ZoneInfo("UTC")).isoformat()
            }
            employees_collection.insert_one(employee)
            if not users_collection.find_one({'email': email}):
                users_collection.insert_one({
                    'email': email,
                    'name': data['name'],
                    'role': data['role'],
                    'created_at': datetime.now(ZoneInfo("UTC")).isoformat()
                })
            logger.info(f"Created employee: {employee_id} with email: {email} and secret key: {secret_key}")
            return jsonify({'message': 'Employee created successfully', 'employee': employee}), 201
        except Exception as e:
            logger.error(f"Error creating employee: {str(e)}")
            logger.error(traceback.format_exc())  # Added for better debugging
            return jsonify({'error': str(e)}), 500

    @app.route('/api/employees/<employee_id>', methods=['PUT'])
    @db_required
    def update_employee(employee_id):
        try:
            data = request.get_json()
            required_fields = ['name', 'phoneNumber', 'vehicleNumber', 'role', 'email', 'secretKey']
            if not data or not all(key in data for key in required_fields):
                return jsonify({'error': 'Missing required fields: name, phoneNumber, vehicleNumber, role, email, secretKey'}), 400
            phone_number = data['phoneNumber']
            if not any(phone_number.startswith(code) for code in VALID_COUNTRY_CODES):
                return jsonify({'error': 'Phone number must include a valid country code (e.g., +91, +1, +971)'}), 400
            code_length = len(next(code for code in VALID_COUNTRY_CODES if phone_number.startswith(code)))
            if len(phone_number) < code_length + 7:
                return jsonify({'error': 'Phone number is too short'}), 400
            email = data['email']
            if not validate_email(email):
                return jsonify({'error': 'Invalid email format'}), 400
            existing_employee = employees_collection.find_one({'email': email, 'employeeId': {'$ne': employee_id}})
            if existing_employee:
                return jsonify({'error': 'Email already exists'}), 400
            secret_key = data['secretKey']
            if not is_valid_secret_key(secret_key):  # FIXED: Global helper now available
                return jsonify({'error': 'Secret key must be exactly 6 digits'}), 400
            existing_secret = employees_collection.find_one({'secretKey': secret_key, 'employeeId': {'$ne': employee_id}})
            if existing_secret:
                return jsonify({'error': 'Secret key already exists. Please choose a unique 6-digit key'}), 400
            updated_employee = {
                'name': data['name'],
                'phoneNumber': phone_number,
                'vehicleNumber': data['vehicleNumber'],
                'role': data['role'],
                'email': email,
                'secretKey': secret_key, # Manual update of secret key
                'updated_at': datetime.now(ZoneInfo("UTC")).isoformat()
            }
            result = employees_collection.update_one(
                {'employeeId': employee_id},
                {'$set': updated_employee}
            )
            if result.matched_count == 0:
                return jsonify({'error': 'Employee not found'}), 404
            users_collection.update_one(
                {'email': email},
                {'$set': {
                    'email': email,
                    'name': data['name'],
                    'role': data['role'],
                    'updated_at': datetime.now(ZoneInfo("UTC")).isoformat()
                }},
                upsert=True
            )
            # Fetch updated employee to return
            updated_full = employees_collection.find_one({'employeeId': employee_id})
            logger.info(f"Updated employee: {employee_id} with email: {email} and secret key: {secret_key}")
            return jsonify({'message': 'Employee updated successfully', 'employee': updated_full}), 200
        except Exception as e:
            logger.error(f"Error updating employee: {str(e)}")
            return jsonify({'error': str(e)}), 500

    # FIXED: Renamed route to avoid name collision with helper function
    @app.route('/api/employees/validate-secret-key', methods=['POST'])
    @db_required
    def validate_employee_secret_key(): # Renamed route function
        try:
            data = request.get_json()
            secret_key = data.get('secretKey')
            if not secret_key:
                return jsonify({'error': 'Secret key required'}), 400
            if not is_valid_secret_key(secret_key):  # FIXED: Global helper now available
                return jsonify({'error': 'Secret key must be exactly 6 digits'}), 400
            employee = employees_collection.find_one({'secretKey': secret_key})
            if not employee:
                return jsonify({'error': 'Invalid secret key'}), 400
            return jsonify({'employeeId': employee['employeeId'], 'name': employee['name']}), 200
        except Exception as e:
            logger.error(f"Error validating secret key: {str(e)}")
            return jsonify({'error': str(e)}), 500

    @app.route('/api/employees/<employee_id>', methods=['DELETE'])
    @db_required
    def delete_employee(employee_id):
        try:
            employee = employees_collection.find_one({'employeeId': employee_id})
            if not employee:
                return jsonify({'error': 'Employee not found'}), 404
            result = employees_collection.delete_one({'employeeId': employee_id})
            if result.deleted_count == 0:
                return jsonify({'error': 'Employee not found'}), 404
            if not employees_collection.find_one({'email': employee['email']}):
                users_collection.delete_one({'email': employee['email']})
            logger.info(f"Deleted employee: {employee_id}")
            return jsonify({'message': 'Employee deleted successfully'}), 200
        except Exception as e:
            logger.error(f"Error deleting employee: {str(e)}")
            return jsonify({"error": str(e)}), 500

    # UPDATED: Endpoint to get trip reports - now fetches matching sales for the employee (case-insensitive deliveryPersonName, orderType='Online Delivery', status != 'Cancelled')
    @app.route('/api/tripreports/<employee_id>', methods=['GET'])
    @db_required
    def get_trip_reports(employee_id):
        try:
            # Fetch employee by ID
            employee = employees_collection.find_one({'employeeId': employee_id})
            if not employee:
                logger.warning(f"Employee not found for ID: {employee_id}")
                return jsonify([]), 200
            employee_name = str(employee.get('name', 'Unknown')).lower() # Ensure str and lower for case-insensitive
            logger.info(f"Fetching trip reports for employee: {employee.get('name')} (ID: {employee_id})")
            # Fetch all sales and filter in Python for case-insensitive matching
            all_sales = sales_collection.find()
            reports = []
            for sale in all_sales:
                # FIXED: Handle None for deliveryPersonName - use or '' to ensure string before .lower()
                delivery_person_name = sale.get('deliveryPersonName') or ''
                if (delivery_person_name.lower() == employee_name and
                    sale.get('orderType') == 'Online Delivery' and
                    sale.get('status') != 'Cancelled'):
                    # Map items to cartItems for frontend compatibility
                    if 'items' in sale:
                        sale['cartItems'] = sale['items']
                        del sale['items'] # Avoid duplication
                    reports.append(sale)
            logger.info(f"Fetched {len(reports)} matching sales as trip reports for {employee.get('name')}")
            # Ensure deliveryPersonName is set if missing (though filter ensures it)
            for report in reports:
                if not report.get('deliveryPersonName'):
                    report['deliveryPersonName'] = employee.get('name')
            return jsonify(convert_objectid_to_str(reports)), 200
        except Exception as e:
            logger.error(f"Error fetching trip reports: {str(e)}")
            logger.error(traceback.format_exc())
            return jsonify({"error": str(e)}), 500
    @app.route('/api/uoms', methods=['GET'])
    @db_required
    def get_uoms():
        try:
            uoms = uoms_collection.find()
            return jsonify(convert_objectid_to_str(uoms)), 200
        except Exception as e:
            return jsonify({'error': f"Failed to fetch UOMs: {str(e)}"}), 500
    @app.route('/api/uoms', methods=['POST'])
    @db_required
    def add_uom():
        try:
            data = request.json
            if not data or 'name' not in data or not data['name'].strip():
                return jsonify({'error': 'Invalid UOM name'}), 400
            if uoms_collection.find_one({'name': data['name']}):
                return jsonify({'error': 'UOM already exists'}), 400
            uom = {
                'name': data['name'].strip(),
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            result = uoms_collection.insert_one(uom)
            inserted_uom = uoms_collection.find_one({'_id': result.inserted_id})
            return jsonify(convert_objectid_to_str(inserted_uom)), 201
        except Exception as e:
            return jsonify({'error': f"Failed to add UOM: {str(e)}"}), 500
    @app.route('/api/uoms/<id>', methods=['PUT'])
    @db_required
    def update_uom(id):
        try:
            data = request.json
            if not data or 'name' not in data or not data['name'].strip():
                return jsonify({'error': 'Invalid UOM name'}), 400
            existing = uoms_collection.find_one({'name': data['name'], '_id': {'$ne': id}})
            if existing:
                return jsonify({'error': 'UOM name already exists'}), 400
            result = uoms_collection.update_one({'_id': id}, {'$set': {'name': data['name'].strip()}})
            if result.matched_count == 0:
                return jsonify({'error': 'UOM not found or no changes'}), 404
            return jsonify({'message': 'UOM updated successfully'}), 200
        except Exception as e:
            return jsonify({'error': f"Failed to update UOM: {str(e)}"}), 500
    @app.route('/api/uoms/<id>', methods=['DELETE'])
    @db_required
    def delete_uom(id):
        try:
            result = uoms_collection.delete_one({'_id': id})
            if result.deleted_count == 0:
                return jsonify({'error': 'UOM not found'}), 404
            return jsonify({'message': 'UOM deleted successfully'}), 200
        except Exception as e:
            return jsonify({'error': f"Failed to delete UOM: {str(e)}"}), 500
    @app.route('/api/purchase_items', methods=['GET'])
    @db_required
    def get_purchase_items():
        try:
            items = purchase_items_collection.find()
            return jsonify(convert_objectid_to_str(items)), 200
        except Exception as e:
            return jsonify({'error': f"Failed to fetch items: {str(e)}"}), 500
    @app.route('/api/purchase_items', methods=['POST'])
    @db_required
    def add_purchase_item():
        try:
            data = request.json
            required_fields = ['company', 'name', 'boxToMaster', 'masterUnit', 'masterToOuter', 'outerUnit', 'outerToNos', 'nosUnit']
            if not all(key in data for key in required_fields):
                return jsonify({'error': 'Missing required fields'}), 400
            item = {
                'company': data['company'],
                'name': data['name'],
                'boxToMaster': float(data['boxToMaster']),
                'masterUnit': data['masterUnit'],
                'masterToOuter': float(data['masterToOuter']),
                'outerUnit': data['outerUnit'],
                'outerToNos': float(data['outerToNos']),
                'nosUnit': data['nosUnit'],
                'conversionFactor': float(data['masterToOuter']) * float(data['outerToNos']),
                'stockMaster': 0,
                'stockOuter': 0,
                'stockNos': 0,
                'soldNos': 0,
                'totalStock': 0,
                'totalPurchased': 0,
                'grams': float(data.get('grams', 0)),
                'suppliers': data.get('suppliers', []),
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            result = purchase_items_collection.insert_one(item)
            inserted_item = purchase_items_collection.find_one({'_id': result.inserted_id})
            return jsonify(convert_objectid_to_str(inserted_item)), 201
        except ValueError:
            return jsonify({'error': 'Invalid numeric value'}), 400
        except Exception as e:
            return jsonify({'error': f"Failed to add item: {str(e)}"}), 500
    @app.route('/api/purchase_items/<id>', methods=['PUT'])
    @db_required
    def update_purchase_item(id):
        try:
            data = request.json
            if not data:
                return jsonify({'error': 'No data provided'}), 400
            old_item = purchase_items_collection.find_one({'_id': id})
            if not old_item:
                return jsonify({'error': 'Item not found'}), 404
            update_data = {}
            if "company" in data:
                update_data["company"] = data["company"]
            if "name" in data:
                update_data["name"] = data["name"]
            if "boxToMaster" in data:
                update_data["boxToMaster"] = float(data["boxToMaster"])
            if "masterUnit" in data:
                update_data["masterUnit"] = data["masterUnit"]
            if "masterToOuter" in data:
                update_data["masterToOuter"] = float(data["masterToOuter"])
            if "outerUnit" in data:
                update_data["outerUnit"] = data["outerUnit"]
            if "outerToNos" in data:
                update_data["outerToNos"] = float(data["outerToNos"])
            if "nosUnit" in data:
                update_data["nosUnit"] = data["nosUnit"]
            if "grams" in data:
                update_data["grams"] = float(data["grams"])
            if "suppliers" in data:
                update_data["suppliers"] = data["suppliers"]
            if any(key in update_data for key in ["masterToOuter", "outerToNos"]):
                masterToOuter = update_data.get("masterToOuter", old_item["masterToOuter"])
                outerToNos = update_data.get("outerToNos", old_item["outerToNos"])
                update_data["conversionFactor"] = masterToOuter * outerToNos
            result = purchase_items_collection.update_one({'_id': id}, {'$set': update_data})
            if result.matched_count == 0:
                return jsonify({'error': 'Item not found'}), 404
            return jsonify({'message': 'Item updated successfully'}), 200
        except ValueError:
            return jsonify({'error': 'Invalid input data'}), 400
        except Exception as e:
            return jsonify({'error': f"Failed to update item: {str(e)}"}), 500
    @app.route('/api/purchase_items/<id>', methods=['DELETE'])
    @db_required
    def delete_purchase_item(id):
        try:
            result = purchase_items_collection.delete_one({'_id': id})
            if result.deleted_count == 0:
                return jsonify({'error': 'Item not found'}), 404
            return jsonify({'message': 'Item deleted successfully'}), 200
        except Exception as e:
            return jsonify({'error': f"Failed to delete item: {str(e)}"}), 500
    @app.route('/api/suppliers', methods=['GET'])
    @db_required
    def get_suppliers():
        try:
            suppliers = suppliers_collection.find()
            return jsonify(convert_objectid_to_str(suppliers)), 200
        except Exception as e:
            return jsonify({'error': f"Failed to fetch suppliers: {str(e)}"}), 500
    @app.route('/api/suppliers', methods=['POST'])
    @db_required
    def add_supplier():
        try:
            data = request.json
            supplier = {
                'company': data.get('company', ''),
                'code': data.get('code', ''),
                'supplier_names': data.get('supplier_names', []),
                'group': data.get('group', ''),
                'country': data.get('country', ''),
                'currency': data.get('currency', ''),
                'taxId': data.get('taxId', ''),
                'taxCategory': data.get('taxCategory', ''),
                'taxWithholdingCategory': data.get('taxWithholdingCategory', ''),
                'contacts': data.get('contacts', []),
                'paymentMode': data.get('paymentMode', ''),
                'paymentTerms': data.get('paymentTerms', ''),
                'creditLimit': float(data.get('creditLimit', 0)),
                'paymentTermsOverride': data.get('paymentTermsOverride', ''),
                'bankDetails': data.get('bankDetails', ''),
                'website': data.get('website', ''),
                'onTimeDelivery': float(data.get('onTimeDelivery', 0)),
                'defectRate': float(data.get('defectRate', 0)),
                'lastPurchaseDate': data.get('lastPurchaseDate', None),
                'lastPurchaseValue': float(data.get('lastPurchaseValue', 0)),
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            result = suppliers_collection.insert_one(supplier)
            inserted_supplier = suppliers_collection.find_one({'_id': result.inserted_id})
            return jsonify(convert_objectid_to_str(inserted_supplier)), 201
        except Exception as e:
            return jsonify({'error': f"Failed to add supplier: {str(e)}"}), 500
    @app.route('/api/suppliers/<id>', methods=['PUT'])
    @db_required
    def update_supplier(id):
        try:
            data = request.json
            if not data:
                return jsonify({'error': 'No data provided'}), 400
            update_fields = {}
            fields = [
                'company', 'code', 'supplier_names', 'group', 'country', 'currency',
                'taxId', 'taxCategory', 'taxWithholdingCategory', 'contacts',
                'paymentMode', 'paymentTerms', 'creditLimit', 'paymentTermsOverride',
                'bankDetails', 'website', 'onTimeDelivery', 'defectRate',
                'lastPurchaseDate', 'lastPurchaseValue'
            ]
            for field in fields:
                if field in data:
                    if field in ['creditLimit', 'onTimeDelivery', 'defectRate', 'lastPurchaseValue']:
                        update_fields[field] = float(data[field])
                    else:
                        update_fields[field] = data[field]
            result = suppliers_collection.update_one({'_id': id}, {'$set': update_fields})
            if result.matched_count == 0:
                return jsonify({'error': 'Supplier not found'}), 404
            return jsonify({'message': 'Supplier updated successfully'}), 200
        except ValueError:
            return jsonify({'error': 'Invalid input data'}), 400
        except Exception as e:
            return jsonify({'error': f"Failed to update supplier: {str(e)}"}), 500
    @app.route('/api/suppliers/<id>', methods=['DELETE'])
    @db_required
    def delete_supplier(id):
        try:
            result = suppliers_collection.delete_one({'_id': id})
            if result.deleted_count == 0:
                return jsonify({'error': 'Supplier not found'}), 404
            return jsonify({'message': 'Supplier deleted successfully'}), 200
        except Exception as e:
            return jsonify({'error': f"Failed to delete supplier: {str(e)}"}), 500
    @app.route('/api/purchase_orders', methods=['GET'])
    @db_required
    def get_purchase_orders():
        try:
            orders = purchase_orders_collection.find()
            return jsonify(convert_objectid_to_str(orders)), 200
        except Exception as e:
            return jsonify({'error': f"Failed to fetch purchase orders: {str(e)}"}), 500
    @app.route('/api/purchase_orders', methods=['POST'])
    @db_required
    def add_purchase_order():
        try:
            data = request.json
            required_fields = ['series', 'date', 'company', 'supplierId', 'name', 'supplierCompany', 'supplierCode', 'supplierGroup', 'supplierCode', 'supplierGroup', 'supplierCode', 'supplierGroup', 'address', 'phone', 'email', 'currency', 'items', 'taxes', 'subtotal', 'totalQuantity', 'totalTaxes', 'grandTotal', 'status']
            if not all(key in data for key in required_fields):
                return jsonify({'error': 'Missing required fields'}), 400
            if purchase_orders_collection.find_one({'series': data['series'] }):
                return jsonify({'error': 'Series already exists'}), 400
            supplier = suppliers_collection.find_one({'_id': data['supplierId']})
            if not supplier:
                return jsonify({'error': 'Supplier not found'}), 404
            items = []
            for item_data in data['items']:
                item_doc = purchase_items_collection.find_one({'_id': item_data['itemId']})
                if not item_doc:
                    return jsonify({'error': f"Item {item_data['itemId']} not found"}), 404
                items.append({
                    'itemId': item_data['itemId'],
                    'quantity': float(item_data['quantity']),
                    'uom': item_data['uom'],
                    'rate': float(item_data.get('rate', 0)),
                    'amount': float(item_data.get('amount', 0))
                })
            order = {
                'series': data['series'],
                'date': datetime.fromisoformat(str(data['date']).replace('Z', '+00:00')),
                'company': data['company'],
                'supplierId': data['supplierId'],
                'name': data['name'],
                'supplierCompany': data['supplierCompany'],
                'supplierCode': data.get('supplierCode', ''),
                'supplierGroup': data.get('supplierGroup', ''),
                'supplierCode': data.get('supplierCode', ''),
                'supplierGroup': data.get('supplierGroup', ''),
                'supplierCode': data.get('supplierCode', ''),
                'supplierGroup': data.get('supplierGroup', ''),
                'address': data['address'],
                'phone': data['phone'],
                'email': data['email'],
                'currency': data['currency'],
                'targetWarehouse': data.get('targetWarehouse', ''),
                'items': items,
                'taxes': data['taxes'],
                'subtotal': float(data['subtotal']),
                'totalQuantity': float(data['totalQuantity']),
                'totalTaxes': float(data['totalTaxes']),
                'grandTotal': float(data['grandTotal']),
                'status': data['status'],
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            result = purchase_orders_collection.insert_one(order)
            inserted_order = purchase_orders_collection.find_one({'_id': result.inserted_id})
            return jsonify(convert_objectid_to_str(inserted_order)), 201
        except ValueError as e:
            return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
        except Exception as e:
            return jsonify({'error': f"Failed to create purchase order: {str(e)}"}), 500
    @app.route('/api/purchase_orders/<id>', methods=['PUT'])
    @db_required
    def update_purchase_order(id):
        try:
            data = request.json
            if not data:
                return jsonify({'error': 'No input data provided'}), 400
            old_order = purchase_orders_collection.find_one({'_id': id})
            if not old_order:
                return jsonify({'error': 'Purchase Order not found'}), 404
            update_data = {}
            for field in ['series', 'date', 'company', 'supplierId', 'name', 'supplierCompany', 'supplierCode', 'supplierGroup', 'supplierCode', 'supplierGroup', 'supplierCode', 'supplierGroup', 'address', 'phone', 'email', 'currency', 'targetWarehouse', 'items', 'taxes', 'subtotal', 'totalQuantity', 'totalTaxes', 'grandTotal', 'status']:
                if field in data:
                    if field == 'date':
                        update_data[field] = datetime.fromisoformat(str(data[field]).replace('Z', '+00:00'))
                    elif field in ['subtotal', 'totalQuantity', 'totalTaxes', 'grandTotal']:
                        update_data[field] = float(data[field])
                    elif field == 'items':
                        items = []
                        for item_data in data[field]:
                            items.append({
                                'itemId': item_data['itemId'],
                                'quantity': float(item_data['quantity']),
                                'uom': item_data['uom'],
                                'rate': float(item_data.get('rate', 0)),
                                'amount': float(item_data.get('amount', 0))
                            })
                        update_data[field] = items
                    else:
                        update_data[field] = data[field]
            if not update_data:
                return jsonify({'error': 'No fields to update'}), 400
            result = purchase_orders_collection.update_one({'_id': id}, {'$set': update_data})
            if result.matched_count == 0:
                return jsonify({'error': 'Purchase Order not found'}), 404
            return jsonify({'message': 'Purchase Order updated successfully'}), 200
        except ValueError as e:
            return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
        except Exception as e:
            return jsonify({'error': f"Failed to update purchase order: {str(e)}"}), 500
    @app.route('/api/purchase_orders/<id>', methods=['DELETE'])
    @db_required
    def delete_purchase_order(id):
        try:
            result = purchase_orders_collection.delete_one({'_id': id})
            if result.deleted_count == 0:
                return jsonify({'error': 'Purchase Order not found'}), 404
            return jsonify({'message': 'Purchase Order deleted successfully'}), 200
        except Exception as e:
            return jsonify({'error': f"Failed to delete purchase order: {str(e)}"}), 500
    @app.route('/api/purchase_receipts', methods=['GET'])
    @db_required
    def get_purchase_receipts():
        try:
            receipts = purchase_receipts_collection.find()
            return jsonify(convert_objectid_to_str(receipts)), 200
        except Exception as e:
            return jsonify({'error': f"Failed to fetch purchase receipts: {str(e)}"}), 500
    @app.route('/api/purchase_receipts', methods=['POST'])
    @db_required
    def add_purchase_receipt():
        try:
            data = request.json
            required_fields = ['series', 'date', 'poId', 'company', 'supplierId', 'name', 'supplierCompany', 'supplierCode', 'supplierGroup', 'supplierCode', 'supplierGroup', 'supplierCode', 'supplierGroup', 'address', 'phone', 'email', 'items', 'taxes', 'subtotal', 'totalTaxes', 'grandTotal', 'status']
            if not all(key in data for key in required_fields):
                return jsonify({'error': 'Missing required fields'}), 400
            if purchase_receipts_collection.find_one({'series': data['series'] }):
                return jsonify({'error': 'Series already exists'}), 400
            po = purchase_orders_collection.find_one({'series': data['poId']})
            if not po:
                return jsonify({'error': 'Purchase Order not found'}), 404
            items = []
            for item_data in data['items']:
                item_doc = purchase_items_collection.find_one({'_id': item_data['itemId']})
                if not item_doc:
                    return jsonify({'error': f"Item {item_data['itemId']} not found"}), 404
                items.append({
                    'itemId': item_data['itemId'],
                    'originalQuantity': float(item_data['originalQuantity']),
                    'acceptedQuantity': float(item_data['acceptedQuantity']),
                    'rejectedQuantity': float(item_data['rejectedQuantity']),
                    'rate': float(item_data.get('rate', 0)),
                    'amount': float(item_data.get('amount', 0)),
                    'unit': item_data['unit']
                })
            receipt = {
                'series': data['series'],
                'date': datetime.fromisoformat(str(data['date']).replace('Z', '+00:00')),
                'poId': data['poId'],
                'company': data['company'],
                'supplierId': data['supplierId'],
                'name': data['name'],
                'supplierCompany': data['supplierCompany'],
                'supplierCode': data.get('supplierCode', ''),
                'supplierGroup': data.get('supplierGroup', ''),
                'supplierCode': data.get('supplierCode', ''),
                'supplierGroup': data.get('supplierGroup', ''),
                'supplierCode': data.get('supplierCode', ''),
                'supplierGroup': data.get('supplierGroup', ''),
                'address': data['address'],
                'phone': data['phone'],
                'email': data['email'],
                'currency': data['currency'],
                'items': items,
                'taxes': data['taxes'],
                'subtotal': float(data['subtotal']),
                'totalTaxes': float(data['totalTaxes']),
                'grandTotal': float(data['grandTotal']),
                'status': data['status'],
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            result = purchase_receipts_collection.insert_one(receipt)
            inserted_receipt = purchase_receipts_collection.find_one({'_id': result.inserted_id})
            if data['status'] == 'Submitted':
                for item in items:
                    item_obj = purchase_items_collection.find_one({'_id': item['itemId']})
                    add_master = 0
                    add_outer = 0
                    add_nos = 0
                    if item['unit'] == 'master':
                        add_master = item['acceptedQuantity']
                    elif item['unit'] == 'outer':
                        add_outer = item['acceptedQuantity']
                    elif item['unit'] == 'nos':
                        add_nos = item['acceptedQuantity']
                    total_added_in_nos = (add_master * item_obj['masterToOuter'] * item_obj['outerToNos']) + (add_outer * item_obj['outerToNos']) + add_nos
                    purchase_items_collection.update_one(
                        {'_id': item['itemId']},
                        {'$inc': {
                            'stockMaster': add_master,
                            'stockOuter': add_outer,
                            'stockNos': add_nos,
                            'totalStock': total_added_in_nos,
                            'totalPurchased': total_added_in_nos
                        }}
                    )
            return jsonify(convert_objectid_to_str(inserted_receipt)), 201
        except ValueError as e:
            return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
        except Exception as e:
            return jsonify({'error': f"Failed to create purchase receipt: {str(e)}"}), 500
    @app.route('/api/purchase_receipts/<series>', methods=['PUT'])
    @db_required
    def update_purchase_receipt(series):
        try:
            data = request.json
            if not data:
                return jsonify({'error': 'No input data provided'}), 400
            old_receipt = purchase_receipts_collection.find_one({'series': series})
            if not old_receipt:
                return jsonify({'error': 'Purchase Receipt not found'}), 404
            was_submitted = old_receipt['status'] == 'Submitted'
            new_status = data.get('status', old_receipt['status'])
            if was_submitted and new_status == 'Submitted':
                pass
            elif was_submitted and new_status != 'Submitted':
                for item in old_receipt['items']:
                    item_obj = purchase_items_collection.find_one({'_id': item['itemId']})
                    sub_master = 0
                    sub_outer = 0
                    sub_nos = 0
                    if item['unit'] == 'master':
                        sub_master = item['acceptedQuantity']
                    elif item['unit'] == 'outer':
                        sub_outer = item['acceptedQuantity']
                    elif item['unit'] == 'nos':
                        sub_nos = item['acceptedQuantity']
                    total_sub_in_nos = (sub_master * item_obj['masterToOuter'] * item_obj['outerToNos']) + (sub_outer * item_obj['outerToNos']) + sub_nos
                    purchase_items_collection.update_one(
                        {'_id': item['itemId']},
                        {'$inc': {
                            'stockMaster': -sub_master,
                            'stockOuter': -sub_outer,
                            'stockNos': -sub_nos,
                            'totalStock': -total_sub_in_nos,
                            'totalPurchased': -total_sub_in_nos
                        }}
                    )
            elif not was_submitted and new_status == 'Submitted':
                items = data.get('items', old_receipt['items'])
                for item in items:
                    item_obj = purchase_items_collection.find_one({'_id': item['itemId']})
                    add_master = 0
                    add_outer = 0
                    add_nos = 0
                    if item['unit'] == 'master':
                        add_master = item['acceptedQuantity']
                    elif item['unit'] == 'outer':
                        add_outer = item['acceptedQuantity']
                    elif item['unit'] == 'nos':
                        add_nos = item['acceptedQuantity']
                    total_added_in_nos = (add_master * item_obj['masterToOuter'] * item_obj['outerToNos']) + (add_outer * item_obj['outerToNos']) + add_nos
                    purchase_items_collection.update_one(
                        {'_id': item['itemId']},
                        {'$inc': {
                            'stockMaster': add_master,
                            'stockOuter': add_outer,
                            'stockNos': add_nos,
                            'totalStock': total_added_in_nos,
                            'totalPurchased': total_added_in_nos
                        }}
                    )
            update_fields = {}
            for field in ['date', 'poId', 'company', 'supplierId', 'name', 'supplierCompany', 'supplierCode', 'supplierGroup', 'supplierCode', 'supplierGroup', 'supplierCode', 'supplierGroup', 'address', 'phone', 'email', 'currency', 'items', 'taxes', 'subtotal', 'totalTaxes', 'grandTotal', 'status']:
                if field in data:
                    if field == 'date':
                        update_fields[field] = datetime.fromisoformat(str(data[field]).replace('Z', '+00:00'))
                    elif field in ['subtotal', 'totalTaxes', 'grandTotal']:
                        update_fields[field] = float(data[field])
                    elif field == 'items':
                        items = []
                        for item_data in data[field]:
                            items.append({
                                'itemId': item_data['itemId'],
                                'originalQuantity': float(item_data['originalQuantity']),
                                'acceptedQuantity': float(item_data['acceptedQuantity']),
                                'rejectedQuantity': float(item_data['rejectedQuantity']),
                                'rate': float(item_data.get('rate', 0)),
                                'amount': float(item_data.get('amount', 0)),
                                'unit': item_data['unit']
                            })
                        update_fields[field] = items
                    else:
                        update_fields[field] = data[field]
            if not update_fields:
                return jsonify({'error': 'No fields to update'}), 400
            result = purchase_receipts_collection.update_one({'series': series}, {'$set': update_fields})
            if result.matched_count == 0:
                return jsonify({'error': 'Purchase Receipt not found'}), 404
            return jsonify({'message': 'Purchase Receipt updated successfully'}), 200
        except ValueError as e:
            return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
        except Exception as e:
            return jsonify({'error': f"Failed to update purchase receipt: {str(e)}"}), 500
    @app.route('/api/purchase_receipts/<series>', methods=['DELETE'])
    @db_required
    def delete_purchase_receipt(series):
        try:
            old_receipt = purchase_receipts_collection.find_one({'series': series})
            if not old_receipt:
                return jsonify({'error': 'Purchase Receipt not found'}), 404
            if old_receipt['status'] == 'Submitted':
                for item in old_receipt['items']:
                    item_obj = purchase_items_collection.find_one({'_id': item['itemId']})
                    sub_master = 0
                    sub_outer = 0
                    sub_nos = 0
                    if item['unit'] == 'master':
                        sub_master = item['acceptedQuantity']
                    elif item['unit'] == 'outer':
                        sub_outer = item['acceptedQuantity']
                    elif item['unit'] == 'nos':
                        sub_nos = item['acceptedQuantity']
                    total_sub_in_nos = (sub_master * item_obj['masterToOuter'] * item_obj['outerToNos']) + (sub_outer * item_obj['outerToNos']) + sub_nos
                    purchase_items_collection.update_one(
                        {'_id': item['itemId']},
                        {'$inc': {
                            'stockMaster': -sub_master,
                            'stockOuter': -sub_outer,
                            'stockNos': -sub_nos,
                            'totalStock': -total_sub_in_nos,
                            'totalPurchased': -total_sub_in_nos
                        }}
                    )
            result = purchase_receipts_collection.delete_one({'series': series})
            if result.deleted_count == 0:
                return jsonify({'error': 'Purchase Receipt not found'}), 404
            return jsonify({'message': 'Purchase Receipt deleted successfully'}), 200
        except Exception as e:
            return jsonify({'error': f"Failed to delete purchase receipt: {str(e)}"}), 500
    @app.route('/api/purchase_invoices', methods=['GET'])
    @db_required
    def get_purchase_invoices():
        try:
            invoices = purchase_invoices_collection.find()
            return jsonify(convert_objectid_to_str(invoices)), 200
        except Exception as e:
            return jsonify({'error': f"Failed to fetch purchase invoices: {str(e)}"}), 500
    @app.route('/api/purchase_invoices', methods=['POST'])
    @db_required
    def add_purchase_invoice():
        try:
            data = request.json
            required_fields = ['series', 'date', 'company', 'supplierId', 'name', 'supplierCompany', 'supplierCode', 'supplierGroup', 'supplierCode', 'supplierGroup', 'supplierCode', 'supplierGroup', 'address', 'phone', 'email', 'poId', 'prId', 'currency', 'items', 'taxes', 'totalQuantity', 'subtotal', 'taxesAdded', 'grandTotal', 'status']
            if not all(key in data for key in required_fields):
                return jsonify({'error': 'Missing required fields'}), 400
            if purchase_invoices_collection.find_one({'series': data['series'] }):
                return jsonify({'error': 'Series already exists'}), 400
            pr = purchase_receipts_collection.find_one({'series': data['prId']})
            if not pr:
                return jsonify({'error': 'Purchase Receipt not found'}), 404
            items = []
            for item_data in data['items']:
                item_doc = purchase_items_collection.find_one({'_id': item_data['itemId']})
                if not item_doc:
                    return jsonify({'error': f"Item {item_data['itemId']} not found"}), 404
                items.append({
                    'itemId': item_data['itemId'],
                    'acceptedQuantity': float(item_data['acceptedQuantity']),
                    'rate': float(item_data.get('rate', 0)),
                    'amount': float(item_data.get('amount', 0)),
                    'unit': item_data['unit']
                })
            invoice = {
                'series': data['series'],
                'date': datetime.fromisoformat(str(data['date']).replace('Z', '+00:00')),
                'company': data['company'],
                'supplierId': data['supplierId'],
                'name': data['name'],
                'supplierCompany': data['supplierCompany'],
                'supplierCode': data.get('supplierCode', ''),
                'supplierGroup': data.get('supplierGroup', ''),
                'supplierCode': data.get('supplierCode', ''),
                'supplierGroup': data.get('supplierGroup', ''),
                'supplierCode': data.get('supplierCode', ''),
                'supplierGroup': data.get('supplierGroup', ''),
                'address': data['address'],
                'phone': data['phone'],
                'email': data['email'],
                'poId': data['poId'],
                'prId': data['prId'],
                'currency': data['currency'],
                'items': items,
                'taxes': data['taxes'],
                'totalQuantity': float(data['totalQuantity']),
                'subtotal': float(data['subtotal']),
                'taxesAdded': float(data['taxesAdded']),
                'grandTotal': float(data['grandTotal']),
                'status': data['status'],
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            result = purchase_invoices_collection.insert_one(invoice)
            inserted_invoice = purchase_invoices_collection.find_one({'_id': result.inserted_id})
            if data['status'] == 'Submitted':
                suppliers_collection.update_one(
                    {'_id': data['supplierId']},
                    {'$set': {
                        'lastPurchaseDate': invoice['date'],
                        'lastPurchaseValue': invoice['grandTotal']
                    }}
                )
            return jsonify(convert_objectid_to_str(inserted_invoice)), 201
        except ValueError as e:
            return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
        except Exception as e:
            return jsonify({'error': f"Failed to create purchase invoice: {str(e)}"}), 500
    @app.route('/api/purchase_invoices/<series>', methods=['PUT'])
    @db_required
    def update_purchase_invoice(series):
        try:
            data = request.json
            if not data:
                return jsonify({'error': 'No input data provided'}), 400
            old_invoice = purchase_invoices_collection.find_one({'series': series})
            if not old_invoice:
                return jsonify({'error': 'Purchase Invoice not found'}), 404
            update_fields = {}
            for field in ['date', 'company', 'supplierId', 'name', 'supplierCompany', 'supplierCode', 'supplierGroup', 'supplierCode', 'supplierGroup', 'supplierCode', 'supplierGroup', 'address', 'phone', 'email', 'poId', 'prId', 'currency', 'items', 'taxes', 'totalQuantity', 'subtotal', 'taxesAdded', 'grandTotal', 'status']:
                if field in data:
                    if field == 'date':
                        update_fields[field] = datetime.fromisoformat(str(data[field]).replace('Z', '+00:00'))
                    elif field in ['totalQuantity', 'subtotal', 'taxesAdded', 'grandTotal']:
                        update_fields[field] = float(data[field])
                    elif field == 'items':
                        items = []
                        for item_data in data[field]:
                            items.append({
                                'itemId': item_data['itemId'],
                                'acceptedQuantity': float(item_data['acceptedQuantity']),
                                'rate': float(item_data.get('rate', 0)),
                                'amount': float(item_data.get('amount', 0)),
                                'unit': item_data['unit']
                            })
                        update_fields[field] = items
                    else:
                        update_fields[field] = data[field]
            if not update_fields:
                return jsonify({'error': 'No fields to update'}), 400
            result = purchase_invoices_collection.update_one({'series': series}, {'$set': update_fields})
            if result.matched_count == 0:
                return jsonify({'error': 'Purchase Invoice not found'}), 404
            new_status = data.get('status', old_invoice['status'])
            supplier_id = data.get('supplierId', old_invoice.get('supplierId'))
            if new_status == 'Submitted' and supplier_id:
                suppliers_collection.update_one(
                    {'_id': supplier_id},
                    {'$set': {
                        'lastPurchaseDate': data.get('date', old_invoice['date']),
                        'lastPurchaseValue': float(data.get('grandTotal', old_invoice['grandTotal']))
                    }}
                )
            return jsonify({'message': 'Purchase Invoice updated successfully'}), 200
        except ValueError as e:
            return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
        except Exception as e:
            return jsonify({'error': f"Failed to update purchase invoice: {str(e)}"}), 500
    @app.route('/api/purchase_invoices/<series>', methods=['DELETE'])
    @db_required
    def delete_purchase_invoice(series):
        try:
            result = purchase_invoices_collection.delete_one({'series': series})
            if result.deleted_count == 0:
                return jsonify({'error': 'Purchase Invoice not found'}), 404
            return jsonify({'message': 'Purchase Invoice deleted successfully'}), 200
        except Exception as e:
            return jsonify({'error': f"Failed to delete purchase invoice: {str(e)}"}), 500
    @app.route('/api/purchase_sales', methods=['GET'])
    @db_required
    def get_purchase_sales():
        try:
            sales = purchase_sales_collection.find()
            return jsonify(convert_objectid_to_str(sales)), 200
        except Exception as e:
            return jsonify({'error': f"Failed to fetch sales: {str(e)}"}), 500
    @app.route('/api/purchase_sales', methods=['POST'])
    @db_required
    def add_purchase_sale():
        try:
            data = request.json
            required_fields = ['itemId', 'quantity']
            if not all(key in data for key in required_fields):
                return jsonify({'error': 'Missing required fields'}), 400
            item = purchase_items_collection.find_one({'_id': data['itemId']})
            if not item:
                return jsonify({'error': 'Item not found'}), 404
            quantity = float(data['quantity'])
            if quantity > item['totalStock']:
                return jsonify({'error': 'Insufficient stock'}), 400
            sale = {
                'itemId': data['itemId'],
                'quantity': quantity,
                'date': datetime.now(timezone.utc).isoformat(),
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            result = purchase_sales_collection.insert_one(sale)
            purchase_items_collection.update_one(
                {'_id': data['itemId']},
                {'$inc': {
                    'soldNos': quantity,
                    'totalStock': -quantity
                }}
            )
            inserted_sale = purchase_sales_collection.find_one({'_id': result.inserted_id})
            return jsonify(convert_objectid_to_str(inserted_sale)), 201
        except ValueError as e:
            return jsonify({'error': f"Invalid data format: {str(e)}"}), 400
        except Exception as e:
            return jsonify({'error': f"Failed to record sale: {str(e)}"}), 500
        
    @app.route('/api/supplier_groups', methods=['GET'])
    @db_required
    def get_supplier_groups():
        try:
            groups = supplier_group_collection.find()
            return jsonify(convert_objectid_to_str(groups)), 200
        except Exception as e:
            return jsonify({'error': f"Failed to fetch supplier groups: {str(e)}"}), 500
    @app.route('/api/supplier_groups', methods=['POST'])
    @db_required
    def add_supplier_group():
        try:
            data = request.json
            if not data or 'group_name' not in data or not data['group_name'].strip():
                return jsonify({'error': 'Invalid group name'}), 400
            if supplier_group_collection.find_one({'group_name': data['group_name']}):
                return jsonify({'error': 'Group already exists'}), 400
            group = {
                'group_name': data['group_name'].strip(),
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            result = supplier_group_collection.insert_one(group)
            inserted_group = supplier_group_collection.find_one({'_id': result.inserted_id})
            return jsonify(convert_objectid_to_str(inserted_group)), 201
        except Exception as e:
            return jsonify({'error': f"Failed to add supplier group: {str(e)}"}), 500
    @app.route('/api/supplier_groups/<id>', methods=['PUT'])
    @db_required
    def update_supplier_group(id):
        try:
            data = request.json
            if not data or 'group_name' not in data or not data['group_name'].strip():
                return jsonify({'error': 'Invalid group name'}), 400
            existing = supplier_group_collection.find_one({'group_name': data['group_name'], '_id': {'$ne': id}})
            if existing:
                return jsonify({'error': 'Group name already exists'}), 400
            result = supplier_group_collection.update_one({'_id': id}, {'$set': {'group_name': data['group_name'].strip()}})
            if result.matched_count == 0:
                return jsonify({'error': 'Group not found or no changes'}), 404
            return jsonify({'message': 'Supplier group updated successfully'}), 200
        except Exception as e:
            return jsonify({'error': f"Failed to update supplier group: {str(e)}"}), 500
    @app.route('/api/supplier_groups/<id>', methods=['DELETE'])
    @db_required
    def delete_supplier_group(id):
        try:
            result = supplier_group_collection.delete_one({'_id': id})
            if result.deleted_count == 0:
                return jsonify({'error': 'Group not found'}), 404
            return jsonify({'message': 'Supplier group deleted successfully'}), 200
        except Exception as e:
            return jsonify({'error': f"Failed to delete supplier group: {str(e)}"}), 500
    @app.route('/api/print_settings/active', methods=['GET'])
    @db_required
    def get_active_print_settings():
        try:
            setting = print_settings_collection.find_one({"active": True})
            if not setting:
                return jsonify({"error": "No active print setting found"}), 404
            return jsonify(convert_objectid_to_str(setting)), 200
        except Exception as e:
            logger.error(f"Error fetching active print setting: {str(e)}")
            return jsonify({"error": "Internal server error"}), 500
    @app.route('/api/print_settings/deactivate_all', methods=['PUT'])
    @db_required
    def deactivate_all_print_settings():
        try:
            print_settings_collection.update_many({}, {"$set": {"active": False}})
            return jsonify({"message": "All print settings deactivated successfully"})
        except Exception as e:
            logger.error(f"Error deactivating print settings: {str(e)}")
            return jsonify({"error": "Internal server error"}), 500
    @app.route('/api/print_settings', methods=['GET'])
    @db_required
    def get_all_print_settings():
        try:
            settings = print_settings_collection.find()
            return jsonify(convert_objectid_to_str(settings)), 200
        except Exception as e:
            logger.error(f"Error fetching print settings: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/print_settings', methods=['POST'])
    @db_required
    def create_print_settings():
        try:
            data = request.get_json()
            if not data:
                return jsonify({"error": "No data provided"}), 400
            data['active'] = data.get('active', False)
            data['created_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
            result = print_settings_collection.insert_one(data)
            if not print_settings_collection.find_one({"_id": {"$ne": result.inserted_id}, "active": True}):
                print_settings_collection.update_many({"_id": {"$ne": result.inserted_id}}, {"$set": {"active": False}})
            logger.info(f"Print settings created with ID: {result.inserted_id}")
            return jsonify({"message": "Print settings created successfully", "id": result.inserted_id}), 201
        except Exception as e:
            logger.error(f"Error creating print settings: {str(e)}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/print_settings/<id>', methods=['GET', 'PUT', 'DELETE', 'OPTIONS'])
    @db_required
    def print_setting(id):
        if request.method == 'OPTIONS':
            response = jsonify({"success": True})
            response.headers['Access-Control-Allow-Origin'] = '*'
            response.headers['Access-Control-Allow-Methods'] = 'GET, PUT, DELETE, OPTIONS'
            response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
            return response, 200
        if request.method == 'GET':
            setting = print_settings_collection.find_one({"_id": id})
            if not setting:
                return jsonify({"error": "Setting not found"}), 404
            return jsonify(convert_objectid_to_str(setting)), 200
        if request.method == 'PUT':
            data = request.get_json()
            if not data:
                return jsonify({"error": "No data provided"}), 400
            update_data = {k: v for k, v in data.items() if k != '_id' and k != 'active'}
            result = print_settings_collection.update_one({"_id": id}, {"$set": update_data})
            if result.matched_count == 0:
                return jsonify({"error": "Setting not found"}), 404
            return jsonify({"message": "Print settings updated successfully"}), 200
        if request.method == 'DELETE':
            result = print_settings_collection.delete_one({"_id": id})
            if result.deleted_count == 0:
                return jsonify({"error": "Setting not found"}), 404
            return jsonify({"message": "Print settings deleted successfully"}), 200
    @app.route('/api/print_settings/set_active/<id>', methods=['PUT'])
    @db_required
    def set_active_print_settings(id):
        try:
            print_settings_collection.update_many({}, {"$set": {"active": False}})
            result = print_settings_collection.update_one({"_id": id}, {"$set": {"active": True}})
            if result.matched_count == 0:
                return jsonify({"error": "Setting not found"}), 404
            return jsonify({"message": "Active print settings set successfully"}), 200
        except Exception as e:
            logger.error(f"Error setting active print settings: {str(e)}")
            return jsonify({"error": "Internal server error"}), 500
    @app.route('/api/upload-combo-image', methods=['POST'])
    @db_required
    def upload_combo_image():
        try:
            if 'file' not in request.files:
                logger.error("No file provided for combo image upload")
                return jsonify({"error": "No file provided"}), 400
            
            file = request.files['file']
            if file.filename == '':
                logger.error("No file selected for combo image upload")
                return jsonify({"error": "No file selected"}), 400
            
            if file:
                filename = secure_filename(file.filename)
                if not filename:
                    logger.error("Invalid filename for combo image upload")
                    return jsonify({"error": "Invalid filename"}), 400
                
                # UPDATED: Use dynamic COMBO_IMAGES_DIR (writable via UPLOAD_FOLDER in EXE)
                upload_dir = COMBO_IMAGES_DIR
                os.makedirs(upload_dir, exist_ok=True)  # Safe create if not exists
                
                filepath = os.path.join(upload_dir, filename)
                
                # Check if file already exists, append timestamp if needed to avoid overwrite
                if os.path.exists(filepath):
                    name, ext = os.path.splitext(filename)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"{name}_{timestamp}{ext}"
                    filepath = os.path.join(upload_dir, filename)
                    logger.info(f"Filename collision detected, renamed to: {filename}")
                
                file.save(filepath)
                logger.info(f"Combo image uploaded successfully: {filename} to {upload_dir}")
                return jsonify({"filename": filename}), 200
            
            return jsonify({"error": "Upload failed - no file processed"}), 500
        
        except Exception as e:
            error_msg = f"Error uploading combo image: {str(e)}\n{traceback.format_exc()}"
            logger.error(error_msg)
            return jsonify({"error": str(e)}), 500

    # UPDATED: Serve combo images from dynamic directory
    @app.route('/api/combo-images/<filename>')
    def serve_combo_image(filename):
        try:
            # UPDATED: Use dynamic COMBO_IMAGES_DIR parent for send_from_directory
            # send_from_directory expects the directory containing the files
            upload_dir = COMBO_IMAGES_DIR
            return send_from_directory(upload_dir, filename)
        except FileNotFoundError:
            logger.warning(f"Combo image not found: {filename}")
            return "Image not found", 404
        except Exception as e:
            logger.error(f"Error serving combo image {filename}: {str(e)}\n{traceback.format_exc()}")
            return "Server error", 500

    @app.route('/api/combo-offer', methods=['GET'])
    @db_required
    def get_combo_offers():
        try:
            # NEW: Run cleanup before fetching (in addition to scheduler)
            clean_expired_combo_offers()
            offers = combo_offers_collection.find()
            current_time = datetime.now(timezone.utc)  # FIXED: Use timezone.utc for awareness
            offers_list = []
            for offer in offers:
                if 'offer_end_time' in offer and offer['offer_end_time']:
                    try:
                        end_time_str = str(offer['offer_end_time'])
                        if end_time_str.endswith('Z'):
                            end_time_str = end_time_str.replace('Z', '+00:00')
                        end_time = datetime.fromisoformat(end_time_str)
                        # FIXED: Ensure end_time is timezone-aware (UTC)
                        if end_time.tzinfo is None:
                            end_time = end_time.replace(tzinfo=timezone.utc)
                        if current_time > end_time:
                            combo_offers_collection.delete_one({'_id': offer['_id']})
                            logger.info(f"Deleted expired combo offer on fetch: {offer['_id']} (ended {end_time})")
                            continue
                    except (ValueError, TypeError) as e:
                        logger.error(f"Invalid offer_end_time for combo offer {offer['_id']}: {str(e)}")
                offer_str = convert_objectid_to_str(offer)
                offers_list.append(offer_str)
            logger.info(f"Fetched {len(offers_list)} active combo offers")
            return jsonify(offers_list), 200
        except Exception as e:
            logger.error(f"Error fetching combo offers: {str(e)}\n{traceback.format_exc()}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/combo-offer/<offer_id>', methods=['GET'])
    @db_required
    def get_combo_offer(offer_id):
        try:
            offer = combo_offers_collection.find_one({'_id': offer_id})
            if not offer:
                logger.warning(f"Combo offer not found: {offer_id}")
                return jsonify({"error": "Combo offer not found"}), 404
            current_time = datetime.now(timezone.utc)  # FIXED: Use timezone.utc
            if 'offer_end_time' in offer and offer['offer_end_time']:
                try:
                    end_time_str = str(offer['offer_end_time'])
                    if end_time_str.endswith('Z'):
                        end_time_str = end_time_str.replace('Z', '+00:00')
                    end_time = datetime.fromisoformat(end_time_str)
                    # FIXED: Ensure awareness
                    if end_time.tzinfo is None:
                        end_time = end_time.replace(tzinfo=timezone.utc)
                    if current_time > end_time:
                        combo_offers_collection.delete_one({'_id': offer['_id']})
                        logger.info(f"Deleted expired combo offer on single fetch: {offer_id} (ended {end_time})")
                        return jsonify({"error": "Combo offer not found (expired)"}), 404
                except (ValueError, TypeError) as e:
                    logger.error(f"Invalid offer_end_time for combo offer {offer_id}: {str(e)}")
            offer = convert_objectid_to_str(offer)
            logger.info(f"Fetched combo offer: {offer_id}")
            return jsonify(offer), 200
        except Exception as e:
            logger.error(f"Error fetching combo offer {offer_id}: {str(e)}\n{traceback.format_exc()}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/combo-offer', methods=['POST'])
    @db_required
    def create_combo_offer():
        try:
            data = request.json
            if not data:
                logger.error("No data provided for combo offer creation")
                return jsonify({"error": "No data provided"}), 400
            required_fields = ['description', 'total_price', 'items']
            for field in required_fields:
                if field not in data:
                    logger.error(f"Missing required field: {field}")
                    return jsonify({"error": f"Missing required field: {field}"}), 400
                value = data[field]
                if field == 'description':
                    if not isinstance(value, str) or not value.strip():
                        logger.error(f"Empty or invalid string field: {field}")
                        return jsonify({"error": f"Field '{field}' must be a non-empty string"}), 400
                elif field == 'total_price':
                    if not isinstance(value, (int, float)) or value < 0:
                        logger.error(f"Invalid total_price: {value}")
                        return jsonify({"error": "Field 'total_price' must be a non-negative number"}), 400
                elif field == 'items':
                    if not isinstance(value, list) or not value:
                        logger.error(f"Empty or invalid items list")
                        return jsonify({"error": "Field 'items' must be a non-empty list"}), 400
            if 'offer_price' in data:
                value = data['offer_price']
                if not isinstance(value, (int, float)) or value < 0:
                    logger.error(f"Invalid offer_price: {value}")
                    return jsonify({"error": "Field 'offer_price' must be a non-negative number"}), 400
            # NEW: Validate images field
            if 'images' in data:
                if not isinstance(data['images'], list):
                    logger.error("Invalid images field: must be a list")
                    return jsonify({"error": "Field 'images' must be a list of strings"}), 400
                for img in data['images']:
                    if not isinstance(img, str):
                        logger.error("Invalid image filename in images list")
                        return jsonify({"error": "Images must be a list of string filenames"}), 400
            if 'offer_start_time' in data and data['offer_start_time'] and 'offer_end_time' in data and data['offer_end_time']:
                try:
                    offer_start_time = datetime.fromisoformat(str(data['offer_start_time']).replace('Z', '+00:00'))
                    offer_end_time = datetime.fromisoformat(str(data['offer_end_time']).replace('Z', '+00:00'))
                    if offer_start_time >= offer_end_time:
                        logger.error("offer_start_time must be before offer_end_time")
                        return jsonify({"error": "Offer start time must be before offer end time"}), 400
                except (ValueError, TypeError) as e:
                    logger.error(f"Invalid offer time format: {str(e)}")
                    return jsonify({"error": f"Invalid offer time format: {str(e)}"}), 400
            data['created_at'] = datetime.now(timezone.utc).isoformat()  # FIXED: Use timezone.utc
            offer_id = combo_offers_collection.insert_one(data).inserted_id
            logger.info(f"Combo offer created with ID: {offer_id}")
            return jsonify({'message': 'Combo offer created successfully!', 'id': offer_id}), 201
        except Exception as e:
            logger.error(f"Error creating combo offer: {str(e)}\n{traceback.format_exc()}")
            return jsonify({'error': str(e)}), 500

    @app.route('/api/combo-offer/<offer_id>', methods=['PUT'])
    @db_required
    def update_combo_offer(offer_id):
        try:
            data = request.json
            if not data:
                logger.error("No data provided for combo offer update")
                return jsonify({"error": "No data provided"}), 400
            if '_id' in data:
                del data['_id']
            for field in data:
                value = data[field]
                if field in ['total_price', 'offer_price']:
                    if not isinstance(value, (int, float)) or value < 0:
                        logger.error(f"Invalid {field}: {value}")
                        return jsonify({"error": f"Field '{field}' must be a non-negative number"}), 400
                if field == 'description':
                    if not isinstance(value, str) or not value.strip():
                        logger.error(f"Empty or invalid string field: {field}")
                        return jsonify({"error": f"Field '{field}' must be a non-empty string"}), 400
                if field == 'items':
                    if not isinstance(value, list) or not value:
                        logger.error(f"Empty or invalid items list")
                        return jsonify({"error": "Field 'items' must be a non-empty list"}), 400
                # NEW: Validate images if present
                if field == 'images':
                    if not isinstance(value, list):
                        logger.error("Invalid images field: must be a list")
                        return jsonify({"error": "Field 'images' must be a list of strings"}), 400
                    for img in value:
                        if not isinstance(img, str):
                            logger.error("Invalid image filename in images list")
                            return jsonify({"error": "Images must be a list of string filenames"}), 400
            if 'offer_start_time' in data and data['offer_start_time'] and 'offer_end_time' in data and data['offer_end_time']:
                try:
                    offer_start_time = datetime.fromisoformat(str(data['offer_start_time']).replace('Z', '+00:00'))
                    offer_end_time = datetime.fromisoformat(str(data['offer_end_time']).replace('Z', '+00:00'))
                    if offer_start_time >= offer_end_time:
                        logger.error("offer_start_time must be before offer_end_time")
                        return jsonify({"error": "Offer start time must be before offer end time"}), 400
                except (ValueError, TypeError) as e:
                    logger.error(f"Invalid offer time format: {str(e)}")
                    return jsonify({"error": f"Invalid offer time format: {str(e)}"}), 400
            data['modified_at'] = datetime.now(timezone.utc).isoformat()  # FIXED: Use timezone.utc
            result = combo_offers_collection.update_one({'_id': offer_id}, {'$set': data})
            if result.matched_count == 0:
                logger.warning(f"Combo offer not found for update: {offer_id}")
                return jsonify({"error": "Combo offer not found"}), 404
            logger.info(f"Combo offer updated: {offer_id}")
            return jsonify({"message": "Combo offer updated successfully"}), 200
        except Exception as e:
            logger.error(f"Error updating combo offer {offer_id}: {str(e)}\n{traceback.format_exc()}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/combo-offer/<offer_id>', methods=['DELETE'])
    @db_required
    def delete_combo_offer(offer_id):
        try:
            result = combo_offers_collection.delete_one({'_id': offer_id})
            if result.deleted_count == 0:
                logger.warning(f"Combo offer not found for deletion: {offer_id}")
                return jsonify({"error": "Combo offer not found"}), 404
            logger.info(f"Combo offer deleted: {offer_id}")
            return jsonify({"message": "Combo offer deleted successfully"}), 200
        except Exception as e:
            logger.error(f"Error deleting combo offer {offer_id}: {str(e)}\n{traceback.format_exc()}")
            return jsonify({"error": str(e)}), 500
    @app.route('/api/save-vat', methods=['POST'])
    @db_required
    def save_vat():
        try:
            data = request.get_json()
            vat = data.get('vat')
            if vat is None:
                return jsonify({"error": "VAT amount required"}), 400
            try:
                vat = float(vat)
            except:
                return jsonify({"error": "Invalid VAT amount"}), 400
            vat_collection.replace_one({"_id": "vat_settings"}, {"_id": "vat_settings", "vat": vat}, upsert=True)
            return jsonify({"message": "VAT saved successfully"}), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    @app.route('/api/get-vat', methods=['GET'])
    @db_required
    def get_vat():
        try:
            settings = vat_collection.find_one({"_id": "vat_settings"})
            vat = settings.get("vat", 10) if settings else 10
            return jsonify({"vat": vat}), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 500
def get_all_print_settings():
    try:
        settings = print_settings_collection.find()
        return jsonify(convert_objectid_to_str(settings)), 200
    except Exception as e:
        logger.error(f"Error fetching print settings: {str(e)}")
        return jsonify({"error": str(e)}), 500
def create_print_settings():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400
        data['active'] = data.get('active', False)
        data['created_at'] = datetime.now(ZoneInfo("UTC")).isoformat()
        result = print_settings_collection.insert_one(data)
        if not print_settings_collection.find_one({"_id": {"$ne": result.inserted_id}, "active": True}):
            print_settings_collection.update_many({"_id": {"$ne": result.inserted_id}}, {"$set": {"active": False}})
        logger.info(f"Print settings created with ID: {result.inserted_id}")
        return jsonify({"message": "Print settings created successfully", "id": result.inserted_id}), 201
    except Exception as e:
        logger.error(f"Error creating print settings: {str(e)}")
        return jsonify({"error": str(e)}), 500
def manage_offers():
    """Check all items and update offer status based on current time."""
    try:
        current_time = datetime.now(ZoneInfo("UTC"))
        items = items_collection.find()
        for item in items:
            item_id = item['_id']
            offer_start_time = item.get('offer_start_time')
            offer_end_time = item.get('offer_end_time')
            should_unset = False
            if offer_start_time and offer_end_time:
                try:
                    start_time = datetime.fromisoformat(str(offer_start_time).replace('Z', '+00:00'))
                    end_time = datetime.fromisoformat(str(offer_end_time).replace('Z', '+00:00'))
                    if current_time > end_time:
                        should_unset = True
                        logger.info(f"Offer expired for item {item.get('item_name')} (ID: {item_id})")
                    elif start_time > end_time:
                        should_unset = True
                        logger.warning(f"Invalid offer times for item {item_id}: start after end")
                    else:
                        logger.debug(f"Offer for item {item.get('item_name')} (ID: {item_id}) is active or pending")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid offer time format for item {item_id}: {str(e)}")
                    should_unset = True
            elif offer_end_time:
                try:
                    end_time = datetime.fromisoformat(str(offer_end_time).replace('Z', '+00:00'))
                    if current_time > end_time:
                        should_unset = True
                        logger.info(f"Offer expired for item {item.get('item_name')} (ID: {item_id})")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid offer_end_time for item {item_id}: {str(e)}")
                    should_unset = True
            if should_unset:
                items_collection.update_one(
                    {'_id': item_id},
                    {'$unset': {'offer_price': "", 'offer_start_time': "", 'offer_end_time': ""}}
                )
                logger.info(f"Unset offer fields for item {item.get('item_name')} (ID: {item_id})")
    except Exception as e:
        logger.error(f"Error in manage_offers: {str(e)}")
def manage_combo_offers():
    """Check all combo offers and delete them when end time is reached."""
    try:
        current_time = datetime.now(ZoneInfo("UTC"))
        offers = combo_offers_collection.find()
        for offer in offers:
            offer_id = offer['_id']
            offer_end_time = offer.get('offer_end_time')
            if offer_end_time:
                try:
                    end_time = datetime.fromisoformat(str(offer_end_time).replace('Z', '+00:00'))
                    if current_time > end_time:
                        combo_offers_collection.delete_one({'_id': offer_id})
                        logger.info(f"Deleted expired combo offer: {offer_id}")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid offer_end_time for combo offer {offer_id}: {str(e)}")
    except Exception as e:
        logger.error(f"Error in manage_combo_offers: {str(e)}")
def schedule_tasks():
    if schedule:
        schedule.every(1).minutes.do(manage_offers)
        schedule.every(1).minutes.do(manage_combo_offers)
        while True:
            schedule.run_pending()
            time.sleep(1)
def start_scheduler():
    scheduler_thread = threading.Thread(target=schedule_tasks, daemon=True)
    scheduler_thread.start()
    logger.info("Automatic backup, offer, and combo offer scheduler started")
    # Schedule backup with initial interval
    settings = get_system_settings()
    interval = settings.get('backup_interval_hours', 6)
    if schedule:
        schedule.every(interval).hours.do(create_backup).tag('backup')
        logger.info(f"Scheduled automatic backups every {interval} hours")

@app.route('/api/company-details', methods=['POST', 'GET', 'OPTIONS'])
@db_required
def manage_company_details():
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    if request.method == 'POST':
        try:
            data = request.get_json()
            if not data:
                logger.error("No data provided in POST request")
                return jsonify({"error": "No data provided"}), 400
            # UPDATED: Handle new contact structure - ensure each contact has phoneCountryCode, whatsappCountryCode, and websites array
            contacts = data.get('contacts', [])
            for contact in contacts:
                if 'phoneCountryCode' not in contact:
                    contact['phoneCountryCode'] = '+91' # Default
                if 'whatsappCountryCode' not in contact:
                    contact['whatsappCountryCode'] = '+91' # Default
                if 'websites' not in contact:
                    contact['websites'] = [] # Ensure array for multiple websites
            # NEW: Ensure specialTimings is an array
            special_timings = data.get('specialTimings', [])
            if not isinstance(special_timings, list):
                special_timings = []
            
            # Sanitize times to 24-hour format
            opening_time = convert_to_24h(data.get('openingTime', ''))
            closing_time = convert_to_24h(data.get('closingTime', ''))
            
            for special in special_timings:
                if 'startTime' in special:
                    special['startTime'] = convert_to_24h(special['startTime'])
                if 'endTime' in special:
                    special['endTime'] = convert_to_24h(special['endTime'])

            company_data = {
                '_id': str(uuid.uuid4()),
                'restaurantName': data.get('restaurantName', ''),
                'ownerName': data.get('ownerName', ''),
                # NEW: Company Licence field
                'companyLicence': data.get('companyLicence', ''),
                'businessType': data.get('businessType', ''),
                'otherBusinessType': data.get('otherBusinessType', ''),
                'taxType': data.get('taxType', ''),
                'taxPercentage': data.get('taxPercentage', ''),
                'taxNumber': data.get('taxNumber', ''),
                'fssaiNumber': data.get('fssaiNumber', ''),
                'panNumber': data.get('panNumber', ''),
                'openingTime': opening_time,
                'closingTime': closing_time,
                'totalTime': data.get('totalTime', ''),
                # NEW: Special Timings
                'specialTimings': special_timings,
                # UPDATED: Addresses now use dynamic fields: country, field1, field2, field3, flat_villa_no, building_name
                'addresses': data.get('addresses', [{'country': '', 'field1': '', 'field2': '', 'field3': '', 'flat_villa_no': '', 'building_name': ''}]),
                'contacts': contacts, # Use processed contacts
                'bankName': data.get('bankName', ''),
                'accountHolderName': data.get('accountHolderName', ''),
                'accountNumber': data.get('accountNumber', ''),
                'ifscCode': data.get('ifscCode', ''),
                'upiId': data.get('upiId', ''),
                'currencyType': data.get('currencyType', ''), # From settings or manual
                'created_at': datetime.now(ZoneInfo("Asia/Kolkata")).isoformat()
            }
            logger.info(f"Saving company details: {company_data}")
            result = company_details_collection.insert_one(company_data)
            logger.info(f"Company details saved with ID: {result.inserted_id}")
            # Return the full saved data
            return jsonify({
                "message": "Company details saved successfully",
                "id": result.inserted_id,
                "companyDetails": company_data
            }), 201
        except Exception as e:
            logger.error(f"Error saving company details: {str(e)}")
            return jsonify({"error": f"Failed to save company details: {str(e)}"}), 500
    if request.method == 'GET':
        try:
            details = list(company_details_collection.find())
            # Convert MongoDB documents to JSON-serializable format
            serialized_details = []
            for detail in details:
                detail['_id'] = str(detail['_id'])
                # UPDATED: Ensure backward compatibility - add defaults if missing fields
                if 'contacts' in detail:
                    for contact in detail['contacts']:
                        if 'phoneCountryCode' not in contact:
                            contact['phoneCountryCode'] = '+91'
                        if 'whatsappCountryCode' not in contact:
                            contact['whatsappCountryCode'] = '+91'
                        if 'websites' not in contact:
                            contact['websites'] = []
                # NEW: Ensure specialTimings is array if missing
                if 'specialTimings' not in detail:
                    detail['specialTimings'] = []
                # NEW: Ensure companyLicence is string if missing
                if 'companyLicence' not in detail:
                    detail['companyLicence'] = ''
                serialized_details.append(detail)
            logger.info(f"Retrieved company details: {serialized_details}")
            return jsonify({"companyDetails": serialized_details}), 200
        except Exception as e:
            logger.error(f"Error retrieving company details: {str(e)}")
            return jsonify({"error": f"Failed to retrieve company details: {str(e)}"}), 500
def generate_order_number(order_type):
    if order_counters_collection is None:
        raise Exception("Database not initialized correctly. Order counters collection is missing.")
    counter_doc = order_counters_collection.find_one_and_update(
        {'_id': order_type},
        {'$inc': {'count': 1}},
        upsert=True,
        return_document=True
    )
    return f"{order_type}-{counter_doc['count']:04d}"
@app.route('/api/activeorders', methods=['POST'])
@db_required
def save_active_order():
    try:
        data = request.get_json()
        order_type = data.get('orderType', 'Dine In')
        order_no = generate_order_number(order_type)
        cart_items = data.get('cartItems', [])
        for item in cart_items:
            required_kitchens = set()
            if item.get('kitchen'):
                required_kitchens.add(item['kitchen'])
            for addon_name, qty in item.get('addonQuantities', {}).items():
                if qty > 0 and 'addonVariants' in item and addon_name in item['addonVariants']:
                    if item['addonVariants'][addon_name].get('kitchen'):
                        required_kitchens.add(item['addonVariants'][addon_name]['kitchen'])
            for combo_name, qty in item.get('comboQuantities', {}).items():
                if qty > 0 and 'comboVariants' in item and combo_name in item['comboVariants']:
                    if item['comboVariants'][combo_name].get('kitchen'):
                        required_kitchens.add(item['comboVariants'][combo_name]['kitchen'])
            item['requiredKitchens'] = list(required_kitchens)
            item['kitchenStatuses'] = {kitchen: 'Pending' for kitchen in required_kitchens}
            item['served'] = False
        active_order = {
            'orderId': str(uuid.uuid4()),
            'orderNo': order_no,
            'customerName': data.get('customerName', 'N/A'),
            'tableNumber': data.get('tableNumber', 'N/A'),
            'chairsBooked': data.get('chairsBooked', []),
            'phoneNumber': data.get('phoneNumber', ''),
            'deliveryAddress': data.get('deliveryAddress', {}),
            'whatsappNumber': data.get('whatsappNumber', ''),
            'email': data.get('email', ''),
            'cartItems': cart_items,
            'timestamp': data.get('timestamp', datetime.now(timezone.utc).isoformat()),
            'orderType': order_type,
            'status': 'Pending', # New status field
            'paid': False,
            'created_at': datetime.now(timezone.utc),
            'deliveryPersonId': data.get('deliveryPersonId', ''),
            'deliveryPersonName': data.get('deliveryPersonName', ''),
            'pickedUpTime': data.get('pickedUpTime', None),
        }
        activeorders_collection.insert_one(active_order)
        kitchen_saved_collection.insert_one(active_order.copy())
        logger.info(f"Created order: {active_order['orderId']} with order number: {order_no}")
        return jsonify({'success': True, 'orderId': active_order['orderId'], 'orderNo': order_no}), 201
    except Exception as e:
        logger.error(f"Error saving active order: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
@app.route('/api/activeorders', methods=['GET'])
@db_required
def get_active_orders():
    try:
        orders = activeorders_collection.find()
        return jsonify(convert_objectid_to_str(list(orders))), 200
    except Exception as e:
        logger.error(f"Error fetching active orders: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
@app.route('/api/activeorders/<order_id>/items/<item_id>/mark-prepared', methods=['POST'])
@db_required
def mark_item_prepared_active(order_id, item_id):
    try:
        data = request.get_json()
        kitchen = data.get('kitchen')
        if not kitchen:
            return jsonify({'success': False, 'error': 'Kitchen not provided'}), 400
        for collection in [activeorders_collection, kitchen_saved_collection]:
            order = collection.find_one({'orderId': order_id})
            if not order:
                return jsonify({'success': False, 'error': 'Order not found'}), 404
            found = False
            for item in order['cartItems']:
                if item['id'] == item_id:
                    if 'kitchenStatuses' not in item:
                        item['kitchenStatuses'] = {}
                    if item['kitchenStatuses'].get(kitchen) in ['Prepared', 'PickedUp']:
                        return jsonify({'success': False, 'error': 'Kitchen already marked as prepared or picked up'}), 400
                    item['kitchenStatuses'][kitchen] = 'Prepared'
                    found = True
                    break
            if not found:
                return jsonify({'success': False, 'error': 'Item not found'}), 404
            collection.replace_one({'orderId': order_id}, order)
        logger.info(f"Marked item {item_id} in order {order_id} as Prepared for kitchen {kitchen}")
        return jsonify({'success': True, 'status': 'Prepared'}), 200
    except Exception as e:
        logger.error(f"Error in mark-prepared: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/api/activeorders/<order_id>/items/<item_id>/mark-pickedup', methods=['POST'])
@db_required
def mark_item_pickedup_active(order_id, item_id):
    try:
        data = request.get_json()
        kitchen = data.get('kitchen')
        if not kitchen:
            return jsonify({'success': False, 'error': 'Kitchen not provided'}), 400
        for collection in [activeorders_collection, kitchen_saved_collection]:
            order = collection.find_one({'orderId': order_id})
            if not order:
                return jsonify({'success': False, 'error': 'Order not found'}), 404
            found = False
            for item in order['cartItems']:
                if item['id'] == item_id:
                    if 'kitchenStatuses' not in item:
                        item['kitchenStatuses'] = {}
                    status = item['kitchenStatuses'].get(kitchen)
                    if status == 'Pending':
                        logger.warning(f"Item {item_id} in order {order_id} was Pending, setting to Prepared automatically for kitchen {kitchen}")
                        item['kitchenStatuses'][kitchen] = 'Prepared'
                    elif status != 'Prepared':
                        return jsonify({'success': False, 'error': 'Item must be prepared before picking up'}), 400
                    item['kitchenStatuses'][kitchen] = 'PickedUp'
                    found = True
                    break
            if not found:
                return jsonify({'success': False, 'error': 'Item not found'}), 404
            collection.replace_one({'orderId': order_id}, order)
        order = activeorders_collection.find_one({'orderId': order_id})
        picked_up_data = {
            'customerName': order.get('customerName', 'Unknown'),
            'tableNumber': order.get('tableNumber', 'N/A'),
            'items': order.get('cartItems', []),
            'pickupTime': datetime.now(timezone.utc).isoformat(),
            'orderType': order.get('orderType', 'Dine In')
        }
        picked_up_collection.insert_one(picked_up_data)
        logger.info(f"Marked item {item_id} in order {order_id} as PickedUp for kitchen {kitchen}")
        return jsonify({'success': True, 'status': 'PickedUp'}), 200
    except Exception as e:
        logger.error(f"Error in mark-pickedup: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/api/activeorders/<order_id>/items/<item_id>/mark-served', methods=['POST'])
@db_required
def mark_item_served(order_id, item_id):
    try:
        data = request.get_json()
        served = data.get('served', True)
        if served is None:
            return jsonify({'success': False, 'error': 'Served status not provided'}), 400
        for collection in [activeorders_collection, kitchen_saved_collection]:
            order = collection.find_one({'orderId': order_id})
            if not order:
                return jsonify({'success': False, 'error': 'Order not found'}), 404
            found = False
            for item in order['cartItems']:
                if item['id'] == item_id:
                    if not all(s == 'PickedUp' for s in item['kitchenStatuses'].values()):
                        return jsonify({'success': False, 'error': 'Item must be picked up before serving'}), 400
                    item['served'] = bool(served)
                    found = True
                    break
            if not found:
                return jsonify({'success': False, 'error': 'Item not found'}), 404
            collection.replace_one({'orderId': order_id}, order)
        logger.info(f"Marked item {item_id} in order {order_id} as {'Served' if served else 'Unserved'}")
        return jsonify({'success': True, 'served': served}), 200
    except Exception as e:
        logger.error(f"Error in mark-served: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/api/activeorders/<order_id>/items/<item_id>', methods=['DELETE'])
@db_required
def delete_order_item(order_id, item_id):
    try:
        for collection in [activeorders_collection, kitchen_saved_collection]:
            order = collection.find_one({'orderId': order_id})
            if not order:
                return jsonify({'success': False, 'error': 'Order not found'}), 404
            order['cartItems'] = [i for i in order['cartItems'] if i['id'] != item_id]
            if not order['cartItems']:
                collection.delete_one({'orderId': order_id})
            else:
                collection.replace_one({'orderId': order_id}, order)
        logger.info(f"Deleted item {item_id} from order {order_id}")
        return jsonify({'success': True}), 200
    except Exception as e:
        logger.error(f"Error deleting item: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
@app.route('/api/activeorders/<order_id>', methods=['PUT'])
@db_required
def update_active_order(order_id):
    try:
        data = request.get_json()
        if '_id' in data:
            del data['_id']
        order_in_db = activeorders_collection.find_one({'orderId': order_id})
        if not order_in_db:
            logger.warning(f"Order not found for update: {order_id}")
            return jsonify({'error': 'Order not found'}), 404
        old_statuses_map = {
            item['id']: {
                'kitchenStatuses': item.get('kitchenStatuses', {}),
                'served': item.get('served', False)
            }
            for item in order_in_db.get('cartItems', []) if 'id' in item
        }
        if 'cartItems' in data:
            for item in data['cartItems']:
                item_id = item.get('id')
                if item_id and item_id in old_statuses_map:
                    old_data = old_statuses_map[item_id]
                    item['kitchenStatuses'] = old_data['kitchenStatuses']
                    item['served'] = old_data['served']
                else:
                    required_kitchens = set()
                    if item.get('kitchen'):
                        required_kitchens.add(item['kitchen'])
                    for addon_name, qty in item.get('addonQuantities', {}).items():
                        if qty > 0 and 'addonVariants' in data and addon_name in data['addonVariants']:
                            if data['addonVariants'][addon_name].get('kitchen'):
                                required_kitchens.add(data['addonVariants'][addon_name]['kitchen'])
                    for combo_name, qty in item.get('comboQuantities', {}).items():
                        if qty > 0 and 'comboVariants' in data and combo_name in data['comboVariants']:
                            if data['comboVariants'][combo_name].get('kitchen'):
                                required_kitchens.add(data['comboVariants'][combo_name]['kitchen'])
                    item['requiredKitchens'] = list(required_kitchens)
                    item['kitchenStatuses'] = {kitchen: 'Pending' for kitchen in required_kitchens}
                    item['served'] = False
        updated_order = {**order_in_db, **data}
        if 'cartItems' in data:
            updated_order['cartItems'] = data['cartItems']
        if 'deliveryPersonId' in updated_order and updated_order['deliveryPersonId']:
            employee = employees_collection.find_one({'employeeId': updated_order['deliveryPersonId']})
            if not employee:
                logger.warning(f"Delivery person not found: {updated_order['deliveryPersonId']}")
                return jsonify({'error': 'Delivery person not found'}), 404
            updated_order['status'] = 'assigned' # Set to assigned instead of deleting
            updated_order['deliveryPersonName'] = employee.get('name', 'N/A')
            # Optional: Move to tripreports if needed, but per requirement, keep in activeorders with status
            logger.info(f"Assigned delivery person {updated_order['deliveryPersonId']} ({updated_order['deliveryPersonName']}) to order {order_id}")
        if updated_order.get('paid', False) and all(item.get('served', False) for item in updated_order.get('cartItems', [])) and updated_order.get('orderType') != 'Online Delivery':
            updated_order['status'] = 'Completed'
        result = activeorders_collection.replace_one({'orderId': order_id}, updated_order)
        kitchen_result = kitchen_saved_collection.replace_one({'orderId': order_id}, updated_order)
        updated_order = activeorders_collection.find_one({'orderId': order_id})
        if result.matched_count > 0 or kitchen_result.matched_count > 0:
            logger.info(f"Updated order: {order_id}")
            return jsonify({'success': True, 'message': 'Order updated', 'order': convert_objectid_to_str(updated_order)}), 200
        logger.info(f"No changes made to order: {order_id}")
        return jsonify({'success': True, 'message': 'No changes made', 'order': convert_objectid_to_str(updated_order)}), 200
    except Exception as e:
        logger.error(f"Error updating active order: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
@app.route('/api/activeorders/<order_id>', methods=['DELETE'])
@db_required
def delete_order(order_id):
    try:
        result = activeorders_collection.delete_one({'orderId': order_id})
        kitchen_result = kitchen_saved_collection.delete_one({'orderId': order_id})
        if result.deleted_count > 0 or kitchen_result.deleted_count > 0:
            logger.info(f"Deleted order: {order_id}")
            return jsonify({'success': True}), 200
        logger.warning(f"Order not found for deletion: {order_id}")
        return jsonify({'error': 'Order not found'}), 404
    except Exception as e:
        logger.error(f"Error deleting order: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
@app.route('/api/activeorders/<order_id>/mark-delivered', methods=['PUT'])
@db_required
def mark_order_delivered(order_id):
    try:
        # Find order in activeorders with status 'assigned'
        order = activeorders_collection.find_one({'orderId': order_id, 'status': 'assigned'})
        if not order:
            return jsonify({'success': False, 'error': 'Order not found or not assigned'}), 404
        # Insert a copy to trip_reports
        trip_report = order.copy()
        trip_report['delivered_at'] = datetime.now(timezone.utc).isoformat()
        trip_report['_id'] = str(uuid.uuid4()) # New ID for trip report
        tripreports_collection.insert_one(trip_report)
        # Delete from activeorders and kitchen_saved
        activeorders_collection.delete_one({'orderId': order_id})
        kitchen_saved_collection.delete_one({'orderId': order_id})
        logger.info(f"Order {order_id} marked as delivered and moved to trip_reports")
        return jsonify({'success': True, 'message': 'Order marked as delivered'}), 200
    except Exception as e:
        logger.error(f"Error marking order delivered: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/api/kitchen-saved', methods=['POST'])
@db_required
def save_kitchen_order():
    try:
        data = request.get_json()
        if not data or 'orderId' not in data:
            return jsonify({'success': False, 'error': 'No data or orderId provided'}), 400
        order_id = data['orderId']
        existing_order = kitchen_saved_collection.find_one({'orderId': order_id})
        cart_items = data.get('cartItems', [])
        for item in cart_items:
            required_kitchens = set()
            if 'kitchen' in item:
                required_kitchens.add(item['kitchen'])
            for addon_name, qty in item.get('addonQuantities', {}).items():
                if qty > 0 and 'addonVariants' in item and addon_name in item['addonVariants']:
                    addon = item['addonVariants'][addon_name]
                    if 'kitchen' in addon:
                        required_kitchens.add(addon['kitchen'])
            for combo_name, qty in item.get('comboQuantities', {}).items():
                if qty > 0 and 'comboVariants' in item and combo_name in item['comboVariants']:
                    combo = item['comboVariants'][combo_name]
                    if 'kitchen' in combo:
                        required_kitchens.add(combo['kitchen'])
            item['requiredKitchens'] = list(required_kitchens)
            item['kitchenStatuses'] = item.get('kitchenStatuses', {kitchen: 'Pending' for kitchen in required_kitchens})
        order = {
            'orderId': order_id,
            'customerName': data.get('customerName', 'N/A'),
            'tableNumber': data.get('tableNumber', 'N/A'),
            'chairsBooked': data.get('chairsBooked', []),
            'pickupTime': data.get('pickupTime', ''),
            'deliveryAddress': data.get('deliveryAddress', {}),
            'whatsappNumber': data.get('whatsappNumber', ''),
            'email': data.get('email', ''),
            'cartItems': cart_items,
            'timestamp': data.get('timestamp', datetime.now(timezone.utc).isoformat()),
            'orderType': data.get('orderType', 'Dine In'),
            'status': data.get('status', 'Pending'),
            'createdAt': datetime.now(timezone.utc).isoformat(),
            'pickedUpTime': data.get('pickedUpTime', None)
        }
        if existing_order:
            kitchen_saved_collection.update_one(
                {'orderId': order_id},
                {'$set': order}
            )
            logger.info(f"Updated kitchen order: {order_id}")
        else:
            kitchen_saved_collection.insert_one(order)
            logger.info(f"Created kitchen order: {order_id}")
        return jsonify({'success': True, 'order_id': order_id}), 201
    except Exception as e:
        logger.error(f"Error in /api/kitchen-saved POST: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/kitchen-saved', methods=['GET'])
@db_required
def get_kitchen_orders():
    try:
        orders = kitchen_saved_collection.find()
        return jsonify({'success': True, 'orders': convert_objectid_to_str(list(orders))}), 200
    except Exception as e:
        logger.error(f"Error in /api/kitchen-saved GET: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/kitchen-saved/<order_id>', methods=['DELETE'])
@db_required
def delete_kitchen_order(order_id):
    try:
        result = kitchen_saved_collection.delete_one({'orderId': order_id})
        if result.deleted_count == 0:
            logger.warning(f"Order not found: {order_id}")
            return jsonify({'success': False, 'error': 'Order not found'}), 404
        logger.info(f"Order deleted: {order_id}")
        return jsonify({'success': True, 'message': 'Order deleted successfully'}), 200
    except Exception as e:
        logger.error(f"Error deleting order {order_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/kitchen-saved/<order_id>/items/<item_id>/mark-prepared', methods=['POST'])
@db_required
def mark_item_prepared(order_id, item_id):
    try:
        data = request.get_json()
        kitchen = data.get('kitchen')
        if not kitchen:
            return jsonify({'success': False, 'error': 'Kitchen not provided'}), 400
        order = kitchen_saved_collection.find_one({'orderId': order_id})
        if not order:
            return jsonify({'success': False, 'error': 'Order not found'}), 404
        item = next((item for item in order['cartItems'] if item['id'] == item_id), None)
        if not item:
            return jsonify({'success': False, 'error': 'Item not found'}), 404
        if not item.get('requiredKitchens') or kitchen not in item['requiredKitchens']:
            return jsonify({'success': False, 'error': 'Kitchen not required for this item'}), 400
        if not item.get('kitchenStatuses'):
            item['kitchenStatuses'] = {k: 'Pending' for k in item['requiredKitchens']}
        if item['kitchenStatuses'][kitchen] in ['Prepared', 'PickedUp']:
            return jsonify({'success': False, 'error': 'Kitchen already marked as prepared or picked up'}), 400
        item['kitchenStatuses'][kitchen] = 'Prepared'
        kitchen_saved_collection.update_one(
            {'orderId': order_id, 'cartItems.id': item_id},
            {'$set': {'cartItems.$.kitchenStatuses': item['kitchenStatuses']}}
        )
        activeorders_collection.update_one(
            {'orderId': order_id, 'cartItems.id': item_id},
            {'$set': {'cartItems.$.kitchenStatuses': item['kitchenStatuses']}}
        )
        return jsonify({'success': True, 'status': 'Prepared'}), 200
    except Exception as e:
        logger.error(f"Error in /api/kitchen-saved/{order_id}/items/{item_id}/mark-prepared: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/api/employee-designations', methods=['GET', 'POST'])
@db_required  # Assuming this decorator checks DB connection
def employee_designations():
    if request.method == 'GET':
        try:
            if employee_designations_collection is None:
                logger.error("employee_designations_collection not initialized")
                return jsonify({"error": "Database not ready"}), 503
            designations = employee_designations_collection.find()
            return jsonify([{'id': str(t['_id']), 'name': t['name'], 'description': t.get('description', ''), 'reportTo': t.get('reportTo', '')} for t in designations]), 200
        except Exception as e:
            logger.error(f"Error fetching employee designations: {str(e)}")
            return jsonify({"error": str(e)}), 500
    elif request.method == 'POST':
        try:
            data = request.get_json()
            if not isinstance(data, dict):
                return jsonify({"error": "JSON data must be an object"}), 400
            name = data.get('name')
            if not name or not name.strip():
                return jsonify({"error": "Name is required"}), 400
            
            description = data.get('description', '').strip()
            report_to = data.get('reportTo', '').strip()

            # Check if name already exists in designations collection only
            existing = employee_designations_collection.find_one({"name": name.strip()})
            if existing:
                return jsonify({"error": "Designation name already exists"}), 400
            new_designation = {
                "_id": str(uuid.uuid4()),
                "name": name.strip(),
                "description": description,
                "reportTo": report_to,
                "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }
            if employee_designations_collection is None:
                logger.error("employee_designations_collection not initialized")
                return jsonify({"error": "Database not ready"}), 503
            employee_designations_collection.insert_one(new_designation)
            logger.info(f"Employee designation created: {name}")
            return jsonify(new_designation), 201
        except Exception as e:
            logger.error(f"Error creating employee designation: {str(e)}")
            return jsonify({"error": str(e)}), 500

@app.route('/api/employee-designations/<designation_id>', methods=['PUT', 'DELETE'])
@db_required
def manage_employee_designation(designation_id):
    if request.method == 'PUT':
        try:
            data = request.get_json()
            if not isinstance(data, dict):
                return jsonify({"error": "JSON data must be an object"}), 400
            name = data.get('name')
            if not name or not name.strip():
                return jsonify({"error": "Name is required"}), 400
            
            description = data.get('description', '').strip()
            report_to = data.get('reportTo', '').strip()

            # Check if new name already exists in designations only
            existing = employee_designations_collection.find_one({"name": name.strip()})
            if existing and str(existing['_id']) != designation_id:
                return jsonify({"error": "Employee designation name already exists"}), 400
            result = employee_designations_collection.update_one(
                {'_id': designation_id},
                {'$set': {'name': name.strip(), 'description': description, 'reportTo': report_to}}
            )
            if result.modified_count == 0:
                return jsonify({"error": "Employee designation not found"}), 404
            logger.info(f"Employee designation updated: {designation_id} to {name}")
            return jsonify({"message": "Employee designation updated successfully"}), 200
        except Exception as e:
            logger.error(f"Error updating employee designation: {str(e)}")
            return jsonify({"error": str(e)}), 500
    elif request.method == 'DELETE':
        try:
            # Check if designation is used in employees
            designation_doc = employee_designations_collection.find_one({'_id': designation_id})
            if not designation_doc:
                return jsonify({"error": "Employee designation not found"}), 404
            used_in_employees = worker_collection.find({'employeeDesignation': designation_doc['name']})
            if list(used_in_employees):
                return jsonify({"error": "Cannot delete: This designation is used by existing employees"}), 400
            result = employee_designations_collection.delete_one({'_id': designation_id})
            if result.deleted_count == 0:
                return jsonify({"error": "Employee designation not found"}), 404
            logger.info(f"Employee designation deleted: {designation_doc['name']}")
            return jsonify({"message": "Employee designation deleted successfully"}), 200
        except Exception as e:
            logger.error(f"Error deleting employee designation: {str(e)}")
            return jsonify({"error": str(e)}), 500

# Separate Employee Types Endpoints
@app.route('/api/employee-types', methods=['GET', 'POST'])
@db_required
def employee_types():
    if request.method == 'GET':
        try:
            if employee_type_collection is None:
                logger.error("employee_type_collection not initialized")
                return jsonify({"error": "Database not ready"}), 503
            types = employee_type_collection.find()
            return jsonify([{'id': str(t['_id']), 'name': t['name'], 'description': t.get('description', ''), 'salaryRange': t.get('salaryRange', ''), 'designation': t.get('designation', ''), 'reportTo': t.get('reportTo', ''), 'grade': t.get('grade', ''), 'branch': t.get('branch', '')} for t in types]), 200
        except Exception as e:
            logger.error(f"Error fetching employee types: {str(e)}")
            return jsonify({"error": str(e)}), 500
    elif request.method == 'POST':
        try:
            data = request.get_json()
            if not isinstance(data, dict):
                return jsonify({"error": "JSON data must be an object"}), 400
            name = data.get('name')
            if not name or not name.strip():
                return jsonify({"error": "Name is required"}), 400
            
            description = data.get('description', '').strip()
            salary_range = data.get('salaryRange', '').strip()
            designation = data.get('designation', '').strip()
            report_to = data.get('reportTo', '').strip()
            grade = data.get('grade', '').strip()
            branch = data.get('branch', '').strip()

            # Check if name already exists in types collection only
            existing = employee_type_collection.find_one({"name": name.strip()})
            if existing:
                return jsonify({"error": "Employee type name already exists"}), 400
            new_type = {
                "_id": str(uuid.uuid4()),
                "name": name.strip(),
                "description": description,
                "salaryRange": salary_range,
                "designation": designation,
                "reportTo": report_to,
                "grade": grade,
                "branch": branch,
                "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }
            if employee_type_collection is None:
                logger.error("employee_type_collection not initialized")
                return jsonify({"error": "Database not ready"}), 503
            employee_type_collection.insert_one(new_type)
            logger.info(f"Employee type created: {name}")
            return jsonify(new_type), 201
        except Exception as e:
            logger.error(f"Error creating employee type: {str(e)}")
            return jsonify({"error": str(e)}), 500

@app.route('/api/employee-types/<type_id>', methods=['PUT', 'DELETE'])
@db_required
def manage_employee_type(type_id):
    if request.method == 'PUT':
        try:
            data = request.get_json()
            if not isinstance(data, dict):
                return jsonify({"error": "JSON data must be an object"}), 400
            name = data.get('name')
            if not name or not name.strip():
                return jsonify({"error": "Name is required"}), 400
            
            description = data.get('description', '').strip()
            salary_range = data.get('salaryRange', '').strip()
            designation = data.get('designation', '').strip()
            report_to = data.get('reportTo', '').strip()
            grade = data.get('grade', '').strip()
            branch = data.get('branch', '').strip()

            # Check if new name already exists in types only
            existing = employee_type_collection.find_one({"name": name.strip()})
            if existing and str(existing['_id']) != type_id:
                return jsonify({"error": "Employee type name already exists"}), 400
            result = employee_type_collection.update_one(
                {'_id': type_id},
                {'$set': {'name': name.strip(), 'description': description, 'salaryRange': salary_range, 'designation': designation, 'reportTo': report_to, 'grade': grade, 'branch': branch}}
            )
            if result.modified_count == 0:
                return jsonify({"error": "Employee type not found"}), 404
            logger.info(f"Employee type updated: {type_id} to {name}")
            return jsonify({"message": "Employee type updated successfully"}), 200
        except Exception as e:
            logger.error(f"Error updating employee type: {str(e)}")
            return jsonify({"error": str(e)}), 500
    elif request.method == 'DELETE':
        try:
            # Check if type is used in employees
            type_doc = employee_type_collection.find_one({'_id': type_id})
            if not type_doc:
                return jsonify({"error": "Employee type not found"}), 404
            used_in_employees = worker_collection.find({'employeeType': type_doc['name']})
            if list(used_in_employees):
                return jsonify({"error": "Cannot delete: This type is used by existing employees"}), 400
            result = employee_type_collection.delete_one({'_id': type_id})
            if result.deleted_count == 0:
                return jsonify({"error": "Employee type not found"}), 404
            logger.info(f"Employee type deleted: {type_doc['name']}")
            return jsonify({"message": "Employee type deleted successfully"}), 200
        except Exception as e:
            logger.error(f"Error deleting employee type: {str(e)}")
            return jsonify({"error": str(e)}), 500


@app.route('/api/add-employee', methods=['GET', 'POST'])
@db_required
def add_employee():
    mode = config.get("mode", "server")
    if mode == 'client':
        if request.method == 'GET':
            try:
                server_url = f"http://{config['server_ip']}:8000/api/add-employee"
                response = requests.get(server_url, timeout=5)
                if response.status_code == 200:
                    return jsonify(response.json()), 200
                else:
                    return jsonify({"error": "Proxy fetch failed"}), response.status_code
            except Exception as e:
                logger.error(f"Proxy GET error: {e}")
                return jsonify({"error": f"Proxy error: {str(e)}"}), 500
        elif request.method == 'POST':
            try:
                server_url = f"http://{config['server_ip']}:8000/api/add-employee"
                response = requests.post(server_url, json=request.get_json(), timeout=5)
                if response.status_code in [200, 201]:
                    return jsonify(response.json()), response.status_code
                else:
                    try:
                        err_data = response.json()
                        return jsonify(err_data), response.status_code
                    except:
                        return jsonify({"error": "Proxy create failed"}), response.status_code
            except Exception as e:
                logger.error(f"Proxy POST error: {e}")
                return jsonify({"error": f"Proxy error: {str(e)}"}), 500
    if request.method == 'GET':
        try:
            # FIXED: SQLiteCollection doesn't support $ne efficiently or at all for some cases.
            # Fetch all and filter in python.
            all_employees = worker_collection.find()
            employees = [e for e in all_employees if e.get('isDraft') is not True]
            converted_employees = []
            for emp in employees:
                if '_id' in emp:
                    emp['_id'] = str(emp['_id'])
                converted_employees.append(emp)
            return jsonify(converted_employees), 200
        except Exception as e:
            logger.error(f"Error fetching employees: {str(e)}")
            logger.error(traceback.format_exc())
            return jsonify({"error": str(e)}), 500
    elif request.method == 'POST':
        try:
            data = request.get_json()
            if not isinstance(data, dict):
                return jsonify({"error": "JSON data must be an object"}), 400
            # Detailed validation
            required_fields = ['name', 'phoneNumber', 'email', 'username', 'password', 'employeeDesignation', 'employeeType']
            # Allow empty password if it's somehow not critical (but it usually is for login)
            missing_fields = []
            for field in required_fields:
                if field not in data or (isinstance(data[field], str) and not data[field].strip()):
                    missing_fields.append(field)
       
            if missing_fields:
                return jsonify({"error": f"Missing required fields: {', '.join(missing_fields)}"}), 400
            # Check if email exists
            if worker_collection.find_one({"email": data['email'], "isDraft": {"$ne": True}}):
                 return jsonify({"error": "Employee with this email already exists"}), 400
            if users_collection.find_one({"email": data['email']}):
                 return jsonify({"error": "User with this email already exists"}), 400
            # Generate ID
            existing_employees = list(worker_collection.find({"isDraft": {"$ne": True}}))
            next_id = len(existing_employees) + 1
            employee_id = f"EMP{next_id:03d}"
            # Validate dates
            date_of_birth = data.get('dateOfBirth', '')
            if date_of_birth:
                try:
                    datetime.fromisoformat(date_of_birth.split('T')[0])
                except ValueError:
                    logger.warning(f"Invalid DOB: {date_of_birth}")
            id_expiry = data.get('idExpiry', '')
            # Safe float conversion
            def safe_float(val, default=0.0):
                try:
                    return float(val) if val else default
                except (ValueError, TypeError):
                    return default
            basic = safe_float(data.get('basicSalary'))
            hra = safe_float(data.get('hra'))
            ta = safe_float(data.get('ta'))
            oa = safe_float(data.get('oa'))
            total_salary = data.get('totalSalary')
            if not total_salary:
                total_salary = basic + hra + ta + oa
            else:
                 total_salary = safe_float(total_salary)
            # Password Hashing
            try:
                hashed_password = bcrypt.hashpw(str(data['password']).encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            except Exception as e:
                logger.error(f"Password hashing failed: {e}")
                return jsonify({"error": "Invalid password format"}), 400
            new_employee = {
                "_id": str(uuid.uuid4()),
                "employeeId": employee_id,
                "name": data['name'],
                "phoneNumber": data['phoneNumber'],
                "email": data['email'],
                "gender": data.get('gender', ''),
                "dateOfBirth": date_of_birth,
                "dateOfJoining": data.get('dateOfJoining', datetime.now(ZoneInfo("UTC")).date().isoformat()),
                "company": data.get('company', 'POS 8'),
                "status": data.get('status', 'Active'),
                "salutation": data.get('salutation', ''),
                "maritalStatus": data.get('maritalStatus', ''),
                "address": data.get('address', ''),
                "idNumber": data.get('idNumber', ''),
                "idExpiry": id_expiry,
                "employeeDesignation": data['employeeDesignation'],
                "employeeType": data['employeeType'],
                "bankName": data.get('bankName', ''),
                "accountHolderName": data.get('accountHolderName', ''),
                "accountNumber": data.get('accountNumber', ''),
                "ifscCode": data.get('ifscCode', ''),
                "basicSalary": basic,
                "hra": hra,
                "ta": ta,
                "oa": oa,
                "totalSalary": total_salary,
                "nationality": data.get('nationality', ''),
                "education": data.get('education', ''),
                "previousExperience": data.get('previousExperience', ''),
                "skills": data.get('skills', ''),
                "healthInfo": data.get('healthInfo', ''),
                "familyDetails": data.get('familyDetails', ''),
                "username": data['username'],
                "password": hashed_password,
                "profileImage": data.get('profileImage', ''),
                "created_at": datetime.now(ZoneInfo("UTC")).isoformat(),
                "isDraft": False
            }
            worker_collection.insert_one(new_employee)
            new_user = {
                "email": data['email'],
                "password": hashed_password,
                "role": data['employeeDesignation'].lower(),
                "username": data['username'],
                "firstName": data['name'],
                "phone_number": data['phoneNumber'],
                "company": new_employee['company'],
                "pos_profile": "POS-001",
                "status": new_employee['status'],
                "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }
            users_collection.insert_one(new_user)
            logger.info(f"Employee created: {data['name']} ({data['email']}) with ID {employee_id}")
            return jsonify({"message": "Employee created successfully", "id": new_employee['_id'], "employeeId": employee_id}), 201
        except Exception as e:
            logger.error(f"Error creating employee: {str(e)}")
            logger.error(traceback.format_exc())
            return jsonify({"error": f"Internal Server Error: {str(e)}"}), 500

@app.route('/api/add-employee/<emp_id>', methods=['GET', 'PUT', 'DELETE'])
@db_required
def manage_employee(emp_id):
    mode = config.get("mode", "server")
    if mode == 'client':
        server_ip = config.get('server_ip', 'localhost') # Fix explicit IP
        server_url = f"http://{server_ip}:8000/api/add-employee/{emp_id}"
        if request.method == 'GET':
            try:
                response = requests.get(server_url, timeout=5)
                return jsonify(response.json()), response.status_code
            except Exception as e:
                return jsonify({"error": f"Proxy GET error: {str(e)}"}), 500
        elif request.method == 'PUT':
            try:
                data = request.get_json()
                response = requests.put(server_url, json=data, timeout=5)
                return jsonify(response.json()), response.status_code
            except Exception as e:
                return jsonify({"error": f"Proxy PUT error: {str(e)}"}), 500
        elif request.method == 'DELETE':
            try:
                response = requests.delete(server_url, timeout=5)
                return jsonify(response.json()), response.status_code
            except Exception as e:
                return jsonify({"error": f"Proxy DELETE error: {str(e)}"}), 500
    if request.method == 'GET':
        try:
            # Fetch single employee by ID, filter out drafts
            emp = worker_collection.find_one({'_id': emp_id})
            if not emp:
                return jsonify({"error": "Employee not found"}), 404
            if emp.get('isDraft') is True:
                return jsonify({"error": "Employee not found"}), 404
            if '_id' in emp:
                emp['_id'] = str(emp['_id'])
            # NEW: Fetch assigned schedule details for Attendance & Leaves
            assignment = employee_schedule_assign_collection.find_one({'employee_id': emp_id})
            assigned_schedule = None
            if assignment:
                schedule_id = assignment.get('schedule_id')
                if schedule_id:
                    rule = schedule_master_collection.find_one({'_id': schedule_id})
                    if rule:
                        shift_id = rule.get('shift_id')
                        shift = shift_master_collection.find_one({'_id': shift_id}) if shift_id else None
                        # Convert _id to str if present
                        if '_id' in rule:
                            rule['_id'] = str(rule['_id'])
                        if shift and '_id' in shift:
                            shift['_id'] = str(shift['_id'])
                        if '_id' in assignment:
                            assignment['_id'] = str(assignment['_id'])
                        # UPDATED: Build assigned_schedule object - Handle multiple time_slots for split shifts
                        # Build default_shift string with all slots
                        if shift and 'time_slots' in shift and shift['time_slots']:
                            slot_str = ', '.join([
                                f"{s.get('start_time', 'N/A')}-{s.get('end_time', 'N/A')}{ ' (O)' if s.get('is_overnight', False) else '' }"
                                for s in shift['time_slots']
                            ])
                            default_shift = f"{shift.get('schedule_name', 'N/A')} ({slot_str})"
                        else:
                            # Fallback to single if no time_slots (legacy)
                            default_shift = f"{shift.get('schedule_name', 'N/A')} {shift.get('start_time', 'N/A')} to {shift.get('end_time', 'N/A')}" if shift else 'N/A'
                        # Merge special days: Rule specials + Assignment overrides (prefer assignment if exists)
                        all_special_days = rule.get('special_days', [])
                        assignment_specials = assignment.get('special_day_assignments', [])
                        merged_specials = []
                        special_map = {f"{sd.get('date', '')}-{sd.get('description', '')}": sd for sd in assignment_specials} # Key for quick lookup
                        for sd in all_special_days:
                            key = f"{sd.get('date', '')}-{sd.get('description', '')}"
                            if key in special_map:
                                merged_specials.append({**sd, **special_map[key], 'is_observed': True}) # Override with assignment
                            else:
                                merged_specials.append({**sd, 'is_observed': sd.get('type') == 'Holiday'}) # Default observed for holidays
                        # Add any assignment-only specials
                        for sd in assignment_specials:
                            key = f"{sd.get('date', '')}-{sd.get('description', '')}"
                            if not any(f"{existing.get('date', '')}-{existing.get('description', '')}" == key for existing in all_special_days):
                                merged_specials.append(sd)
                       
                        assigned_schedule = {
                            'schedule_name': rule.get('schedule_name', 'N/A'), # Used as Holiday List code e.g., HL-IND-KL-2025
                            'start_date': rule.get('start_date'), # NEW: Full period for weekly off generation
                            'end_date': rule.get('end_date'), # NEW
                            'default_shift': default_shift,
                            'working_days': rule.get('working_days', []),
                            'weekly_off': rule.get('weekly_off', []), # NEW: For weekly off list
                            'special_days': all_special_days, # Original rule specials (for reference)
                            'special_day_assignments': merged_specials, # Merged for employee-specific
                            'assignment_notes': assignment.get('notes', ''),
                        }
            emp['assigned_schedule'] = assigned_schedule
            return jsonify(emp), 200
        except Exception as e:
            logger.error(f"Error fetching employee: {str(e)}")
            logger.error(traceback.format_exc())
            return jsonify({"error": str(e)}), 500
    elif request.method == 'PUT':
        try:
            data = request.get_json()
            if not isinstance(data, dict):
                return jsonify({"error": "JSON data must be an object"}), 400
           
            existing_emp = worker_collection.find_one({'_id': emp_id})
            if not existing_emp:
                return jsonify({"error": "Employee not found"}), 404
            # Check email uniqueness (excluding self)
            if 'email' in data and data['email'] != existing_emp.get('email'):
                if worker_collection.find_one({"email": data['email'], "_id": {"$ne": emp_id}, "isDraft": {"$ne": True}}):
                     return jsonify({"error": "Email already in use by another employee"}), 400
           
            # Safe float conversion helper
            def safe_float(val, default=0.0):
                try:
                    return float(val) if val else default
                except (ValueError, TypeError):
                    return default
            # Prepare update fields
            update_fields = {
                "name": data.get('name', existing_emp.get('name')),
                "phoneNumber": data.get('phoneNumber', existing_emp.get('phoneNumber')),
                "email": data.get('email', existing_emp.get('email')),
                "gender": data.get('gender', existing_emp.get('gender')),
                "dateOfBirth": data.get('dateOfBirth', existing_emp.get('dateOfBirth')),
                "dateOfJoining": data.get('dateOfJoining', existing_emp.get('dateOfJoining')),
                "company": data.get('company', existing_emp.get('company')),
                "status": data.get('status', existing_emp.get('status')),
                "salutation": data.get('salutation', existing_emp.get('salutation')),
                "maritalStatus": data.get('maritalStatus', existing_emp.get('maritalStatus')),
                "address": data.get('address', existing_emp.get('address')),
                "idNumber": data.get('idNumber', existing_emp.get('idNumber')),
                "idExpiry": data.get('idExpiry', existing_emp.get('idExpiry')),
                "employeeDesignation": data.get('employeeDesignation', existing_emp.get('employeeDesignation')),
                "employeeType": data.get('employeeType', existing_emp.get('employeeType')),
                # Bank Details
                "bankName": data.get('bankName', existing_emp.get('bankName', '')),
                "accountHolderName": data.get('accountHolderName', existing_emp.get('accountHolderName', '')),
                "accountNumber": data.get('accountNumber', existing_emp.get('accountNumber', '')),
                "ifscCode": data.get('ifscCode', existing_emp.get('ifscCode', '')),
                # Salary
                "basicSalary": safe_float(data.get('basicSalary')),
                "hra": safe_float(data.get('hra')),
                "ta": safe_float(data.get('ta')),
                "oa": safe_float(data.get('oa')),
                # Extra
                "nationality": data.get('nationality', existing_emp.get('nationality')),
                "education": data.get('education', existing_emp.get('education')),
                "previousExperience": data.get('previousExperience', existing_emp.get('previousExperience')),
                "skills": data.get('skills', existing_emp.get('skills')),
                "healthInfo": data.get('healthInfo', existing_emp.get('healthInfo')),
                "familyDetails": data.get('familyDetails', existing_emp.get('familyDetails')),
                "username": data.get('username', existing_emp.get('username')),
                "updated_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }
            # Recalculate total salary if components provided, else use provided, else sum
            basic = update_fields['basicSalary']
            hra = update_fields['hra']
            ta = update_fields['ta']
            oa = update_fields['oa']
            if data.get('totalSalary'):
                update_fields['totalSalary'] = safe_float(data.get('totalSalary'))
            else:
                update_fields['totalSalary'] = basic + hra + ta + oa
            # Handle password update
            if 'password' in data and data['password']:
                try:
                    hashed_password = bcrypt.hashpw(str(data['password']).encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                    update_fields['password'] = hashed_password
                except Exception as e:
                    return jsonify({"error": "Invalid password format"}), 400
           
            # Handle profile image
            if 'profileImage' in data:
                update_fields['profileImage'] = data['profileImage']
            worker_collection.update_one({'_id': emp_id}, {'$set': update_fields})
           
            # Update user collection if relevant fields changed
            user_update = {}
            if 'email' in update_fields: user_update['email'] = update_fields['email']
            if 'username' in update_fields: user_update['username'] = update_fields['username']
            if 'name' in update_fields: user_update['firstName'] = update_fields['name']
            if 'phoneNumber' in update_fields: user_update['phone_number'] = update_fields['phoneNumber']
            if 'employeeDesignation' in update_fields: user_update['role'] = update_fields['employeeDesignation'].lower()
            if 'status' in update_fields: user_update['status'] = update_fields['status']
            if 'password' in update_fields: user_update['password'] = update_fields['password']
           
            # Try to find associated user by email (original or new)
            # Strategy: find by original email if possible, else current.
            # Since email might change, we should rely on what was in existing_emp
            users_collection.update_one({'email': existing_emp.get('email')}, {'$set': user_update})
            return jsonify({"message": "Employee updated successfully", "id": emp_id}), 200
        except Exception as e:
            logger.error(f"Error updating employee: {str(e)}")
            logger.error(traceback.format_exc())
            return jsonify({"error": f"Internal Server Error: {str(e)}"}), 500
    elif request.method == 'DELETE':
        try:
            existing_emp = worker_collection.find_one({'_id': emp_id})
            if not existing_emp:
                return jsonify({"error": "Employee not found"}), 404
           
            # Delete employee
            worker_collection.delete_one({'_id': emp_id})
           
            # Delete associated user
            if existing_emp.get('email'):
                users_collection.delete_one({'email': existing_emp['email']})
           
            # Optional: Delete/Archive assignments?
            # Keeping it simple: Delete assignments
            employee_schedule_assign_collection.delete_many({'employee_id': emp_id})
           
            return jsonify({"message": "Employee deleted successfully"}), 200
        except Exception as e:
            logger.error(f"Error deleting employee: {str(e)}")
            return jsonify({"error": f"Internal Server Error: {str(e)}"}), 500# 1. SHIFT MASTER


# FIXED: Updated /api/attendance GET handler to support month + employeeId filtering
# FIXED: Updated /api/attendance GET handler to support month + employeeId filtering
# UPDATED: In PUT, now also updates dailySalary if provided
# ATTENDANCE (Full, with logic for weekoff, special holiday, extended hours)
@app.route('/api/attendance', methods=['GET', 'POST', 'PUT', 'DELETE'])
@db_required
def attendance():
    mode = config.get("mode", "server")
    if mode == 'client':
        server_ip = config.get('server_ip', 'localhost')
        server_url = f"http://{server_ip}:8000/api/attendance"
        if request.method == 'GET':
            params = request.args
            response = requests.get(server_url, params=params, timeout=10)
            return jsonify(response.json()), response.status_code
        elif request.method == 'POST':
            data = request.get_json()
            response = requests.post(server_url, json=data, timeout=10)
            return jsonify(response.json()), response.status_code
        elif request.method == 'PUT':
            data = request.get_json()
            response = requests.put(server_url, json=data, timeout=10)
            return jsonify(response.json()), response.status_code
        elif request.method == 'DELETE':
            data = request.get_json()
            response = requests.delete(server_url, json=data, timeout=10)
            return jsonify(response.json()), response.status_code
        return jsonify({"error": "Method not supported"}), 405
    if request.method == 'GET':
        try:
            employee_id = request.args.get('employee_id')
            date = request.args.get('date')
            month = request.args.get('month')
            # Case 1: employee_id only (no date/month) - Get active assignment for auto-populate
            if employee_id and not date and not month:
                # Validate employee
                emp = worker_collection.find_one({'_id': employee_id})
                if not emp:
                    return jsonify({"error": "Employee not found"}), 404
                # FIXED: Python-side filter/sort for SQLite compatibility
                all_assignments = list(employee_schedule_assign_collection.find({'employee_id': employee_id}))
                active_assignments = [a for a in all_assignments if a.get('is_active') is True]
                active_assignments.sort(key=lambda x: x.get('assigned_date', '1900-01-01'), reverse=True)
                if not active_assignments:
                    return jsonify({"error": "No active schedule assignment found for employee"}), 404
                assignment = active_assignments[0]
                schedule_id = assignment['schedule_id']
                schedule = schedule_master_collection.find_one({'_id': schedule_id})
                if not schedule:
                    return jsonify({"error": "Schedule not found"}), 404
                shift_id = schedule.get('shift_id', '')
                shift = shift_master_collection.find_one({'_id': shift_id}) if shift_id else None
                if not shift:
                    return jsonify({"error": "Shift not found"}), 404
                response_data = {
                    'assignment': convert_objectid_to_str(assignment),
                    'schedule': convert_objectid_to_str(schedule),
                    'shift': convert_objectid_to_str(shift)
                }
                return jsonify(response_data), 200
            # Case 2: employee_id + date - Existing or compute daily (for auto-populate with date)
            elif employee_id and date:
                # Validate employee
                emp = worker_collection.find_one({'_id': employee_id})
                if not emp:
                    return jsonify({"error": "Employee not found"}), 404
                # Check existing attendance
                existing = attendance_collection.find_one({
                    'employee_id': employee_id,
                    'attendance_date': date
                })
                if existing:
                    # Populate employee for consistency
                    employee = worker_collection.find_one({'_id': employee_id})
                    populated = convert_objectid_to_str(existing)
                    populated['employee'] = convert_objectid_to_str(employee) if employee else None
                    return jsonify(populated), 200
                # FIXED: Python-side filter/sort for SQLite
                all_assignments = list(employee_schedule_assign_collection.find({'employee_id': employee_id}))
                active_assignments = [a for a in all_assignments if a.get('is_active') is True]
                active_assignments.sort(key=lambda x: x.get('assigned_date', '1900-01-01'), reverse=True)
                if not active_assignments:
                    return jsonify({"error": "No active schedule assignment found"}), 404
                assignment = active_assignments[0]
                schedule_id = assignment['schedule_id']
                schedule = schedule_master_collection.find_one({'_id': schedule_id})
                if not schedule:
                    return jsonify({"error": "Schedule not found"}), 404
                # Parse date for weekday/special checks
                try:
                    dt = datetime.strptime(date, '%Y-%m-%d')
                    weekday = dt.weekday() # 0=Mon, 6=Sun
                    weekday_name = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][weekday]
                except ValueError:
                    return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
                # Defaults
                status = 'Present'
                special_day_type = 'None'
                notes = ''
                shift_id = schedule.get('shift_id', '')
                # Check special_day_assignments in assignment (for extended/halfday custom shifts)
                special_assign = next((s for s in assignment.get('special_day_assignments', [])
                                       if s.get('date') == date), None)
                custom_start = None
                custom_end = None
                custom_overnight = None
                if special_assign:
                    shift_id = special_assign.get('shift_id', shift_id)
                    notes = special_assign.get('notes', '') or special_assign.get('description', '')
                    special_day_type = special_assign.get('type', special_day_type) # e.g., 'Extended', 'HalfDay'
                    if special_day_type == 'Extended':
                        status = 'Extended'
                        custom_start = special_assign.get('extended_start')
                        custom_end = special_assign.get('extended_end')
                        custom_overnight = True
                    elif special_day_type == 'HalfDay':
                        status = 'HalfDay'
                        custom_start = special_assign.get('start_time')
                        custom_end = special_assign.get('end_time')
                        # overnight remains from shift
                # Check weekly_off
                weekly_off = schedule.get('weekly_off', [])
                if isinstance(weekly_off, list) and weekday_name in weekly_off:
                    status = 'WeeklyOff'
                    special_day_type = 'WeeklyOff'
                # Check special_days in schedule (holidays/extended)
                special_days = schedule.get('special_days', [])
                special_day = next((sd for sd in special_days if sd.get('date') == date), None)
                if special_day:
                    day_type = special_day.get('type', 'Holiday') # 'Holiday', 'Extended', 'HalfDay'
                    status = day_type if day_type != 'Holiday' else 'Holiday'
                    special_day_type = day_type
                    notes = special_day.get('description', '') or notes
                    if day_type == 'Extended':
                        shift_id = special_day.get('shift_id', shift_id)
                        custom_start = special_day.get('extended_start')
                        custom_end = special_day.get('extended_end')
                        custom_overnight = True
                    elif day_type == 'HalfDay':
                        shift_id = special_day.get('shift_id', shift_id)
                        custom_start = special_day.get('start_time')
                        custom_end = special_day.get('end_time')
                        # overnight from shift
                # Get shift
                if not shift_id:
                    return jsonify({"error": "No shift_id found"}), 400
                shift = shift_master_collection.find_one({'_id': shift_id})
                if not shift:
                    return jsonify({"error": "Shift not found"}), 404
                planned_start_time = shift.get('start_time')
                planned_end_time = shift.get('end_time')
                is_overnight = shift.get('is_overnight', False)
                # Fallback to time_slots[0] if top-level missing
                if not planned_start_time and shift.get('time_slots') and len(shift.get('time_slots', [])) > 0:
                    first_slot = shift['time_slots'][0]
                    planned_start_time = first_slot.get('start_time')
                    planned_end_time = first_slot.get('end_time')
                    is_overnight = first_slot.get('is_overnight', False)
                # Override with custom if present
                if custom_start:
                    planned_start_time = custom_start
                if custom_end:
                    planned_end_time = custom_end
                if custom_overnight is not None:
                    is_overnight = custom_overnight
                response_data = {
                    'auto_filled': True,
                    'schedule_id': str(schedule_id),
                    'shift_id': str(shift_id),
                    'status': status,
                    'special_day_type': special_day_type,
                    'planned_start_time': planned_start_time,
                    'planned_end_time': planned_end_time,
                    'is_overnight': is_overnight,
                    'notes': notes
                }
                return jsonify(response_data), 200
            # Case 3: Other filters (month, date without employee, or NO filters = fetch ALL)
            filter_dict = {}
            fetch_all = not employee_id and not date and not month # NEW: Detect no filters to fetch all
            if employee_id:
                filter_dict['employee_id'] = employee_id
            if date:
                filter_dict['attendance_date'] = date
            elif month:
                # For month, use Python filter since regex might not be supported in SQLite find
                filter_dict = {} # Will filter after fetch if needed
                month_filter = month # Handle below
            # FIXED: Always fetch as list, then Python filter/sort for SQLite compatibility
            if fetch_all:
                # Fetch all records
                raw_records = list(attendance_collection.find({}))
            else:
                raw_records = list(attendance_collection.find(filter_dict))
            # Apply month filter if needed (Python-side)
            if month and not date:
                raw_records = [rec for rec in raw_records if str(rec.get('attendance_date', '')).startswith(month + '-')]
            # Python sort: desc by attendance_date
            raw_records.sort(key=lambda x: x.get('attendance_date', '1900-01-01'), reverse=True)
            # Populate employee for each record (join)
            records = []
            for rec in raw_records:
                populated = convert_objectid_to_str(rec)
                emp_id = rec.get('employee_id')
                if emp_id:
                    employee = worker_collection.find_one({'_id': emp_id})
                    populated['employee'] = convert_objectid_to_str(employee) if employee else None
                else:
                    populated['employee'] = None
                records.append(populated)
            return jsonify(records), 200
        except Exception as e:
            logger.error(f"Error fetching attendance: {str(e)}")
            logger.error(traceback.format_exc())
            return jsonify({"error": str(e)}), 500
    elif request.method == 'POST':
        try:
            data = request.get_json()
            if not isinstance(data, dict):
                return jsonify({"error": "Invalid JSON"}), 400
            required_fields = ['employee_id', 'attendance_date', 'status', 'schedule_id', 'shift_id']
            if not all(field in data for field in required_fields):
                return jsonify({"error": f"Missing required fields: {', '.join(required_fields)}"}), 400
            # Duplicate Check: Check for existing record with same Emp + Date + Start Time
            # This allows split shifts (different start times) but prevents exact duplicate entries
            existing = attendance_collection.find_one({
                "employee_id": data['employee_id'],
                "attendance_date": data['attendance_date'],
                "planned_start_time": data.get('planned_start_time')
            })
            if existing:
                return jsonify({"error": "Attendance record already exists for this employee, date, and time slot."}), 400
            # Validate FKs
            if not worker_collection.find_one({"_id": data['employee_id']}):
                return jsonify({"error": "Invalid employee_id"}), 400
            if not schedule_master_collection.find_one({"_id": data['schedule_id']}):
                return jsonify({"error": "Invalid schedule_id"}), 400
            if not shift_master_collection.find_one({"_id": data['shift_id']}):
                return jsonify({"error": "Invalid shift_id"}), 400
            # Fetch shift for defaults
            # Fetch shift for defaults
            shift = shift_master_collection.find_one({"_id": data['shift_id']})
            
            # Determine base times from shift (top-level or first slot)
            base_start = shift.get('start_time')
            base_end = shift.get('end_time')
            base_overnight = shift.get('is_overnight', False)
            if not base_start and shift.get('time_slots') and len(shift.get('time_slots', [])) > 0:
                first_slot = shift['time_slots'][0]
                base_start = first_slot.get('start_time')
                base_end = first_slot.get('end_time')
                base_overnight = first_slot.get('is_overnight', False)

            planned_start_time = data.get('planned_start_time') or base_start
            planned_end_time = data.get('planned_end_time') or base_end
            is_overnight = data.get('is_overnight', base_overnight)
            # Auto-compute minutes if times provided
            worked_minutes = data.get('worked_minutes', 0)
            overtime_minutes = data.get('overtime_minutes', 0)
            late_minutes = data.get('late_minutes', 0)
            early_exit_minutes = data.get('early_exit_minutes', 0)
            if data.get('actual_check_in') and data.get('actual_check_out') and planned_start_time and planned_end_time:
                def parse_time_to_minutes(time_str):
                    if not time_str:
                        return 0
                    h, m = map(int, time_str.split(':'))
                    return h * 60 + m
                in_mins = parse_time_to_minutes(data.get('actual_check_in'))
                out_mins = parse_time_to_minutes(data.get('actual_check_out'))
                plan_start = parse_time_to_minutes(planned_start_time)
                plan_end = parse_time_to_minutes(planned_end_time)
                # Compute planned_duration
                planned_duration = plan_end - plan_start if plan_end >= plan_start else plan_end + 1440 - plan_start
                # Simple calc; for overnight, assume out > in or adjust if needed (e.g., +24h if out < in)
                if is_overnight and out_mins < in_mins:
                    out_mins += 24 * 60
                worked = max(0, out_mins - in_mins)
                late = max(0, in_mins - plan_start)
                # For early: adjust plan_end if overnight
                early_plan_end = plan_end + (1440 if is_overnight and plan_end < plan_start else 0)
                early = max(0, early_plan_end - out_mins)
                overtime = max(0, worked - planned_duration)
                worked_minutes = worked
                overtime_minutes = overtime
                late_minutes = late
                early_exit_minutes = early
            # If status is off/holiday/absent/leave, force 0 minutes
            if data['status'] in ['WeeklyOff', 'Holiday', 'Absent', 'Leave']:
                worked_minutes = overtime_minutes = late_minutes = early_exit_minutes = 0
            new_record = {
                "_id": str(uuid.uuid4()),
                "employee_id": data['employee_id'],
                "attendance_date": data['attendance_date'],
                "schedule_id": data['schedule_id'],
                "shift_id": data['shift_id'],
                "status": data['status'],
                "planned_start_time": planned_start_time,
                "planned_end_time": planned_end_time,
                "actual_check_in": data.get('actual_check_in', ''),
                "actual_check_out": data.get('actual_check_out', ''),
                "worked_minutes": worked_minutes,
                "overtime_minutes": overtime_minutes,
                "late_minutes": late_minutes,
                "early_exit_minutes": early_exit_minutes,
                "is_overnight": is_overnight,
                "special_day_type": data.get('special_day_type', 'None'),
                "notes": data.get('notes', ''),
                "created_at": datetime.now(ZoneInfo("UTC")).isoformat(),
                "updated_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }
            attendance_collection.insert_one(new_record)
            # Populate for response
            populated = convert_objectid_to_str(new_record)
            emp_id = new_record['employee_id']
            employee = worker_collection.find_one({'_id': emp_id})
            populated['employee'] = convert_objectid_to_str(employee) if employee else None
            logger.info(f"Attendance created for {data['employee_id']} on {data['attendance_date']}")
            return jsonify({"message": "Attendance created successfully", "record": populated}), 201
        except Exception as e:
            logger.error(f"Error creating attendance: {str(e)}")
            logger.error(traceback.format_exc())
            return jsonify({"error": str(e)}), 500
    elif request.method == 'PUT':
        try:
            data = request.get_json()
            if not data or '_id' not in data:
                return jsonify({"error": "Missing _id"}), 400
            current = attendance_collection.find_one({"_id": data['_id']})
            if not current:
                return jsonify({"error": "Record not found"}), 404
            # Fetch shift for recompute
            # Fetch shift for recompute
            shift = shift_master_collection.find_one({"_id": current['shift_id']})
            
            # Determine base times from shift (top-level or first slot)
            base_start = shift.get('start_time')
            base_end = shift.get('end_time')
            base_overnight = shift.get('is_overnight', False)
            if not base_start and shift.get('time_slots') and len(shift.get('time_slots', [])) > 0:
                first_slot = shift['time_slots'][0]
                base_start = first_slot.get('start_time')
                base_end = first_slot.get('end_time')
                base_overnight = first_slot.get('is_overnight', False)

            planned_start_time = data.get('planned_start_time', current.get('planned_start_time') or base_start)
            planned_end_time = data.get('planned_end_time', current.get('planned_end_time') or base_end)
            is_overnight = data.get('is_overnight', current.get('is_overnight', base_overnight))
            # Recompute if times changed
            if 'actual_check_in' in data or 'actual_check_out' in data:
                def parse_time_to_minutes(time_str):
                    if not time_str:
                        return 0
                    h, m = map(int, time_str.split(':'))
                    return h * 60 + m
                in_mins = parse_time_to_minutes(data.get('actual_check_in', current['actual_check_in']))
                out_mins = parse_time_to_minutes(data.get('actual_check_out', current['actual_check_out']))
                plan_start = parse_time_to_minutes(planned_start_time)
                plan_end = parse_time_to_minutes(planned_end_time)
                # Compute planned_duration
                planned_duration = plan_end - plan_start if plan_end >= plan_start else plan_end + 1440 - plan_start
                # Overnight handling
                if is_overnight and out_mins < in_mins:
                    out_mins += 24 * 60
                worked = max(0, out_mins - in_mins)
                late = max(0, in_mins - plan_start)
                # For early: adjust plan_end if overnight
                early_plan_end = plan_end + (1440 if is_overnight and plan_end < plan_start else 0)
                early = max(0, early_plan_end - out_mins)
                overtime = max(0, worked - planned_duration)
                data['worked_minutes'] = worked
                data['overtime_minutes'] = overtime
                data['late_minutes'] = late
                data['early_exit_minutes'] = early
            # If status off, zero minutes
            if data.get('status', current['status']) in ['WeeklyOff', 'Holiday', 'Absent', 'Leave']:
                data['worked_minutes'] = data['overtime_minutes'] = data['late_minutes'] = data['early_exit_minutes'] = 0
            # Update fields
            update_set = {"$set": {
                "status": data.get('status', current['status']),
                "planned_start_time": planned_start_time,
                "planned_end_time": planned_end_time,
                "actual_check_in": data.get('actual_check_in', current['actual_check_in']),
                "actual_check_out": data.get('actual_check_out', current['actual_check_out']),
                "worked_minutes": data.get('worked_minutes', current['worked_minutes']),
                "overtime_minutes": data.get('overtime_minutes', current['overtime_minutes']),
                "late_minutes": data.get('late_minutes', current['late_minutes']),
                "early_exit_minutes": data.get('early_exit_minutes', current['early_exit_minutes']),
                "is_overnight": data.get('is_overnight', current['is_overnight']),
                "special_day_type": data.get('special_day_type', current['special_day_type']),
                "notes": data.get('notes', current['notes']),
                "updated_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }}
            result = attendance_collection.update_one({"_id": data['_id']}, update_set)
            if result.modified_count == 0:
                return jsonify({"error": "No changes or record not found"}), 404
            updated = attendance_collection.find_one({"_id": data['_id']})
            # Populate for response
            populated = convert_objectid_to_str(updated)
            emp_id = updated['employee_id']
            employee = worker_collection.find_one({'_id': emp_id})
            populated['employee'] = convert_objectid_to_str(employee) if employee else None
            return jsonify({"message": "Attendance updated successfully", "record": populated}), 200
        except Exception as e:
            logger.error(f"Error updating attendance: {str(e)}")
            return jsonify({"error": str(e)}), 500
    elif request.method == 'DELETE':
        try:
            data = request.get_json()
            if not data or '_id' not in data:
                return jsonify({"error": "Missing _id"}), 400
            result = attendance_collection.delete_one({"_id": data['_id']})
            if result.deleted_count == 0:
                return jsonify({"error": "Record not found"}), 404
            return jsonify({"message": "Attendance deleted successfully"}), 200
        except Exception as e:
            logger.error(f"Error deleting attendance: {str(e)}")
            return jsonify({"error": str(e)}), 500
    return jsonify({"error": "Method not allowed"}), 405
        
@app.route('/api/working-days', methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'])
@db_required
def working_days_handler():
    if request.method == 'OPTIONS':
        response = jsonify({"success": True})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200
    if request.method == 'GET':
        try:
            year = request.args.get('year')
            month = request.args.get('month')
            if not year:
                return jsonify({"error": "Year is required"}), 400
            year_str = str(year)
            holidays = []
            if month:
                # Fetch for specific month
                month_str = str(month)
                doc = working_days_collection.find_one({"year": year_str, "month": month_str})
                holidays = doc.get("holidays", []) if doc else []
                total_working_days = doc.get("totalWorkingDays", 0) if doc else 0
                return jsonify({"holidays": holidays, "totalWorkingDays": total_working_days}), 200
            else:
                # Fetch all holidays for the year (aggregate across months)
                cursor = working_days_collection.find({"year": year_str})
                for doc in cursor:
                    holidays.extend(doc.get("holidays", []))
                return jsonify({"holidays": holidays}), 200
        except Exception as e:
            logger.error(f"Error fetching working days: {str(e)}")
            return jsonify({"error": str(e)}), 500
    elif request.method == 'POST':
        try:
            data = request.get_json()
            year = str(data.get('year')) # Ensure string
            month = str(data.get('month')) # Ensure string, padded already in frontend
            holidays = data.get('holidays', [])
            total_working_days = data.get('totalWorkingDays')
            if not year or not month:
                return jsonify({"error": "Year and month are required"}), 400
            doc = {
                "year": year,
                "month": month,
                "holidays": holidays,
                "totalWorkingDays": total_working_days,
                "updated_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }
            # Use replace_one with upsert to update or insert
            result = working_days_collection.replace_one(
                {"year": year, "month": month},
                doc,
                upsert=True
            )
            # FIXED: Use matched_count to check if new insert or update (avoids upserted_id attribute error in older PyMongo versions)
            if result.matched_count == 0:
                logger.info(f"New working days document inserted for {year}-{month}")
            else:
                logger.info(f"Working days updated for {year}-{month}")
            return jsonify({"message": "Working days saved successfully"}), 200
        except Exception as e:
            logger.error(f"Error saving working days: {str(e)}")
            return jsonify({"error": str(e)}), 500
    elif request.method == 'PUT':
        # Optional: For partial updates, e.g., update specific holidays
        try:
            data = request.get_json()
            year = str(data.get('year'))
            month = str(data.get('month'))
            if not year or not month:
                return jsonify({"error": "Year and month are required"}), 400
            # Update holidays array
            result = working_days_collection.update_one(
                {"year": year, "month": month},
                {"$set": {"holidays": data.get('holidays', []), "updated_at": datetime.now(ZoneInfo("UTC")).isoformat()}}
            )
            if result.modified_count > 0:
                return jsonify({"message": "Working days updated successfully"}), 200
            return jsonify({"message": "No changes made"}), 200
        except Exception as e:
            logger.error(f"Error updating working days: {str(e)}")
            return jsonify({"error": str(e)}), 500
    elif request.method == 'DELETE':
        # Optional: Delete for a specific year-month
        try:
            year = request.args.get('year')
            month = request.args.get('month')
            if not year or not month:
                return jsonify({"error": "Year and month are required"}), 400
            result = working_days_collection.delete_one({"year": str(year), "month": str(month)})
            if result.deleted_count > 0:
                return jsonify({"message": "Working days deleted successfully"}), 200
            return jsonify({"message": "No document found to delete"}), 200
        except Exception as e:
            logger.error(f"Error deleting working days: {str(e)}")
            return jsonify({"error": str(e)}), 500
    else:
        return jsonify({"error": "Method not allowed"}), 405

@app.route('/api/salary-slip', methods=['POST'])
@db_required
def salary_slip():
    mode = config.get("mode", "server")
    if mode == 'client':
        server_url = f"http://{config['server_ip']}:8000/api/salary-slip"
        response = requests.post(server_url, json=request.get_json())
        if response.status_code in [200, 201]:
            return jsonify(response.json()), response.status_code
        else:
            return jsonify({"error": "Proxy save failed"}), response.status_code
    
    if request.method == 'POST':
        try:
            data = request.get_json()
            if not isinstance(data, dict):
                return jsonify({"error": "JSON data must be an object"}), 400
            
            required_fields = ['employeeId', 'month', 'grossSalary', 'totalSalary', 'netPay']
            if not all(field in data for field in required_fields):
                return jsonify({"error": "Missing required fields"}), 400
            
            # Fetch employee details including bank info
            emp = worker_collection.find_one({"_id": data['employeeId']})
            if not emp:
                return jsonify({"error": "Employee not found"}), 404
            
            new_slip = {
                "_id": str(uuid.uuid4()),
                "employeeId": data['employeeId'],
                "employeeName": emp.get('name', ''),
                "month": data['month'],
                "grossSalary": float(data['grossSalary']),
                "fullCount": data.get('fullCount', 0),
                "offCount": data.get('offCount', 0),
                "leaveWithoutPay": data.get('leaveWithoutPay', 0),
                "absentCount": data.get('absentCount', 0),
                "paymentDays": data.get('paymentDays', 0),
                "totalSalary": float(data['totalSalary']),
                "deductions": float(data.get('deductions', 0)),
                "netPay": float(data['netPay']),
                "dailyRate": float(data.get('dailyRate', 0)),
                "grossPay": float(data.get('grossPay', 0)),
                "grossYearToDate": float(data.get('grossYearToDate', 0)),
                "totalDeductions": float(data.get('totalDeductions', 0)),
                "employeeType": emp.get('employeeType', ''),
                "employeeIdCode": emp.get('employeeId', ''),
                # Bank Details
                "bankName": emp.get('bankName', ''),
                "accountHolderName": emp.get('accountHolderName', ''),
                "accountNumber": emp.get('accountNumber', ''),
                "ifscCode": emp.get('ifscCode', ''),
                # Earnings and Deductions arrays
                "earnings": data.get('earnings', []),
                "deductions": data.get('deductions', []),
                "created_at": datetime.now(ZoneInfo("UTC")).isoformat()
            }
            
            if salary_slips_collection is None:
                logger.error("salary_slips_collection not initialized")
                return jsonify({"error": "Database not ready"}), 503
            
            # Check if slip exists for employee and month
            existing = salary_slips_collection.find_one({"employeeId": data['employeeId'], "month": data['month']})
            if existing:
                return jsonify({"error": "Salary slip already exists for this employee and month"}), 400
            
            salary_slips_collection.insert_one(new_slip)
            logger.info(f"Salary slip saved for {new_slip['employeeName']} - {data['month']}")
            return jsonify({"message": "Salary slip saved successfully", "slip": new_slip}), 201
        except Exception as e:
            logger.error(f"Error saving salary slip: {str(e)}")
            return jsonify({"error": str(e)}), 500
# NEW: Endpoint for /api/worker (alias or additional, as per request)
@app.route('/api/worker', methods=['POST', 'GET'])
@db_required
def worker():
    return add_employee() # Reuse the same logic
# --- Brand Management API Endpoints ---

@app.route('/api/brands', methods=['GET'])
@db_required
def get_brands():
    try:
        brands = brands_collection.find()
        return jsonify(convert_objectid_to_str(brands)), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch brands: {str(e)}"}), 500

@app.route('/api/brands', methods=['POST'])
@db_required
def add_brand():
    try:
        data = request.json
        if not data or 'name' not in data or not data['name'].strip():
            return jsonify({'error': 'Invalid brand name'}), 400
        
        # Check if brand already exists
        if brands_collection.find_one({'name': data['name'].strip()}):
            return jsonify({'error': 'Brand already exists'}), 400
        
        brand = {
            'name': data['name'].strip(),
            'created_at': datetime.now(timezone.utc).isoformat()
        }
        result = brands_collection.insert_one(brand)
        inserted_brand = brands_collection.find_one({'_id': result.inserted_id})
        return jsonify(convert_objectid_to_str(inserted_brand)), 201
    except Exception as e:
        return jsonify({'error': f"Failed to add brand: {str(e)}"}), 500

@app.route('/api/brands/<id>', methods=['PUT'])
@db_required
def update_brand(id):
    try:
        data = request.json
        if not data or 'name' not in data or not data['name'].strip():
            return jsonify({'error': 'Invalid brand name'}), 400
        
        # Check if another brand with the same name exists
        existing = brands_collection.find_one({'name': data['name'].strip(), '_id': {'$ne': id}})
        if existing:
            return jsonify({'error': 'Brand name already exists'}), 400
        
        result = brands_collection.update_one({'_id': id}, {'$set': {'name': data['name'].strip()}})
        if result.matched_count == 0:
            return jsonify({'error': 'Brand not found or no changes'}), 404
        
        return jsonify({'message': 'Brand updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to update brand: {str(e)}"}), 500

@app.route('/api/brands/<id>', methods=['DELETE'])
@db_required
def delete_brand(id):
    try:
        result = brands_collection.delete_one({'_id': id})
        if result.deleted_count == 0:
            return jsonify({'error': 'Brand not found'}), 404
        
        return jsonify({'message': 'Brand deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to delete brand: {str(e)}"}), 500
    
# --- Scheduling System API Endpoints ---

# 1. SHIFT MASTER
@app.route('/api/schedules', methods=['GET'])
@db_required
def get_shifts():
    try:
        shifts = shift_master_collection.find()
        return jsonify(convert_objectid_to_str(shifts)), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch shifts: {str(e)}"}), 500

@app.route('/api/schedules', methods=['POST'])
@db_required
def add_shift():
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Handle single slot backward compatibility or array
        time_slots_single = {
            'start_time': data.get('start_time'),
            'end_time': data.get('end_time'),
            'is_overnight': data.get('is_overnight', False)
        }
        time_slots = data.get('time_slots', [])
        if not time_slots:
            if time_slots_single['start_time'] and time_slots_single['end_time']:
                time_slots = [time_slots_single]
            else:
                return jsonify({'error': 'At least one time slot with start and end time is required'}), 400
        
        # Validate each slot
        for slot in time_slots:
            if not slot.get('start_time') or not slot.get('end_time'):
                return jsonify({'error': 'Each time slot must have start_time and end_time'}), 400
        
        shift = {
            'schedule_name': data.get('schedule_name'),
            'time_slots': time_slots,
            'description': data.get('description', ''),
            'created_at': datetime.now(timezone.utc).isoformat()
        }
        result = shift_master_collection.insert_one(shift)
        inserted_shift = shift_master_collection.find_one({'_id': result.inserted_id})
        return jsonify(convert_objectid_to_str(inserted_shift)), 201
    except Exception as e:
        return jsonify({'error': f"Failed to add shift: {str(e)}"}), 500

@app.route('/api/schedules/<id>', methods=['PUT'])
@db_required
def update_shift(id):
    try:
        data = request.json
        time_slots = data.get('time_slots', [])
        if not time_slots:
            return jsonify({'error': 'time_slots array is required'}), 400
        
        # Validate each slot
        for slot in time_slots:
            if not slot.get('start_time') or not slot.get('end_time'):
                return jsonify({'error': 'Each time slot must have start_time and end_time'}), 400
        
        update_data = {
            'schedule_name': data.get('schedule_name'),
            'time_slots': time_slots,
            'description': data.get('description', ''),
        }
        result = shift_master_collection.update_one({'_id': id}, {'$set': update_data})
        if result.matched_count == 0:
            return jsonify({'error': 'Shift not found'}), 404
        return jsonify({'message': 'Shift updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to update shift: {str(e)}"}), 500

@app.route('/api/schedules/<id>', methods=['DELETE'])
@db_required
def delete_shift(id):
    try:
        result = shift_master_collection.delete_one({'_id': id})
        if result.deleted_count == 0:
            return jsonify({'error': 'Shift not found'}), 404
        return jsonify({'message': 'Shift deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to delete shift: {str(e)}"}), 500

# SCHEDULE RULE MASTER routes remain unchanged
@app.route('/api/schedule-rules', methods=['GET'])
@db_required
def get_schedule_rules():
    try:
        rules = schedule_master_collection.find()
        return jsonify(convert_objectid_to_str(rules)), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch schedule rules: {str(e)}"}), 500

@app.route('/api/schedule-rules', methods=['POST'])
@db_required
def add_schedule_rule():
    try:
        data = request.json
        rule = {
            'schedule_name': data.get('schedule_name'),
            'start_date': data.get('start_date'),
            'end_date': data.get('end_date'),
            'working_days': data.get('working_days', []),
            'weekly_off': data.get('weekly_off', []),
            'shift_id': data.get('shift_id'),
            'special_days': data.get('special_days', []),
            'created_at': datetime.now(timezone.utc).isoformat()
        }
        result = schedule_master_collection.insert_one(rule)
        inserted_rule = schedule_master_collection.find_one({'_id': result.inserted_id})
        return jsonify(convert_objectid_to_str(inserted_rule)), 201
    except Exception as e:
        return jsonify({'error': f"Failed to add schedule rule: {str(e)}"}), 500

@app.route('/api/schedule-rules/<id>', methods=['PUT'])
@db_required
def update_schedule_rule(id):
    try:
        data = request.json
        update_data = {
            'schedule_name': data.get('schedule_name'),
            'start_date': data.get('start_date'),
            'end_date': data.get('end_date'),
            'working_days': data.get('working_days', []),
            'weekly_off': data.get('weekly_off', []),
            'shift_id': data.get('shift_id'),
            'special_days': data.get('special_days', [])
        }
        result = schedule_master_collection.update_one({'_id': id}, {'$set': update_data})
        if result.matched_count == 0:
             return jsonify({'error': 'Rule not found'}), 404
        return jsonify({'message': 'Rule updated successfully'}), 200
    except Exception as e:
         return jsonify({'error': f"Failed to update rule: {str(e)}"}), 500
      
@app.route('/api/schedule-rules/<id>', methods=['DELETE'])
@db_required
def delete_schedule_rule(id):
    try:
        result = schedule_master_collection.delete_one({'_id': id})
        if result.deleted_count == 0:
            return jsonify({'error': 'Rule not found'}), 404
        return jsonify({'message': 'Rule deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to delete rule: {str(e)}"}), 500

# SCHEDULE ASSIGNMENT routes - minor update for auto-populate special_days with time_slots awareness (but since shift_id, no change needed)
@app.route('/api/schedule-assignments', methods=['GET'])
@db_required
def get_assignments():
    try:
        assignments = employee_schedule_assign_collection.find()
        return jsonify(convert_objectid_to_str(assignments)), 200
    except Exception as e:
        return jsonify({'error': f"Failed to fetch assignments: {str(e)}"}), 500

@app.route('/api/schedule-assignments', methods=['POST'])
@db_required
def add_assignment():
    try:
        data = request.json
        assignment = {
            'employee_id': data.get('employee_id'),
            'schedule_id': data.get('schedule_id'),
            'assigned_date': data.get('assigned_date'),
            'is_active': data.get('is_active', True),
            'special_day_assignments': data.get('special_day_assignments', []), # From form
            'notes': data.get('notes', ''),
            'created_at': datetime.now(timezone.utc).isoformat()
        }
        # NEW: If no special_day_assignments provided, fetch from rule and auto-set is_observed: true for Holidays
        if not assignment['special_day_assignments']:
            rule = schedule_master_collection.find_one({'_id': assignment['schedule_id']})
            if rule and rule.get('special_days'):
                assignment['special_day_assignments'] = [
                    {**sd, 'is_observed': sd.get('type') == 'Holiday'} for sd in rule['special_days']
                ]
        result = employee_schedule_assign_collection.insert_one(assignment)
        inserted_assign = employee_schedule_assign_collection.find_one({'_id': result.inserted_id})
        return jsonify(convert_objectid_to_str(inserted_assign)), 201
    except Exception as e:
        return jsonify({'error': f"Failed to add assignment: {str(e)}"}), 500

@app.route('/api/schedule-assignments/<id>', methods=['PUT'])
@db_required
def update_assignment(id):
    try:
        data = request.json
        update_data = {
            'employee_id': data.get('employee_id'),
            'schedule_id': data.get('schedule_id'),
            'assigned_date': data.get('assigned_date'),
            'is_active': data.get('is_active', True),
            'special_day_assignments': data.get('special_day_assignments', []),
            'notes': data.get('notes', '')
        }
        # NEW: Same auto-set logic if empty
        if not update_data['special_day_assignments']:
            rule = schedule_master_collection.find_one({'_id': update_data['schedule_id']})
            if rule and rule.get('special_days'):
                update_data['special_day_assignments'] = [
                    {**sd, 'is_observed': sd.get('type') == 'Holiday'} for sd in rule['special_days']
                ]
        result = employee_schedule_assign_collection.update_one({'_id': id}, {'$set': update_data})
        if result.matched_count == 0:
            return jsonify({'error': 'Assignment not found'}), 404
        return jsonify({'message': 'Assignment updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to update assignment: {str(e)}"}), 500

@app.route('/api/schedule-assignments/<id>', methods=['DELETE'])
@db_required
def delete_assignment(id):
    try:
        result = employee_schedule_assign_collection.delete_one({'_id': id})
        if result.deleted_count == 0:
            return jsonify({'error': 'Assignment not found'}), 404
        return jsonify({'message': 'Assignment deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to delete assignment: {str(e)}"}), 500


# --- Item-wise Purchase Report API Endpoint ---
@app.route('/api/reports/item-wise-purchase', methods=['GET'])
@db_required
def get_item_wise_purchase_report():
    try:
        # Get query parameters for filtering
        date_from = request.args.get('dateFrom')
        date_to = request.args.get('dateTo')
        item_id = request.args.get('itemId')
        supplier_id = request.args.get('supplierId')
        
        # Fetch all submitted purchase invoices
        invoices = purchase_invoices_collection.find({'status': 'Submitted'})
        
        # Filter by date range if provided
        if date_from or date_to:
            filtered_invoices = []
            for invoice in invoices:
                invoice_date = invoice.get('date')
                if isinstance(invoice_date, str):
                    invoice_date = datetime.fromisoformat(invoice_date.replace('Z', '+00:00'))
                
                include = True
                if date_from:
                    from_date = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
                    if invoice_date < from_date:
                        include = False
                if date_to:
                    to_date = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
                    if invoice_date > to_date:
                        include = False
                
                if include:
                    filtered_invoices.append(invoice)
            invoices = filtered_invoices
        
        # Aggregate data by item
        item_data = {}
        
        for invoice in invoices:
            # Filter by supplier if provided
            if supplier_id and invoice.get('supplierId') != supplier_id:
                continue
            
            for item in invoice.get('items', []):
                item_id_key = item.get('itemId')
                
                # Filter by item if provided
                if item_id and item_id_key != item_id:
                    continue
                
                if item_id_key not in item_data:
                    # Get item details
                    item_doc = purchase_items_collection.find_one({'_id': item_id_key})
                    if not item_doc:
                        continue
                    
                    item_data[item_id_key] = {
                        'itemId': item_id_key,
                        'itemName': item_doc.get('name', 'Unknown'),
                        'brand': item_doc.get('company', ''),
                        'totalQuantity': 0,
                        'totalAmount': 0,
                        'currency': invoice.get('currency', 'AED'),
                        'suppliers': {}
                    }
                
                # Add quantity and amount
                quantity = float(item.get('acceptedQuantity', 0))
                amount = float(item.get('amount', 0))
                
                item_data[item_id_key]['totalQuantity'] += quantity
                item_data[item_id_key]['totalAmount'] += amount
                
                # Track supplier breakdown
                supp_id = invoice.get('supplierId')
                
                # Get supplier name from suppliers collection
                supplier_doc = suppliers_collection.find_one({'_id': supp_id})
                supp_name = supplier_doc.get('company', 'Unknown') if supplier_doc else 'Unknown'
                
                if supp_id not in item_data[item_id_key]['suppliers']:
                    item_data[item_id_key]['suppliers'][supp_id] = {
                        'supplierId': supp_id,
                        'supplierName': supp_name,
                        'quantity': 0,
                        'amount': 0
                    }
                
                item_data[item_id_key]['suppliers'][supp_id]['quantity'] += quantity
                item_data[item_id_key]['suppliers'][supp_id]['amount'] += amount
        
        # Convert to list and format suppliers as array
        result = []
        for item_id_key, data in item_data.items():
            data['suppliers'] = list(data['suppliers'].values())
            result.append(data)
        
        # Sort by item name
        result.sort(key=lambda x: x['itemName'])
        
        return jsonify(convert_objectid_to_str(result)), 200
    except Exception as e:
        logger.error(f"Error generating item-wise purchase report: {str(e)}")
        return jsonify({'error': f"Failed to generate report: {str(e)}"}), 500

# Catch-all for React app
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_react_app(path):
    if path.startswith('api/'):
        return jsonify({"error": "API route not found"}), 404
    if path != "" and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    else:
        if os.path.exists(os.path.join(app.static_folder, 'index.html')):
            return send_from_directory(app.static_folder, 'index.html')
        else:
            return "<h1>Backend is running</h1><p>Frontend not found. Ensure the 'dist' folder contains the built React app.</p>", 404
if __name__ == '__main__':
    connect_to_sqlite()
    if conn and schedule:
        start_scheduler()
    logger.info(f"Serving static files from: {app.static_folder}")
    if getattr(sys, 'frozen', False):
        logger.info("Running as frozen executable, using Waitress")
        import waitress
        waitress.serve(app, host='0.0.0.0', port=8000, threads=8)
    else:
        app.run(host='0.0.0.0', port=8000, debug=True) 